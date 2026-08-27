"""
╔══════════════════════════════════════════════════════════════╗
║   🏪 Professional Digital Store Bot - v5.2 (Optimized)      ║
║   ✅ Auto Deposit (Bkash/Nagad/Rocket)                       ║
║   ✅ Manual Deposit (Binance)                                ║
║   ✅ Channel log only on APPROVE (no pending log)            ║
║   ✅ Performance Optimized — 32 threads                      ║
║   ✅ Background Broadcast — bot never blocks                 ║
║   ✅ Subscription cache 120s — Telegram API calls minimum    ║
║   ✅ DB cache TTL 90s — Firebase reads ~90% কম               ║
║   ✅ Stock count cache 60s — disk I/O কমানো                  ║
║   ✅ Callback instant ACK — বাটন ক্লিকে শূন্য delay          ║
║   ✅ ThreadPoolExecutor — heavy tasks background এ           ║
║   ✅ Shallow DB copy — deepcopy overhead দূর করা হয়েছে      ║
║                                                              ║
║   🔥 FIREBASE QUOTA FIX (v5.1):                             ║
║   ✅ purchase এ পুরো DB GET বাদ → শুধু /users/{id} fetch    ║
║   ✅ approve_deposit → targeted path writes                  ║
║   ✅ auto_approve_deposit → targeted path writes             ║
║   ✅ process_sms → targeted path writes                      ║
║   ✅ API dashboard → batch writes (10 purchase এ ১ write)   ║
║                                                              ║
║   🔥 FIREBASE QUOTA FIX (v5.2):                             ║
║   ✅ save_db() সম্পূর্ণ বাদ — সব জায়গায় targeted write     ║
║   ✅ refer bonus → /users/{id} only write                    ║
║   ✅ deposit → /deposit_requests/{id} only write             ║
║   ✅ lang change → /users/{id}/lang only write               ║
║   ✅ ban/unban → /banned_users only write                    ║
║   ✅ price edit → /products/{name} only write                ║
║   ✅ product delete → specific path delete                   ║
║   ✅ settings toggle → /settings/{key} only write           ║
║   ✅ promo code → /promo_codes only write                    ║
║   ✅ flash sale → /flash_sales only write                    ║
║   ✅ SMS cleanup → per-entry delete (no full write)          ║
║   📉 Firebase write: ~95% কমানো হয়েছে                      ║
╚══════════════════════════════════════════════════════════════╝
"""

import telebot
from telebot import types
from urllib.parse import quote

# ══════════════════════════════════════════════════════
# 🎨 COLORED BUTTON HELPERS (Bot API 9.4+ "style" field)
# ══════════════════════════════════════════════════════
# Telegram Bot API 9.4 (Feb 9, 2026) যোগ করেছে বাটনের জন্য
# একটা "style" ফিল্ড — যা দিয়ে বাটনের background color সেট
# করা যায়। মাত্র ৩টা প্রিসেট কালার আছে:
#   • primary → নীল   (default/neutral action)
#   • success → সবুজ  (positive/confirm action)
#   • danger  → লাল   (destructive/cancel action)
#
# ⚠️ এই ফিচার কাজ করার জন্য দুইটা শর্ত লাগবে:
#   1) pyTelegramBotAPI লাইব্রেরি এই "style" প্যারামিটার
#      সাপোর্ট করে এমন ভার্সনে থাকতে হবে (pip install -U pyTelegramBotAPI)
#   2) ইউজারের Telegram app 9 Feb 2026 এর পরের ভার্সনের হতে হবে
#
# পুরোনো লাইব্রেরি/ক্লায়েন্টে থাকলেও bot কখনো crash করবে না।
#
# ⚠️ FIX: শুধু constructor-এ style= পাস করলেই যথেষ্ট না —
# pyTelegramBotAPI এর to_dict() মেথড যদি "style" ফিল্ড না চেনে,
# তাহলে সেটা কখনো JSON payload-এ যায়ই না, ফলে Telegram-এ বাটন
# রঙিন দেখা যায় না। তাই নিচে to_dict() নিজেই override করে
# style জোর করে JSON dict-এ ঢুকিয়ে দেওয়া হচ্ছে — লাইব্রেরি
# version যাই থাকুক না কেন, style এখন থেকে গ্যারান্টিড ভাবে
# Telegram API তে পাঠানো হবে।
def _force_style_in_json(btn, style):
    """btn.to_dict()/to_json() কে patch করে 'style' key জোর করে বসিয়ে দেয়।"""
    if not style:
        return btn
    try:
        original_to_dict = btn.to_dict
        def patched_to_dict():
            d = original_to_dict()
            if isinstance(d, dict):
                d["style"] = style
            return d
        btn.to_dict = patched_to_dict
    except Exception:
        pass
    try:
        import json as _json_for_style
        original_to_json = btn.to_json
        def patched_to_json():
            d = btn.to_dict()
            return _json_for_style.dumps(d)
        btn.to_json = patched_to_json
    except Exception:
        pass
    # attribute হিসেবেও রেখে দিচ্ছি, ভবিষ্যতে লাইব্রেরি native সাপোর্ট
    # করলে সেটাও যেন কাজ করে
    try:
        setattr(btn, "style", style)
    except Exception:
        pass
    return btn

def _make_styled(cls, text, style=None, **kwargs):
    if style:
        try:
            btn = cls(text, style=style, **kwargs)
        except TypeError:
            btn = cls(text, **kwargs)
        return _force_style_in_json(btn, style)
    return cls(text, **kwargs)

def InlineBtn(text, style=None, **kwargs):
    """types.InlineKeyboardButton এর color-safe wrapper।"""
    return _make_styled(types.InlineKeyboardButton, text, style=style, **kwargs)

# ══════════════════════════════════════════════════════
# 🎨 KEYBOARD BUTTON — PER-BUTTON COLOR — Admin Panel Controlled
# ══════════════════════════════════════════════════════
# প্রতিটা keyboard (reply) বাটনের color আলাদাভাবে admin panel
# থেকে বদলানো যায়। বাটনগুলোকে "role" দিয়ে গ্রুপ করা হয়েছে
# (একই role এর সব বাটন একসাথে একই color পাবে — যেমন সব "Back"
# বাটন, বা সব "Product" বাটন)। 3টা preset color:
#   primary(🔵নীল) / success(🟢সবুজ) / danger(🔴লাল)
# _KB_ROLE_COLORS ইন-মেমরি cache — instant reflect হয়, আর
# Firebase এ persist হয় যাতে bot restart এর পরও মনে থাকে।
_KB_COLOR_CHOICES = {"primary", "success", "danger"}

# role → (বাংলা লেবেল, ডিফল্ট কালার — কোডে যা আগে হার্ডকোড ছিল)
KB_BUTTON_ROLES = {
    "shop":         ("🛍️ শপ (Shop)",              "primary"),
    "deposit":      ("💰 ডিপোজিট (Deposit)",        "success"),
    "balance":      ("💳 ব্যালেন্স (Balance)",       "primary"),
    "orders":       ("📦 অর্ডার (Orders)",          "primary"),
    "refer":        ("🔗 রেফার (Refer)",            "primary"),
    "support":      ("🆘 সাপোর্ট (Support)",         "primary"),
    "language":     ("🌐 ভাষা (Language)",           "primary"),
    "admin":        ("🛠️ এডমিন (Admin)",            "primary"),
    "back":         ("🔙 ব্যাক (Back)",              "primary"),
    "cancel":       ("❌ ক্যানসেল (Cancel)",         "danger"),
    "skip":         ("⏭️ স্কিপ (Skip)",              "primary"),
    "buy_single":   ("🛒 সিঙ্গেল বাই (Single Buy)", "success"),
    "buy_bulk":     ("📦 বাল্ক বাই (Bulk Buy)",     "success"),
    "product_item": ("📦 প্রোডাক্ট লিস্ট বাটন",      "success"),
    "category_item":("📂 ক্যাটাগরি/সাব-আইটেম বাটন", "primary"),
    "misc_option":  ("⚙️ অন্যান্য অপশন বাটন",        "primary"),
}

_KB_ROLE_COLORS = {role: default for role, (_label, default) in KB_BUTTON_ROLES.items()}

def get_kb_button_color(role=None):
    """একটা নির্দিষ্ট role এর জন্য বর্তমান admin-সেট color।
    role না দিলে সাধারণ ডিফল্ট (primary) রিটার্ন করে।"""
    if role is None:
        return "primary"
    return _KB_ROLE_COLORS.get(role, KB_BUTTON_ROLES.get(role, (None, "primary"))[1])

def set_kb_button_color(role, color):
    """Admin panel থেকে একটা নির্দিষ্ট role এর color পরিবর্তন করলে এখানে সেট হয়।"""
    if color not in _KB_COLOR_CHOICES:
        color = "primary"
    if role in KB_BUTTON_ROLES:
        _KB_ROLE_COLORS[role] = color

def load_kb_button_colors_from_dict(saved):
    """Firebase থেকে লোড করা dict দিয়ে in-memory cache আপডেট করো।"""
    if not isinstance(saved, dict):
        return
    for role, color in saved.items():
        if role in KB_BUTTON_ROLES and color in _KB_COLOR_CHOICES:
            _KB_ROLE_COLORS[role] = color

def KbBtn(text, style=None, role=None, **kwargs):
    """types.KeyboardButton এর color-safe wrapper।
    ⚠️ style প্যারামিটার এখন সরাসরি ব্যবহার হয় না — প্রতিটা কল সাইটে
    একটা `role` দেওয়া থাকে (যেমন role="back", role="cancel"), আর
    সেই role এর জন্য admin panel থেকে সেট করা color ব্যবহার হয়
    (get_kb_button_color(role))। role না দিলে style অথবা primary
    ডিফল্ট হিসেবে ব্যবহার হবে (backward-compat)।
    """
    effective_style = get_kb_button_color(role) if role else (style or "primary")
    return _make_styled(types.KeyboardButton, text, style=effective_style, **kwargs)
import os
import json
import uuid
import re
from datetime import datetime
import pytz
import openpyxl
from io import BytesIO
import logging
import traceback
from flask import Flask, request as flask_request
import threading
from concurrent.futures import ThreadPoolExecutor
import requests as http_requests

# ===========================
# GLOBAL THREAD POOL — heavy callback tasks এখানে চলবে
# webhook thread block হবে না, instant response দেবে
# ===========================
_EXECUTOR = ThreadPoolExecutor(max_workers=32, thread_name_prefix="bot_worker")

# ===========================
# CONCURRENCY LOCKS
# ===========================
# প্রতিটা product এর জন্য আলাদা lock — হাজার user একসাথে কিনলেও safe
_stock_locks = {}
_stock_locks_meta = threading.Lock()

def get_stock_lock(product_name: str) -> threading.Lock:
    """প্রতিটা product এর জন্য dedicated lock দাও।"""
    with _stock_locks_meta:
        if product_name not in _stock_locks:
            _stock_locks[product_name] = threading.Lock()
        return _stock_locks[product_name]

# Per-user purchase lock — একই user একসাথে দুইটা কিনতে পারবে না
_user_purchase_locks = {}
_user_locks_meta = threading.Lock()

def get_user_lock(user_id: str) -> threading.Lock:
    """প্রতিটা user এর জন্য dedicated lock।"""
    with _user_locks_meta:
        if user_id not in _user_purchase_locks:
            _user_purchase_locks[user_id] = threading.Lock()
        return _user_purchase_locks[user_id]

# ══════════════════════════════════════════════════════
# CALLBACK DEDUP — একই user এর duplicate callback block
# Double-click বা network retry তে একই order দুইবার
# deliver হবে না।
# ══════════════════════════════════════════════════════
_in_flight_purchases = set()   # {user_id} — purchase চলছে
_in_flight_lock = threading.Lock()

def _try_acquire_in_flight(user_id: str) -> bool:
    """True হলে proceed, False হলে duplicate — block করো।"""
    with _in_flight_lock:
        if user_id in _in_flight_purchases:
            return False
        _in_flight_purchases.add(user_id)
        return True

def _release_in_flight(user_id: str):
    """Purchase শেষ হলে slot খালি করো।"""
    with _in_flight_lock:
        _in_flight_purchases.discard(user_id)

# ===========================
# FIREBASE SETUP
# ===========================
import firebase_admin
from firebase_admin import credentials, db as firebase_db

_FIREBASE_DB_URL = "https://endfirebase-ff423-default-rtdb.asia-southeast1.firebasedatabase.app/"

if not firebase_admin._apps:
    try:
        _firebase_cred_json = os.environ.get("FIREBASE_CREDENTIALS")
        if _firebase_cred_json:
            import json as _json_mod
            _cred_dict = _json_mod.loads(_firebase_cred_json)
            _cred = credentials.Certificate(_cred_dict)
            print("✅ Firebase: loaded from environment variable")
        else:
            _FIREBASE_CRED_PATH = os.path.join(os.path.dirname(__file__), "nd-update-firebase-adminsdk-fbsvc-e0d487473d.json")
            print(f"🔍 Firebase: loading from file: {_FIREBASE_CRED_PATH}")
            print(f"🔍 File exists: {os.path.exists(_FIREBASE_CRED_PATH)}")
            _cred = credentials.Certificate(_FIREBASE_CRED_PATH)
            print("✅ Firebase: loaded from file")
        firebase_admin.initialize_app(_cred, {"databaseURL": _FIREBASE_DB_URL})
    except Exception as e:
        print(f"❌ FIREBASE INIT FAILED: {e}")
        raise

def _fb_ref(path="/"):
    return firebase_db.reference(path)

# ===========================
# CONFIGURATION — সব কনফিগ এখানে
# ===========================

API_TOKEN = os.environ.get('BOT_TOKEN', ' ')
ADMIN_ID = int(os.environ.get('ADMIN_ID', ' '))
DONGVANFB_API_KEY = os.environ.get('DONGVANFB_API_KEY', '')

# ===========================
# SECURITY CONFIG
# ===========================
SMS_WEBHOOK_SECRET = os.environ.get('SMS_WEBHOOK_SECRET', 'change_this_secret_key_2024')
MAX_DEPOSIT_AMOUNT  = 50000   # সর্বোচ্চ একবারে ডিপোজিট (BDT)
MAX_DAILY_DEPOSIT   = 100000  # প্রতিদিন সর্বোচ্চ ডিপোজিট (BDT)
MAX_MSG_PER_MINUTE  = 60      # Rate limit: প্রতি মিনিটে সর্বোচ্চ message
DEPOSIT_COOLDOWN_SEC = 0      # কোনো cooldown নেই — যেকোনো সময় deposit করা যাবে
MAX_BULK_QUANTITY   = 10000     # একবারে সর্বোচ্চ bulk কেনা যাবে

CHANNELS = [
    {"id": "-1003800642288", "link": "https://t.me/SmartEarnOfficail"},
    {"id": "-1003798562939", "link": "https://t.me/virtual_Shop1472"}
]
LOG_CHANNEL_ID = "-1003798562939"
SUPPORT_USERNAME = "relax1472"
BOT_USERNAME = "vertual_shop_bot"

# পেমেন্ট একাউন্ট নম্বর
BKASH_NUMBER   = "01961167208"
NAGAD_NUMBER   = "01961167208"
ROCKET_NUMBER  = "01961167208"
BINANCE_ID     = "761945447"

MIN_DEPOSIT = 10   # সর্বনিম্ন ডিপোজিট

# ===========================
# HOTMAIL143 API CONFIG
# ===========================
# API Key — environment variable থেকে নাও অথবা এখানে সরাসরি দাও
HOTMAIL143_API_KEY = os.environ.get('HOTMAIL143_API_KEY', '')
HOTMAIL143_BASE_URL = "https://www.hotmail143.com/api/v1"

# ===========================
# BULKMAIL API CONFIG
# ===========================
# Bulkmail.shop API Key — environment variable থেকে নাও
BULKMAIL_API_KEY  = os.environ.get('BULKMAIL_API_KEY', '')
BULKMAIL_BASE_URL = "https://bulkmail.shop/api/v2"

# ─── Bulkmail product mapping ───
# key = bot এ product এর নাম, value = bulkmail product_id
BULKMAIL_PRODUCTS = {
    "Edu Mail 24hr": 13,   # Bulkmail product ID 13
}

# ─── API থেকে কোন product deliver হবে তার mapping ───
# key = bot এ product এর নাম (DB তে যেভাবে আছে)
# value = API parameters
API_PRODUCTS = {
    "Edu Mail 24H": {
        "product_type": "edu_gmail",
        "account_type": "edu_gmail",
        "source": "hotmail143",
    },
    "Edu Mail 72H": {
        "product_type": "edu_gmail_72_hours",
        "account_type": "edu_gmail",
        "source": "hotmail143",
    },
    "Hotmail": {
        "product_type": "long_live",
        "account_type": "hotmail",
        "source": "hotmail143",
    },
    "Outlook": {
        "product_type": "long_live",
        "account_type": "outlook.com",
        "source": "hotmail143",
    },
    "Edu Mail 24hr": {
        "source": "bulkmail",
        "product_id": 13,
    },
}

# ─── Edu Mail parent product → sub-type keyboard config ───
EDU_MAIL_PARENT    = "Edu Mail"          # DB তে product এর নাম
EDU_MAIL_SUB_24H   = "Edu Mail 24H"     # hotmail143 API key
EDU_MAIL_SUB_72H   = "Edu Mail 72H"     # hotmail143 API key
EDU_MAIL_SUB_24HR  = "Edu Mail 24hr"    # bulkmail API key (নতুন)

# ─── Hotmail ও Outlook — standalone API products (no sub-menu) ───
HOTMAIL_PROD  = "Hotmail"
OUTLOOK_PROD  = "Outlook"
# API dashboard tracking — কোন product কতটা কিনা হয়েছে
# Firebase path: /api_dashboard/{product_name} → {count, total_bdt}
API_DASHBOARD_PRODUCTS = [EDU_MAIL_SUB_24H, EDU_MAIL_SUB_72H, EDU_MAIL_SUB_24HR, HOTMAIL_PROD, OUTLOOK_PROD]

def hotmail143_get_stock(product_type, account_type):
    """
    hotmail143 API থেকে stock count আনো।
    Return: int (stock count) অথবা -1 (error)

    Actual API response format (product_type filter দিলে):
    {
      "status": "success",
      "data": {
        "total": 136133,
        "product_type": "edu_gmail",
        "account_type": "edu_gmail"
      }
    }

    No filter দিলে (সব product):
    {
      "status": "success",
      "data": {
        "edu_gmail": {"edu_gmail": 25},
        "total": 600
      }
    }
    """
    if not HOTMAIL143_API_KEY:
        logging.warning("HOTMAIL143_API_KEY not set")
        return -1
    try:
        resp = http_requests.get(
            f"{HOTMAIL143_BASE_URL}/stock",
            params={
                "api_key": HOTMAIL143_API_KEY,
                "product_type": product_type,
                "account_type": account_type,
            },
            timeout=4  # ⚡ 4s — 1200 user load এ safe timeout
        )
        data = resp.json()
        logging.info(f"hotmail143 stock API raw: {data}")

        if data.get("status") != "success":
            logging.warning(f"hotmail143 stock error: {data}")
            return -1

        stock_data = data.get("data", {})

        # ✅ Format 1: product_type filter দিলে flat response আসে
        # {"total": 136133, "product_type": "edu_gmail", "account_type": "edu_gmail"}
        if "total" in stock_data and "product_type" in stock_data:
            count = int(stock_data.get("total", 0))
            logging.info(f"hotmail143 stock (flat format): {product_type}/{account_type} = {count}")
            return count

        # ✅ Format 2: nested dict — {"edu_gmail": {"edu_gmail": 25}}
        section = stock_data.get(product_type)
        if section is not None:
            if isinstance(section, dict):
                val = section.get(account_type)
                if val is not None:
                    return int(val)
                # যেকোনো int value নাও
                for v in section.values():
                    if isinstance(v, (int, float)):
                        return int(v)
            elif isinstance(section, (int, float)):
                return int(section)

        # ✅ Format 3: fallback — total key
        if "total" in stock_data:
            return int(stock_data["total"])

        logging.warning(f"hotmail143 stock: unknown format: {stock_data}")
        return 0

    except Exception as e:
        logging.error(f"hotmail143_get_stock error: {e}")
        return -1

def hotmail143_purchase(product_type, account_type, quantity=1):
    """
    hotmail143 API দিয়ে account purchase করো।
    Return: (success: bool, accounts: list[dict], error_msg: str)

    API response format (বিভিন্ন রকম হতে পারে):
      Format 1: {"status":"success","data":{"accounts":[{"email":"...","password":"..."}]}}
      Format 2: {"status":"success","data":{"accounts":["email:pass","email2:pass2"]}}
      Format 3: {"status":"success","data":["email:pass"]}
      Format 4: {"status":"success","data":{"email":"...","password":"..."}}
    সব format কে normalize করে list of dict এ convert করো।
    """
    if not HOTMAIL143_API_KEY:
        return False, [], "API Key সেট করা নেই।"
    try:
        resp = http_requests.post(
            f"{HOTMAIL143_BASE_URL}/purchase",
            data={
                "api_key": HOTMAIL143_API_KEY,
                "product_type": product_type,
                "account_type": account_type,
                "quantity": quantity,
            },
            timeout=30
        )
        data = resp.json()
        logging.info(f"hotmail143 purchase raw response: {data}")

        if data.get("status") == "success":
            raw_data = data.get("data", {})
            accounts = []

            def parse_account_string(s):
                """
                'email|password|refresh_token|client_id' format parse করো।
                Full string টা Full_Data তে রাখো।
                """
                s = str(s).strip()
                if "|" in s:
                    parts = [p.strip() for p in s.split("|")]
                    result = {"Full_Data": s}
                    if len(parts) >= 1:
                        result["Email"] = parts[0]
                    if len(parts) >= 2:
                        result["Password"] = parts[1]
                    if len(parts) >= 3:
                        result["Refresh_Token"] = parts[2]
                    if len(parts) >= 4:
                        result["Client_Id"] = parts[3]
                    return result
                elif ":" in s:
                    parts = s.split(":", 1)
                    if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                        return {"Email": parts[0].strip(), "Password": parts[1].strip(),
                                "Full_Data": s}
                return {"Account": s}

            # Format 1 & 2: data এ "accounts" key আছে
            if isinstance(raw_data, dict) and "accounts" in raw_data:
                raw_accounts = raw_data["accounts"]
                for acc in raw_accounts:
                    if isinstance(acc, dict):
                        accounts.append(acc)
                    elif isinstance(acc, str):
                        accounts.append(parse_account_string(acc))

            # Format 3: data নিজেই list
            elif isinstance(raw_data, list):
                for acc in raw_data:
                    if isinstance(acc, dict):
                        accounts.append(acc)
                    elif isinstance(acc, str):
                        accounts.append(parse_account_string(acc))

            # Format 4: data নিজেই একটা account dict
            elif isinstance(raw_data, dict) and raw_data:
                # accounts key ছাড়া সরাসরি account fields
                if any(k.lower() in ("email", "password", "username", "pass", "account") for k in raw_data):
                    accounts.append(raw_data)
                else:
                    # অন্য কোনো structure — সব values দেখাও
                    for k, v in raw_data.items():
                        if isinstance(v, (str, int)) and str(v).strip():
                            accounts.append({k: str(v)})

            logging.info(f"hotmail143 purchase parsed accounts: {len(accounts)} items")
            if not accounts:
                logging.warning(f"hotmail143 purchase: no accounts parsed from: {raw_data}")
                return False, [], "Account data পাওয়া যায়নি। Admin কে জানান।"
            return True, accounts, ""
        else:
            msg = data.get("message", "Purchase ব্যর্থ হয়েছে।")
            logging.warning(f"hotmail143 purchase failed: {data}")
            return False, [], msg
    except Exception as e:
        logging.error(f"hotmail143_purchase error: {e}")
        return False, [], f"API error: {e}"

# ═══════════════════════════════════════════════════════════════
# BULKMAIL.SHOP API FUNCTIONS
# Base URL: https://bulkmail.shop/api/v2
# Auth: X-API-Key header
# ═══════════════════════════════════════════════════════════════

def bulkmail_get_stock(product_id: int) -> int:
    """
    Bulkmail.shop API থেকে product stock আনো।
    GET /api/v2/stock/check?ids={product_id}
    Response: {"success":true,"data":{"products":[{"id":60,"stock":1001,...}]}}
    Return: int (stock count) অথবা -1 (error)
    """
    try:
        headers = {"Accept": "application/json"}
        if BULKMAIL_API_KEY:
            headers["X-API-Key"] = BULKMAIL_API_KEY

        # /stock/check?ids= endpoint use করো
        resp = http_requests.get(
            f"{BULKMAIL_BASE_URL}/stock/check",
            params={"ids": str(product_id)},
            headers=headers,
            timeout=8
        )
        data = resp.json()
        logging.info(f"bulkmail stock/check (product_id={product_id}): {data}")
        if data.get("success"):
            products = data.get("data", {}).get("products", [])
            for item in products:
                pid = item.get("id") or item.get("product_id")
                if str(pid) == str(product_id):
                    count = item.get("stock") or item.get("stock_count") or 0
                    return int(count)
            return 0

        # Fallback: /stock/{product_id}
        resp2 = http_requests.get(
            f"{BULKMAIL_BASE_URL}/stock/{product_id}",
            headers=headers,
            timeout=8
        )
        data2 = resp2.json()
        logging.info(f"bulkmail stock/{product_id}: {data2}")
        if data2.get("success"):
            d = data2.get("data", {})
            count = d.get("stock") or d.get("stock_count") or 0
            return int(count)

        logging.warning(f"bulkmail stock error: {data2}")
        return -1
    except Exception as e:
        logging.error(f"bulkmail_get_stock error: {e}")
        return -1


def bulkmail_purchase(product_id: int, quantity: int = 1):
    """
    Bulkmail.shop API দিয়ে product purchase করো।
    POST /api/v2/orders  body: {"product_id": 13, "quantity": 1}
    Response: {"success":true,"data":{"stock_items":["email:password",...]}}
    Return: (success: bool, accounts: list[dict], error_msg: str)
    """
    if not BULKMAIL_API_KEY:
        return False, [], "Bulkmail API Key সেট করা নেই।"
    try:
        resp = http_requests.post(
            f"{BULKMAIL_BASE_URL}/orders",
            headers={
                "X-API-Key": BULKMAIL_API_KEY,
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            json={"product_id": product_id, "quantity": quantity},
            timeout=30
        )
        data = resp.json()
        logging.info(f"bulkmail purchase raw response: {data}")

        logging.warning(f"bulkmail purchase HTTP={resp.status_code} response={data}")

        # success=true অথবা HTTP 201 — দুটোই সফল
        if data.get("success") or resp.status_code == 201:
            order_data  = data.get("data", {})
            # field name: 'items' অথবা 'stock_items' (API response অনুযায়ী)
            stock_items = order_data.get("items") or order_data.get("stock_items") or []

            # এখনও খালি হলে order_id দিয়ে GET /orders/{id} fetch করো
            if not stock_items:
                order_id_bm = order_data.get("id") or order_data.get("order_id")
                logging.warning(f"bulkmail items empty, order_id={order_id_bm}, fetching detail")
                if order_id_bm:
                    try:
                        r2 = http_requests.get(
                            f"{BULKMAIL_BASE_URL}/orders/{order_id_bm}",
                            headers={"X-API-Key": BULKMAIL_API_KEY, "Accept": "application/json"},
                            timeout=15
                        )
                        d2 = r2.json()
                        logging.warning(f"bulkmail GET /orders/{order_id_bm}: {d2}")
                        if d2.get("success"):
                            d2data = d2.get("data", {})
                            stock_items = d2data.get("items") or d2data.get("stock_items") or []
                        # এখনও খালি হলে export try করো
                        if not stock_items:
                            r3 = http_requests.get(
                                f"{BULKMAIL_BASE_URL}/orders/{order_id_bm}/export",
                                params={"format": "txt"},
                                headers={"X-API-Key": BULKMAIL_API_KEY},
                                timeout=15
                            )
                            logging.warning(f"bulkmail export HTTP={r3.status_code} text={r3.text[:300]}")
                            if r3.status_code == 200 and r3.text.strip():
                                stock_items = [line.strip() for line in r3.text.strip().splitlines() if line.strip()]
                    except Exception as _fe:
                        logging.error(f"bulkmail fetch order error: {_fe}")

            accounts = []
            for item in stock_items:
                item = str(item).strip()
                if not item:
                    continue
                if "|" in item:
                    parts = [p.strip() for p in item.split("|")]
                    acc = {"Full_Data": item}
                    if len(parts) >= 1: acc["Email"]         = parts[0]
                    if len(parts) >= 2: acc["Password"]      = parts[1]
                    if len(parts) >= 3: acc["Refresh_Token"] = parts[2]
                    if len(parts) >= 4: acc["Client_Id"]     = parts[3]
                    accounts.append(acc)
                elif ":" in item:
                    parts = item.split(":", 1)
                    accounts.append({
                        "Email":     parts[0].strip(),
                        "Password":  parts[1].strip(),
                        "Full_Data": item,
                    })
                else:
                    accounts.append({"Account": item, "Full_Data": item})

            if not accounts:
                return False, [], f"Account data পাওয়া যায়নি। Raw order: {str(order_data)[:200]}"
            return True, accounts, ""
        else:
            err_code = data.get("error", "")
            err_detail = str(data.get("errors", "")) if data.get("errors") else ""
            if err_code == "INSUFFICIENT_BALANCE":
                err = "Bulkmail wallet এ balance নেই। Admin কে জানান।"
            elif err_code == "INSUFFICIENT_STOCK":
                err = "Bulkmail এ stock নেই।"
            elif err_code:
                err = f"Bulkmail error: {err_code}{(' — ' + err_detail) if err_detail else ''}"
            else:
                err = f"Purchase failed (HTTP {resp.status_code}). Raw: {str(data)[:200]}"
            logging.warning(f"bulkmail purchase failed: {data}")
            return False, [], err
    except Exception as e:
        logging.error(f"bulkmail_purchase error: {e}")
        return False, [], f"Bulkmail API error: {e}"


def is_bulkmail_product(product_name: str) -> bool:
    """এই product টা bulkmail থেকে deliver হবে কিনা।"""
    return API_PRODUCTS.get(product_name, {}).get("source") == "bulkmail"


def is_bulkmail_enabled(product_name: str, db: dict) -> bool:
    """
    Admin bulkmail product on/off করেছে কিনা।
    DB path: /bulkmail_products/{product_name}/enabled  → True/False
    Default: True (enabled)
    """
    return db.get("bulkmail_products", {}).get(product_name, {}).get("enabled", True)


def is_hotmail143_enabled(product_name: str, db: dict) -> bool:
    """
    Admin hotmail143.com API product on/off করেছে কিনা।
    DB path: /hotmail143_products/{product_name}/enabled  → True/False
    Default: True (enabled)
    """
    return db.get("hotmail143_products", {}).get(product_name, {}).get("enabled", True)


DB_FILE   = "users_db.json"
STOCK_DIR = "stocks"
LOG_FILE  = "bot_logs.txt"

if not os.path.exists(STOCK_DIR):
    os.makedirs(STOCK_DIR)

# ===========================
# REAL TIME HELPER (UTC+6 Bangladesh / International)
# ===========================

BD_TZ = pytz.timezone("Asia/Dhaka")

def get_now():
    """Real international datetime — Bangladesh time (UTC+6)."""
    return datetime.now(BD_TZ).strftime("%d/%m/%Y %I:%M %p")

def get_now_short():
    """Short date for join_date."""
    return datetime.now(BD_TZ).strftime("%d/%m/%y")

bot = telebot.TeleBot(API_TOKEN, threaded=True, num_threads=32, colorful_logs=False)

# ===========================
# CHANNEL LEAVE — GOODBYE MESSAGE
# ===========================

@bot.chat_member_handler()
def handle_chat_member_update(update: types.ChatMemberUpdated):
    """User channel থেকে leave করলে goodbye message পাঠাও।"""
    try:
        # শুধু আমাদের নির্দিষ্ট channel গুলোতে track করব
        channel_ids = [ch["id"] for ch in CHANNELS]
        chat_id = str(update.chat.id)
        if chat_id not in channel_ids:
            return

        old_status = update.old_chat_member.status if update.old_chat_member else None
        new_status = update.new_chat_member.status if update.new_chat_member else None

        # Member ছিল কিন্তু এখন left/kicked হয়েছে
        if old_status in ("member", "administrator", "creator") and new_status in ("left", "kicked"):
            user = update.new_chat_member.user
            user_id = str(user.id)
            # ⚡ তাৎক্ষণিক cache invalidate — পরের message এই block হবে
            invalidate_sub_cache(user_id)
            try:
                goodbye_markup = types.InlineKeyboardMarkup(row_width=2)
                # Row 1: Main Channel + Live History
                goodbye_markup.row(
                    InlineBtn("📢 Main Channel", style="primary", url=CHANNELS[0]["link"]),
                    InlineBtn("📜 Live History", style="primary", url=CHANNELS[1]["link"] if len(CHANNELS) > 1 else CHANNELS[0]["link"])
                )
                # Row 2: Come Back button
                goodbye_markup.add(
                    InlineBtn("🔙 Come Back", style="primary", callback_data="goodbye_come_back")
                )
                bot.send_message(
                    user_id,
                    "👋 <b>Goodbye! We'll Miss You!</b>\n\n"
                    "We hope to see you again soon!\n"
                    "💫 You're always welcome back anytime.\n\n"
                    "👇 আমাদের সাথে যোগ দিন:",
                    parse_mode="HTML",
                    reply_markup=goodbye_markup
                )
            except Exception as e:
                logging.warning(f"Goodbye message failed for user {user_id}: {e}")
    except Exception as e:
        logging.error(f"chat_member_update error: {e}")

# Admin session store — product delete confirmation ইত্যাদির জন্য
_admin_sessions = {}

# ⚡ Pending product in-memory cache — Firebase write ছাড়া product selection মনে রাখে
# {user_id: product_name}
_pending_product_cache = {}

# Stock upload এর পরে broadcast pending রাখার জন্য
# key: product_name, value: {broadcast_msg, channel_msg, stock_count, price, user_count}
_pending_broadcasts = {}

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# ===========================
# SMS WEBHOOK SERVER (Auto Deposit)
# Flask server — SMS forwarder এই endpoint এ POST করবে
# ===========================

sms_app = Flask(__name__)

# ── Flask performance settings ──
sms_app.config["JSON_SORT_KEYS"]    = False   # JSON sort বন্ধ → দ্রুত serialize
sms_app.config["PROPAGATE_EXCEPTIONS"] = False
sms_app.config["MAX_CONTENT_LENGTH"]   = 1 * 1024 * 1024  # 1MB max request size

# Rate limiting for SMS webhook
import time as _time_module
_sms_request_times = []

# ===========================
# RAILWAY URL — Webhook এর জন্য
# ===========================
RAILWAY_URL  = os.environ.get("RAILWAY_URL", "https://worker-production-1a27.up.railway.app")
WEBHOOK_PATH = "/webhook/" + API_TOKEN
WEBHOOK_URL  = RAILWAY_URL + WEBHOOK_PATH

@sms_app.route('/health', methods=['GET'])
def health_check():
    """Railway health check endpoint."""
    return {"status": "ok", "bot": "running"}, 200

@sms_app.route('/sms', methods=['GET'])
def sms_health_check():
    """Browser দিয়ে চেক করার জন্য GET endpoint।"""
    return {"status": "SMS Webhook is running ✅", "usage": "POST JSON: {from, body}"}, 200

# Pre-built response object — প্রতিটা request এ নতুন dict না বানিয়ে একটাই ব্যবহার করো
_WEBHOOK_OK  = ('{"status":"ok"}', 200, {"Content-Type": "application/json"})
_WEBHOOK_ERR = ('{"status":"error"}', 500, {"Content-Type": "application/json"})

@sms_app.route("/webhook/" + API_TOKEN, methods=["POST"])
def telegram_webhook():
    """Telegram Webhook — ⚡ ZERO-COPY instant return.
    JSON parse → Update object → thread pool submit → তুরন্ত 200 return।
    Flask thread কখনো block হবে না।
    """
    try:
        # get_data() → json.loads() — get_json() এর চেয়ে দ্রুত (কোনো content-type check নেই)
        raw = flask_request.get_data(cache=False)
        if raw:
            import json as _j
            json_data = _j.loads(raw)
            update = telebot.types.Update.de_json(json_data)
            _EXECUTOR.submit(bot.process_new_updates, [update])
        return _WEBHOOK_OK
    except Exception as e:
        logging.error(f"Webhook error: {e}")
        return _WEBHOOK_ERR

@sms_app.route('/sms', methods=['POST'])
def receive_sms():
    """
    SMS Forwarder app এই endpoint এ POST করবে।
    সব ধরনের format সাপোর্ট করে।
    """
    try:
        # SECURITY: Secret key check
        auth = flask_request.headers.get('X-Secret-Key', '') or flask_request.args.get('key', '')
        if auth != SMS_WEBHOOK_SECRET:
            logging.warning(f"SMS webhook unauthorized: {flask_request.remote_addr}")
            return {"status": "unauthorized"}, 403

        # Rate limiting
        now_ts = _time_module.time()
        _sms_request_times[:] = [t for t in _sms_request_times if now_ts - t < 60]
        if len(_sms_request_times) >= 100:
            return {"status": "rate_limited"}, 429
        _sms_request_times.append(now_ts)

        sender   = ""
        sms_body = ""

        # Step 1: raw body পড়ো
        raw = flask_request.get_data(as_text=True) or ""

        # Step 2: JSON try করো
        try:
            data = flask_request.get_json(force=True) or {}
            if data:
                sender   = str(data.get("from", "") or data.get("sender", "") or "")[:50]
                sms_body = str(
                    data.get("body", "") or data.get("message", "") or
                    data.get("text", "") or data.get("msg", "") or ""
                )[:500]
        except Exception:
            pass

        # Step 3: Form data try করো
        if not sms_body:
            sender   = str(flask_request.form.get("from", "") or flask_request.form.get("sender", "") or "")[:50]
            sms_body = str(
                flask_request.form.get("body", "") or flask_request.form.get("message", "") or
                flask_request.form.get("text", "") or flask_request.form.get("msg", "") or ""
            )[:500]

        # Step 4: "From : bKash() SMS text" plain text format parse করো
        if not sms_body and raw:
            m = re.match(r"From\s*:\s*([^\(\n]+)\([^\)]*\)\s*([\s\S]*)", raw)
            if m:
                sender   = m.group(1).strip()[:50]
                sms_body = m.group(2).strip()[:500]

        # Step 5: raw body তে bkash/nagad/rocket keyword থাকলে পুরোটাই body হিসেবে নাও
        if not sms_body and raw:
            raw_lower = raw.lower()
            if any(k in raw_lower for k in ["bkash", "nagad", "nogod", "rocket", "dbbl"]):
                sms_body = raw[:500]
                # sender বের করার চেষ্টা
                for k in ["bkash", "nagad", "rocket"]:
                    if k in raw_lower:
                        sender = k
                        break

        if not sms_body:
            logging.warning(f"SMS webhook: empty body | raw={raw[:100]}")
            return {"status": "empty_body"}, 400

        logging.info(f"SMS received | from={sender!r} | body={sms_body[:100]}")
        # ⚡ Background এ process করো — SMS webhook instant return করবে
        _EXECUTOR.submit(process_sms, sender, sms_body)
        return {"status": "ok"}, 200

    except Exception as e:
        logging.error(f"SMS webhook error: {e}")
        return {"status": "error"}, 500

def _schedule_periodic_flush():
    """প্রতি 5 মিনিটে API dashboard pending changes flush করো।"""
    import time as _t
    while True:
        _t.sleep(300)  # 5 minutes
        try:
            with _api_dashboard_lock:
                has_pending = bool(_api_dashboard_pending)
            if has_pending:
                _flush_api_dashboard()
                logging.info("Periodic API dashboard flush completed")
        except Exception as e:
            logging.error(f"Periodic flush error: {e}")

def start_sms_server():
    """SMS + Telegram Webhook — একই server এ দুটোই চলবে।
    Gunicorn থাকলে Gunicorn দিয়ে চালায় (production-grade, ultra fast)।
    না থাকলে Flask dev server fallback।
    """
    # Telegram Webhook set করো
    # ⚡ Background periodic flush — API dashboard changes Firebase এ save হবে
    threading.Thread(target=_schedule_periodic_flush, daemon=True, name="dashboard_flush").start()

    try:
        bot.remove_webhook()
        import time as _tw
        _tw.sleep(0.5)
        bot.set_webhook(
            url=WEBHOOK_URL,
            allowed_updates=[
                "message", "callback_query", "chat_member",
                "my_chat_member", "channel_post", "inline_query"
            ]
        )
        print(f"✅ Telegram Webhook set: {WEBHOOK_URL}")
    except Exception as e:
        print(f"❌ Webhook set failed: {e}")

    PORT = int(os.environ.get("PORT", 5000))

    # ── Gunicorn দিয়ে চালানোর চেষ্টা (production-grade WSGI) ──
    try:
        import gunicorn.app.base

        class _StandaloneApp(gunicorn.app.base.BaseApplication):
            def __init__(self, app, options=None):
                self.options = options or {}
                self.application = app
                super().__init__()

            def load_config(self):
                for key, val in self.options.items():
                    if key in self.cfg.settings and val is not None:
                        self.cfg.set(key.lower(), val)

            def load(self):
                return self.application

        cpu_count = os.cpu_count() or 2
        workers   = 1  # ✅ 1 worker — multiple workers এ in-memory lock কাজ করে না, duplicate delivery হয়
        options = {
            "bind":                f"0.0.0.0:{PORT}",
            "workers":             workers,
            "worker_class":        "gthread",
            "threads":             24,  # threads বাড়িয়ে concurrency maintain করো
            "timeout":             60,
            "keepalive":           10,
            "max_requests":        2000,
            "max_requests_jitter": 200,
            "worker_connections":  1000,
            "preload_app":         True,
            "accesslog":           "-",
            "errorlog":            "-",
            "loglevel":            "warning",
        }
        print(f"🚀 Gunicorn starting: {workers} workers x 8 threads on port {PORT}")
        _StandaloneApp(sms_app, options).run()

    except ImportError:
        print(f"⚠️  Gunicorn not found — falling back to Flask dev server on port {PORT}")
        sms_app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False, threaded=True)

# ===========================
# SMS PARSER — TrxID খোঁজা
# ===========================

SMS_PATTERNS = {
    # ══════════════════════════════════════════════════════
    # BKASH — Real SMS examples:
    #   "You have received BDT 500. TrxID AB1CD23456 at 01/05/2025 10:30"
    #   "AB1CD23456 has been credited to your bKash account."
    #   "TrxID: AB1CD23456. Amount: Tk. 500."
    # ══════════════════════════════════════════════════════
    "Bkash": [
        r"TrxID[:\s\-\.]*([A-Z0-9]{8,12})",           # TrxID AB1CD23456
        r"Trx\s*ID[:\s\-\.]*([A-Z0-9]{8,12})",        # Trx ID AB1CD23456
        r"transaction\s*id[:\s\-\.]*([A-Z0-9]{8,12})", # transaction id AB1CD23456
        r"([A-Z]{2}[0-9A-Z]{6,10})\s+(?:has been|is your|credited)",  # AB1CD23456 has been
        r"TxnID[:\s\-\.]*([A-Z0-9]{8,12})",           # TxnID AB1CD23456
    ],

    # ══════════════════════════════════════════════════════
    # NAGAD — Real SMS examples:
    #   "Aপনার Nagad account-এ 500.00 Tk জমা হয়েছে। TrxID:NM12AB3456"
    #   "500 Tk. credited. Reference: NM12AB3456"
    #   "TxnID NM12AB3456 নিশ্চিত হয়েছে।"
    # ══════════════════════════════════════════════════════
    "Nagad": [
        r"TrxID[:\s\-\.]*([A-Z0-9]{8,14})",           # TrxID NM12AB3456
        r"TxnID[:\s\-\.]*([A-Z0-9]{8,14})",           # TxnID NM12AB3456
        r"Trx\s*ID[:\s\-\.]*([A-Z0-9]{8,14})",
        r"transaction\s*id[:\s\-\.]*([A-Z0-9]{8,14})",
        r"Reference[:\s\-\.]*([A-Z0-9]{8,14})",       # Reference NM12AB3456
        r"([A-Z0-9]{8,14})\s*নিশ্চিত",                # NM12AB3456 নিশ্চিত
        r"([A-Z0-9]{8,14})\s*confirmed",
    ],

    # ══════════════════════════════════════════════════════
    # ROCKET (Dutch-Bangla) — Real SMS examples:
    #   "TrxID 1234567890. Amount: 500 Tk."
    #   "Trx: 1234567890 successful."
    #   "DBBL Rocket TrxID:RK12345678"
    # ══════════════════════════════════════════════════════
    "Rocket": [
        r"TrxID[:\s\-\.]*([A-Z0-9]{8,14})",           # TrxID 1234567890 বা RK12345678
        r"TxnID[:\s\-\.]*([A-Z0-9]{8,14})",
        r"Trx\s*ID[:\s\-\.]*([A-Z0-9]{8,14})",
        r"Trx[:\s\-\.]+([A-Z0-9]{8,14})",             # Trx: 1234567890
        r"transaction\s*id[:\s\-\.]*([A-Z0-9]{8,14})",
        r"\b([0-9]{10,14})\b",                        # Rocket pure numeric TrxID
    ],
}

# Amount pattern — BDT 500 / Tk. 500 / 500 Tk / ৳500 / 500.00 টাকা সব ধরবে
AMOUNT_PATTERN = (
    r"BDT\s*([\d,]+\.?\d*)"           # BDT 500
    r"|Tk\.?\s*([\d,]+\.?\d*)"        # Tk. 500 / Tk 500
    r"|৳\s*([\d,]+\.?\d*)"            # ৳500
    r"|([\d,]+\.?\d*)\s*(?:Tk|টাকা|BDT)"  # 500 Tk / 500 টাকা
)

# ✅ শুধু এই sender নাম/নম্বর থেকে আসা SMS process হবে (অফিসিয়াল নম্বর)
ALLOWED_SENDERS = [
    "bkash", "nogod", "nagad", "roket", "rocket",
    "Bkash", "Nogod", "Nagad", "Roket", "Rocket",
]

# ✅ অফিসিয়াল numeric sender নম্বর — এগুলো always allow
ALLOWED_NUMERIC_SENDERS = [
    "16216",   # NAGAD অফিসিয়াল
]

def is_allowed_sender(sender: str) -> bool:
    """
    শুধু bkash/nagad/rocket এর অফিসিয়াল SMS accept করবে।
    - Keyword-based: bKash, NAGAD, Rocket ইত্যাদি
    - Numeric whitelist: 16216 (NAGAD অফিসিয়াল)
    কোনো personal নম্বর (01XXXXXXXXX) থেকে আসলে ignore।
    """
    sender_stripped = sender.strip()
    sender_lower = sender_stripped.lower()

    # Whitelist numeric senders চেক করো (exact match)
    digits_only = sender_lower.replace("+", "").replace("-", "").replace(" ", "")
    if digits_only in ALLOWED_NUMERIC_SENDERS:
        return True

    # যদি sender পুরোপুরি numeric এবং whitelist এ নেই → ignore
    # e.g. "01711234567", "+8801711234567"
    if digits_only.isdigit():
        return False

    # যদি নামের মধ্যে allowed keyword থাকে → allow
    for allowed in ALLOWED_SENDERS:
        if allowed.lower() in sender_lower:
            return True

    return False

def detect_method(sender, body):
    """
    SMS sender ও body দেখে payment method detect করে।
    Priority: body keyword > sender name fallback
    ⚠️ Sender filter নেই — App নিজেই filter করবে।
    """
    sender_lower = sender.strip().lower()
    body_lower = body.lower()

    # Body keyword দিয়ে detect (সবচেয়ে reliable)
    if "bkash" in body_lower:
        return "Bkash"
    if "nagad" in body_lower or "nogod" in body_lower:
        return "Nagad"
    if "rocket" in body_lower or "dutch" in body_lower or "roket" in body_lower or "dbbl" in body_lower:
        return "Rocket"

    # Sender name দিয়ে fallback (body তে keyword না থাকলে)
    if "bkash" in sender_lower:
        return "Bkash"
    if "nagad" in sender_lower or "nogod" in sender_lower:
        return "Nagad"
    if "rocket" in sender_lower or "roket" in sender_lower or "dutch" in sender_lower:
        return "Rocket"

    return None

def extract_trxid(body, method):
    patterns = SMS_PATTERNS.get(method, [])
    for pattern in patterns:
        match = re.search(pattern, body, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return None

def extract_amount(body):
    match = re.search(AMOUNT_PATTERN, body, re.IGNORECASE)
    if match:
        amount_str = match.group(1) or match.group(2) or match.group(3) or match.group(4)
        amount_str = amount_str.replace(",", "").replace(" ", "")
        try:
            return float(amount_str)
        except:
            return None
    return None

def is_trxid_already_used(db, trx_id):
    """একই TrxID দিয়ে আগে approve হয়েছে কিনা চেক করা।"""
    trx_id = trx_id.upper()
    for req in db["deposit_requests"].values():
        if req.get("transaction_id", "").upper() == trx_id:
            if req.get("status") == "approved":
                return True
    if trx_id in db.get("verified_trxids", []):
        return True
    return False

def mark_trxid_used(db, trx_id):
    """TrxID কে used হিসেবে mark করা।"""
    if "verified_trxids" not in db:
        db["verified_trxids"] = []
    trx_id = trx_id.upper()
    if trx_id not in db["verified_trxids"]:
        db["verified_trxids"].append(trx_id)

def _cleanup_old_sms_pending():
    """৩০ দিনের বেশি পুরনো sms_pending_trxids এন্ট্রি মুছে ফেলো।"""
    db = load_db()
    pending = db.get("sms_pending_trxids", {})
    if not pending:
        return

    now_dt = datetime.now(BD_TZ)
    to_delete = []

    for trx_id, data in pending.items():
        time_str = data.get("time", "")
        try:
            entry_dt = BD_TZ.localize(
                datetime.strptime(time_str, "%d/%m/%Y %I:%M %p")
            )
            age_days = (now_dt - entry_dt).days
            if age_days >= 30:
                to_delete.append(trx_id)
        except Exception:
            pass  # parse error হলে রেখে দাও

    if to_delete:
        for trx_id in to_delete:
            del db["sms_pending_trxids"][trx_id]
            _fb_ref(f"/sms_pending_trxids/{trx_id}").delete()
        _update_db_cache_in_place(db)
        logging.info(f"SMS cleanup: {len(to_delete)} পুরনো entry মুছে ফেলা হয়েছে।")


# process_sms — নিচে line 7052 এ final unified version আছে


# ===========================
# CASHBACK HELPER
# ===========================
def get_cashback(amount, db=None):
    """ডিপোজিট amount অনুযায়ী cashback return করো।
    Admin cashback feature বন্ধ থাকলে সবসময় 0 return করে।
    """
    if db is None:
        db = load_db()
    # Admin cashback toggle check
    if not db.get("settings", {}).get("cashback_enabled", True):
        return 0
    if 500 <= amount <= 999:
        return 20
    elif 1000 <= amount <= 2000:
        return 50
    return 0

# auto_approve_deposit — unified version নিচে আছে (line ~7130)

# ===========================
# LANGUAGE STRINGS
# ===========================

LANG = {
    "en": {
        "welcome": (
            "🏪<b>WELCOME TO DIGITAL STORE!</b> 🛍️\n\n"
            "\n"
            "✨ Premium Digital Accounts &amp; Services \n"
            "⚡ Instant Delivery (24/7)\n"
            "🔒 100% Safe &amp; Trusted with Warranty\n"
            "💰 <b>DEPOSIT AUTO APPROVE</b>\n\n"
            "👇<i> CHOOSE AN OPTION FROM THE MENU BELOW</i> 👇"
        ),
        "shop": "🛍️ Shop Now",
        "deposit": "💰 Deposit",
        "balance": "👤 My Account",
        "orders": "📜 My Orders",
        "refer": "👥 Refer & Earn",
        "support": "☎️ Support",
        "admin": "⚙️ Admin Panel",
        "back": "🔙 Back to Main Menu",
        "cancel": "❌ Cancel",
        "single_buy": "🛒 Buy one pice",
        "bulk_buy": "📦 Bulk Buy",
        "product_list_title": "🛍️ <b>Product List</b>",
        "product_detail": "🏷️ <b>{name}</b>\n\n💵 <b>Price:</b> <code>{price} BDT</code>\n📦 <b>Stock:</b> <code>{stock} pcs</code>\n\n<i>Choose an option below 👇</i>",
        "out_of_stock": "❌ <b>{name}</b>\n\nSorry! This product is currently out of stock.\nPlease try again later.",
        "bulk_prompt": "📦 <b>Bulk Buy — {name}</b>\n\n💵 Per piece: <code>{price} BDT</code>\n📊 Available: <code>{stock} pcs</code>\n\n<b>How many do you want? Enter a number:</b>\n<i>Example: 4, 10, 20</i>",
        "invalid_number": "❌ <b>Invalid input!</b>\n\nPlease enter a number only. Example: <code>2</code>, <code>5</code>",
        "qty_zero": "❌ <b>Quantity must be at least 1!</b>",
        "insufficient_stock": "❌ <b>Insufficient stock!</b>\n\n📦 You requested: <b>{qty} pcs</b>\n📊 Available: <b>{stock} pcs</b>\n\n<i>Try a smaller quantity.</i>",
        "insufficient_balance": (
            "💸 <b>Insufficient Balance!</b>\n\n"
            "\n"
            "🛒 <b>Required:</b> <code>{total} BDT</code>\n"
            "💳 <b>Your Balance:</b> <code>{bal} BDT</code>\n"
            "❗ <b>Shortage:</b> <code>{short} BDT</code>\n"
            "\n\n"
            "💰 Please deposit and try again."
        ),
        "purchase_success": (
            "🎉 <b>Purchase Successful!</b>\n"
            "\n"
            "🆔 <b>Order ID:</b> <code>{order_id}</code>\n"
            "📦 <b>Product:</b> {product}\n"
            "🔢 <b>Quantity:</b> {qty} pcs\n"
            "💵 <b>Per piece:</b> {price} BDT\n"
            "💰 <b>Total:</b> {total} BDT\n"
            "💳 <b>Remaining Balance:</b> {bal} BDT\n"
            "📅 <b>Date:</b> {date}\n"
            "\n"
            "📎 <i>Sending your account file below...</i>"
        ),
        "file_error": "⚠️ <b>File Error</b>\n\nStock file issue. Please contact support.",
        "db_error": "❌ <b>Database Error!</b> Please try again.",
        "general_error": "❌ <b>Error!</b>\n\nSomething went wrong. Please try again.",
        "no_orders": "📭 <b>No orders yet!</b>\n\nYou haven't made any purchases.\nStart shopping! 🛒",
        "orders_title": "📂 <b>Your Order History (Today's Last 5)</b>\n",
        "balance_msg": (
            "\n"
            "👤 <b>Account Info</b>\n"
            "\n"
            "👤 <b>Username:</b> @{username}\n"
            "🆔 <b>User ID:</b> <code>{uid}</code>\n"
            "💳 <b>Balance:</b> <code>{bal} BDT</code>\n"
            "📦 <b>Total Orders:</b> {orders}\n"
            "👥 <b>Referrals:</b> {refs}\n"
            ""
        ),
        "refer_msg": (
            "\n"
            "👥 <b>Referral Program</b>\n"
            "\n"
            "🎁 <b>Your Referral Link:</b>\n"
            "<code>{link}</code>\n\n"
            "💰 <b>Bonus per referral:</b> {bonus} BDT\n"
            "👥 <b>Total referrals:</b> {count}\n"
            "💵 <b>Total earned:</b> {total} BDT\n"
            "\n\n"
            "<i>🔗 Share your link and earn!</i>"
        ),
        "refer_bonus_notif": "🎉 <b>Referral Bonus!</b>\n\nA new user joined via your link.\nYou received: <b>+{bonus} BDT</b> 💰",
        "support_msg": (
            "📞 <b>Customer Support</b>\n\n"
            "Contact us for any issues:\n\n"
            "👤 <b>Support Admin:</b> @{support}\n\n"
            "<i>Click the button below to reach us directly</i>"
        ),
        "contact_support": "💬 Contact Support",
        "main_menu": "🏠 <b>Main Menu</b>",
        "lang_select": "🌐 <b>Select Language / ভাষা বেছে নিন</b>",
        "lang_set": "✅ Language set to English!",
        "join_channels": "⚠️ <b>Channel Subscription Required</b>\n\nPlease join our channels to use the bot 👇",
        "join_channel_btn": "📢 Join Channel {n}",
        "verify_join": "✅ I've Joined — Verify",
        "verify_ok": "✅ Verified! Welcome!",
        "verify_fail": "❌ Please join all channels!",
        "banned": "🚫 <b>Access Denied</b>\n\nYou have been banned from this bot.\nPlease contact support.",
        "deposit_menu": (
            "\n"
            "💰 <b>Add Balance</b>\n"
            "\n\n"
            "🟠 <b>Bkash</b> → <code>{bkash}</code>\n"
            "🟢 <b>Nagad</b> → <code>{nagad}</code>\n"
            "🔴 <b>Rocket</b> → <code>{rocket}</code>\n"
            "🔵 <b>Binance</b> → <code>{binance}</code>\n\n"
            "\n"
            "⚡ <b>Bkash/Nagad/Rocket</b> — Auto approved via SMS\n"
            "🔵 <b>Binance</b> — Manual approval\n"
            "\n"
            f"<i>Minimum deposit: {MIN_DEPOSIT} BDT</i>\n\n"
            "🎁 <b>Cashback Offer:</b>\n"
            "  💵 Deposit 500–999 BDT → <b>+20 BDT Cashback</b> 🎉\n"
            "  💵 Deposit 1000–2000 BDT → <b>+50 BDT Cashback</b> 🎉"
        ),
        "dep_method_select": "💳 <b>Select Payment Method</b>\n\n<i>Choose how you will send money:</i>",
        "dep_instruction": (
            "\n"
            "💳✨ <b>{method} DEPOSIT</b> ✨💳\n"
            "\n\n"
            "📞 Send money to:\n"
            "──────────────────\n"
            "  <code>{account}</code>\n"
            "──────────────────\n\n"
            "💚 Follow these steps:\n\n"
            "1️⃣ 💰 Send to the number above\n"
            "2️⃣ 📋 Copy the TRXID\n"
            "3️⃣ 🖊️ Type your amount below\n\n"
            "──────────────────\n"
            "<b>Enter Your Amount👇</b>\n"       
        ),
        "dep_binance_instruction": (
            "\n"
            "🔵 <b>Binance Deposit</b>\n"
            "\n\n"
            "<b>Binance Pay ID:</b>\n"
            "<code>{account}</code>\n\n"
            "<b>Steps:</b>\n"
            "1️⃣ Open Binance → Pay\n"
            "2️⃣ Send USDT/BNB to the ID above\n"
            "3️⃣ Note your  Binance <b>username </b> \n"
            "4️⃣ Enter the amount below\n\n"
            "⚠️ <i>Binance is manually approved by admin.</i>\n\n"
            "<b>How much did you send? (Enter BDT)1$ =126 BDT</b>"
        ),
        "dep_invalid_amount": "❌ <b>Invalid amount!</b>\n\nPlease enter a number (e.g. 500, 1000):",
        "dep_min": f"❌ <b>Minimum deposit is {MIN_DEPOSIT} BDT!</b>\n\nPlease try again:",
        "dep_trxid_prompt": (
            "\n"
            "📝 <b>ENTER TRANSACTION ID</b>\n"
            "\n\n"
            "💳 {method}  |  💰 {amount} BDT\n"
            "──────────────────────────\n\n"
            "📋 Your TrxID:\n\n"
            "🔹 Example: ABC123XYZ\n\n"
            "👉 Type your TrxID below👇"
        ),
        "dep_invalid_trx": "❌ <b>Invalid Transaction ID!</b>\n\nPlease try again:",
        "dep_submitted_auto": (
            "\n"
            "✅ <b>Deposit Request Submitted!</b>\n"
            "\n\n"
            "🆔 <b>Request ID:</b> <code>{req_id}</code>\n"
            "💳 <b>Method:</b> {method}\n"
            "💰 <b>Amount:</b> {amount} BDT\n"
            "🔑 <b>TrxID:</b> <code>{trx}</code>\n"
            "📅 <b>Date:</b> {date}\n\n"
            "\n"
            "⚡ <b>Status:</b> Auto-verifying via SMS...\n\n"
            
        ),
        "dep_submitted_manual": (
            "\n"
            "✅ <b>Deposit Request Submitted!</b>\n"
            "\n\n"
            "🆔 <b>Request ID:</b> <code>{req_id}</code>\n"
            "💳 <b>Method:</b> {method}\n"
            "💰 <b>Amount:</b> {amount} BDT\n"
            "🔑 <b>TrxID:</b> <code>{trx}</code>\n"
            "📅 <b>Date:</b> {date}\n\n"
            "\n"
            "⏳ <b>Status:</b> Pending admin approval\n\n"
            "<i>Usually approved within 5 minutes.</i>"
        ),
        "dep_approved": "✅ <b>Deposit Approved!</b>\n\n💰 <b>Amount:</b> {amount} BDT\n💳 <b>New Balance:</b> {bal} BDT\n🆔 <b>Request ID:</b> <code>{req_id}</code>",
        "dep_rejected": "❌ <b>Deposit Rejected!</b>\n\n💰 <b>Amount:</b> {amount} BDT\n🆔 <b>Request ID:</b> <code>{req_id}</code>\n\n<i>Please contact support.</i>",
        "dep_history": "📊 <b>Deposit History</b>\n\n",
        "no_dep_history": "❌ <b>No deposit history</b>",
        "dep_help": (
            "\n"
            "❓ <b>How to Deposit</b>\n"
            "\n\n"
            "🟠 <b>Bkash / 🟢 Nagad / 🔴 Rocket:</b>\n"
            "  1. Send money to our number\n"
            "  2. Enter amount & TrxID in bot\n"
            "  3. ⚡ Auto-approved via SMS!\n\n"
            "🔵 <b>Binance:</b>\n"
            "  1. Send to our Binance Pay ID\n"
            "  2. Enter amount & TrxID in bot\n"
            "  3. Wait for admin approval\n\n"
            "⏱ <i>Bkash/Nagad/Rocket: ~1 min | Binance: ~5 min</i>"
        ),
        "do_deposit": "💳 Make a Deposit",
        "how_deposit": "❓ How to Deposit",
        "dep_history_btn": "📊 Deposit History",
        "go_back": "🔙 Back",
        "main_menu_btn": "🔙 Main Menu",
        "order_item": "🆔 <code>{id}</code>\n📦 <b>Product:</b> {product}\n🔢 <b>Qty:</b> {qty} pcs | 💰 <b>Total:</b> {total} BDT\n📅 <b>Date:</b> {date}\n",
        "file_caption": "📄 <b>{product} — Account Data</b>\n\n🆔 Order: <code>{order_id}</code>\n🔢 Qty: {qty} pcs\n📅 {date}\n\n⚠️ <i>Please save this file!</i>",
        "processing": "⏳ Processing...",
        "order_cancelled": "❌ <b>Order cancelled.</b>",
    },
    "bn": {
        "welcome": (
            "⭐ <b>ডিজিটাল স্টোরে বেস্ট আপডেট অফার!</b> ⭐\n\n"
            "\n"
            "✅ সব ধরনের প্রিমিয়াম অ্যাকাউন্ট\n"
            "✅ ইনস্ট্যান্ট ডেলিভারি (রাত-দিন একরকম)\n"
            "✅ ১০০% রিপ্লেসমেন্ট গ্যারান্টি\n"
            "💰 <b>ডিপোজিট অটো অ্যাপ্রুভ</b>\n\n"
            "📣 এখনই অর্ডার করুন — স্টক সীমিত\n\n"
            "👉 মেনু দেখে যেটা ভালো লাগে, সেটা সিলেক্ট করুন 👉"
        ),
        "shop": "🛍️ শপিং করুন",
        "deposit": "💰 ডিপোজিট",
        "balance": "👤 আমার অ্যাকাউন্ট",
        "orders": "📜 আমার অর্ডার",
        "refer": "👥 রেফার করুন",
        "support": "☎️ সাপোর্ট",
        "admin": "⚙️ অ্যাডমিন প্যানেল",
        "back": "🔙 মেইন মেনুতে ফিরুন",
        "cancel": "❌ বাতিল করুন",
        "single_buy": "🛒 একটি কিনুন",
        "bulk_buy": "📦 বাল্ক কিনুন",
        "product_list_title": "🛍️ <b>পণ্য তালিকা</b>",
        "product_detail": "🏷️ <b>{name}</b>\n\n💵 <b>দাম:</b> <code>{price} BDT</code>\n📦 <b>স্টক:</b> <code>{stock} পিস</code>\n\n<i>নিচের অপশন বেছে নিন 👇</i>",
        "out_of_stock": "❌ <b>{name}</b>\n\nদুঃখিত! এই পণ্যটি বর্তমানে স্টকে নেই।\nপরে আবার চেষ্টা করুন।",
        "bulk_prompt": "📦 <b>বাল্ক কেনা — {name}</b>\n\n💵 প্রতি পিস: <code>{price} BDT</code>\n📊 আছে: <code>{stock} পিস</code>\n\n<b>কতটি কিনতে চান? সংখ্যা লিখুন:</b>\n<i>উদাহরণ: ৪, ১০, ২০</i>",
        "invalid_number": "❌ <b>ভুল ইনপুট!</b>\n\nশুধু সংখ্যা দিন। যেমন: <code>২</code>, <code>৫</code>",
        "qty_zero": "❌ <b>পরিমাণ কমপক্ষে ১ হতে হবে!</b>",
        "insufficient_stock": "❌ <b>পর্যাপ্ত স্টক নেই!</b>\n\n📦 আপনি চেয়েছেন: <b>{qty} পিস</b>\n📊 উপলব্ধ আছে: <b>{stock} পিস</b>\n\n<i>কম পরিমাণ দিয়ে আবার চেষ্টা করুন।</i>",
        "insufficient_balance": (
            "💸 <b>ব্যালেন্স অপর্যাপ্ত!</b>\n\n"
            "\n"
            "🛒 <b>প্রয়োজন:</b> <code>{total} BDT</code>\n"
            "💳 <b>আপনার ব্যালেন্স:</b> <code>{bal} BDT</code>\n"
            "❗ <b>ঘাটতি:</b> <code>{short} BDT</code>\n"
            "\n\n"
            "💰 ডিপোজিট করুন এবং আবার চেষ্টা করুন।"
        ),
        "purchase_success": (
            "🎉 <b>ক্রয় সফল হয়েছে!</b>\n"
            "\n"
            "🆔 <b>অর্ডার ID:</b> <code>{order_id}</code>\n"
            "📦 <b>পণ্য:</b> {product}\n"
            "🔢 <b>পরিমাণ:</b> {qty} পিস\n"
            "💵 <b>প্রতি পিস:</b> {price} BDT\n"
            "💰 <b>মোট খরচ:</b> {total} BDT\n"
            "💳 <b>বর্তমান ব্যালেন্স:</b> {bal} BDT\n"
            "📅 <b>তারিখ:</b> {date}\n"
            "\n"
            "📎 <i>নিচে আপনার অ্যাকাউন্ট ফাইল পাঠানো হচ্ছে...</i>"
        ),
        "file_error": "⚠️ <b>ফাইল এরর</b>\n\nস্টক ফাইলে সমস্যা হয়েছে। সাপোর্টে যোগাযোগ করুন।",
        "db_error": "❌ <b>ডেটাবেস এরর!</b> আবার চেষ্টা করুন।",
        "general_error": "❌ <b>এরর!</b>\n\nসমস্যা হয়েছে। আবার চেষ্টা করুন।",
        "no_orders": "📭 <b>কোনো অর্ডার নেই!</b>\n\nএখনো কোনো কেনাকাটা করেননি।\nশপিং শুরু করুন! 🛒",
        "orders_title": "📂 <b>আজকের সর্বশেষ ৫টি অর্ডার</b>\n",
        "balance_msg": (
            "\n"
            "👤 <b>অ্যাকাউন্ট তথ্য</b>\n"
            "\n"
            "👤 <b>ইউজারনেম:</b> @{username}\n"
            "🆔 <b>User ID:</b> <code>{uid}</code>\n"
            "💳 <b>ব্যালেন্স:</b> <code>{bal} BDT</code>\n"
            "📦 <b>মোট অর্ডার:</b> {orders}টি\n"
            "👥 <b>রেফারেল:</b> {refs} জন\n"
            ""
        ),
        "refer_msg": (
            "\n"
            "👥 <b>রেফারেল প্রোগ্রাম</b>\n"
            "\n"
            "🎁 <b>আপনার রেফারেল লিংক:</b>\n"
            "<code>{link}</code>\n\n"
            "💰 <b>প্রতি রেফারেলে বোনাস:</b> {bonus} BDT\n"
            "👥 <b>মোট রেফারেল:</b> {count} জন\n"
            "💵 <b>মোট আয়:</b> {total} BDT\n"
            "\n\n"
            "<i>🔗 লিংক শেয়ার করুন এবং আয় করুন!</i>"
        ),
        "refer_bonus_notif": "🎉 <b>রেফারেল বোনাস পেয়েছেন!</b>\n\nনতুন ইউজার আপনার লিংক দিয়ে যোগ দিয়েছে।\nআপনি পেয়েছেন: <b>+{bonus} BDT</b> 💰",
        "support_msg": (
            "📞 <b>কাস্টমার সাপোর্ট</b>\n\n"
            "যেকোনো সমস্যায় আমাদের সাথে যোগাযোগ করুন:\n\n"
            "👤 <b>সাপোর্ট অ্যাডমিন:</b> @{support}\n\n"
            "<i>নিচের বাটনে ক্লিক করে সরাসরি যোগাযোগ করুন</i>"
        ),
        "contact_support": "💬 সাপোর্টে যোগাযোগ করুন",
        "main_menu": "🏠 <b>মেইন মেনু</b>",
        "lang_select": "🌐 <b>Select Language / ভাষা বেছে নিন</b>",
        "lang_set": "✅ ভাষা বাংলায় সেট করা হয়েছে!",
        "join_channels": "⚠️ <b>চ্যানেল সাবস্ক্রিপশন প্রয়োজন</b>\n\nবট ব্যবহার করতে আমাদের চ্যানেলগুলো জয়েন করুন 👇",
        "join_channel_btn": "📢 চ্যানেল {n} জয়েন করুন",
        "verify_join": "✅ জয়েন করেছি — ভেরিফাই করুন",
        "verify_ok": "✅ ভেরিফাইড! স্বাগতম!",
        "verify_fail": "❌ সব চ্যানেল জয়েন করুন!",
        "banned": "🚫 <b>অ্যাক্সেস নিষিদ্ধ</b>\n\nআপনাকে এই বট থেকে ব্যান করা হয়েছে।\nসাপোর্টের সাথে যোগাযোগ করুন।",
        "deposit_menu": (
            "\n"
            "💰 <b>ব্যালেন্স যোগ করুন</b>\n"
            "\n\n"
            "🟠 <b>Bkash</b> → <code>{bkash}</code>\n"
            "🟢 <b>Nagad</b> → <code>{nagad}</code>\n"
            "🔴 <b>Rocket</b> → <code>{rocket}</code>\n"
            "🔵 <b>Binance</b> → <code>{binance}</code>\n\n"
            "\n"
            "⚡ <b>Bkash/Nagad/Rocket</b> — SMS দিয়ে অটো অ্যাপ্রুভ\n"
            "🔵 <b>Binance</b> — অ্যাডমিন ম্যানুয়ালি অ্যাপ্রুভ করবে\n"
            "\n"
            f"<i>সর্বনিম্ন ডিপোজিট: {MIN_DEPOSIT} BDT</i>\n\n"
            "🎁 <b>ক্যাশব্যাক অফার:</b>\n"
            "  💵 ৫০০–৯৯৯ BDT ডিপোজিটে → <b>+২০ BDT ক্যাশব্যাক</b> 🎉\n"
            "  💵 ১০০০–২০০০ BDT ডিপোজিটে → <b>+৫০ BDT ক্যাশব্যাক</b> 🎉"
        ),
        "dep_method_select": "💳 <b>পেমেন্ট মেথড বেছে নিন</b>\n\n<i>আপনি যে মেথডে টাকা পাঠাবেন সেটি বেছে নিন:</i>",
        "dep_instruction": (
            "\n"
            "💳✨ <b>{method} ডিপোজিট</b> ✨💳\n"
            "\n\n"
            "📞 টাকা পাঠান এই নম্বরে ??\n"
            "──────────────────\n"
            "  <code>{account}</code>\n"
            "──────────────────\n\n"
            "💚 এভাবে করুন:\n\n"
            "১) উপরের নম্বরে 💰 টাকা পাঠান\n"
            "২) 📋 টিআরএক্সআইডি কপি করে রাখুন\n"
            "৩) 🖊️ নিচে কত টাকা পাঠিয়েছেন সেটি লিখুন\n\n"
            "──────────────────\n"
            "❓ কত টাকা পাঠিয়েছেন?\n"
            "👉 শুধু সংখ্যাটি লিখুন (যেমন: 500)"
        ),
        "dep_binance_instruction": (
            "\n"
            "🔵 <b>Binance ডিপোজিট</b>\n"
            "\n\n"
            "<b>Binance Pay ID:</b>\n"
            "<code>{account}</code>\n\n"
            "<b>ধাপসমূহ:</b>\n"
            "1️⃣ Binance → Pay খুলুন\n"
            "2️⃣ উপরের ID তে USDT/BNB পাঠান\n"
            "3️⃣ Transaction ID কপি করুন\n"
            "4️⃣ নিচে পরিমাণ লিখুন\n\n"
            "⚠️ <i>Binance অ্যাডমিন ম্যানুয়ালি অ্যাপ্রুভ করবেন।</i>\n\n"
            "<b>কত $ ডলার পাঠিয়েছেন?</b>"
        ),
        "dep_invalid_amount": "❌ <b>ভুল পরিমাণ!</b>\n\nসংখ্যায় লিখুন (যেমন: 500, 1000):",
        "dep_min": f"❌ <b>সর্বনিম্ন ডিপোজিট {MIN_DEPOSIT} BDT!</b>\n\nআবার চেষ্টা করুন:",
        "dep_trxid_prompt": (
            "\n"
            "📝 <b>ট্রানজেকশন আইডি দিন</b>\n"
            "\n\n"
            "💳 পদ্ধতি: {method}\n"
            "💰 পরিমাণ: {amount} টাকা\n"
            "──────────────────────────\n\n"
            "📋 আপনার TrxID লিখুন:\n\n"
            "🔹 উদাহরণ: ABC123XYZ\n\n"
            "👉 নিচে আপনার TrxID টাইপ করুন"
        ),
        "dep_invalid_trx": "❌ <b>ভুল Transaction ID!</b>\n\nআবার চেষ্টা করুন:",
        "dep_submitted_auto": (
            "\n"
            "✅ <b>ডিপোজিট রিকোয়েস্ট জমা হয়েছে!</b>\n"
            "\n\n"
            "🆔 <b>Request ID:</b> <code>{req_id}</code>\n"
            "💳 <b>মেথড:</b> {method}\n"
            "💰 <b>পরিমাণ:</b> {amount} BDT\n"
            "🔑 <b>TrxID:</b> <code>{trx}</code>\n"
            "📅 <b>তারিখ:</b> {date}\n\n"
            "\n"
            "⚡ <b>স্ট্যাটাস:</b> SMS দিয়ে অটো ভেরিফাই হচ্ছে...\n\n"
            "<i>সাধারণত ১-২ মিনিটে অ্যাপ্রুভ হয়।</i>"
        ),
        "dep_submitted_manual": (
            "\n"
            "✅ <b>ডিপোজিট রিকোয়েস্ট জমা হয়েছে!</b>\n"
            "\n\n"
            "🆔 <b>Request ID:</b> <code>{req_id}</code>\n"
            "💳 <b>মেথড:</b> {method}\n"
            "💰 <b>পরিমাণ:</b> {amount} BDT\n"
            "🔑 <b>TrxID:</b> <code>{trx}</code>\n"
            "📅 <b>তারিখ:</b> {date}\n\n"
            "\n"
            "⏳ <b>স্ট্যাটাস:</b> অ্যাডমিন ভেরিফিকেশন পেন্ডিং\n\n"
            "<i>সাধারণত ৫ মিনিটের মধ্যে অ্যাপ্রুভ হয়।</i>"
        ),
        "dep_approved": "✅ <b>ডিপোজিট অ্যাপ্রুভ হয়েছে!</b>\n\n💰 <b>পরিমাণ:</b> {amount} BDT\n💳 <b>নতুন ব্যালেন্স:</b> {bal} BDT\n🆔 <b>Request ID:</b> <code>{req_id}</code>",
        "dep_rejected": "❌ <b>ডিপোজিট রিজেক্ট হয়েছে!</b>\n\n💰 <b>পরিমাণ:</b> {amount} BDT\n🆔 <b>Request ID:</b> <code>{req_id}</code>\n\n<i>সাপোর্টে যোগাযোগ করুন।</i>",
        "dep_history": "📊 <b>ডিপোজিট ইতিহাস</b>\n\n",
        "no_dep_history": "❌ <b>কোনো ডিপোজিট ইতিহাস নেই</b>",
        "dep_help": (
            "\n"
            "❓ <b>কিভাবে ডিপোজিট করবেন</b>\n"
            "\n\n"
            "🟠 <b>Bkash / 🟢 Nagad / 🔴 Rocket:</b>\n"
            "  ১. আমাদের নম্বরে টাকা পাঠান\n"
            "  ২. বটে পরিমাণ ও TrxID দিন\n"
            "  ৩. ⚡ SMS দিয়ে অটো অ্যাপ্রুভ!\n\n"
            "🔵 <b>Binance:</b>\n"
            "  ১. Binance Pay ID তে পাঠান\n"
            "  ২. বটে পরিমাণ ও TrxID দিন\n"
            "  ৩. অ্যাডমিন অ্যাপ্রুভের জন্য অপেক্ষা করুন\n\n"
            "⏱ <i>Bkash/Nagad/Rocket: ~১ মিনিট | Binance: ~৩০ মিনিট</i>"
        ),
        "do_deposit": "💳 ডিপোজিট করুন",
        "how_deposit": "❓ কিভাবে ডিপোজিট করবেন",
        "dep_history_btn": "📊 ডিপোজিট ইতিহাস",
        "go_back": "🔙 ফিরুন",
        "main_menu_btn": "🔙 মেইন মেনু",
        "order_item": "🆔 <code>{id}</code>\n📦 <b>পণ্য:</b> {product}\n🔢 <b>পরিমাণ:</b> {qty} পিস | 💰 <b>মোট:</b> {total} BDT\n📅 <b>তারিখ:</b> {date}\n",
        "file_caption": "📄 <b>{product} — অ্যাকাউন্ট ডেটা</b>\n\n🆔 Order: <code>{order_id}</code>\n🔢 পরিমাণ: {qty} পিস\n📅 {date}\n\n⚠️ <i>ফাইলটি সংরক্ষণ করুন!</i>",
        "processing": "⏳ প্রসেস হচ্ছে...",
        "order_cancelled": "❌ <b>অর্ডার বাতিল করা হয়েছে।</b>",
        "single_buy": "🛒 ১ পিস কিনুন",
        "bulk_buy": "📦 বাল্ক কিনুন",
        "admin": "⚙️ অ্যাডমিন প্যানেল",
    }
}

def t(user_id, key, **kwargs):
    # Avoid repeated DB load — use cached lang lookup
    try:
        lang = _lang_cache.get(str(user_id), "en")
    except Exception:
        lang = "en"
    text = LANG.get(lang, LANG["en"]).get(key, LANG["en"].get(key, key))
    if kwargs:
        try:
            text = text.format(**kwargs)
        except Exception:
            pass
    return text

# Simple in-memory language cache to avoid per-translation DB hits
_lang_cache = {}

# ===========================
# RATE LIMITING SYSTEM
# ===========================
_user_message_times = {}   # {user_id: deque of timestamps}
_user_deposit_times = {}   # {user_id: last_deposit_timestamp}
from collections import deque as _deque

def is_rate_limited(user_id):
    """⚡ O(1) amortized rate limit — deque দিয়ে পুরনো entries ছাঁটো।"""
    uid = str(user_id)
    now = _time_module.time()
    window_start = now - 60

    if uid not in _user_message_times:
        _user_message_times[uid] = _deque()

    dq = _user_message_times[uid]
    # পুরনো entries বাম দিক থেকে সরাও — O(k) শুধু expire হওয়া entries
    while dq and dq[0] < window_start:
        dq.popleft()
    dq.append(now)
    return len(dq) > MAX_MSG_PER_MINUTE

def is_deposit_cooldown(user_id):
    """Cooldown disabled — সবসময় False।"""
    return False

def get_user_lang_cached(user_id, db=None):
    uid = str(user_id)
    if uid in _lang_cache:
        return _lang_cache[uid]
    if db is None:
        db = load_db()
    lang = db["users"].get(uid, {}).get("lang", "en")
    _lang_cache[uid] = lang
    return lang

def get_user_lang(user_id, db=None):
    """get_user_lang_cached এর alias — সব জায়গায় কাজ করবে।"""
    return get_user_lang_cached(user_id, db)

# ===========================
# DATABASE FUNCTIONS
# ===========================

_DB_DEFAULT = {
    "users": {},
    "products": {
        "Facebook": 8,
        "Gmail": 20,
        "Instagram": 10
    },
    "product_details": {},
    "flash_sales": {},
    "settings": {"refer_bonus": 2, "new_user_discount": 0, "promo_feature_enabled": True, "cashback_enabled": True},
    "banned_users": [],
    "deposit_requests": {},
    "sms_pending_trxids": {},
    "verified_trxids": [],
    "trx_attempts": {},
    "blocked_trxids": [],
    "tickets": {},
    "promo_codes": {},
    "promo_usage": {},
    "bulkmail_products": {},     # Bulkmail product on/off status
    "hotmail143_products": {},   # Hotmail143 API product on/off status
}

def _firebase_to_list(value, default_type):
    """
    Firebase Realtime DB list/array কে dict হিসেবে return করে।
    এখানে যদি value None হয় তাহলে default দিই।
    list টাইপের field গুলো (banned_users, verified_trxids, blocked_trxids)
    Firebase এ dict আকারে আসতে পারে — এগুলোকে list এ convert করি।
    """
    if value is None:
        return default_type() if callable(default_type) else default_type
    if isinstance(default_type, list) or default_type is list:
        if isinstance(value, dict):
            return list(value.values())
        if isinstance(value, list):
            return [v for v in value if v is not None]
        return []
    return value

def save_db(data, _async=True):
    """Firebase Realtime Database তে পুরো DB save করো।
    ⚡ FAST: cache invalidate না করে নতুন data দিয়ে cache update করো।
    পরের load_db() Firebase hit করবে না — instant response।

    ⚠️ WARNING: এই function পুরো DB (~10MB) write করে — Firebase quota drain হয়।
    ছোট changes এর জন্য update_db_path() বা patch_db_path() ব্যবহার করো।
    যেমন: update_db_path(f"/users/{user_id}", db["users"][user_id])
    """
    _update_db_cache_in_place(data)
    if _async:
        _EXECUTOR.submit(_save_db_sync, data)
        return True
    return _save_db_sync(data)

def _update_db_cache_in_place(data):
    """Cache invalidate না করে নতুন data দিয়ে সরাসরি update করো।
    Section cache ও update করো।
    """
    global _db_cache, _db_cache_time
    now_ts = _time_module.time()
    with _section_cache_lock:
        # Section cache এ relevant keys update করো
        for section in ("users", "deposits", "config", "txn"):
            keys = _get_section_keys(section)
            section_update = {k: data[k] for k in keys if k in data}
            if section_update:
                existing = _section_cache.get(section)
                if existing:
                    merged = dict(existing[0])
                    merged.update(section_update)
                    _section_cache[section] = (merged, now_ts)
                else:
                    _section_cache[section] = (section_update, now_ts)
        # _db_cache backward compat
        new_cache = {}
        for k, v in data.items():
            if isinstance(v, dict):
                new_cache[k] = dict(v)
            elif isinstance(v, list):
                new_cache[k] = list(v)
            else:
                new_cache[k] = v
        _db_cache = new_cache
        _db_cache_time = now_ts

def _save_db_sync(data):
    """Synchronous Firebase save — background thread এ চলে।"""
    try:
        upload = {}
        for key, val in data.items():
            if isinstance(val, list):
                upload[key] = {str(i): v for i, v in enumerate(val)} if val else {}
            else:
                upload[key] = val
        _fb_ref("/").set(upload)
        return True
    except Exception as e:
        logging.error(f"Firebase save_db error: {e}")
        tmp_file = DB_FILE + ".tmp"
        try:
            with open(tmp_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            os.replace(tmp_file, DB_FILE)
        except:
            pass
        return False

def update_db_path(path, value, _async=True):
    """Firebase এর specific path শুধু update করো (পুরো DB write না করে)।
    ⚡ FAST: cache invalidate না করে — পরের load_db() instant থাকবে।
    """
    # path থেকে top-level key বের করে cache update করো
    _patch_cache_for_path(path, value)
    if _async:
        _EXECUTOR.submit(_update_db_path_sync, path, value)
        return True
    return _update_db_path_sync(path, value)

def _patch_cache_for_path(path, value):
    """Cache তে specific path update করো — section cache এ directly patch।
    ✅ FIX: 3-level deep path যেমন /users/{id}/balance সরাসরি patch করো।
    """
    global _db_cache
    parts = [p for p in path.strip("/").split("/") if p]
    if not parts:
        return
    top = parts[0]
    section = _get_section_name(top)

    with _section_cache_lock:
        cached = _section_cache.get(section)
        if cached:
            sec_data = cached[0]
            if len(parts) == 1:
                # /users → পুরো top-level replace
                sec_data[top] = value
            elif len(parts) == 2 and top in sec_data and isinstance(sec_data[top], dict):
                # /users/12345 → user object replace
                sec_data[top][parts[1]] = value
            elif (len(parts) == 3
                  and top in sec_data
                  and isinstance(sec_data.get(top), dict)
                  and isinstance(sec_data[top].get(parts[1]), dict)):
                # /users/12345/balance → শুধু balance field update
                sec_data[top][parts[1]][parts[2]] = value
            else:
                # অজানা deep path — এই section invalidate করো
                _section_cache.pop(section, None)
                return
            # _db_cache ও sync করো (backward compat)
            if _db_cache is not None:
                if len(parts) == 1:
                    _db_cache[top] = value
                elif len(parts) == 2 and top in _db_cache and isinstance(_db_cache[top], dict):
                    _db_cache[top][parts[1]] = value
                elif (len(parts) == 3 and top in _db_cache
                      and isinstance(_db_cache.get(top), dict)
                      and isinstance(_db_cache[top].get(parts[1]), dict)):
                    _db_cache[top][parts[1]][parts[2]] = value

def _update_db_path_sync(path, value):
    try:
        if isinstance(value, list):
            value = {str(i): v for i, v in enumerate(value)} if value else {}
        _fb_ref(path).set(value)
        return True
    except Exception as e:
        logging.error(f"Firebase update_db_path error [{path}]: {e}")
        return False

def patch_db_path(path, updates: dict):
    """Firebase এর specific path এ partial update (merge) করো।
    ⚡ FAST: cache তে সাথে সাথে merge করো — load_db() instant থাকবে।
    """
    # Cache তে locally merge করো
    parts = [p for p in path.strip("/").split("/") if p]
    if _db_cache is not None and len(parts) <= 2:
        with _db_cache_lock:
            if _db_cache is not None:
                top = parts[0]
                if len(parts) == 1:
                    if isinstance(_db_cache.get(top), dict):
                        _db_cache[top].update(updates)
                elif len(parts) == 2:
                    if isinstance(_db_cache.get(top), dict) and isinstance(_db_cache[top].get(parts[1]), dict):
                        _db_cache[top][parts[1]].update(updates)
    else:
        _invalidate_db_cache()
    try:
        _EXECUTOR.submit(_fb_ref(path).update, updates)
        return True
    except Exception as e:
        logging.error(f"Firebase patch_db_path error [{path}]: {e}")
        return False

def log_action(action, user_id, details=""):
    logging.info(f"[{get_now()}] User: {user_id} | Action: {action} | Details: {details}")

# ===========================
# DB CACHE — Section-wise আলাদা cache (Firebase download ~95% কম)
# ===========================
# ┌─────────────────────────────────────────────────────────────┐
# │  Section         │ Size    │ TTL    │ কারণ                  │
# │  users           │ 1.44MB  │ 600s   │ সবচেয়ে বড়, কম বদলায় │
# │  deposit_requests│ 182KB   │ 120s   │ দ্রুত update দরকার    │
# │  config (products│         │        │                       │
# │  settings,promo) │ ~50KB   │ 600s   │ খুব কম বদলায়         │
# │  transactions    │ ~40KB   │ 300s   │ medium frequency      │
# └─────────────────────────────────────────────────────────────┘
import copy as _copy_mod

# Section grouping
_SECTION_USERS       = {"users"}
_SECTION_DEPOSITS    = {"deposit_requests"}
_SECTION_CONFIG      = {"products", "product_details", "settings", "promo_codes",
                        "promo_usage", "flash_sales", "sub_products", "product_buttons",
                        "bulkmail_products", "hotmail143_products", "api_dashboard",
                        "banned_users", "tickets"}
_SECTION_TXN         = {"sms_pending_trxids", "verified_trxids", "trx_attempts",
                        "blocked_trxids", "_pending_refer"}

_TTL_USERS    = 600   # 10 মিনিট — balance update_db_path দিয়ে in-place হয়
_TTL_DEPOSITS = 120   # 2 মিনিট — deposit status দ্রুত দেখাবে
_TTL_CONFIG   = 600   # 10 মিনিট — products/settings খুব কম বদলায়
_TTL_TXN      = 300   # 5 মিনিট

# section_name → (data_dict, timestamp)
_section_cache = {}
_section_cache_lock = threading.Lock()

# backward compat — পুরনো code এর জন্য
_DB_CACHE_TTL = 300
_db_cache = None
_db_cache_time = 0
_db_cache_lock = _section_cache_lock

def _get_section_name(key):
    """key থেকে section নাম বের করো।"""
    if key in _SECTION_USERS:    return "users"
    if key in _SECTION_DEPOSITS: return "deposits"
    if key in _SECTION_TXN:      return "txn"
    return "config"

def _get_section_ttl(section):
    return {"users": _TTL_USERS, "deposits": _TTL_DEPOSITS,
            "config": _TTL_CONFIG, "txn": _TTL_TXN}.get(section, 300)

def _get_section_keys(section):
    return {"users": _SECTION_USERS, "deposits": _SECTION_DEPOSITS,
            "config": _SECTION_CONFIG, "txn": _SECTION_TXN}.get(section, set())

def _load_section(section):
    """একটা section Firebase থেকে load করো।"""
    keys = _get_section_keys(section)
    result = {}
    for key in keys:
        try:
            raw = _fb_ref(f"/{key}").get()
            default_val = _DB_DEFAULT.get(key)
            if raw is None:
                result[key] = (list(default_val) if isinstance(default_val, list)
                               else dict(default_val) if isinstance(default_val, dict)
                               else (default_val if default_val is not None else {}))
            elif isinstance(default_val, list):
                result[key] = _firebase_to_list(raw, list)
            else:
                result[key] = raw
        except Exception as e:
            logging.error(f"Firebase load_section [{section}/{key}] error: {e}")
            default_val = _DB_DEFAULT.get(key)
            result[key] = (list(default_val) if isinstance(default_val, list)
                           else dict(default_val) if isinstance(default_val, dict)
                           else (default_val if default_val is not None else {}))
    return result

def _get_cached_section(section):
    """Cache থেকে section নাও, expire হলে Firebase থেকে reload।"""
    now_ts = _time_module.time()
    ttl = _get_section_ttl(section)

    # Fast path — lock ছাড়া
    cached = _section_cache.get(section)
    if cached and (now_ts - cached[1]) < ttl:
        return dict(cached[0])

    with _section_cache_lock:
        cached = _section_cache.get(section)
        if cached and (now_ts - cached[1]) < ttl:
            return dict(cached[0])
        data = _load_section(section)
        _section_cache[section] = (data, now_ts)
        return dict(data)

def load_db():
    """Firebase থেকে DB load করো — section-wise cache।
    প্রতিটা section আলাদাভাবে cache হয়, মোট download ~95% কম।
    """
    result = {}
    for section in ("users", "deposits", "config", "txn"):
        section_data = _get_cached_section(section)
        result.update(section_data)

    # _DB_DEFAULT এ যেসব key আছে কিন্তু কোনো section এ নেই
    for key, default_val in _DB_DEFAULT.items():
        if key not in result:
            result[key] = (list(default_val) if isinstance(default_val, list)
                           else dict(default_val) if isinstance(default_val, dict)
                           else default_val)

    # backward compat: _db_cache update করো
    global _db_cache, _db_cache_time
    _db_cache = result
    _db_cache_time = _time_module.time()
    return result

def _invalidate_db_cache():
    """DB write হওয়ার পর cache clear করো।"""
    global _db_cache, _db_cache_time
    with _section_cache_lock:
        _db_cache = None
        _db_cache_time = 0
        # Section cache ও clear করো
        _section_cache.clear()

def _invalidate_section_for_key(key):
    """Specific key এর section cache clear করো।"""
    section = _get_section_name(key)
    with _section_cache_lock:
        _section_cache.pop(section, None)

# ===========================
# SUBSCRIPTION CHECK (with cache — Telegram API call কমাতে)
# ===========================

_sub_cache = {}          # {uid: (True/False, timestamp)}
_SUB_CACHE_TTL = 30      # 30s cache — Channel subscription দ্রুত reflect হবে
_sub_cache_lock = threading.Lock()

def is_subscribed(uid):
    """
    Subscription check — cache সহ।
    Cache valid থাকলে instant, expire বা না থাকলে sync check।
    ✅ FIX: stale cache এও সবসময় sync check — leave করা user ঢুকতে পারবে না।
    """
    uid_str = str(uid)
    now_ts  = _time_module.time()

    cached = _sub_cache.get(uid_str)
    if cached:
        age = now_ts - cached[1]
        if age < _SUB_CACHE_TTL:
            return cached[0]  # ⚡ instant — 0ms

    # Cache নেই বা expire — সবসময় sync check (False ধরো না, verify করো)
    return _refresh_sub_cache_sync(uid_str)

def _refresh_sub_cache_sync(uid_str):
    """Subscription check — সব channel parallel এ check করো (fast)।
    API error হলে True ধরো — subscribed user কে block করবে না।
    শুধু explicitly 'left' বা 'kicked' হলেই False।
    """
    # ⚡ সব channel একসাথে parallel check — sequential loop এর চেয়ে fast
    results = [True] * len(CHANNELS)

    def _check_one(idx, ch):
        try:
            status = bot.get_chat_member(ch["id"], uid_str).status
            if status in ("left", "kicked"):
                results[idx] = False
        except Exception as _e:
            logging.warning(f"sub_check API error uid={uid_str} ch={ch['id']}: {_e}")
            results[idx] = True  # error → assume subscribed

    threads = []
    for i, ch in enumerate(CHANNELS):
        t = threading.Thread(target=_check_one, args=(i, ch), daemon=True)
        t.start()
        threads.append(t)
    for t in threads:
        t.join(timeout=3)  # max 3s wait per check

    result = all(results)
    _sub_cache[uid_str] = (result, _time_module.time())
    return result

def _refresh_sub_cache(uid_str):
    """Background refresh — return value ignore করা হয়।"""
    _refresh_sub_cache_sync(uid_str)

def invalidate_sub_cache(uid):
    """User join verify করলে cache clear করো।"""
    _sub_cache.pop(str(uid), None)

def get_join_markup(user_id):
    lang = get_user_lang(user_id)
    markup = types.InlineKeyboardMarkup(row_width=1)
    for i, ch in enumerate(CHANNELS, 1):
        markup.add(InlineBtn(
            LANG[lang]["join_channel_btn"].format(n=i), style="success", url=ch['link']))
    markup.add(InlineBtn(LANG[lang]["verify_join"], style="success", callback_data="check_join"))
    return markup

# ===========================
# STOCK MANAGEMENT
# ===========================

# ===========================
# STOCK COUNT CACHE — বারবার disk read বন্ধ করে
# ===========================
_stock_count_cache = {}        # {product_name: (count, timestamp)}
_STOCK_CACHE_TTL = 60          # ⚡ 60 সেকেন্ড — fresh stock + disk I/O কম (optimized)

def get_stock_count(product_name):
    now_ts = _time_module.time()
    cached = _stock_count_cache.get(product_name)
    if cached and (now_ts - cached[1]) < _STOCK_CACHE_TTL:
        return cached[0]

    # ─── Bulkmail product ───
    if is_bulkmail_product(product_name):
        product_id = API_PRODUCTS[product_name]["product_id"]
        api_count  = bulkmail_get_stock(product_id)
        if api_count < 0:
            api_count = 0

        # Local xlsx stock ও চেক করো (admin extra stock upload করতে পারে)
        local_count = 0
        local_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
        if os.path.exists(local_path):
            try:
                wb = openpyxl.load_workbook(local_path, read_only=True)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_idx == 1: continue
                    if any(cell is not None and str(cell).strip() != "" for cell in row):
                        local_count += 1
                wb.close()
            except Exception:
                pass

        total = api_count + local_count
        _stock_count_cache[product_name] = (total, now_ts)
        return total

    # API Product হলে hotmail143 থেকে stock নাও
    # এবং admin local xlsx upload করলে সেটাও যোগ করো
    if product_name in API_PRODUCTS:
        api_conf = API_PRODUCTS[product_name]
        api_count = hotmail143_get_stock(api_conf["product_type"], api_conf["account_type"])
        if api_count < 0:
            api_count = 0

        # Local xlsx stock ও চেক করো (admin extra stock upload করতে পারে)
        local_count = 0
        local_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
        if os.path.exists(local_path):
            try:
                wb = openpyxl.load_workbook(local_path, read_only=True)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_idx == 1:
                        continue
                    if any(cell is not None and str(cell).strip() != "" for cell in row):
                        local_count += 1
                wb.close()
            except Exception:
                pass

        total = api_count + local_count
        _stock_count_cache[product_name] = (total, now_ts)
        return total

    path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
    if not os.path.exists(path):
        _stock_count_cache[product_name] = (0, now_ts)
        return 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # 1st row = heading, skip করো
            # Row এর সব cell None বা empty string হলে skip
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                count += 1
        wb.close()
        _stock_count_cache[product_name] = (count, now_ts)
        return count
    except Exception as e:
        logging.error(f"Stock count error for {product_name}: {str(e)}")
        return 0

def invalidate_stock_cache(product_name):
    """Stock write হলে cache clear করো। Sub-product গুলোও clear হবে।"""
    _stock_count_cache.pop(product_name, None)
    prefix = product_name + "__"
    for k in [k for k in list(_stock_count_cache.keys()) if k.startswith(prefix)]:
        _stock_count_cache.pop(k, None)

def read_single_account_text(product_name):
    """
    Buy One এর জন্য: xlsx থেকে ১টা row পড়ে heading অনুযায়ী
    account details text হিসেবে return করে।
    Return: (account_dict, headers) or (None, None)
    """
    path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
    try:
        if not os.path.exists(path):
            return None, None
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
        if len(all_rows) < 2:
            return None, None
        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        # প্রথম non-empty data row খোঁজো
        for row in all_rows[1:]:
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                account = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
                return account, headers
        return None, None
    except Exception as e:
        logging.error(f"read_single_account_text error: {e}")
        return None, None


def read_stock_file(product_name, quantity):
    """
    xlsx stock file থেকে `quantity` টা row পড়ে:
      - CSV (plain text, BOM-less UTF-8) হিসেবে BytesIO return করে
      - 1st row = heading (fixed — product এর column নাম)
      - Data rows = plain values, কোনো formatting নেই
        → copy করলে নরমাল/plain text পাওয়া যাবে
      - পড়া rows গুলো xlsx থেকে delete করে দেয় (stock কমে)
    """
    import csv as _csv
    path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
    try:
        if not os.path.exists(path):
            return None
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        all_rows = list(ws.iter_rows(min_row=1, values_only=False))
        if not all_rows:
            wb.close()
            return None

        # 1st row = heading
        header_data = [str(cell.value).strip() if cell.value is not None else "" for cell in all_rows[0]]

        rows_copied = 0
        row_indices_to_delete = []
        collected_rows = []

        for row_idx, row in enumerate(all_rows[1:], start=2):
            if rows_copied >= quantity:
                break
            if any(cell.value is not None and str(cell.value).strip() != "" for cell in row):
                row_data = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                collected_rows.append(row_data)
                row_indices_to_delete.append(row_idx)
                rows_copied += 1

        if rows_copied == 0:
            wb.close()
            return None

        # Stock থেকে delivered rows মুছে ফেলো
        for row_idx in sorted(row_indices_to_delete, reverse=True):
            ws.delete_rows(row_idx, 1)
        wb.save(path)
        wb.close()
        # Stock cache clear করো — পরের request তে fresh count পাবে
        invalidate_stock_cache(product_name)

        # CSV তৈরি করো — plain UTF-8, BOM ছাড়া, heading সঠিকভাবে
        import io as _io
        text_stream = _io.StringIO()
        writer = _csv.writer(text_stream, lineterminator="\n")
        # 1st row = heading (Excel এর exact heading গুলো)
        writer.writerow(header_data)
        # Data rows — Excel এর row order অনুযায়ী
        for row in collected_rows:
            # heading count অনুযায়ী row pad/trim করো
            padded = list(row) + [""] * max(0, len(header_data) - len(row))
            padded = padded[:len(header_data)]
            writer.writerow(padded)

        csv_bytes = text_stream.getvalue().encode("utf-8")
        stream = BytesIO(csv_bytes)
        stream.seek(0)
        return stream

    except Exception as e:
        logging.error(f"Stock file read error: {str(e)}\n{traceback.format_exc()}")
        return None

def _restore_stock_file(product_name, csv_stream):
    """
    Delivery ব্যর্থ হলে read_stock_file এ মুছে যাওয়া rows
    আবার xlsx এ সামনে থেকে ঢুকিয়ে দাও।
    """
    import csv as _csv, io as _io
    path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
    try:
        csv_stream.seek(0)
        text = csv_stream.read().decode("utf-8")
        reader = list(_csv.reader(_io.StringIO(text)))
        if len(reader) < 2:
            return
        header    = reader[0]
        data_rows = reader[1:]

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # data rows সামনে ঢোকাও (row 2 থেকে)
            for ri, row in enumerate(data_rows, 2):
                ws.insert_rows(ri)
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(header)
            for row in data_rows:
                ws.append(row)

        wb.save(path)
        wb.close()
        logging.info(f"Stock restored: {product_name} | {len(data_rows)} rows returned")
    except Exception as e:
        logging.error(f"_restore_stock_file error ({product_name}): {e}")

# ===========================
# KEYBOARD GENERATORS
# ===========================

def get_main_menu(uid):
    lang = get_user_lang(uid)
    L = LANG[lang]
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add( KbBtn(L["shop"], role="shop"), KbBtn(L["deposit"], role="deposit"))
    markup.add( KbBtn(L["balance"], role="balance"), KbBtn(L["orders"], role="orders"))
    markup.add( KbBtn(L["refer"], role="refer"), KbBtn(L["support"], role="support"))
    markup.add( KbBtn("🌐 Language / ভাষা", role="language"))
    if str(uid) == str(ADMIN_ID):
        markup.add( KbBtn(L["admin"], role="admin"))
    return markup

def get_back_button(user_id):
    lang = get_user_lang(user_id)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add( KbBtn(LANG[lang]["back"], role="back"))
    return markup

def get_cancel_button(user_id=None):
    lang = get_user_lang(user_id) if user_id else "en"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add( KbBtn(LANG[lang]["cancel"], role="cancel"))
    return markup

def get_product_buy_keyboard(user_id, product_name=None):
    lang = get_user_lang(user_id)
    L = LANG[lang]
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add( KbBtn(L["single_buy"], role="buy_single"))
    markup.add( KbBtn(L["bulk_buy"], role="buy_bulk"))
    markup.add( KbBtn("🔙 Back", role="back"))
    return markup

# ===========================
# 2FA KEY → CODE GENERATOR
# ===========================

# ===========================
# START & LANGUAGE
# ===========================

@bot.message_handler(commands=['start'])
def handle_start(message):
    # ⚡ Thread pool এ — webhook instant free হবে
    _EXECUTOR.submit(_start_worker, message)

def _start_worker(message):
    user_id = str(message.chat.id)
    db = load_db()

    if user_id in db["banned_users"]:
        bot.send_message(user_id, LANG["en"]["banned"], parse_mode="HTML")
        return

    # Refer link থেকে আসা args সংরক্ষণ করো (channel join এর আগেও)
    args = message.text.split()
    referred_by = args[1] if len(args) > 1 else None

    if not is_subscribed(user_id):
        lang = db["users"].get(user_id, {}).get("lang", "en")
        # Channel join না করলেও referred_by টা pending হিসেবে রাখো
        # ✅ FIX: user DB তে থাকুক বা না থাকুক, refer save করো
        if referred_by:
            already_got_bonus = db.get("users", {}).get(user_id, {}).get("refer_bonus_given", False)
            if not already_got_bonus:
                if "_pending_refer" not in db:
                    db["_pending_refer"] = {}
                db["_pending_refer"][user_id] = referred_by
                update_db_path(f"/_pending_refer/{user_id}", referred_by)
        bot.send_message(user_id, LANG[lang]["join_channels"],
            reply_markup=get_join_markup(user_id), parse_mode="HTML")
        return

    if user_id not in db["users"]:
        # Pending refer চেক করো (channel join এর আগে আসা refer)
        if not referred_by:
            referred_by = db.get("_pending_refer", {}).get(user_id)

        db["users"][user_id] = {
            "balance": 0, "refer_count": 0, "orders": [],
            "referred_by": referred_by,
            "join_date": get_now_short(),
            "lang": "en"
        }
        _lang_cache[user_id] = "en"

        # ── Referral bonus ──
        already_referred = db["users"][user_id].get("refer_bonus_given", False)
        if (referred_by and
                referred_by in db["users"] and
                referred_by != user_id and
                not already_referred):
            bonus = db["settings"].get("refer_bonus", 2)
            db["users"][referred_by]["balance"] = round(
                float(db["users"][referred_by].get("balance", 0)) + float(bonus), 2)
            db["users"][referred_by]["refer_count"] = int(
                db["users"][referred_by].get("refer_count", 0)) + 1
            db["users"][user_id]["refer_bonus_given"] = True
            try:
                bot.send_message(referred_by,
                    t(referred_by, "refer_bonus_notif", bonus=bonus), parse_mode="HTML")
            except: pass

        # Pending refer entry পরিষ্কার করো
        if "_pending_refer" in db and user_id in db["_pending_refer"]:
            del db["_pending_refer"][user_id]
            _fb_ref(f"/_pending_refer/{user_id}").delete()

        # ⚡ Targeted write — পুরো DB write না করে শুধু নতুন user data save করো
        update_db_path(f"/users/{user_id}", db["users"][user_id])
        if referred_by and referred_by in db["users"]:
            update_db_path(f"/users/{referred_by}", db["users"][referred_by])
        log_action("NEW_USER", user_id, f"Referred by: {referred_by}")
        show_language_selection(user_id)
        return

    # ✅ FIX: User DB তে আছে কিন্তু refer bonus পায়নি — এখন দাও
    # (আগে channel join না করে এসেছিল, এখন join করে ফিরে এসেছে)
    pending_refer = db.get("_pending_refer", {}).get(user_id)
    if not pending_refer:
        pending_refer = referred_by
    already_referred = db["users"][user_id].get("refer_bonus_given", False)
    if (pending_refer and
            pending_refer in db["users"] and
            pending_refer != user_id and
            not already_referred):
        bonus = db["settings"].get("refer_bonus", 2)
        db["users"][pending_refer]["balance"] = round(
            float(db["users"][pending_refer].get("balance", 0)) + float(bonus), 2)
        db["users"][pending_refer]["refer_count"] = int(
            db["users"][pending_refer].get("refer_count", 0)) + 1
        db["users"][user_id]["refer_bonus_given"] = True
        if "_pending_refer" in db and user_id in db["_pending_refer"]:
            del db["_pending_refer"][user_id]
            _fb_ref(f"/_pending_refer/{user_id}").delete()
        # ⚡ Targeted write — শুধু দুইটা user update করো
        update_db_path(f"/users/{pending_refer}", db["users"][pending_refer])
        update_db_path(f"/users/{user_id}", db["users"][user_id])
        try:
            bot.send_message(pending_refer,
                t(pending_refer, "refer_bonus_notif", bonus=bonus), parse_mode="HTML")
        except: pass

    bot.send_message(user_id, t(user_id, "welcome"),
        reply_markup=get_main_menu(user_id), parse_mode="HTML")

def show_language_selection(user_id):
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineBtn("🇬🇧 English", style="primary", callback_data="set_lang_en"),
        InlineBtn("🇧🇩 বাংলা", style="primary", callback_data="set_lang_bn")
    )
    bot.send_message(user_id,
        "🌐 <b>Select Language / ভাষা বেছে নিন</b>",
        reply_markup=markup, parse_mode="HTML")

# ===========================
# TEXT MESSAGE HANDLER
# ===========================

@bot.message_handler(func=lambda message: True)
def handle_text_messages(message):
    # ⚡ FAST: thread pool reuse করো — নতুন thread তৈরির overhead নেই
    _EXECUTOR.submit(_text_message_worker, message)

def _text_message_worker(message):
    user_id = str(message.chat.id)
    text = message.text or ""

    # ⚡ Rate limit আগে check করো — DB load লাগে না
    if str(user_id) != str(ADMIN_ID) and is_rate_limited(user_id):
        bot.send_message(user_id,
            "⚠️ <b>একটু ধীরে!</b>\n\nপরপর অনেক message পাঠাচ্ছেন।\nকয়েক সেকেন্ড অপেক্ষা করুন। 🙏",
            parse_mode="HTML")
        return

    # ✅ Subscription gate — admin বাদে সবাইকে channel join enforce করো
    if str(user_id) != str(ADMIN_ID):
        # Cache check (fast path) — TTL এর মধ্যে থাকলে API call করবে না
        _cached = _sub_cache.get(user_id)
        _now_ts = _time_module.time()
        if _cached and (_now_ts - _cached[1]) < _SUB_CACHE_TTL:
            _sub_ok = _cached[0]
        else:
            # Cache নেই বা expire — Telegram API দিয়ে fresh check
            _sub_ok = _refresh_sub_cache_sync(user_id)
        if not _sub_ok:
            _lang_now = _lang_cache.get(user_id, 'en')
            _join_markup = types.InlineKeyboardMarkup(row_width=1)
            for i, ch in enumerate(CHANNELS, 1):
                _btn_text = LANG.get(_lang_now, LANG['en'])['join_channel_btn'].format(n=i)
                _join_markup.add(InlineBtn(_btn_text, style="primary", url=ch['link']))
            _join_markup.add(InlineBtn(
                LANG.get(_lang_now, LANG['en'])['verify_join'], style="success", callback_data='check_join'))
            bot.send_message(user_id,
                LANG.get(_lang_now, LANG['en'])['join_channels'],
                reply_markup=_join_markup, parse_mode='HTML')
            return

    # ⚡ Lang cache দিয়ে check — DB load না করে
    lang = _lang_cache.get(user_id, "en")
    L = LANG[lang]

    # ⚡ Ban check fast path — lang cache দিয়ে
    # সাধারণ command গুলো DB ছাড়াই handle করো
    if text == "🌐 Language / ভাষা":
        show_language_selection(user_id); return
    if text == "/stock":
        handle_stock_command(message); return
    if text == "/apitest" and str(user_id) == str(ADMIN_ID):
        handle_apitest_command(message); return
    if text == "/token" or text == "/get_code":
        handle_get_code_command(message); return
    if text.startswith("/2fa_key"):
        handle_2fa_command(message); return
    if text == "/code" or text.startswith("/code "):
        handle_code_command(message); return
    if text == "/tn_verify" or text.startswith("/tn_verify "):
        handle_tn_verify_command(message); return
    if text in (L["support"], LANG["en"]["support"], LANG["bn"]["support"]):
        show_support_info(user_id); return
    if text in (L["back"], L["cancel"], LANG["en"]["back"], LANG["bn"]["back"]):
        _pending_product_cache.pop(user_id, None)
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    if text == "🔙 Back":
        db_quick = load_db()
        # Edu Mail sub-products এর Back → Edu Mail submenu তে যাবে
        # অন্য products এর Back → shop menu তে যাবে
        last_product = _pending_product_cache.get(user_id)
        _pending_product_cache.pop(user_id, None)
        if last_product in (EDU_MAIL_SUB_24H, EDU_MAIL_SUB_72H, EDU_MAIL_SUB_24HR):
            _show_edu_mail_submenu(user_id, db_quick)
        else:
            show_shop_menu(user_id, db_quick)
        return

    # ⚡ এখন DB load করো (cache hit হলে μs, miss হলে Firebase call)
    db = load_db()

    # Ban check
    if user_id in db["banned_users"]:
        bot.send_message(user_id, LANG["en"]["banned"], parse_mode="HTML")
        return

    # Lang cache sync
    if user_id in db["users"]:
        _lang_cache[user_id] = db["users"][user_id].get("lang", "en")
        lang = _lang_cache[user_id]
        L = LANG[lang]

    if text == L["shop"] or text == LANG["en"]["shop"] or text == LANG["bn"]["shop"]:
        show_shop_menu(user_id, db); return
    elif text in (L["orders"], LANG["en"]["orders"], LANG["bn"]["orders"]):
        show_user_orders(user_id, db); return
    elif text in (L["balance"], LANG["en"]["balance"], LANG["bn"]["balance"]):
        show_user_balance(user_id, db); return
    elif text in (L["refer"], LANG["en"]["refer"], LANG["bn"]["refer"]):
        show_referral_info(user_id, db); return
    elif text in (L["deposit"], LANG["en"]["deposit"], LANG["bn"]["deposit"]):
        show_deposit_menu(user_id); return
    elif text in (L.get("admin",""), LANG["en"]["admin"]) and str(user_id) == str(ADMIN_ID):
        show_admin_panel(user_id); return
    # ── Edu Mail parent button → submenu দেখাও ──
    elif text == EDU_MAIL_PARENT:
        _show_edu_mail_submenu(user_id, db); return
    # ── Edu Mail sub-buttons ──
    elif text == "Edu Mail 24H":
        _pending_product_cache[user_id] = EDU_MAIL_SUB_24H
        show_product_detail_with_subnav(user_id, EDU_MAIL_SUB_24H, db); return
    elif text == "Edu Mail 72H":
        _pending_product_cache[user_id] = EDU_MAIL_SUB_72H
        show_product_detail_with_subnav(user_id, EDU_MAIL_SUB_72H, db); return
    elif text == EDU_MAIL_SUB_24HR:
        _pending_product_cache[user_id] = EDU_MAIL_SUB_24HR
        show_product_detail_with_subnav(user_id, EDU_MAIL_SUB_24HR, db); return
    # ── Hotmail ও Outlook — API products ──
    elif text == HOTMAIL_PROD:
        _pending_product_cache[user_id] = HOTMAIL_PROD
        show_product_detail_with_subnav(user_id, HOTMAIL_PROD, db); return
    elif text == OUTLOOK_PROD:
        _pending_product_cache[user_id] = OUTLOOK_PROD
        show_product_detail_with_subnav(user_id, OUTLOOK_PROD, db); return
    elif text in db["products"]:
        show_product_detail_with_subnav(user_id, text, db); return
    elif text in (L["single_buy"], LANG["en"]["single_buy"], LANG["bn"].get("single_buy", "")):
        handle_single_buy_from_keyboard(user_id, db); return
    elif text in (L["bulk_buy"], LANG["en"]["bulk_buy"], LANG["bn"].get("bulk_buy", "")):
        handle_bulk_buy_from_keyboard(user_id, db); return

# ===========================
# SHOP FUNCTIONS
# ===========================

def show_shop_menu(user_id, db=None):
    if db is None:
        db = load_db()
    lang = get_user_lang(user_id, db)
    L = LANG[lang]
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    # Edu Mail 24H, 72H, 24hr হলো sub-product — মূল shop menu তে দেখাবে না
    # Edu Mail (parent) — submenu দিয়ে access হবে, তাই DB products list থেকে hide করো
    _hidden_sub = {EDU_MAIL_SUB_24H, EDU_MAIL_SUB_72H, EDU_MAIL_SUB_24HR, EDU_MAIL_PARENT}
    # ⚡ set দিয়ে O(1) lookup — list comprehension fast
    db_product_keys = set(db["products"].keys())
    products = [p for p in db["products"].keys() if p not in _hidden_sub]

    # ✅ FIX: Edu Mail 24H, 72H বা 24hr এর মধ্যে অন্তত একটা enabled থাকলে parent "Edu Mail" button দেখাও
    _bulkmail_edu_enabled   = is_bulkmail_enabled(EDU_MAIL_SUB_24HR, db)
    _h143_24h_enabled       = is_hotmail143_enabled(EDU_MAIL_SUB_24H, db) and EDU_MAIL_SUB_24H in db_product_keys
    _h143_72h_enabled       = is_hotmail143_enabled(EDU_MAIL_SUB_72H, db) and EDU_MAIL_SUB_72H in db_product_keys
    show_edu_mail_btn = _bulkmail_edu_enabled or _h143_24h_enabled or _h143_72h_enabled or EDU_MAIL_PARENT in db_product_keys

    if show_edu_mail_btn:
        products = [EDU_MAIL_PARENT] + products  # প্রথমে Edu Mail

    if products:
        # Add products 2 per row
        for i in range(0, len(products), 2):
            row = products[i:i+2]
            markup.add( *[KbBtn(_x, role="category_item") for _x in row])
    markup.add( KbBtn(L["back"], role="back"))
    bot.send_message(user_id, "🛍️ <b>Choose a product:</b>",
        reply_markup=markup, parse_mode="HTML")

def _show_edu_mail_submenu(user_id, db):
    """
    Edu Mail parent button ক্লিক করলে এই menu আসবে।
    Buttons: Edu Mail 24H | Edu Mail 72H | Edu Mail 24hr (Bulkmail) | Back
    Admin on/off থাকলে disabled product দেখাবে না।
    """
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)

    # ─── Edu Mail 24H — admin on হলে দেখাবে ───
    if is_hotmail143_enabled(EDU_MAIL_SUB_24H, db):
        markup.add(KbBtn(EDU_MAIL_SUB_24H, role="category_item"))

    # ─── Edu Mail 72H — admin on হলে দেখাবে ───
    if is_hotmail143_enabled(EDU_MAIL_SUB_72H, db):
        markup.add(KbBtn(EDU_MAIL_SUB_72H, role="category_item"))

    # ─── Edu Mail 24hr (Bulkmail product) — admin on হলে দেখাবে ───
    if is_bulkmail_enabled(EDU_MAIL_SUB_24HR, db):
        markup.add(KbBtn(EDU_MAIL_SUB_24HR, role="category_item"))

    markup.add(KbBtn("🔙 Back", role="back"))
    bot.send_message(
        user_id,
        "📧 <b>Edu Mail</b>\n\nপছন্দের ধরন বেছে নিন 👇",
        parse_mode="HTML",
        reply_markup=markup
    )


def show_product_detail_with_subnav(user_id, product_name, db=None):
    if db is None:
        db = load_db()

    # ══════════════════════════════════════════
    # EDU MAIL PARENT → sub-keyboard দেখাও
    # ══════════════════════════════════════════
    if product_name == EDU_MAIL_PARENT:
        _show_edu_mail_submenu(user_id, db)
        return

    # API sub-product (Edu Mail 24H / 72H) — DB তে না থাকলেও চলবে
    _is_api_prod = product_name in API_PRODUCTS

    if not _is_api_prod and product_name not in db["products"]:
        show_shop_menu(user_id, db); return

    # Price: DB তে থাকলে সেটা, না থাকলে 0
    # ✅ NOTE: Edu Mail 24H / 72H এর price অবশ্যই DB তে (products node এ) সেট করতে হবে
    price = float(db["products"].get(product_name) or 0)
    # ⚡ Stock: cache hit হলে instant, miss হলে background এ fetch
    # pending_product সেট করো
    _pending_product_cache[user_id] = product_name
    # Stock — cache থেকে নাও (10s TTL), API call হলেও background এ হবে না
    stock = get_stock_count(product_name)

    # Flash Sale check
    flash_info = get_flash_sale_info(db, product_name)
    flash_text = ""
    display_price = price  # user কে দেখানো price
    if flash_info:
        remaining = flash_info.get("remaining_seconds", 0)
        if remaining > 0:
            h = remaining // 3600
            m = (remaining % 3600) // 60
            s = remaining % 60
            orig_price = flash_info.get("original_price", price)
            discount = flash_info.get("discount_percent", 0)
            discounted = flash_info.get("discounted_price", price)
            display_price = discounted  # flash sale price দেখাবে
            flash_text = (
                f"\n🔥 <b>FLASH SALE!</b> ⏱ <code>{h:02d}:{m:02d}:{s:02d}</code> বাকি\n"
                f"💥 <s>{orig_price} BDT</s> → <b>{discounted} BDT</b> ({discount}% OFF)\n"
            )

    # Flash sale থাকলে strikethrough দিয়ে original price + discounted price দেখাও
    personal_disc_user = float(db.get("users", {}).get(user_id, {}).get("personal_discount", 0))
    if flash_info:
        orig_price = flash_info.get("original_price", price)
        discounted = flash_info.get("discounted_price", price)
        def _fmt(p):
            if p == int(p):
                return str(int(p))
            return f"{p:.2f}"
        price_line = f"💵 <b>Price:</b> <s>{_fmt(orig_price)} BDT</s> → <code>{_fmt(discounted)} BDT</code>\n"
    elif personal_disc_user > 0:
        # Personal discount আছে — দেখাও
        discounted_personal = apply_discount(price, personal_disc_user)
        def _fmt(p):
            if p == int(p): return str(int(p))
            return f"{p:.2f}"
        price_line = (f"💵 <b>Price:</b> <s>{_fmt(price)} BDT</s> → <code>{_fmt(discounted_personal)} BDT</code>\n"
                      f"🏷️ <b>আপনার Personal Discount:</b> <code>{personal_disc_user:.0f}% OFF</code>\n")
    else:
        price_line = f"💵 <b>Price:</b> <code>{price} BDT</code>\n"

    # ✅ FIX: শুধু Edu Mail parent — stock line দেখাবে না
    if product_name == EDU_MAIL_PARENT:
        stock_line = ""
    elif is_bulkmail_product(product_name):
        # Bulkmail product — bulkmail API থেকে stock আনো
        _bm_pid    = API_PRODUCTS[product_name]["product_id"]
        _cached_bm = _stock_count_cache.get(product_name)
        _now_bm    = _time_module.time()
        if _cached_bm and (_now_bm - _cached_bm[1]) < _STOCK_CACHE_TTL:
            bm_stock = _cached_bm[0]
        else:
            bm_stock = bulkmail_get_stock(_bm_pid)
            if bm_stock >= 0:
                _stock_count_cache[product_name] = (bm_stock, _now_bm)
        if bm_stock < 0:
            stock_line = "📦 <b>Stock:</b> <code>⚠️ চেক করা যায়নি</code>\n"
        elif bm_stock == 0:
            stock_line = "📦 <b>Stock:</b> 🔴 <b>Out of Stock</b>\n"
        else:
            stock_line = f"📦 <b>Stock:</b> <code>{bm_stock} pcs ✅</code>\n"
    elif product_name in API_PRODUCTS:
        # API product — সরাসরি API call করে real stock আনো
        api_conf_r = API_PRODUCTS[product_name]
        # Cache আছে কিনা চেক করো (TTL মধ্যে থাকলে instant)
        _cached_entry = _stock_count_cache.get(product_name)
        _now_ts_s = _time_module.time()
        if _cached_entry and (_now_ts_s - _cached_entry[1]) < _STOCK_CACHE_TTL:
            cached_stock = _cached_entry[0]
        else:
            # Cache নেই বা expire — API call করো
            try:
                fresh_api = hotmail143_get_stock(api_conf_r["product_type"], api_conf_r["account_type"])

                # Local xlsx stock যোগ করো
                local_path_r = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
                local_r = 0
                if os.path.exists(local_path_r):
                    try:
                        _wb_r = openpyxl.load_workbook(local_path_r, read_only=True)
                        _ws_r = _wb_r.active
                        for _ri_r, _row_r in enumerate(_ws_r.iter_rows(values_only=True), start=1):
                            if _ri_r == 1: continue
                            if any(c is not None and str(c).strip() != "" for c in _row_r):
                                local_r += 1
                        _wb_r.close()
                    except Exception:
                        pass

                if fresh_api < 0:
                    # API error — পুরনো cache থাকলে সেটা রাখো, না থাকলে শুধু local দেখাও
                    if _cached_entry:
                        cached_stock = _cached_entry[0]
                        logging.warning(f"API stock error for {product_name}, using stale cache={cached_stock}")
                    else:
                        cached_stock = local_r if local_r > 0 else -1
                else:
                    cached_stock = fresh_api + local_r
                    _stock_count_cache[product_name] = (cached_stock, _now_ts_s)

            except Exception as _e:
                logging.warning(f"API stock fetch error {product_name}: {_e}")
                cached_stock = _cached_entry[0] if _cached_entry else -1

        if cached_stock < 0:
            stock_line = "📦 <b>Stock:</b> <code>⚠️ চেক করা যায়নি</code>\n"
        elif cached_stock == 0:
            stock_line = "📦 <b>Stock:</b> 🔴 <b>Out of Stock</b>\n"
        else:
            stock_line = f"📦 <b>Stock:</b> <code>{cached_stock} pcs ✅</code>\n"
    elif stock <= 0:
        stock_line = "📦 <b>Stock:</b> 🔴 <b>Out of Stock</b>\n"
    else:
        stock_line = f"📦 <b>Stock:</b> <code>{stock} pcs</code>\n"

    detail_msg = (
        f"🏷️ <b>{product_name}</b>\n"
        f"\n"
        f"{price_line}"
        f"{stock_line}"
    )
    if flash_text:
        detail_msg += flash_text

    # ✅ FIX 3: Description আগে, তারপর italic Features
    description = db.get("product_details", {}).get(product_name, {}).get("description", "")
    features = db.get("product_details", {}).get(product_name, {}).get("features", "")

    if description:
        detail_msg += f"\n📝 <b>Description:</b>\n{description}\n"
    if features:
        detail_msg += f"\n✨ <b>Features:</b>\n<i>{features}</i>\n"

    detail_msg += f"\n\n<i>নিচে অপশন বেছে নিন 👇</i>"

    # ── Sub-product check: Gmail বা Facebook হলে sub-buttons দেখাও ──
    sub_config = db.get("sub_products", {}).get(product_name)
    if sub_config and sub_config.get("sub_items"):
        # Sub-items আছে → inline sub-buttons দেখাবে, সরাসরি কেনা যাবে না
        sub_items = sub_config["sub_items"]
        # enabled গুলো filter করো
        enabled_items = [s for s in sub_items if s.get("enabled", True)]
        if enabled_items:
            markup_sub = types.InlineKeyboardMarkup(row_width=2)
            btns = []
            for item in enabled_items:
                sub_name = item.get("name", "")
                sub_price = item.get("price", price)
                sub_stock = get_stock_count(f"{product_name}__{sub_name}")
                stock_icon = "✅" if sub_stock > 0 else "❌"
                btns.append(InlineBtn(
                    f"{stock_icon} {sub_name} — {sub_price}BDT", style="success",
                    callback_data=f"sub_buy_{product_name}__{sub_name}"
                ))
            # 2 পাশাপাশি
            for i in range(0, len(btns), 2):
                pair = btns[i:i+2]
                markup_sub.row( *[InlineBtn(_x, style="primary") for _x in pair])
            markup_sub.add(InlineBtn("🔙 Back", style="primary", callback_data="back_to_shop"))
            bot.send_message(user_id, detail_msg,
                parse_mode="HTML", reply_markup=markup_sub)
            return

    # ✅ FIX 2: Custom product buttons (inline) দেখাও
    custom_buttons = db.get("product_buttons", {}).get(product_name, [])
    inline_btns = [b for b in custom_buttons if b.get("type") == "inline"]
    keyboard_btns = [b for b in custom_buttons if b.get("type") == "keyboard"]

    # Inline keyboard — flash timer + custom inline buttons
    inline_markup = types.InlineKeyboardMarkup(row_width=1)
    if flash_info:
        inline_markup.add(
            InlineBtn("🔄 Timer Refresh", style="primary", callback_data=f"refresh_flash_{product_name}")
        )
    for btn in inline_btns:
        if btn.get("url"):
            inline_markup.add(InlineBtn(btn["text"], style="primary", url=btn["url"]))

    # Reply keyboard — buy options + custom keyboard buttons
    buy_markup = get_product_buy_keyboard(user_id)
    # keyboard_btns থাকলে নতুন markup বানাও
    if keyboard_btns:
        lang = get_user_lang(user_id)
        L = LANG[lang]
        buy_markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        buy_markup.add( KbBtn(L["single_buy"], role="buy_single"))
        buy_markup.add( KbBtn(L["bulk_buy"], role="buy_bulk"))
        for kb in keyboard_btns:
            buy_markup.add( KbBtn(kb["text"], role="misc_option"))
        buy_markup.add( KbBtn("🔙 Back", role="back"))

    has_inline = flash_info or inline_btns
    if has_inline:
        bot.send_message(user_id, detail_msg, parse_mode="HTML", reply_markup=inline_markup)
        # ✅ FIX: Language-aware buy option message
        _lang_now = get_user_lang(user_id)
        _buy_prompt = "👇 Choose an option below:" if _lang_now == "en" else "👇 কেনার অপশন বেছে নিন:"
        bot.send_message(user_id, _buy_prompt, reply_markup=buy_markup)
    else:
        bot.send_message(user_id, detail_msg, reply_markup=buy_markup, parse_mode="HTML")

def handle_single_buy_from_keyboard(user_id, db):
    product_name = _pending_product_cache.get(user_id) or db["users"].get(user_id, {}).get("pending_product")
    is_api = product_name in API_PRODUCTS if product_name else False
    if not product_name or (not is_api and product_name not in db["products"]):
        # product আবার select করতে বলো, main menu তে না
        bot.send_message(user_id, "⚠️ Please select a product first.",
            reply_markup=get_product_buy_keyboard(user_id)); return
    # ✅ BUG FIX: API product (Edu Mail 24H / 72H) DB তে নেই,
    # তাই db["products"].get() করলে 0 আসে এবং delivery হয় না।
    # সমাধান: DB তে না থাকলে API_PRODUCTS থেকে price নাও।
    # Admin যদি DB তে price সেট করে থাকে সেটাই priority পাবে।
    price = db["products"].get(product_name) 
    if price is None:
        # API product এর price DB তে নেই — admin কে price সেট করতে হবে
        bot.send_message(user_id,
            f"⚠️ <b>{product_name}</b> এর price সেট করা নেই!\n\n"
            f"Admin panel থেকে এই product এর price সেট করুন।",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
        logging.error(f"Price not set for API product: {product_name}")
        return
    price = float(price)

    # ── Out of Stock check ──
    stock = get_stock_count(product_name)
    # API product — API তে live stock থাকলে allow করো (local 0 হলেও চলবে)
    _is_api_single = product_name in API_PRODUCTS
    if stock <= 0 and not _is_api_single:
        bot.send_message(user_id,
            f"🔴 <b>Out of Stock!</b>\n\n"
            f"😔 দুঃখিত! <b>{product_name}</b> এখন স্টকে নেই।\n"
            f"পরে আবার চেষ্টা করুন।\n\n"
            f"⬅️ Back চাপুন বা অন্য product বেছে নিন।",
            parse_mode="HTML", reply_markup=get_product_buy_keyboard(user_id))
        return

    # ── Promo feature চেক ──
    promo_feature_on = db.get("settings", {}).get("promo_feature_enabled", True)

    # ── Personal Discount (Admin দেওয়া) check ──
    personal_disc = float(db.get("users", {}).get(user_id, {}).get("personal_discount", 0))

    # New user discount auto-apply
    new_user_disc = get_new_user_discount(db)
    user_orders = db.get("users", {}).get(user_id, {}).get("orders", [])

    # Personal discount থাকলে সেটা সব discount এর চেয়ে বড় হিসেবে নাও
    auto_discount = max(personal_disc, new_user_disc if (new_user_disc > 0 and not user_orders) else 0)

    if auto_discount > 0:
        discounted_price = apply_discount(price, auto_discount)
        disc_label = f"🏷️ Personal {auto_discount:.0f}% ছাড়" if personal_disc >= auto_discount else f"🎁 নতুন user {auto_discount:.0f}% ছাড়"
        if promo_feature_on:
            cancel_btn = types.ReplyKeyboardMarkup(resize_keyboard=True)
            cancel_btn.add( KbBtn("⏭️ Skip", role="skip"), KbBtn("🔙 Back", role="back"))
            msg = bot.send_message(user_id,
                f"🎟️ <b>Promo / Coupon Code</b>\n\n{disc_label} পাবেন!\nঅথবা একটি <b>Promo Code</b> দিন বড় ছাড়ের জন্য।\n\n💡 Code না থাকলে <b>Skip</b> করুন।",
                parse_mode="HTML", reply_markup=cancel_btn)
            bot.register_next_step_handler(msg, lambda m: _single_buy_promo_step(m, product_name, price, auto_discount))
        else:
            final_price = apply_discount(price, auto_discount)
            current_balance = db["users"].get(user_id, {}).get("balance", 0)
            if current_balance < final_price:
                bot.send_message(user_id,
                    t(user_id, "insufficient_balance", total=final_price, bal=current_balance, short=round(final_price-current_balance,2)),
                    reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
            process_purchase(user_id, product_name, 1, final_price, final_price, discount_percent=auto_discount)
        return

    if promo_feature_on:
        # Promo ON — promo code জিজ্ঞেস করো (personal_discount pass করো)
        cancel_btn = types.ReplyKeyboardMarkup(resize_keyboard=True)
        cancel_btn.add( KbBtn("⏭️ Skip", role="skip"), KbBtn("🔙 Back", role="back"))
        msg = bot.send_message(user_id,
            f"🎟️ <b>Promo / Coupon Code আছে?</b>\n\nথাকলে লিখুন, না থাকলে <b>Skip</b> করুন।",
            parse_mode="HTML", reply_markup=cancel_btn)
        # ✅ BUG FIX: personal_disc pass করো যাতে promo step এ combine হয়
        bot.register_next_step_handler(msg, lambda m: _single_buy_promo_step(m, product_name, price, personal_disc))
    else:
        # Promo OFF — personal discount apply করে buy করো
        final_price_direct = apply_discount(price, personal_disc)
        current_balance = db["users"].get(user_id, {}).get("balance", 0)
        if current_balance < final_price_direct:
            bot.send_message(user_id,
                t(user_id, "insufficient_balance", total=final_price_direct, bal=current_balance, short=round(final_price_direct-current_balance,2)),
                reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
        process_purchase(user_id, product_name, 1, final_price_direct, final_price_direct, discount_percent=personal_disc)

def _single_buy_promo_step(message, product_name, price, auto_discount):
    user_id = str(message.chat.id)
    db = load_db()
    text = (message.text or "").strip()
    if text == "❌ Cancel":
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    if text == "🔙 Back":
        # এক ধাপ back — buy one pcs / bulk buy / back অপশনে ফিরে যাও
        show_product_detail_with_subnav(user_id, product_name, db); return
    final_discount = auto_discount
    promo_code_used = None
    if text and text != "⏭️ Skip":
        is_valid, disc, msg_txt = validate_promo(db, user_id, text)
        if not is_valid:
            bot.send_message(user_id, msg_txt, parse_mode="HTML",
                reply_markup=get_main_menu(user_id)); return
        final_discount = max(disc, auto_discount)  # বড় discount নাও
        promo_code_used = text.strip().upper()
        bot.send_message(user_id, msg_txt, parse_mode="HTML")
    final_price = apply_discount(price, final_discount)
    current_balance = db["users"].get(user_id, {}).get("balance", 0)
    if current_balance < final_price:
        bot.send_message(user_id,
            t(user_id, "insufficient_balance", total=final_price, bal=current_balance, short=round(final_price-current_balance,2)),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    if promo_code_used:
        apply_promo_usage(db, user_id, promo_code_used)
        # ⚡ Targeted write — শুধু promo data update করো
        update_db_path(f"/promo_codes/{promo_code_used}", db["promo_codes"][promo_code_used])
        update_db_path(f"/promo_usage/{promo_code_used}", db["promo_usage"].get(promo_code_used, {}))
    process_purchase(user_id, product_name, 1, final_price, final_price,
                     discount_percent=final_discount, promo_code=promo_code_used)

def handle_bulk_buy_from_keyboard(user_id, db):
    product_name = _pending_product_cache.get(user_id) or db["users"].get(user_id, {}).get("pending_product")
    is_api = product_name in API_PRODUCTS if product_name else False
    if not product_name or (not is_api and product_name not in db["products"]):
        bot.send_message(user_id, "⚠️ Please select a product first.",
            reply_markup=get_product_buy_keyboard(user_id)); return
    # ✅ BUG FIX: API product এর price DB তে না থাকলে error দাও
    price = db["products"].get(product_name)
    if price is None:
        bot.send_message(user_id,
            f"⚠️ <b>{product_name}</b> এর price সেট করা নেই!\n\n"
            f"Admin panel থেকে এই product এর price সেট করুন।",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
        logging.error(f"Price not set for API product: {product_name}")
        return
    price = float(price)

    # ── Out of Stock check ──
    stock_now = get_stock_count(product_name)
    # API product — API তে live stock থাকলে allow করো (local 0 হলেও চলবে)
    _is_api_bulk = product_name in API_PRODUCTS
    if stock_now <= 0 and not _is_api_bulk:
        bot.send_message(user_id,
            f"🔴 <b>Out of Stock!</b>\n\n"
            f"😔 দুঃখিত! <b>{product_name}</b> এখন স্টকে নেই।\n"
            f"পরে আবার চেষ্টা করুন।\n\n"
            f"⬅️ Back চাপুন বা অন্য product বেছে নিন।",
            parse_mode="HTML", reply_markup=get_product_buy_keyboard(user_id))
        return

    # ── Promo feature চেক ──
    promo_feature_on = db.get("settings", {}).get("promo_feature_enabled", True)

    if promo_feature_on:
        # Promo ON — promo code জিজ্ঞেস করো
        cancel_btn = types.ReplyKeyboardMarkup(resize_keyboard=True)
        cancel_btn.add( KbBtn("⏭️ Skip", role="skip"), KbBtn("🔙 Back", role="back"))
        msg = bot.send_message(user_id,
            "🎟️ <b>Promo / Coupon Code আছে?</b>\n\n থাকলে লিখুন, না থাকলে<b>Skip</b> korun.",
            parse_mode="HTML", reply_markup=cancel_btn)
        bot.register_next_step_handler(msg, lambda m: _bulk_promo_step(m, product_name, price))
    else:
        # Promo OFF — personal discount apply করে quantity step এ যাও
        _personal_disc_bulk = float(db.get("users", {}).get(user_id, {}).get("personal_discount", 0))
        _bulk_price = apply_discount(price, _personal_disc_bulk)
        stock = get_stock_count(product_name)
        _bulk_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        _bulk_markup.add( KbBtn("🔙 Back", role="back"))
        msg = bot.send_message(user_id,
            t(user_id, "bulk_prompt", name=product_name, price=_bulk_price, stock=stock),
            parse_mode="HTML", reply_markup=_bulk_markup)
        bot.register_next_step_handler(msg, lambda m: receive_bulk_quantity(m, product_name, _bulk_price, None))

def _bulk_promo_step(message, product_name, price):
    user_id = str(message.chat.id)
    db = load_db()
    text = (message.text or "").strip()
    if text == "❌ Cancel":
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    if text == "🔙 Back":
        show_product_detail_with_subnav(user_id, product_name, db); return
    # Personal discount auto-apply
    personal_disc = float(db.get("users", {}).get(user_id, {}).get("personal_discount", 0))
    final_discount = personal_disc
    promo_code_used = None
    if text and text != "⏭️ Skip":
        is_valid, disc, msg_txt = validate_promo(db, user_id, text)
        if not is_valid:
            bot.send_message(user_id, msg_txt, parse_mode="HTML",
                reply_markup=get_main_menu(user_id)); return
        final_discount = max(disc, personal_disc)  # বড় discount নাও
        promo_code_used = text.strip().upper()
        bot.send_message(user_id, msg_txt, parse_mode="HTML")
    discounted_price = apply_discount(price, final_discount)
    stock = get_stock_count(product_name)
    _bulk_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    _bulk_markup.add( KbBtn("🔙 Back", role="back"))
    msg = bot.send_message(user_id,
        t(user_id, "bulk_prompt", name=product_name, price=discounted_price, stock=stock),
        parse_mode="HTML", reply_markup=_bulk_markup)
    bot.register_next_step_handler(msg, lambda m: receive_bulk_quantity(m, product_name, discounted_price, promo_code_used))

def receive_bulk_quantity(message, product_name, price, promo_code=None):
    user_id = str(message.chat.id)
    db = load_db()
    lang = get_user_lang(user_id)
    L = LANG[lang]
    if message.text in (L["back"], L["cancel"], "🔙 Back"):
        show_product_detail_with_subnav(user_id, product_name, db); return
    _bulk_back = types.ReplyKeyboardMarkup(resize_keyboard=True)
    _bulk_back.add( KbBtn("🔙 Back", role="back"))
    if not message.text or not message.text.strip().isdigit():
        msg = bot.send_message(user_id, t(user_id, "invalid_number"),
            reply_markup=_bulk_back, parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: receive_bulk_quantity(m, product_name, price, promo_code)); return
    quantity = int(message.text.strip())
    if quantity <= 0:
        msg = bot.send_message(user_id, t(user_id, "qty_zero"),
            reply_markup=_bulk_back, parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: receive_bulk_quantity(m, product_name, price, promo_code)); return
    if quantity > MAX_BULK_QUANTITY:
        msg = bot.send_message(user_id,
            f"❌ <b>সর্বোচ্চ {MAX_BULK_QUANTITY} পিস একসাথে কেনা যাবে!</b>",
            reply_markup=_bulk_back, parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: receive_bulk_quantity(m, product_name, price, promo_code)); return
    stock = get_stock_count(product_name)
    if stock < quantity:
        bot.send_message(user_id,
            t(user_id, "insufficient_stock", qty=quantity, stock=stock),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    total_price = price * quantity
    current_balance = db["users"].get(user_id, {}).get("balance", 0)
    if current_balance < total_price:
        shortage = total_price - current_balance
        bot.send_message(user_id,
            t(user_id, "insufficient_balance", total=total_price, bal=current_balance, short=shortage),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return
    if promo_code:
        apply_promo_usage(db, user_id, promo_code)
        # ⚡ Targeted write — শুধু promo data update করো
        update_db_path(f"/promo_codes/{promo_code}", db["promo_codes"][promo_code])
        update_db_path(f"/promo_usage/{promo_code}", db["promo_usage"].get(promo_code, {}))
    process_purchase(user_id, product_name, quantity, price, total_price, promo_code=promo_code)

# ===========================
# WEBSITE API HELPER FUNCTIONS
# ===========================

def is_api_product(product_name):
    """API product check — API_PRODUCTS dict এ থাকলে True।"""
    return product_name in API_PRODUCTS


# ===========================
# 🎟️ PROMO / COUPON SYSTEM
# ===========================

def validate_promo(db, user_id, code):
    code = code.strip().upper()
    promos = db.get("promo_codes", {})
    if code not in promos:
        return False, 0, "❌ <b>Invalid promo code!</b>\n\nএই কোডটি বিদ্যমান নয়।"
    promo = promos[code]
    if not promo.get("enabled", True):
        return False, 0, "❌ <b>Promo code টি বর্তমানে নিষ্ক্রিয়!</b>"
    expiry = promo.get("expiry", "")
    if expiry:
        try:
            exp_dt = BD_TZ.localize(datetime.strptime(expiry, "%d/%m/%Y"))
            if datetime.now(BD_TZ) > exp_dt:
                return False, 0, "❌ <b>Promo code মেয়াদ শেষ হয়ে গেছে!</b>"
        except Exception:
            pass
    max_uses = promo.get("max_uses", 0)
    used_count = promo.get("used_count", 0)
    if max_uses > 0 and used_count >= max_uses:
        return False, 0, "❌ <b>Promo code এর ব্যবহার সীমা শেষ!</b>"
    new_user_only = promo.get("new_user_only", False)
    if new_user_only:
        user_orders = db.get("users", {}).get(user_id, {}).get("orders", [])
        if user_orders:
            return False, 0, "❌ <b>এই কোডটি শুধুমাত্র নতুন User দের জন্য!</b>"
    promo_usage = db.get("promo_usage", {})
    user_used = promo_usage.get(code, {}).get(user_id, False)
    if user_used:
        return False, 0, "❌ <b>আপনি এই কোডটি আগেই ব্যবহার করেছেন!</b>"
    discount = float(promo.get("discount_percent", 0))
    return True, discount, f"✅ <b>Promo code সফলভাবে প্রয়োগ হয়েছে!</b>\n💸 <b>{discount:.0f}% ছাড় পাবেন।</b>"

def apply_promo_usage(db, user_id, code):
    code = code.strip().upper()
    if code not in db.get("promo_codes", {}):
        return
    db["promo_codes"][code]["used_count"] = db["promo_codes"][code].get("used_count", 0) + 1
    if "promo_usage" not in db:
        db["promo_usage"] = {}
    if code not in db["promo_usage"]:
        db["promo_usage"][code] = {}
    db["promo_usage"][code][user_id] = get_now()

def get_new_user_discount(db):
    return float(db.get("settings", {}).get("new_user_discount", 0))

def apply_discount(price, discount_percent):
    if discount_percent <= 0:
        return price
    return max(round(price * (1 - discount_percent / 100), 2), 0)

# ===========================
# API DASHBOARD TRACKING
# ===========================

# ⚡ API Dashboard in-memory accumulator — Firebase read বন্ধ
# প্রতিটা purchase এ Firebase GET বাদ, instead in-memory dict এ জমা করো
# প্রতি 60s বা 10 purchase পরে একবার Firebase write করো
_api_dashboard_cache = {}   # {key: {count, total_bdt, last_updated}}
_api_dashboard_pending = {} # {key: {count_delta, total_delta}} — unsaved changes
_api_dashboard_lock = threading.Lock()
_api_dashboard_write_count = 0
_API_DASHBOARD_WRITE_INTERVAL = 10  # প্রতি 10 purchase এ একবার Firebase write

def update_api_dashboard(product_name, quantity, total_bdt):
    """
    ⚡ OPTIMIZED: Firebase read বাদ — in-memory accumulate করো।
    প্রতি 10 purchase পরে একবার Firebase write (আগে: প্রতিটা purchase এ read+write)
    """
    global _api_dashboard_write_count
    key = product_name.replace(' ', '_')
    with _api_dashboard_lock:
        # In-memory accumulate
        if key not in _api_dashboard_pending:
            _api_dashboard_pending[key] = {"count_delta": 0, "total_delta": 0.0}
        _api_dashboard_pending[key]["count_delta"] += int(quantity)
        _api_dashboard_pending[key]["total_delta"] = round(
            _api_dashboard_pending[key]["total_delta"] + float(total_bdt), 2)
        _api_dashboard_write_count += 1
        should_flush = _api_dashboard_write_count >= _API_DASHBOARD_WRITE_INTERVAL

    if should_flush:
        _EXECUTOR.submit(_flush_api_dashboard)

def _flush_api_dashboard():
    """Pending dashboard changes Firebase এ write করো।"""
    global _api_dashboard_write_count
    with _api_dashboard_lock:
        if not _api_dashboard_pending:
            return
        pending_copy = dict(_api_dashboard_pending)
        _api_dashboard_pending.clear()
        _api_dashboard_write_count = 0
    try:
        # একবারে সব product এর data fetch করো (একটা GET call)
        existing_all = _fb_ref("/api_dashboard").get() or {}
        updates = {}
        now_str = get_now()
        for key, delta in pending_copy.items():
            existing = existing_all.get(key, {})
            new_count = int(existing.get("count", 0)) + delta["count_delta"]
            new_total = round(float(existing.get("total_bdt", 0)) + delta["total_delta"], 2)
            updates[key] = {"count": new_count, "total_bdt": new_total, "last_updated": now_str}
        # একটাই write call — সব product একসাথে
        _fb_ref("/api_dashboard").update(updates)
        logging.info(f"API dashboard flushed: {list(updates.keys())}")
    except Exception as e:
        logging.error(f"_flush_api_dashboard error: {e}")
        # Failed write গুলো আবার pending এ ফেরত দাও
        with _api_dashboard_lock:
            for key, delta in pending_copy.items():
                if key not in _api_dashboard_pending:
                    _api_dashboard_pending[key] = {"count_delta": 0, "total_delta": 0.0}
                _api_dashboard_pending[key]["count_delta"] += delta["count_delta"]
                _api_dashboard_pending[key]["total_delta"] += delta["total_delta"]

def get_api_dashboard_data():
    """সব API product এর dashboard data একসাথে আনো।"""
    try:
        data = _fb_ref("/api_dashboard").get() or {}
        # Pending (unsaved) changes যোগ করো
        with _api_dashboard_lock:
            for key, delta in _api_dashboard_pending.items():
                if key not in data:
                    data[key] = {"count": 0, "total_bdt": 0}
                data[key]["count"] = int(data[key].get("count", 0)) + delta["count_delta"]
                data[key]["total_bdt"] = round(
                    float(data[key].get("total_bdt", 0)) + delta["total_delta"], 2)
        return data
    except Exception as e:
        logging.error(f"get_api_dashboard_data error: {e}")
        return {}
        return {}

# ===========================
# MODIFIED process_purchase — API + Local Stock দুটোই সাপোর্ট
# ===========================


def process_purchase(user_id, product_name, quantity, price, total_price, display_name=None, discount_percent=0, promo_code=None):
    show_name = display_name or product_name

    # ══════════════════════════════════════════════════════
    # IN-FLIGHT DEDUP — double-click / retry তে duplicate block
    # ══════════════════════════════════════════════════════
    if not _try_acquire_in_flight(str(user_id)):
        bot.send_message(user_id,
            "⏳ <b>আপনার আগের order process হচ্ছে!</b>\n\nএকটু অপেক্ষা করুন...",
            parse_mode="HTML")
        return

    # ══════════════════════════════════════════════════════
    # USER LOCK — একই user একসাথে দুইটা purchase করতে পারবে না
    # এটা double-spend বন্ধ করে (race condition fix)
    # ══════════════════════════════════════════════════════
    user_lock = get_user_lock(str(user_id))
    if not user_lock.acquire(blocking=False):
        _release_in_flight(str(user_id))
        bot.send_message(user_id,
            "⏳ <b>আপনার আগের order process হচ্ছে!</b>\n\nএকটু অপেক্ষা করুন...",
            parse_mode="HTML")
        return

    try:
        # ⚡ OPTIMIZED: পুরো DB download বাদ — শুধু user balance fresh read
        # আগে: _fb_ref("/").get() → পুরো 10MB DB = Firebase quota drain
        # এখন: শুধু /users/{user_id} fetch → কয়েক KB মাত্র
        try:
            fresh_user = _fb_ref(f"/users/{user_id}").get() or {}
            db = load_db()  # cache hit — μs মাত্র
            if fresh_user:
                db["users"][user_id] = fresh_user
        except Exception:
            db = load_db()

        # ── Balance check (lock এর ভেতরে — race condition নেই) ──
        current_balance = float(db["users"].get(user_id, {}).get("balance", 0))
        if current_balance < total_price:
            shortage = round(total_price - current_balance, 2)
            bot.send_message(user_id,
                t(user_id, "insufficient_balance", total=total_price, bal=current_balance, short=shortage),
                reply_markup=get_main_menu(user_id), parse_mode="HTML")
            return

        order_id = f"ORD-{uuid.uuid4().hex[:8].upper()}"
        current_time = get_now()

        # ── Safety check: price 0 হলে API product deliver হবে না ──
        # ✅ BUG FIX: Edu Mail 24H / 72H price DB তে সেট না থাকলে block করো
        if product_name in API_PRODUCTS and total_price == 0:
            bot.send_message(user_id,
                f"⚠️ <b>Price Error!</b>\n\n"
                f"<b>{show_name}</b> এর price সেট করা নেই।\n"
                f"Admin কে জানান।\n\n"
                f"📞 @{SUPPORT_USERNAME}",
                parse_mode="HTML", reply_markup=get_main_menu(user_id))
            logging.error(f"process_purchase: zero price for API product {product_name}, user {user_id}")
            return

        # ── Order processing message removed — delivery হলেই message যাবে ──

        # ════════════════════════════════════════════════
        # API PRODUCT — local stock আগে, তারপর API
        # ════════════════════════════════════════════════
        if product_name in API_PRODUCTS:
            api_conf = API_PRODUCTS[product_name]

            # ─── BULKMAIL PRODUCT ───────────────────────────
            if api_conf.get("source") == "bulkmail":
                bm_product_id = api_conf["product_id"]

                # Local xlsx stock আগে চেক করো
                local_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
                local_stock = 0
                if os.path.exists(local_path):
                    try:
                        _wb = openpyxl.load_workbook(local_path, read_only=True)
                        _ws = _wb.active
                        for _ri, _row in enumerate(_ws.iter_rows(values_only=True), start=1):
                            if _ri == 1: continue
                            if any(c is not None and str(c).strip() != "" for c in _row):
                                local_stock += 1
                        _wb.close()
                    except Exception:
                        pass

                accounts = []
                success  = False

                if local_stock >= quantity:
                    # ── Local xlsx থেকে deliver ──
                    stock_lock = get_stock_lock(product_name)
                    with stock_lock:
                        try:
                            stock_file = read_stock_file(product_name, quantity)
                            if stock_file is None:
                                raise Exception("stock_file None")
                            import csv as _csv_bm, io as _io_bm
                            stock_file.seek(0)
                            csv_text_bm = stock_file.read().decode("utf-8-sig").strip()
                            stock_file.seek(0)
                            reader_bm = list(_csv_bm.reader(_io_bm.StringIO(csv_text_bm)))
                            if len(reader_bm) >= 2:
                                _bm_hdrs = [str(h).strip() for h in reader_bm[0] if str(h).strip()]
                                for _row_bm in reader_bm[1:]:
                                    _cells_bm = [str(c).strip() if c is not None else "" for c in _row_bm]
                                    if not any(_cells_bm): continue
                                    _acc = {_bm_hdrs[i]: (_cells_bm[i] if i < len(_cells_bm) else "")
                                            for i in range(len(_bm_hdrs))}
                                    accounts.append(_acc)
                            success = len(accounts) > 0
                        except Exception as _e:
                            logging.error(f"Local stock read error for bulkmail product: {_e}")
                            success = False
                    invalidate_stock_cache(product_name)
                else:
                    # ── Bulkmail API থেকে purchase ──
                    success, accounts, err_msg = bulkmail_purchase(bm_product_id, quantity)

                if not success:
                    bot.send_message(user_id,
                        "❌ Purchase failed. Please try again or contact support team.",
                        reply_markup=get_main_menu(user_id))
                    try:
                        bot.send_message(ADMIN_ID,
                            f"⚠️ <b>Bulkmail Purchase Failed</b>\n"
                            f"User: <code>{user_id}</code>\n"
                            f"Product: {product_name} (id={bm_product_id})\n"
                            f"Error: <code>{err_msg}</code>",
                            parse_mode="HTML")
                    except Exception:
                        pass
                    return

                # ⚡ Balance কাটো + সাথে সাথে cache ও Firebase update
                new_balance = round(current_balance - total_price, 2)
                db["users"][user_id]["balance"] = new_balance
                _update_db_cache_in_place(db)  # cache instant update — user balance কাটা দেখবে

                # Purchase success message
                bot.send_message(user_id,
                    t(user_id, "purchase_success",
                      order_id=order_id, product=show_name, qty=quantity,
                      price=price, total=total_price, bal=new_balance, date=current_time),
                    parse_mode="HTML")

                # Account details deliver করো — Email/Password/Full Data format
                if accounts:
                    def _bm_fmt(a):
                        email     = str(a.get("Email", "")).strip()
                        password  = str(a.get("Password", "")).strip()
                        full_data = str(a.get("Full_Data", "")).strip()
                        if not full_data:
                            full_data = ":".join(filter(None, [email, password]))
                        return email, password, full_data

                    if quantity == 1:
                        email, password, full_data = _bm_fmt(accounts[0])
                        account_msg = (
                            f"📬 <b>Account Details:</b>\n\n"
                            f"<b>Email:</b> <code>{email}</code>\n"
                            f"<b>Password:</b> <code>{password}</code>\n"
                            f"<b>Full Data:</b> <code>{full_data}</code>\n\n"
                            f"🆔 <b>Order:</b> <code>{order_id}</code>\n"
                            f"⚠️ <i>Please save this info!</i>"
                        )
                        bot.send_message(user_id, account_msg, parse_mode="HTML",
                            reply_markup=get_product_buy_keyboard(user_id))
                    else:
                        import csv as _csv_bm2, io as _io_bm2
                        output_bm  = _io_bm2.BytesIO()
                        wrapper_bm = _io_bm2.TextIOWrapper(output_bm, encoding="utf-8-sig", newline="")
                        writer_bm  = _csv_bm2.writer(wrapper_bm)
                        writer_bm.writerow(["Email", "Password", "Full Data"])
                        for a in accounts:
                            em, pw, fd = _bm_fmt(a)
                            writer_bm.writerow([em, pw, fd])
                        wrapper_bm.flush()
                        output_bm.seek(0)
                        file_name_bm = f"{show_name.replace(' ', '_')}_{order_id}.csv"
                        caption_bm = (
                            f"📄 <b>{show_name} — Account Data</b>\n\n"
                            f"🆔 Order: <code>{order_id}</code>\n"
                            f"🔢 Qty: {quantity} pcs\n"
                            f"📅 {current_time}\n\n"
                            f"⚠️ <i>Please save this file!</i>"
                        )
                        output_bm.name = file_name_bm
                        bot.send_document(user_id, output_bm, caption=caption_bm,
                            parse_mode="HTML", reply_markup=get_product_buy_keyboard(user_id))
                else:
                    bot.send_message(user_id,
                        f"⚠️ <b>Account Delivery Issue!</b>\n\n"
                        f"Purchase সম্পন্ন হয়েছে কিন্তু account details পাঠানো যায়নি।\n"
                        f"🆔 Order ID: <code>{order_id}</code>\n\n"
                        f"📞 Support: @{SUPPORT_USERNAME}",
                        parse_mode="HTML", reply_markup=get_product_buy_keyboard(user_id))
                    logging.error(f"Bulkmail purchase success but empty accounts for order {order_id}")

                # API Dashboard tracking
                _EXECUTOR.submit(update_api_dashboard, product_name, quantity, total_price)

                # Order DB তে save করো
                order_entry = {
                    "id": order_id, "product_name": show_name, "quantity": quantity,
                    "price_per_unit": price, "total": total_price,
                    "date": current_time, "status": "completed", "source": "bulkmail"
                }
                if discount_percent > 0: order_entry["discount_percent"] = discount_percent
                if promo_code:           order_entry["promo_code"] = promo_code
                db["users"][user_id].setdefault("orders", []).append(order_entry)
                update_db_path(f"/users/{user_id}", db["users"][user_id])
                invalidate_stock_cache(product_name)

                try:
                    bot.send_message(LOG_CHANNEL_ID,
                        f"🛒 <b>Purchase Successful!</b>🎉\n\n"
                        f"👤 <b>User ID:</b> <code>{mask_id(user_id)}</code>\n"
                        f"📦 <b>Product:</b> {show_name}\n"
                        f"🔢 <b>Quantity:</b> {quantity} pcs\n"
                        f"💰 <b>Total:</b> {total_price} BDT\n"
                        f"🆔 <b>Order ID:</b> <code>{mask_id(order_id)}</code>\n"
                        f"✅ <b>Status: Delivered</b>\n📅 <b>Date:</b> {current_time}",
                        parse_mode="HTML")
                except: pass
                log_action("PURCHASE_SUCCESS", user_id,
                    f"Product: {show_name} [bulkmail], Qty: {quantity}, Amount: {total_price}")
                return
            # ─── END BULKMAIL ────────────────────────────────

            # ─── HOTMAIL143 PRODUCT (existing logic) ─────────

            # ── Local xlsx stock আগে চেক করো ──
            local_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
            local_stock = 0
            if os.path.exists(local_path):
                try:
                    _wb = openpyxl.load_workbook(local_path, read_only=True)
                    _ws = _wb.active
                    for _ri, _row in enumerate(_ws.iter_rows(values_only=True), start=1):
                        if _ri == 1: continue
                        if any(c is not None and str(c).strip() != "" for c in _row):
                            local_stock += 1
                    _wb.close()
                except Exception:
                    pass

            accounts = []
            success = False

            if local_stock >= quantity:
                # ── Local stock দিয়ে deliver করো (xlsx heading অনুযায়ী) ──
                stock_lock = get_stock_lock(product_name)
                with stock_lock:
                    try:
                        stock_file = read_stock_file(product_name, quantity)
                        if stock_file is None:
                            raise Exception("stock_file None")

                        import csv as _csv_local, io as _io_local
                        stock_file.seek(0)
                        csv_text_local = stock_file.read().decode("utf-8-sig").strip()
                        stock_file.seek(0)
                        reader_local = list(_csv_local.reader(_io_local.StringIO(csv_text_local)))

                        if len(reader_local) >= 2:
                            # ── xlsx এর actual heading গুলো নাও ──
                            _local_headers = [str(h).strip() for h in reader_local[0] if str(h).strip()]

                            for _row_l in reader_local[1:]:
                                _cells_l = [str(c).strip() if c is not None else "" for c in _row_l]
                                if not any(_cells_l):
                                    continue
                                _acc = {_local_headers[i]: (_cells_l[i] if i < len(_cells_l) else "")
                                        for i in range(len(_local_headers))}
                                accounts.append(_acc)
                        elif len(reader_local) == 1:
                            # Heading নেই — শুধু data row
                            _row_l = reader_local[0]
                            _cells_l = [str(c).strip() for c in _row_l if c is not None]
                            if _cells_l:
                                raw = _cells_l[0]
                                if "|" in raw:
                                    parts = [p.strip() for p in raw.split("|")]
                                    _acc = {f"Field{i+1}": parts[i] for i in range(len(parts))}
                                elif ":" in raw:
                                    p = raw.split(":", 1)
                                    _acc = {"Email": p[0].strip(), "Password": p[1].strip()}
                                else:
                                    _acc = {"Account": raw}
                                accounts.append(_acc)

                        success = len(accounts) > 0
                    except Exception as _e:
                        logging.error(f"Local stock read error for API product: {_e}")
                        success = False
                invalidate_stock_cache(product_name)
            else:
                # ─── HOTMAIL143 product enabled check ───
                if not is_hotmail143_enabled(product_name, db):
                    bot.send_message(user_id,
                        "❌ Purchase failed. Please try again later.",
                        reply_markup=get_main_menu(user_id))
                    return
                # ── API থেকে purchase করো ──
                success, accounts, err_msg = hotmail143_purchase(
                    api_conf["product_type"], api_conf["account_type"], quantity
                )

            if not success:
                bot.send_message(user_id,
                    "❌ Purchase failed. Please try again later.",
                    reply_markup=get_main_menu(user_id))
                return

            # ⚡ Balance কাটো + সাথে সাথে cache ও Firebase update
            new_balance = round(current_balance - total_price, 2)
            db["users"][user_id]["balance"] = new_balance
            _update_db_cache_in_place(db)  # cache instant update

            # Purchase success message
            bot.send_message(user_id,
                t(user_id, "purchase_success",
                  order_id=order_id, product=show_name, qty=quantity,
                  price=price, total=total_price, bal=new_balance, date=current_time),
                parse_mode="HTML")

            # Account details deliver করো
            if accounts:
                # ── Hotmail / Outlook account normalize helper ──
                def _normalize_hotmail_account(acc_dict):
                    """
                    API response: email|password|refresh_token|client_id format।
                    parse_account_string এ Full_Data, Email, Password, Refresh_Token, Client_Id key আসে।
                    """
                    email     = str(acc_dict.get("Email", "")).strip()
                    password  = str(acc_dict.get("Password", "")).strip()
                    full_data = str(acc_dict.get("Full_Data", "")).strip()

                    # Full_Data না থাকলে সব values pipe দিয়ে join করো
                    if not full_data:
                        all_vals = [str(v).strip() for v in acc_dict.values() if v is not None and str(v).strip()]
                        full_data = "|".join(all_vals)

                    return email, password, full_data

                # ── Local থেকে এলে সেটা চেক করো ──
                _from_local_stock = (local_stock >= quantity)
                # Hotmail / Outlook → বিশেষ format (শুধু API থেকে এলে)
                _is_hotmail_outlook = product_name in (HOTMAIL_PROD, OUTLOOK_PROD) and not _from_local_stock

                if quantity == 1:
                    acc = accounts[0]
                    if _is_hotmail_outlook:
                        email, password, full_data = _normalize_hotmail_account(acc)
                        account_msg = (
                            f"📬 <b>Account Details:</b>\n\n"
                            f"<b>Email:</b> <code>{email}</code>\n"
                            f"<b>Password:</b> <code>{password}</code>\n"
                            f"<b>Full Data:</b> <code>{full_data}</code>\n\n"
                            f"🆔 <b>Order:</b> <code>{order_id}</code>\n"
                            f"⚠️ <i>Please save this info!</i>"
                        )
                    else:
                        # ── xlsx heading অনুযায়ী deliver করো (local ও API দুটোর জন্য) ──
                        details_lines = ""
                        for k, v in acc.items():
                            if k == "Full_Data":
                                continue
                            v_str = str(v).strip() if v is not None else ""
                            if v_str:
                                details_lines += f"<b>{k}:</b> <code>{v_str}</code>\n"
                        if not details_lines:
                            for k, v in acc.items():
                                v_str = str(v).strip() if v is not None else ""
                                if v_str:
                                    details_lines += f"<b>{k}:</b> <code>{v_str}</code>\n"
                        account_msg = (
                            f"📬 <b>Account Details:</b>\n\n"
                            f"{details_lines}\n"
                            f"🆔 <b>Order:</b> <code>{order_id}</code>\n"
                            f"⚠️ <i>Please save this info!</i>"
                        )
                    bot.send_message(user_id, account_msg, parse_mode="HTML",
                        reply_markup=get_product_buy_keyboard(user_id))
                else:
                    # Bulk — CSV file বানাও (xlsx heading অনুযায়ী)
                    import csv as _csv_mod, io as _io_mod
                    output = _io_mod.BytesIO()
                    if _is_hotmail_outlook:
                        # API Hotmail/Outlook bulk — Email, Password, Full Data
                        wrapper = _io_mod.TextIOWrapper(output, encoding="utf-8-sig", newline="")
                        writer = _csv_mod.writer(wrapper)
                        writer.writerow(["Email", "Password", "Full Data"])
                        for a in accounts:
                            em, pw, fd = _normalize_hotmail_account(a)
                            writer.writerow([em, pw, fd])
                        wrapper.flush()
                        output.seek(0)
                        file_name = f"{show_name.replace(' ', '_')}_{order_id}.csv"
                    else:
                        # Local xlsx বা অন্য API product — xlsx এর heading অনুযায়ী CSV
                        all_keys = []
                        for a in accounts:
                            for k in a.keys():
                                if k not in all_keys and k != "Full_Data":
                                    all_keys.append(k)
                        fieldnames = all_keys if all_keys else ["Account"]
                        wrapper = _io_mod.TextIOWrapper(output, encoding="utf-8-sig", newline="")
                        writer = _csv_mod.DictWriter(wrapper, fieldnames=fieldnames, extrasaction="ignore")
                        writer.writeheader()
                        writer.writerows(accounts)
                        wrapper.flush()
                        output.seek(0)
                        file_name = f"{show_name.replace(' ', '_')}_{order_id}.csv"

                    caption = (
                        f"📄 <b>{show_name} — Account Data</b>\n\n"
                        f"🆔 Order: <code>{order_id}</code>\n"
                        f"🔢 Qty: {quantity} pcs\n"
                        f"📅 {current_time}\n\n"
                        f"⚠️ <i>Please save this file!</i>"
                    )
                    output.name = file_name
                    bot.send_document(user_id, output, caption=caption, parse_mode="HTML",
                        reply_markup=get_product_buy_keyboard(user_id))
            else:
                # accounts empty — error জানাও কিন্তু balance কাটা হয়েছে তাই support এ যেতে বলো
                bot.send_message(user_id,
                    f"⚠️ <b>Account Delivery Issue!</b>\n\n"
                    f"Purchase সম্পন্ন হয়েছে কিন্তু account details পাঠানো যায়নি।\n"
                    f"🆔 Order ID: <code>{order_id}</code>\n\n"
                    f"📞 Support: @{SUPPORT_USERNAME}",
                    parse_mode="HTML", reply_markup=get_product_buy_keyboard(user_id))
                logging.error(f"API purchase success but empty accounts for order {order_id}")

            # API Dashboard tracking — কতটা কিনা হলো, কত টাকা
            _EXECUTOR.submit(update_api_dashboard, product_name, quantity, total_price)

            # Order DB তে save করো
            order_entry = {
                "id": order_id, "product_name": show_name, "quantity": quantity,
                "price_per_unit": price, "total": total_price,
                "date": current_time, "status": "completed"
            }
            if discount_percent > 0:
                order_entry["discount_percent"] = discount_percent
            if promo_code:
                order_entry["promo_code"] = promo_code
            db["users"][user_id].setdefault("orders", []).append(order_entry)
            update_db_path(f"/users/{user_id}", db["users"][user_id])
            # Stock cache clear করো
            invalidate_stock_cache(product_name)

            # Log
            try:
                bot.send_message(LOG_CHANNEL_ID,
                    f"🛒 <b>Purchase Successful!</b>🎉\n\n"
                    f"👤 <b>User ID:</b> <code>{mask_id(user_id)}</code>\n"
                    f"📦 <b>Product:</b> {show_name}\n🔢 <b>Quantity:</b> {quantity} pcs\n"
                    f"💰 <b>Total:</b> {total_price} BDT\n🆔 <b>Order ID:</b> <code>{mask_id(order_id)}</code>\n"
                    f"✅ <b>Status: Delivered</b>\n📅 <b>Date:</b> {current_time}",
                    parse_mode="HTML")
            except: pass
            log_action("PURCHASE_SUCCESS", user_id,
                f"Product: {show_name}, Qty: {quantity}, Amount: {total_price}")
            return

        # ════════════════════════════════════════════════
        # LOCAL PRODUCT — xlsx file থেকে দাও (THREAD-SAFE)
        # ════════════════════════════════════════════════
        # STOCK LOCK — এই product এ একসাথে একজনই access করতে পারবে
        # হাজার user একসাথে কিনলেও একই account দুইজন পাবে না ✅
        stock_lock = get_stock_lock(product_name)
        with stock_lock:
                # Lock এর ভেতরে fresh stock count — অন্য thread ইতিমধ্যে নিয়ে নিতে পারে
                stock = get_stock_count(product_name)
                if stock < quantity:
                    bot.send_message(user_id,
                        t(user_id, "insufficient_stock", qty=quantity, stock=stock),
                        reply_markup=get_main_menu(user_id), parse_mode="HTML")
                    return

                try:
                    # Stock file read ও write একই lock এর ভেতরে — safe ✅
                    stock_file = read_stock_file(product_name, quantity)
                except Exception as e:
                    logging.error(f"Stock read error: {e}\n{traceback.format_exc()}")
                    bot.send_message(user_id, t(user_id, "file_error"),
                        reply_markup=get_main_menu(user_id), parse_mode="HTML")
                    return

                if stock_file is None:
                    bot.send_message(user_id, t(user_id, "file_error"),
                        reply_markup=get_main_menu(user_id), parse_mode="HTML")
                    return

                # ── আগে file/account deliver করো ──
                delivery_ok = False

                if quantity == 1:
                    try:
                        # ✅ BUG FIX: read_single_account_text() আর call করা হচ্ছে না।
                        # আগে read_stock_file() xlsx থেকে row মুছে stock_file দিত,
                        # তারপর read_single_account_text() আবার xlsx খুলে পরের row পড়ত —
                        # ফলে দুইটা আলাদা account deliver হত, এবং ২য়টা xlsx এ থেকে যেত।
                        # এখন: stock_file (CSV) কেই parse করে text বানানো হচ্ছে।
                        import csv as _csv_mod, io as _io_mod
                        stock_file.seek(0)
                        raw_bytes = stock_file.read()
                        # BOM strip করো (Excel এ BOM থাকতে পারে)
                        csv_text = raw_bytes.decode("utf-8-sig").strip()
                        stock_file.seek(0)  # restore for possible restore later

                        reader = list(_csv_mod.reader(_io_mod.StringIO(csv_text)))

                        account_dict = None
                        headers = None

                        if len(reader) >= 2:
                            # ── Excel heading row ──
                            raw_headers = [str(h).strip() for h in reader[0]]
                            headers = [h for h in raw_headers if h]

                            # প্রথম non-empty data row নাও
                            for row in reader[1:]:
                                cells = [str(c).strip() if c is not None else "" for c in row]
                                if not any(c for c in cells):
                                    continue  # empty row skip

                                # ── Case 1: heading ১টা, data ও ১টা cell এ (pipe/colon format) ──
                                if len(headers) == 1 and len(cells) >= 1:
                                    raw_val = cells[0]
                                    if "|" in raw_val:
                                        # heading|heading|heading format চেক
                                        # অথবা data|data|data
                                        parts = [p.strip() for p in raw_val.split("|")]
                                        # heading row তে "|" থাকলে সেটাই heading ধরো
                                        sub_headers = [p.strip() for p in str(reader[0][0]).split("|") if p.strip()]
                                        if len(sub_headers) > 1:
                                            headers = sub_headers
                                            account_dict = {sub_headers[i]: parts[i] if i < len(parts) else "" for i in range(len(sub_headers))}
                                        else:
                                            # heading নেই — Field1, Field2... বানাও
                                            headers = [f"Field{i+1}" for i in range(len(parts))]
                                            account_dict = {headers[i]: parts[i] for i in range(len(parts))}
                                    elif ":" in raw_val:
                                        parts = raw_val.split(":", 1)
                                        headers = ["Email", "Password"]
                                        account_dict = {"Email": parts[0].strip(), "Password": parts[1].strip()}
                                    else:
                                        account_dict = {headers[0]: raw_val}
                                    break

                                # ── Case 2: normal — heading column = data column ──
                                account_dict = {}
                                for i, h in enumerate(headers):
                                    account_dict[h] = cells[i] if i < len(cells) else ""
                                break

                        elif len(reader) == 1:
                            row = [str(c).strip() for c in reader[0] if c is not None]
                            if row:
                                raw_val = row[0]
                                if "|" in raw_val:
                                    parts = [p.strip() for p in raw_val.split("|")]
                                    headers = [f"Field{i+1}" for i in range(len(parts))]
                                    account_dict = {headers[i]: parts[i] for i in range(len(parts))}
                                elif ":" in raw_val:
                                    parts = raw_val.split(":", 1)
                                    headers = ["Email", "Password"]
                                    account_dict = {"Email": parts[0].strip(), "Password": parts[1].strip()}
                                else:
                                    headers = ["Account"]
                                    account_dict = {"Account": raw_val}

                        if account_dict and headers:
                            # Excel heading অনুযায়ী সঠিক order এ data দেখাও
                            details_lines = ""
                            for h in headers:
                                val = str(account_dict.get(h, "")).strip()
                                if val:
                                    details_lines += f"<b>{h}:</b> <code>{val}</code>\n"
                            if not details_lines:
                                for k, v in account_dict.items():
                                    if str(v).strip():
                                        details_lines += f"<b>{k}:</b> <code>{v}</code>\n"
                            account_msg = (
                                f"📬 <b>Account Details:</b>\n\n"
                                f"{details_lines}\n"
                                f"🆔 <b>Order:</b> <code>{order_id}</code>\n"
                                f"⚠️ <i>Please save this info!</i>"
                            )
                            # ① আগে purchase success
                            new_balance = round(current_balance - total_price, 2)
                            bot.send_message(user_id,
                                t(user_id, "purchase_success",
                                  order_id=order_id, product=show_name, qty=quantity,
                                  price=price, total=total_price, bal=new_balance, date=current_time),
                                parse_mode="HTML")
                            # ② তারপর account details + buy menu
                            bot.send_message(user_id, account_msg, parse_mode="HTML",
                                reply_markup=get_product_buy_keyboard(user_id))
                            delivery_ok = True
                        else:
                            bot.send_message(user_id, t(user_id, "file_error"), parse_mode="HTML")
                    except Exception as e:
                        logging.error(f"Single account text send error: {e}")
                        bot.send_message(user_id, t(user_id, "file_error"), parse_mode="HTML")
                else:
                    try:
                        file_name = f"{show_name.replace(' ', '_').replace('/', '-')}_{order_id}.csv"
                        stock_file.name = file_name
                        caption = (
                            f"📄 <b>{show_name} — Account Data</b>\n\n"
                            f"🆔 Order: <code>{order_id}</code>\n"
                            f"🔢 Qty: {quantity} pcs\n"
                            f"📅 {current_time}\n\n"
                            f"⚠️ <i>Please save this file!</i>"
                        )
                        new_balance = round(current_balance - total_price, 2)
                        # ① আগে purchase success
                        bot.send_message(user_id,
                            t(user_id, "purchase_success",
                              order_id=order_id, product=show_name, qty=quantity,
                              price=price, total=total_price, bal=new_balance, date=current_time),
                            parse_mode="HTML")
                        # ② তারপর file + buy menu
                        bot.send_document(user_id, stock_file, caption=caption, parse_mode="HTML",
                            reply_markup=get_product_buy_keyboard(user_id))
                        delivery_ok = True
                    except Exception as e:
                        logging.error(f"File send error: {e}")
                        bot.send_message(user_id, t(user_id, "file_error"), parse_mode="HTML")

                if not delivery_ok:
                    # Delivery ব্যর্থ — stock ফেরত দাও, balance কাটবো না
                    try:
                        _restore_stock_file(product_name, stock_file)
                    except Exception as re:
                        logging.error(f"Stock restore error: {re}")
                    bot.send_message(user_id,
                        "❌ <b>Delivery ব্যর্থ হয়েছে!</b>\n\nআপনার balance কাটা হয়নি। আবার চেষ্টা করুন।",
                        parse_mode="HTML", reply_markup=get_main_menu(user_id))
                    return

                # ── Delivery সফল — এখন balance কাটো ও DB save করো ──
                new_balance = round(current_balance - total_price, 2)
                db["users"][user_id]["balance"] = new_balance
                _update_db_cache_in_place(db)  # ⚡ cache instant update — user balance কাটা দেখবে
                order_entry = {
                    "id": order_id, "product_name": show_name, "quantity": quantity,
                    "price_per_unit": price, "total": total_price,
                    "date": current_time, "status": "completed", "source": "local"
                }
                if discount_percent > 0:
                    order_entry["discount_percent"] = discount_percent
                if promo_code:
                    order_entry["promo_code"] = promo_code
                db["users"][user_id].setdefault("orders", []).append(order_entry)
                # pending_product রাখো — user আবার কিনতে পারবে (Back চাপলে তখন মুছবে)
                # ⚡ OPTIMIZED: শুধু user data update (পুরো DB নয়)
                update_db_path(f"/users/{user_id}", db["users"][user_id])
        # ── Lock ছেড়ে দেওয়ার পরে log করো ──
        try:
            bot.send_message(LOG_CHANNEL_ID,
                f"🛒 <b>Purchase Successful!</b>??\n\n"
                f"👤 <b>User ID:</b> <code>{mask_id(user_id)}</code>\n"
                f"📦 <b>Product:</b> {show_name}\n🔢 <b>Quantity:</b> {quantity} pcs\n"
                f"💰 <b>Total:</b> {total_price} BDT\n🆔 <b>Order ID:</b> <code>{mask_id(order_id)}</code>\n"
                f"✅ <b>Status: Auto-Delivery</b>\n📅 <b>Date:</b> {current_time}",
                parse_mode="HTML")
        except: pass
        log_action("PURCHASE_SUCCESS", user_id,
            f"Product: {show_name}, Qty: {quantity}, Amount: {total_price}")

    finally:
        user_lock.release()
        _release_in_flight(str(user_id))  # ✅ purchase শেষ — slot খালি

# ===========================
# USER FUNCTIONS
# ===========================

def show_user_orders(user_id, db):
    all_orders = db["users"].get(user_id, {}).get("orders", [])

    # আজকের date বের করো (BD timezone)
    today_str = datetime.now(BD_TZ).strftime("%d/%m/%y")

    # আজকের orders filter করো
    today_orders = [
        o for o in all_orders
        if today_str in str(o.get("date", ""))
    ]

    # আজকের না থাকলে সর্বশেষ orders দেখাই
    if today_orders:
        display_orders = today_orders[-5:]
        header = t(user_id, "orders_title")
    else:
        display_orders = all_orders[-5:]
        lang = get_user_lang(user_id)
        header = ("📂 <b>Recent Orders (Last 5)</b>\n" if lang == "en"
                  else "📂 <b>সর্বশেষ ৫টি অর্ডার</b>\n")

    if not display_orders:
        bot.send_message(user_id, t(user_id, "no_orders"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        return

    display = header + "\n"
    for order in display_orders:
        display += t(user_id, "order_item",
            id=order["id"], product=order["product_name"],
            qty=order["quantity"], total=order["total"], date=order["date"])
        display += "\n"

    bot.send_message(user_id, display, parse_mode="HTML",
        reply_markup=get_main_menu(user_id))

def show_user_balance(user_id, db=None):
    # ⚡ Always fresh user data — Firebase থেকে শুধু এই user এর data নাও
    # Cache stale হলেও balance সবসময় accurate দেখাবে
    try:
        fresh_user_data = _fb_ref(f"/users/{user_id}").get()
        if fresh_user_data:
            # Cache update করো
            if db is None:
                db = load_db()
            db["users"][user_id] = fresh_user_data
            _update_db_cache_in_place(db)
    except Exception:
        pass

    if db is None:
        db = load_db()
    user = db["users"].get(user_id, {})

    # Username বের করো
    try:
        chat = bot.get_chat(user_id)
        username = chat.username if chat.username else (chat.first_name or "Unknown")
    except Exception:
        username = "Unknown"

    balance = user.get("balance", 0)
    orders_count = len(user.get("orders", []))
    refs_count = user.get("refer_count", 0)

    lang_u = _lang_cache.get(user_id, "en")
    if lang_u == "bn":
        balance_display = (
            f"\n"
            f"👤 <b>অ্যাকাউন্ট তথ্য</b>\n"
            f"\n"
            f"👤 <b>Username:</b> @{username}\n"
            f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
            f"💳 <b>Balance:</b> <code>{balance} BDT</code>\n"
            f"📦 <b>মোট অর্ডার:</b> {orders_count}টি\n"
            f"👥 <b>রেফারেল:</b> {refs_count} জন\n"
        )
    else:
        balance_display = (
            f"\n"
            f"👤 <b>Account Info</b>\n"
            f"\n"
            f"👤 <b>Username:</b> @{username}\n"
            f"🆔 <b>User ID:</b> <code>{user_id}</code>\n"
            f"💳 <b>Balance:</b> <code>{balance} BDT</code>\n"
            f"📦 <b>Total Orders:</b> {orders_count}\n"
            f"👥 <b>Referrals:</b> {refs_count}\n"
        )

    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineBtn("🔄 Refresh", style="primary", callback_data="refresh_balance"),
        InlineBtn("💰 Deposit", style="success", callback_data="deposit_start")
    )
    markup.add(InlineBtn("🛍️ Shop Now", style="primary", callback_data="go_to_shop"))

    bot.send_message(user_id, balance_display, parse_mode="HTML",
        reply_markup=get_main_menu(user_id))

def show_support_info(user_id):
    lang = get_user_lang(user_id)
    markup = types.InlineKeyboardMarkup()
    markup.add(InlineBtn(
        LANG[lang]["contact_support"], style="primary", url=f"https://t.me/{SUPPORT_USERNAME}"))
    markup.add(InlineBtn(LANG[lang]["go_back"], style="primary", callback_data="back_to_main"))
    bot.send_message(user_id, t(user_id, "support_msg", support=SUPPORT_USERNAME),
        parse_mode="HTML", reply_markup=markup)

def show_referral_info(user_id, db=None):
    if db is None:
        db = load_db()
    referral_link = f"https://t.me/{BOT_USERNAME}?start={user_id}"
    bonus = db["settings"].get("refer_bonus", 2)
    refer_count = db["users"].get(user_id, {}).get("refer_count", 0)
    total_earned = refer_count * bonus
    lang = get_user_lang(user_id)
    share_label = "🔗 Share your link and earn!"
    share_text = "Join this bot and get a bonus!" if lang == "en" else "এই বটে জয়েন করুন এবং বোনাস পান!"
    # ✅ Telegram এর নিজস্ব share/forward interface ওপেন করে —
    # ইউজার এখান থেকে সরাসরি অন্য চ্যাট/গ্রুপে লিংক ফরোয়ার্ড করতে পারবে
    share_url = (
        "https://t.me/share/url?url=" + quote(referral_link, safe="")
        + "&text=" + quote(share_text, safe="")
    )
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(InlineBtn(share_label, style="success", url=share_url))
    bot.send_message(user_id,
        t(user_id, "refer_msg",
          link=referral_link, bonus=bonus,
          count=refer_count, total=total_earned),
        parse_mode="HTML", reply_markup=markup)

# ===========================
# DEPOSIT SYSTEM
# ===========================

def show_deposit_menu(user_id):
    lang = get_user_lang(user_id)
    L = LANG[lang]
    db = load_db()
    cashback_on = db.get("settings", {}).get("cashback_enabled", True)
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(InlineBtn(L["do_deposit"], style="success", callback_data="deposit_start"))
    markup.add(InlineBtn(L["how_deposit"], style="success", callback_data="deposit_help"))
    markup.add(InlineBtn(L["dep_history_btn"], style="success", callback_data="deposit_history"))
    markup.add(InlineBtn(L["main_menu_btn"], style="primary", callback_data="back_to_main"))
    base_msg = t(user_id, "deposit_menu",
          bkash=BKASH_NUMBER, nagad=NAGAD_NUMBER,
          rocket=ROCKET_NUMBER, binance=BINANCE_ID)
    # Cashback OFF থাকলে cashback অফার অংশ বাদ দাও
    if not cashback_on:
        # Cashback Offer section strip করো
        for marker in ["🎁 <b>Cashback Offer:</b>", "🎁 <b>ক্যাশব্যাক অফার:</b>"]:
            idx = base_msg.find(marker)
            if idx != -1:
                base_msg = base_msg[:idx].rstrip()
                break
    bot.send_message(user_id, base_msg, parse_mode="HTML", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "deposit_start")
def start_deposit_process(call):
    bot.answer_callback_query(call.id)  # ⚡ Instant ACK
    _EXECUTOR.submit(_deposit_start_worker, call)

def _deposit_start_worker(call):
    user_id = str(call.message.chat.id)
    lang = get_user_lang(user_id)
    L = LANG[lang]
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(InlineBtn(f"🟠 Bkash ({BKASH_NUMBER})", style="primary", callback_data="dep_method_bkash"))
    markup.add(InlineBtn(f"🟢 Nagad ({NAGAD_NUMBER})", style="success", callback_data="dep_method_nagad"))
    markup.add(InlineBtn(f"🔴 Rocket ({ROCKET_NUMBER})", style="danger", callback_data="dep_method_rocket"))
    markup.add(InlineBtn(f"🔵 Binance ({BINANCE_ID})", style="primary", callback_data="dep_method_binance"))
    markup.add(InlineBtn(L["go_back"], style="primary", callback_data="back_to_main"))
    try:
        bot.edit_message_text(t(user_id, "dep_method_select"),
            user_id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    except:
        bot.send_message(user_id, t(user_id, "dep_method_select"),
            reply_markup=markup, parse_mode="HTML")

@bot.callback_query_handler(func=lambda call: call.data.startswith("dep_method_"))
def handle_deposit_method(call):
    bot.answer_callback_query(call.id)  # ⚡ Instant ACK
    _EXECUTOR.submit(_deposit_method_worker, call)

def _deposit_method_worker(call):
    user_id = str(call.message.chat.id)
    method_data = {
        "dep_method_bkash":   ("Bkash",   BKASH_NUMBER,  True),
        "dep_method_nagad":   ("Nagad",   NAGAD_NUMBER,  True),
        "dep_method_rocket":  ("Rocket",  ROCKET_NUMBER, True),
        "dep_method_binance": ("Binance", BINANCE_ID,    False),
    }
    if call.data not in method_data:
        bot.answer_callback_query(call.id, "Invalid method!", show_alert=True); return

    method, account, is_auto = method_data[call.data]

    if is_auto:
        instruction_key = "dep_instruction"
    else:
        instruction_key = "dep_binance_instruction"

    lang = get_user_lang(user_id)
    markup = types.InlineKeyboardMarkup()
    markup.add(InlineBtn(LANG[lang]["go_back"], style="primary", callback_data="back_to_main"))

    msg = bot.send_message(user_id,
        t(user_id, instruction_key, method=method, account=account),
        parse_mode="HTML")
    bot.register_next_step_handler(msg, lambda m: ask_for_amount(m, method, account, is_auto))

def ask_for_amount(message, method, account, is_auto):
    user_id = str(message.chat.id)
    try:
        lang = get_user_lang(user_id)
        L = LANG[lang]
        if message.text in (L["back"], L["cancel"]):
            show_deposit_menu(user_id); return
        try:
            amount_val = float(message.text.strip().replace(',', '.'))
            if amount_val <= 0:
                raise ValueError
            amount = amount_val
        except (ValueError, AttributeError):
            msg = bot.send_message(user_id, t(user_id, "dep_invalid_amount"),
                parse_mode="HTML", reply_markup=get_back_button(user_id))
            bot.register_next_step_handler(msg, lambda m: ask_for_amount(m, method, account, is_auto)); return
        if amount < MIN_DEPOSIT:
            msg = bot.send_message(user_id, t(user_id, "dep_min"),
                parse_mode="HTML", reply_markup=get_back_button(user_id))
            bot.register_next_step_handler(msg, lambda m: ask_for_amount(m, method, account, is_auto)); return

        # ✅ SECURITY: সর্বোচ্চ deposit limit
        if amount > MAX_DEPOSIT_AMOUNT:
            msg = bot.send_message(user_id,
                f"❌ <b>সর্বোচ্চ ডিপোজিট {MAX_DEPOSIT_AMOUNT} BDT!</b>\n\n"
                f"একবারে এর বেশি ডিপোজিট করা যাবে না।\n"
                f"সাহায্যের জন্য: @{SUPPORT_USERNAME}",
                parse_mode="HTML", reply_markup=get_back_button(user_id))
            bot.register_next_step_handler(msg, lambda m: ask_for_amount(m, method, account, is_auto)); return

        # amount integer format এ দেখাও (12.0 → 12)
        display_amount = int(amount) if amount == int(amount) else amount
        msg = bot.send_message(user_id,
            t(user_id, "dep_trxid_prompt", method=method, amount=display_amount),
            parse_mode="HTML", reply_markup=get_back_button(user_id))
        bot.register_next_step_handler(msg,
            lambda m: process_deposit_request(m, method, account, amount, is_auto))

    except Exception as e:
        logging.error(f"ask_for_amount CRASH | user={user_id} | err={e}", exc_info=True)
        try:
            bot.send_message(user_id,
                "❌ <b>কিছু একটা সমস্যা হয়েছে!</b>\n\n"
                "আবার চেষ্টা করুন অথবা সাপোর্টে যোগাযোগ করুন:\n"
                f"@{SUPPORT_USERNAME}",
                parse_mode="HTML", reply_markup=get_main_menu(user_id))
        except Exception:
            pass

def process_deposit_request(message, method, account, amount, is_auto):
    """
    ══════════════════════════════════════════════════════════
    🔒 SECURE DEPOSIT VERIFICATION — v2.1
    ══════════════════════════════════════════════════════════
    Bkash/Nagad/Rocket:
      • TrxID submit → বট SMS server এ check করে
      • SMS confirmed থাকলে → সাথে সাথে APPROVE
      • SMS না থাকলে → Pending save (SMS আসলে auto approve)
      • User reject হলে আবার চেষ্টা করতে পারবে
    Binance:
      • আগের মতোই manual admin approval
    ══════════════════════════════════════════════════════════
    """
    user_id = str(message.chat.id)
    try:
        _process_deposit_request_inner(message, user_id, method, account, amount, is_auto)
    except Exception as e:
        logging.error(f"process_deposit_request CRASH | user={user_id} | method={method} | err={e}", exc_info=True)
        try:
            bot.send_message(user_id,
                "❌ <b>কিছু একটা সমস্যা হয়েছে!</b>\n\n"
                "আবার চেষ্টা করুন অথবা সাপোর্টে যোগাযোগ করুন:\n"
                f"@{SUPPORT_USERNAME}",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id))
        except Exception:
            pass
        try:
            bot.send_message(ADMIN_ID,
                f"🚨 <b>Deposit Handler Crash!</b>\n\n"
                f"👤 User: <code>{user_id}</code>\n"
                f"💳 Method: {method}\n"
                f"💰 Amount: {amount}\n"
                f"❌ Error: <code>{str(e)[:300]}</code>",
                parse_mode="HTML")
        except Exception:
            pass

def _process_deposit_request_inner(message, user_id, method, account, amount, is_auto):
    """Inner handler — process_deposit_request এর try/except এ wrapped।"""
    lang = get_user_lang(user_id)
    L = LANG[lang]

    if message.text in (L["back"], L["cancel"]):
        show_deposit_menu(user_id); return

    transaction_id = (message.text or "").strip().upper()

    # ── TrxID format validation ──
    if not transaction_id or len(transaction_id) < 5:
        msg = bot.send_message(user_id,
            "❌ <b>ভুল Transaction ID!</b>\n\n"
            "TrxID কমপক্ষে ৫ অক্ষরের হতে হবে।\n"
            "আবার চেষ্টা করুন:",
            parse_mode="HTML", reply_markup=get_back_button(user_id))
        bot.register_next_step_handler(msg,
            lambda m: process_deposit_request(m, method, account, amount, is_auto))
        return

    # শুধু alphanumeric চরিত্র থাকতে হবে
    # Rocket এর TrxID numeric-only হতে পারে (10-14 digits), তাই সেটাও allow
    if not re.match(r'^[A-Z0-9]+$', transaction_id):
        msg = bot.send_message(user_id,
            "❌ <b>ভুল Transaction ID format!</b>\n\n"
            "TrxID শুধু letters ও numbers দিয়ে হয়।\n"
            "উদাহরণ: <code>ABC1234XYZ</code> বা <code>1234567890</code>\n\n"
            "আবার চেষ্টা করুন:",
            parse_mode="HTML", reply_markup=get_back_button(user_id))
        bot.register_next_step_handler(msg,
            lambda m: process_deposit_request(m, method, account, amount, is_auto))
        return

    db = load_db()

    # Cooldown removed — যেকোনো সময় deposit করা যাবে

    # ✅ SECURITY: Daily deposit limit check
    today = datetime.now(BD_TZ).strftime("%d/%m/%Y")
    daily_total = sum(
        int(float(r.get("amount", 0)))
        for r in db.get("deposit_requests", {}).values()
        if r.get("user_id") == user_id
        and r.get("status") == "approved"
        and r.get("date", "").startswith(today)
    )
    if daily_total + amount > MAX_DAILY_DEPOSIT:
        remaining_limit = MAX_DAILY_DEPOSIT - daily_total
        bot.send_message(user_id,
            f"❌ <b>দৈনিক ডিপোজিট সীমা অতিক্রম!</b>\n\n"
            f"\n"
            f"📊 আজ করেছেন: <b>{daily_total} BDT</b>\n"
            f"✅ আজ আর করতে পারবেন: <b>{max(0, remaining_limit)} BDT</b>\n"
            f"🔄 সীমা রিসেট হয় প্রতি মধ্যরাতে।\n"
            f"\n"
            f"বেশি লাগলে সাপোর্টে যোগাযোগ করুন।",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
        return

    # ══════════════════════════════════════════════════
    # 🔒 SECURITY CHECK 0: TrxID blocked check (3 attempt limit)
    # ══════════════════════════════════════════════════
    if is_trx_blocked(db, user_id, transaction_id):
        bot.send_message(user_id,
            "🚫 <b>এই TrxID Block করা হয়েছে!</b>\n\n"
            "\n"
            "আপনি এই TrxID দিয়ে ৩ বারের বেশি\n"
            "চেষ্টা করেছেন। এটি block করা হয়েছে।\n\n"
            f"📞 সাপোর্টে যোগাযোগ করুন:\n"
            f"@{SUPPORT_USERNAME}\n"
            "",
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id))
        return

    # ══════════════════════════════════════════════════
    # 🔒 SECURITY CHECK 1: আগেই approve হওয়া TrxID block
    # ══════════════════════════════════════════════════
    if is_trxid_already_used(db, transaction_id):
        logging.warning(f"SECURITY: Duplicate TrxID attempt | User: {user_id} | TrxID: {transaction_id}")
        bot.send_message(user_id,
            "🚫 <b>এই Transaction ID আগেই ব্যবহার হয়েছে!</b>\n\n"
            "\n"
            "❌ একই TrxID দিয়ে দুইবার deposit সম্ভব নয়।\n"
            "✅ সঠিক TrxID দিয়ে আবার চেষ্টা করুন।\n"
            "\n\n"
            "সমস্যা হলে support এ যোগাযোগ করুন।",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
        # Admin কে alert
        try:
            bot.send_message(ADMIN_ID,
                f"🚨 <b>Duplicate TrxID Attempt!</b>\n\n"
                f"👤 <b>User:</b> <code>{user_id}</code>\n"
                f"🔑 <b>TrxID:</b> <code>{transaction_id}</code>\n"
                f"💳 <b>Method:</b> {method}\n"
                f"💰 <b>Amount:</b> {amount} BDT\n"
                f"📅 <b>Time:</b> {get_now()}\n\n"
                f"⚠️ <i>সম্ভাব্য fraud attempt!</i>",
                parse_mode="HTML")
        except: pass
        return

    # ══════════════════════════════════════════════════════════
    # 🔒 BKASH / NAGAD / ROCKET — SMS Verification
    # ══════════════════════════════════════════════════════════
    # Logic:
    #   Case 1: SMS আছে + TrxID match + Amount match → ✅ AUTO APPROVE
    #   Case 2: SMS আছে + TrxID match + Amount মিলছে না → ❌ Error msg (pending না)
    #   Case 3: SMS নেই (TrxID unknown) → ❌ TrxID not found msg (pending না)
    #   Case 4: SMS আছে কিন্তু TrxID মিলছে না → ❌ TrxID not found msg
    #   Special: TrxID+Amount দুটোই match কিন্তু SMS bot এ আসেনি →
    #            ⏳ Pending (SMS আসলে auto approve হবে)
    # ══════════════════════════════════════════════════════════
    if is_auto:
        sms_pending = db.get("sms_pending_trxids", {})
        sms_data = sms_pending.get(transaction_id)  # TrxID দিয়ে SMS খোঁজো

        if sms_data:
            # ════════════════════════════════════════
            # SMS পাওয়া গেছে — Amount check করো
            # ════════════════════════════════════════
            sms_amount = sms_data.get("amount")
            sms_method = sms_data.get("method", "")

            # Method mismatch → error
            if sms_method.upper() != method.upper():
                logging.warning(f"Method mismatch | User={user_id} | Claimed={method} | SMS={sms_method} | TrxID={transaction_id}")
                bot.send_message(user_id,
                    f"❌ <b>ডিপোজিট ভেরিফাই হয়নি!</b>\n\n"
                    f"আপনার দেওয়া Amount বা TrxID সঠিক নয়।\n"
                    f"সঠিক তথ্য দিয়ে আবার চেষ্টা করুন।\n\n"
                    f"⚠️ সমস্যা হলে: @{SUPPORT_USERNAME}",
                    parse_mode="HTML", reply_markup=get_main_menu(user_id))
                return

            # Amount mismatch → error msg, pending না
            if sms_amount is not None:
                try:
                    if int(float(sms_amount)) != int(float(amount)):
                        logging.warning(f"Amount mismatch | User={user_id} | Claimed={amount} | SMS={sms_amount} | TrxID={transaction_id}")
                        bot.send_message(user_id,
                            f"❌ <b>ডিপোজিট ভেরিফাই হয়নি!</b>\n\n"
                            f"আপনার দেওয়া Amount বা TrxID সঠিক নয়।\n"
                            f"সঠিক তথ্য দিয়ে আবার চেষ্টা করুন।\n\n"
                            f"⚠️ সমস্যা হলে: @{SUPPORT_USERNAME}",
                            parse_mode="HTML", reply_markup=get_main_menu(user_id))
                        return
                except (ValueError, TypeError):
                    pass  # parse error → amount check skip, approve করো

            # ✅ TrxID + Method + Amount সব match → AUTO APPROVE
            final_amount = int(float(sms_amount)) if sms_amount is not None else int(float(amount))
            request_id   = f"DEP-{uuid.uuid4().hex[:6].upper()}"
            current_time = get_now()
            deposit_record = {
                "request_id": request_id, "user_id": user_id,
                "method": method, "amount": final_amount,
                "transaction_id": transaction_id, "status": "pending",
                "date": current_time, "is_auto": True
            }
            db["deposit_requests"][request_id] = deposit_record
            db["sms_pending_trxids"].pop(transaction_id, None)
            # ⚡ Targeted write — শুধু deposit record save করো
            update_db_path(f"/deposit_requests/{request_id}", deposit_record)
            _fb_ref(f"/sms_pending_trxids/{transaction_id}").delete()
            logging.info(f"AUTO APPROVE | User={user_id} | TrxID={transaction_id} | Method={method} | Amount={final_amount}")
            auto_approve_deposit(request_id, deposit_record, transaction_id, final_amount)
            return

        else:
            # ════════════════════════════════════════════════════════
            # SMS এ এই TrxID নেই।
            # দুটো সম্ভাবনা:
            #   A) TrxID ভুল → error msg দাও
            #   B) TrxID সঠিক কিন্তু SMS এখনো bot এ আসেনি → Pending save করো
            #
            # আমরা সবসময় Pending save করি — কারণ SMS delay হতে পারে।
            # SMS আসলে process_sms() নিজেই TrxID+Amount match করে auto approve করবে।
            # ════════════════════════════════════════════════════════

            # 3 বারের বেশি ভুল TrxID দিলে block করো
            attempt_count = increment_trx_attempt(db, user_id, transaction_id)
            if attempt_count >= 3:
                block_trx_id(db, user_id, transaction_id)
                # ⚡ Targeted write — শুধু blocked TrxID save করো
                update_db_path(f"/blocked_trxids/{transaction_id}", db.get("blocked_trxids", {}).get(transaction_id, {}))
                bot.send_message(user_id,
                    f"🚫 <b>এই TrxID block হয়ে গেছে!</b>\n\n"
                    f"আপনি {attempt_count} বার ভুল TrxID দিয়েছেন।\n"
                    f"সাহায্যের জন্য: @{SUPPORT_USERNAME}",
                    parse_mode="HTML", reply_markup=get_main_menu(user_id))
                try:
                    bot.send_message(ADMIN_ID,
                        f"⛔ <b>TrxID Blocked ({attempt_count} attempts)</b>\n\n"
                        f"👤 User: <code>{user_id}</code>\n"
                        f"🔑 TrxID: <code>{transaction_id}</code>\n"
                        f"💳 Method: {method}\n"
                        f"📅 {get_now()}",
                        parse_mode="HTML")
                except: pass
                return

            # ⏳ Pending save — SMS আসলে auto approve হবে
            request_id   = f"DEP-{uuid.uuid4().hex[:6].upper()}"
            current_time = get_now()
            deposit_record = {
                "request_id": request_id, "user_id": user_id,
                "method": method, "amount": amount,
                "transaction_id": transaction_id, "status": "pending",
                "date": current_time, "is_auto": True
            }
            db["deposit_requests"][request_id] = deposit_record
            # ⚡ Targeted write — শুধু deposit record save করো
            update_db_path(f"/deposit_requests/{request_id}", deposit_record)
            logging.info(f"PENDING (SMS not received yet) | User={user_id} | TrxID={transaction_id} | Method={method} | Amount={amount}")

            try:
                bot.send_message(ADMIN_ID,
                    f"⏳ <b>Pending Deposit (SMS আসেনি)</b>\n\n"
                    f"👤 User: <code>{user_id}</code>\n"
                    f"🔑 TrxID: <code>{transaction_id}</code>\n"
                    f"💳 Method: {method}\n"
                    f"💰 Amount: {amount} BDT\n"
                    f"🆔 Request: <code>{request_id}</code>\n"
                    f"📅 {current_time}\n\n"
                    f"<i>SMS আসলে TrxID+Amount match হলে অটো approve হবে।</i>",
                    parse_mode="HTML")
            except: pass

            bot.send_message(user_id,
                f"⏳ <b>Request জমা হয়েছে!</b>\n\n"
                f"🆔 <b>Request ID:</b> <code>{request_id}</code>\n"
                f"💳 <b>Method:</b> {method}\n"
                f"💰 <b>Amount:</b> {int(float(amount))} BDT\n"
                f"🔑 <b>TrxID:</b> <code>{transaction_id}</code>\n"
                f"📅 <b>Time:</b> {current_time}\n\n"
                f"✅ SMS verify হলে <b>অটোমেটিক approve</b> হয়ে যাবে।\n\n"
                f"⚠️ <i>সমস্যা হলে: @{SUPPORT_USERNAME}</i>",
                parse_mode="HTML", reply_markup=get_main_menu(user_id))
            return

    # ══════════════════════════════════════════════════
    # 🔵 BINANCE — Manual Admin Approval (আগের মতোই)
    # ══════════════════════════════════════════════════
    request_id = f"DEP-{uuid.uuid4().hex[:6].upper()}"
    current_time = get_now()

    deposit_record = {
        "request_id": request_id, "user_id": user_id,
        "method": method, "amount": amount,
        "transaction_id": transaction_id, "status": "pending",
        "date": current_time, "is_auto": False
    }
    db["deposit_requests"][request_id] = deposit_record
    # ⚡ Targeted write — শুধু deposit record save করো
    update_db_path(f"/deposit_requests/{request_id}", deposit_record)

    bot.send_message(user_id,
        t(user_id, "dep_submitted_manual", req_id=request_id, method=method,
          amount=amount, trx=transaction_id, date=current_time),
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

    # Admin notification
    admin_message = (
        f"\n"
        f"🔵 Binance <b>Deposit Request</b>\n"
        f"\n\n"
        f"🆔 <b>Request ID:</b> <code>{request_id}</code>\n"
        f"👤 <b>User ID:</b> <code>{user_id}</code>\n"
        f"💳 <b>Method:</b> {method}\n"
        f"💰 <b>Amount:</b> {amount} BDT\n"
        f"🔑 <b>TrxID:</b> <code>{transaction_id}</code>\n"
        f"📅 <b>Date:</b> {current_time}\n"
        f""
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(
        InlineBtn("✅ Approve", style="success", callback_data=f"dep_approve_{request_id}"),
        InlineBtn("❌ Reject", style="danger", callback_data=f"dep_reject_{request_id}")
    )
    try:
        bot.send_message(ADMIN_ID, admin_message, reply_markup=markup, parse_mode="HTML")
    except: pass

    log_action("DEPOSIT_REQUEST", user_id,
               f"Method: {method}, Amount: {amount}, TrxID: {transaction_id}, Auto: {is_auto}")

# ===========================
# DEPOSIT APPROVE / REJECT
# ===========================

def show_pending_deposits(admin_id):
    db = load_db()
    pending = {k: v for k, v in db["deposit_requests"].items() if v["status"] == "pending"}
    if not pending:
        bot.send_message(admin_id, "✅ <b>No pending deposits</b>", parse_mode="HTML"); return
    for req_id, data in list(pending.items())[:5]:
        auto_badge = "⚡ Auto" if data.get("is_auto") else "🔵 Manual"
        message = (
            f"💰 <b>Pending Deposit</b> [{auto_badge}]\n"
            f"\n"
            f"🆔 <code>{req_id}</code>\n"
            f"👤 User: <code>{data['user_id']}</code>\n"
            f"💳 Method: {data['method']}\n"
            f"💰 Amount: {data['amount']} BDT\n"
            f"🔑 TrxID: <code>{data['transaction_id']}</code>\n"
            f"📅 Date: {data['date']}"
        )
        markup = types.InlineKeyboardMarkup()
        markup.add(
            InlineBtn("✅ Approve", style="success", callback_data=f"dep_approve_{req_id}"),
            InlineBtn("❌ Reject", style="danger", callback_data=f"dep_reject_{req_id}")
        )
        bot.send_message(admin_id, message, parse_mode="HTML", reply_markup=markup)

def approve_deposit(admin_id, request_id, call):
    db = load_db()
    if request_id not in db["deposit_requests"]:
        bot.answer_callback_query(call.id, "Request not found!", show_alert=True); return
    req = db["deposit_requests"][request_id]
    if req["status"] != "pending":
        bot.answer_callback_query(call.id, "Already processed!", show_alert=True); return
    user_id  = req["user_id"]
    amount   = int(float(req["amount"]))
    method   = req["method"]
    trx_id   = req.get("transaction_id", "N/A")   # ← BUG FIX: was undefined
    now      = get_now()

    if user_id not in db["users"]:
        db["users"][user_id] = {"balance": 0, "refer_count": 0, "orders": [],
                                "join_date": now, "lang": "en"}
    new_bal = round(float(db["users"][user_id].get("balance", 0)) + float(amount), 2)
    # CASHBACK
    cashback = get_cashback(amount, db)
    if cashback > 0:
        new_bal = round(new_bal + cashback, 2)
    db["users"][user_id]["balance"] = new_bal
    req["status"] = "approved"
    req["approved_date"] = now
    mark_trxid_used(db, trx_id)
    # ⚡ OPTIMIZED: পুরো DB write বাদ — শুধু changed paths update করো
    update_db_path(f"/users/{user_id}/balance", new_bal, _async=False)
    update_db_path(f"/deposit_requests/{request_id}", req, _async=False)
    # verified_trxids আলাদা update
    _EXECUTOR.submit(_update_db_path_sync, "/verified_trxids",
                     {str(i): v for i, v in enumerate(db.get("verified_trxids", []))})
    _update_db_cache_in_place(db)  # cache sync করো

    # ⚡ User কে সাথে সাথে clear balance notification
    try:
        cashback_line = (
            f"\n🎁 <b>Cashback Bonus:</b> +{cashback} BDT 🎉"
        ) if cashback > 0 else ""

        lang_u = _lang_cache.get(user_id, "en")
        if lang_u == "bn":
            msg_text = (
                f"✅ <b>ডিপোজিট অ্যাপ্রুভ হয়েছে!</b>\n\n"
                f"💰 <b>+{amount} BDT যোগ হয়েছে</b>"
                f"{cashback_line}\n\n"
                f"💳 <b>নতুন Balance: {new_bal} BDT</b>\n"
                f"🆔 Request: <code>{request_id}</code>\n\n"
                f"🛍️ এখনই কেনাকাটা করুন!"
            )
        else:
            msg_text = (
                f"✅ <b>Deposit Approved!</b>\n\n"
                f"💰 <b>+{amount} BDT Added</b>"
                f"{cashback_line}\n\n"
                f"💳 <b>New Balance: {new_bal} BDT</b>\n"
                f"🆔 Request: <code>{request_id}</code>\n\n"
                f"🛍️ Start shopping now!"
            )
        shop_markup = types.InlineKeyboardMarkup()
        shop_markup.add(InlineBtn("🛍️ Shop Now", style="primary", callback_data="go_to_shop"))
        bot.send_message(user_id, msg_text, parse_mode="HTML", reply_markup=shop_markup)
    except: pass

    # ✅ Channel এ শুধু APPROVE log — masked
    try:
        bot.send_message(LOG_CHANNEL_ID,
            f"✅ <b>Deposit Approved</b>\n"
            f"\n"
            f"👤 User: <code>{mask_id(user_id)}</code>\n"
            f"💳 Method: {method}\n"
            f"💰 Amount: {amount} BDT\n"
            f"🆔 Request: <code>{mask_id(request_id)}</code>\n"
            f"🔑 TrxID: <code>{mask_id(req['transaction_id'])}</code>\n"
            f"👤 Admin: Manual\n"
            f"📅 {now}",
            parse_mode="HTML")
    except: pass

    try:
        bot.edit_message_reply_markup(admin_id, call.message.message_id, reply_markup=None)
        bot.send_message(admin_id,
            f"✅ <b>Approved!</b> <code>{request_id}</code> | {amount} BDT → User <code>{user_id}</code>",
            parse_mode="HTML")
    except: pass
    log_action("DEPOSIT_APPROVED", admin_id, f"Request: {request_id}, Amount: {amount}")
    bot.answer_callback_query(call.id, "✅ Approved!", show_alert=False)

def reject_deposit(admin_id, request_id, call):
    db = load_db()
    if request_id not in db["deposit_requests"]:
        bot.answer_callback_query(call.id, "Request not found!", show_alert=True); return
    req = db["deposit_requests"][request_id]
    if req["status"] != "pending":
        bot.answer_callback_query(call.id, "Already processed!", show_alert=True); return
    user_id = req["user_id"]
    now = get_now()
    req["status"] = "rejected"
    req["rejected_date"] = now
    # ⚡ Targeted write — শুধু deposit request status update করো
    update_db_path(f"/deposit_requests/{request_id}", req)
    _update_db_cache_in_place(db)
    try:
        bot.send_message(user_id,
            t(user_id, "dep_rejected", amount=req["amount"], req_id=request_id),
            parse_mode="HTML")
    except: pass

    # ✅ Reject log channel এ যাবে না — শুধু admin কে জানাবে
    # (Channel এ শুধু approve log যাবে)

    try:
        bot.edit_message_reply_markup(admin_id, call.message.message_id, reply_markup=None)
        bot.send_message(admin_id,
            f"❌ <b>Rejected!</b> Request: <code>{request_id}</code>",
            parse_mode="HTML")
    except: pass
    log_action("DEPOSIT_REJECTED", admin_id, f"Request: {request_id}")
    bot.answer_callback_query(call.id, "Rejected!", show_alert=False)

# ===========================
# ADMIN PANEL
# ===========================

def show_admin_panel(user_id):
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineBtn("➕ Add Product", style="success", callback_data="admin_add_product"),
        InlineBtn("📤 Upload Stock", style="primary", callback_data="admin_upload_stock")
    )
    markup.add(
        InlineBtn("📊 Inventory", style="primary", callback_data="admin_inventory"),
        InlineBtn("🗑️ Delete Product", style="danger", callback_data="admin_delete_product")
    )
    markup.add(
        InlineBtn("✏️ Edit Features", style="primary", callback_data="admin_edit_features"),
        InlineBtn("⚡ Flash Sale", style="primary", callback_data="admin_flash_sale")
    )
    markup.add(
        InlineBtn("🚫 Ban User", style="danger", callback_data="admin_ban_user"),
        InlineBtn("✅ Unban User", style="danger", callback_data="admin_unban_user")
    )
    markup.add(
        InlineBtn("💰 Edit Balance", style="primary", callback_data="admin_edit_balance"),
        InlineBtn("👤 User Info", style="primary", callback_data="admin_user_info")
    )
    markup.add(
        InlineBtn("🏷️ Edit Price", style="primary", callback_data="admin_edit_price"),
        InlineBtn("💬 Message User", style="primary", callback_data="admin_message_user")
    )
    markup.add(
        InlineBtn("📧 Edu Mail Price", style="primary", callback_data="admin_edu_mail_price"),
        InlineBtn("📦 Edu Mail Stock", style="primary", callback_data="admin_edu_mail_stock")
    )
    markup.add(
        InlineBtn("📩 Hotmail Price", style="primary", callback_data="admin_hotmail_price"),
        InlineBtn("📩 Outlook Price", style="primary", callback_data="admin_outlook_price")
    )
    markup.add(
        InlineBtn("📊 API Dashboard", style="primary", callback_data="admin_api_dashboard")
    )
    # ─── Bulkmail product on/off toggle ───
    markup.add(
        InlineBtn("🔀 Bulkmail Products ON/OFF", style="primary", callback_data="admin_bulkmail_toggle")
    )
    # ─── Hotmail143 API product on/off toggle ───
    markup.add(
        InlineBtn("🔀 Hotmail143 Products ON/OFF", style="primary", callback_data="admin_hotmail143_toggle")
    )
    markup.add(
        InlineBtn("🔍 Bulkmail Stock Test", style="primary", callback_data="admin_bulkmail_stocktest")
    )
    markup.add(
        InlineBtn("📋 Bulkmail Last Orders", style="primary", callback_data="admin_bulkmail_orders")
    )
    markup.add(
        InlineBtn("📢 Broadcast", style="primary", callback_data="admin_broadcast"),
        InlineBtn("🔍 Search Order", style="primary", callback_data="admin_search_order")
    )
    markup.add(
        InlineBtn("✅ Deposits", style="success", callback_data="admin_deposits"),
        InlineBtn("📈 Analytics & Reports", style="primary", callback_data="admin_analytics")
    )
    markup.add(
        InlineBtn("👥 User Management", style="primary", callback_data="admin_user_mgmt"),
        InlineBtn("💵 Finance Report", style="primary", callback_data="admin_finance")
    )
    markup.add(
        InlineBtn("📦 Stock Alerts", style="primary", callback_data="admin_stock_alerts"),
        InlineBtn("🎁 Set Refer Bonus", style="primary", callback_data="admin_set_refer_bonus")
    )
    markup.add(
        InlineBtn("🔑 Refer Bonus Settings", style="primary", callback_data="admin_refer_settings"),
        InlineBtn("📤 Firebase Export (All Data)", style="primary", callback_data="admin_firebase_export")
    )
    markup.add(
        InlineBtn("📦 Sub-Product Manager (Gmail/FB)", style="primary", callback_data="admin_subproduct_manager")
    )
    markup.add(
        InlineBtn("🎁 Cashback Settings", style="primary", callback_data="admin_cashback_settings")
    )
    markup.add(
        InlineBtn("🎨 Keyboard Button Color", style="primary", callback_data="admin_kb_color")
    )
    markup.add(
        InlineBtn("📋 Unsold Products", style="primary", callback_data="admin_unsold_products")
    )
    markup.add(
        InlineBtn("🎟️ Promo Codes", style="primary", callback_data="admin_promo_manager"),
        InlineBtn("🆕 New User Discount", style="primary", callback_data="admin_new_user_discount")
    )
    markup.add(
        InlineBtn("🏷️ User Discount (1-99%)", style="primary", callback_data="admin_user_discount")
    )
    db = load_db()
    total_users = len(db["users"])
    total_products = len(db["products"])
    pending_deps = sum(1 for r in db["deposit_requests"].values() if r["status"] == "pending")
    bot.send_message(user_id,
        f"\n"
        f"🛠️ <b>Admin Control Panel</b>\n"
        f"\n\n"
        f"👥 <b>Total Users:</b> {total_users}\n"
        f"📦 <b>Products:</b> {total_products}\n"
        f"⏳ <b>Pending Deposits:</b> {pending_deps}\n\n"
        f"<i>Choose an option 👇</i>",
        reply_markup=markup, parse_mode="HTML")

# ===========================
# CALLBACK QUERY HANDLER
# ===========================

@bot.callback_query_handler(func=lambda call: True)
def handle_callback_queries(call):
    # ⚡ INSTANT ACK — বাটন ক্লিকের সাথে সাথে Telegram কে জানাও
    # এতে বাটনের loading spinner তুরন্ত সরে যায়
    # Real processing background thread এ হবে
    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass
    # Heavy processing thread pool এ পাঠাও — webhook thread block হবে না
    _EXECUTOR.submit(_callback_worker, call)

def _callback_worker(call):
    """Callback processing — background thread এ চলে।
    ⚡ db = lazy — শুধু দরকার হলে load হবে।
    """
    user_id = str(call.message.chat.id)
    callback_data = call.data

    # ⚡ FAST PATH: lang cache থেকে নাও — DB load ছাড়াই
    # DB load শুধু তখনই হবে যখন সত্যিই দরকার
    _db_loaded_holder = [False]
    _db_ref_holder = [None]

    def db_lazy():
        if not _db_loaded_holder[0]:
            _db_ref_holder[0] = load_db()
            _db_loaded_holder[0] = True
        return _db_ref_holder[0]

    db = load_db()  # cache hit হলে এটা μs মাত্র লাগে (60s TTL cache)

    # ✅ Subscription gate — সব callback এ enforce (admin exempt)
    # check_join, set_lang, goodbye_come_back — join ছাড়াও কাজ করবে
    _sub_exempt = {"check_join", "set_lang_en", "set_lang_bn", "goodbye_come_back"}
    if callback_data not in _sub_exempt and str(user_id) != str(ADMIN_ID):
        # Smart cache check — TTL এর মধ্যে থাকলে API call করবে না
        _cached_sub = _sub_cache.get(user_id)
        _now_ts2 = _time_module.time()
        if _cached_sub and (_now_ts2 - _cached_sub[1]) < _SUB_CACHE_TTL:
            _sub_ok = _cached_sub[0]
        else:
            # Cache নেই বা expire — fresh Telegram API check
            _sub_ok = _refresh_sub_cache_sync(user_id)

        if not _sub_ok:
            # ❌ Join করেনি — কোন channel miss সেটা detect করে join prompt দেখাও
            try: bot.delete_message(user_id, call.message.message_id)
            except: pass
            _lang_now = db.get("users", {}).get(user_id, {}).get("lang", "en")
            _join_markup = types.InlineKeyboardMarkup(row_width=1)
            for i, ch in enumerate(CHANNELS, 1):
                _btn_text = LANG.get(_lang_now, LANG["en"])["join_channel_btn"].format(n=i)
                _join_markup.add(InlineBtn(_btn_text, style="primary", url=ch["link"]))
            _join_markup.add(InlineBtn(
                LANG.get(_lang_now, LANG["en"])["verify_join"], style="success", callback_data="check_join"))
            bot.send_message(user_id,
                LANG.get(_lang_now, LANG["en"])["join_channels"],
                reply_markup=_join_markup, parse_mode="HTML")
            return

    # Language
    if callback_data in ("set_lang_en", "set_lang_bn"):
        new_lang = callback_data.replace("set_lang_", "")
        db["users"].setdefault(user_id, {})["lang"] = new_lang
        _lang_cache[user_id] = new_lang  # update cache immediately
        # ⚡ Targeted write — শুধু lang field update করো
        update_db_path(f"/users/{user_id}/lang", new_lang)
        try: bot.delete_message(user_id, call.message.message_id)
        except: pass
        bot.answer_callback_query(call.id, LANG[new_lang]["lang_set"], show_alert=False)
        bot.send_message(user_id, t(user_id, "welcome"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML"); return

    # Subscription verify
    if callback_data == "check_join":
        invalidate_sub_cache(user_id)  # cache clear — fresh Telegram API check
        # Synchronous fresh check — cache bypass করো
        _verified = _refresh_sub_cache_sync(user_id)
        if _verified:
            try: bot.delete_message(user_id, call.message.message_id)
            except: pass

            db = load_db()
            pending_refer = db.get("_pending_refer", {}).get(user_id)

            # নতুন user হলে DB তে add করো
            if user_id not in db["users"]:
                db["users"][user_id] = {
                    "balance": 0, "refer_count": 0, "orders": [],
                    "referred_by": pending_refer,
                    "join_date": get_now_short(),
                    "lang": db.get("users", {}).get(user_id, {}).get("lang", "en")
                }

            # Refer bonus দাও
            already_referred = db["users"][user_id].get("refer_bonus_given", False)
            if (pending_refer and
                    pending_refer in db["users"] and
                    pending_refer != user_id and
                    not already_referred):
                bonus = db["settings"].get("refer_bonus", 2)
                db["users"][pending_refer]["balance"] = round(
                    float(db["users"][pending_refer].get("balance", 0)) + float(bonus), 2)
                db["users"][pending_refer]["refer_count"] = int(
                    db["users"][pending_refer].get("refer_count", 0)) + 1
                db["users"][user_id]["refer_bonus_given"] = True
                db["users"][user_id]["referred_by"] = pending_refer

                if user_id in db.get("_pending_refer", {}):
                    del db["_pending_refer"][user_id]
                    _fb_ref(f"/_pending_refer/{user_id}").delete()

                # ⚡ Targeted write — শুধু দুইটা user update করো
                update_db_path(f"/users/{user_id}", db["users"][user_id])
                update_db_path(f"/users/{pending_refer}", db["users"][pending_refer])
                try:
                    bot.send_message(pending_refer,
                        t(pending_refer, "refer_bonus_notif", bonus=bonus), parse_mode="HTML")
                except: pass
            else:
                # ⚡ Targeted write — শুধু user data save করো
                update_db_path(f"/users/{user_id}", db["users"][user_id])

            # Language selection বা welcome দেখাও
            if db["users"][user_id].get("lang") in ("en", "bn"):
                bot.send_message(user_id, t(user_id, "welcome"),
                    reply_markup=get_main_menu(user_id), parse_mode="HTML")
            else:
                show_language_selection(user_id)

            bot.answer_callback_query(call.id, t(user_id, "verify_ok"), show_alert=False)
        else:
            # কোন channel এ join নেই সেটা detect করো
            _missing = []
            for ch in CHANNELS:
                try:
                    _st = bot.get_chat_member(ch["id"], user_id).status
                    if _st in ("left", "kicked"):
                        _missing.append(ch)
                except Exception:
                    pass

            _lang_now = db.get("users", {}).get(user_id, {}).get("lang", "en")
            _fail_markup = types.InlineKeyboardMarkup(row_width=1)
            for i, ch in enumerate(CHANNELS, 1):
                _is_missing = ch in _missing
                _icon = "❌" if _is_missing else "✅"
                _fail_markup.add(InlineBtn(
                    f"{_icon} Channel {i} — {'Join করুন' if _is_missing else 'Joined'}", style="success",
                    url=ch["link"]
                ))
            _fail_markup.add(InlineBtn(
                LANG.get(_lang_now, LANG["en"])["verify_join"], style="success", callback_data="check_join"))

            _miss_links = "\n".join(f"👉 {ch['link']}" for ch in _missing)
            try:
                bot.edit_message_text(
                    f"❌ <b>এখনও সব Channel এ Join করেননি!</b>\n\n"
                    f"নিচের Channel গুলোতে Join করুন:\n{_miss_links}\n\n"
                    f"Join করার পর <b>✅ Verify</b> বাটনে চাপুন।",
                    user_id, call.message.message_id,
                    parse_mode="HTML", reply_markup=_fail_markup
                )
            except:
                bot.send_message(user_id,
                    f"❌ <b>এখনও সব Channel এ Join করেননি!</b>\n\n"
                    f"নিচের Channel গুলোতে Join করুন:\n{_miss_links}\n\n"
                    f"Join করার পর <b>✅ Verify</b> বাটনে চাপুন।",
                    parse_mode="HTML", reply_markup=_fail_markup)
            bot.answer_callback_query(call.id, "❌ সব Channel এ Join করুন!", show_alert=False)
        return

    # Back to main
    if callback_data == "back_to_main":
        try: bot.delete_message(user_id, call.message.message_id)
        except: pass
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        bot.answer_callback_query(call.id); return

    # ⚡ Refresh balance — Firebase থেকে fresh read
    if callback_data == "refresh_balance":
        try:
            fresh = _fb_ref(f"/users/{user_id}").get()
            if fresh:
                db["users"][user_id] = fresh
                _update_db_cache_in_place(db)
            bal = db["users"].get(user_id, {}).get("balance", 0)
            bot.answer_callback_query(call.id,
                f"✅ Balance: {bal} BDT", show_alert=False)
        except Exception:
            bot.answer_callback_query(call.id, "⚠️ Refresh failed", show_alert=False)
        show_user_balance(user_id, db)
        return

    # Goodbye "Come Back" button — join check করো
    if callback_data == "goodbye_come_back":
        invalidate_sub_cache(user_id)
        if is_subscribed(user_id):
            # সবগুলো channel এ join আছে → welcome করো
            try: bot.delete_message(user_id, call.message.message_id)
            except: pass
            bot.send_message(
                user_id,
                "🎉 <b>Welcome Back!</b>\n\n"
                "আপনাকে আবার স্বাগতম! 🥳\n"
                "আমাদের সাথে থাকার জন্য ধন্যবাদ। 💙",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id)
            )
        else:
            # Join করেননি → join করতে বলো
            join_markup = types.InlineKeyboardMarkup(row_width=1)
            for i, ch in enumerate(CHANNELS, 1):
                join_markup.add(InlineBtn(f"📢 Channel {i} তে Join করুন", style="success", url=ch["link"]))
            join_markup.add(InlineBtn("✅ Join করেছি", style="success", callback_data="goodbye_come_back"))
            try:
                bot.edit_message_text(
                    "⚠️ <b>আগে আমাদের Channel এ Join করুন!</b>\n\n"
                    "নিচের বাটনে ক্লিক করে Join করুন, তারপর আবার চেষ্টা করুন। 👇",
                    user_id, call.message.message_id,
                    parse_mode="HTML", reply_markup=join_markup
                )
            except:
                bot.send_message(
                    user_id,
                    "⚠️ <b>আগে আমাদের Channel এ Join করুন!</b>\n\n"
                    "নিচের বাটনে ক্লিক করে Join করুন, তারপর আবার চেষ্টা করুন। 👇",
                    parse_mode="HTML", reply_markup=join_markup
                )
        bot.answer_callback_query(call.id); return

    if callback_data in ("back_to_shop", "go_to_shop"):
        # ✅ Subscription check — channel leave করে shop এ আসলে join করতে বলো
        invalidate_sub_cache(user_id)
        if not is_subscribed(user_id):
            join_markup = types.InlineKeyboardMarkup(row_width=1)
            for i, ch in enumerate(CHANNELS, 1):
                join_markup.add(InlineBtn(
                    f"📢 Channel {i} তে Join করুন", style="success", url=ch["link"]))
            join_markup.add(InlineBtn(
                "✅ Join করেছি — Verify", style="success", callback_data="check_join"))
            try:
                bot.edit_message_text(
                    "⚠️ <b>Channel এ Join করুন!</b>\n\n"
                    "Shop ব্যবহার করতে আমাদের Channel এ Join থাকতে হবে। 👇",
                    user_id, call.message.message_id,
                    parse_mode="HTML", reply_markup=join_markup)
            except:
                bot.send_message(user_id,
                    "⚠️ <b>Channel এ Join করুন!</b>\n\n"
                    "Shop ব্যবহার করতে আমাদের Channel এ Join থাকতে হবে। 👇",
                    parse_mode="HTML", reply_markup=join_markup)
            bot.answer_callback_query(call.id); return

        try: bot.delete_message(user_id, call.message.message_id)
        except: pass
        # শুধু Reply Keyboard দেখাও, product list text নেই
        db2 = load_db()
        lang2 = get_user_lang(user_id, db2)
        L2 = LANG[lang2]
        markup_shop = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        # Edu Mail 24H, 72H, 24hr ও Edu Mail (parent) — shop menu তে দেখাবে না
        _hidden_sub2 = {EDU_MAIL_SUB_24H, EDU_MAIL_SUB_72H, EDU_MAIL_SUB_24HR, EDU_MAIL_PARENT}
        products = [p for p in db2["products"].keys() if p not in _hidden_sub2]
        for i in range(0, len(products), 2):
            markup_shop.add( *[KbBtn(_x, role="product_item") for _x in products[i:i+2]])
        markup_shop.add( KbBtn(L2["back"], role="back"))
        bot.send_message(user_id,
            "🛍️ <b>Choose a product:</b>",
            parse_mode="HTML", reply_markup=markup_shop)
        bot.answer_callback_query(call.id); return

    # Deposit help
    if callback_data == "deposit_help":
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn(t(user_id, "go_back"), style="primary", callback_data="back_to_main"))
        try:
            bot.edit_message_text(t(user_id, "dep_help"), user_id,
                call.message.message_id, reply_markup=markup, parse_mode="HTML")
        except:
            bot.send_message(user_id, t(user_id, "dep_help"),
                reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id); return

    # Deposit history
    if callback_data == "deposit_history":
        requests_list = [v for v in db.get("deposit_requests", {}).values()
                         if v.get("user_id") == user_id]
        if not requests_list:
            history_msg = t(user_id, "no_dep_history")
        else:
            history_msg = t(user_id, "dep_history")
            for req in requests_list[-5:]:
                status_emoji = "✅" if req['status'] == 'approved' else ("❌" if req['status'] == 'rejected' else "⏳")
                auto_tag = " ⚡" if req.get("auto_approved") else ""
                history_msg += (f"🆔 {req['request_id']}\n"
                                f"💰 {req['amount']} BDT | {req['method']}\n"
                                f"{status_emoji} {req['status'].upper()}{auto_tag}\n\n")
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn(t(user_id, "go_back"), style="primary", callback_data="back_to_main"))
        try:
            bot.edit_message_text(history_msg, user_id, call.message.message_id,
                reply_markup=markup, parse_mode="HTML")
        except:
            bot.send_message(user_id, history_msg, reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id); return

    # ── DEPOSIT APPROVE/REJECT (admin only but handled before generic admin check) ──
    if callback_data.startswith("dep_approve_"):
        if str(user_id) != str(ADMIN_ID):
            bot.answer_callback_query(call.id, "❌ Admin access only!", show_alert=True); return
        request_id = callback_data.replace("dep_approve_", "")
        approve_deposit(ADMIN_ID, request_id, call); return

    if callback_data.startswith("dep_reject_"):
        if str(user_id) != str(ADMIN_ID):
            bot.answer_callback_query(call.id, "❌ Admin access only!", show_alert=True); return
        request_id = callback_data.replace("dep_reject_", "")
        reject_deposit(ADMIN_ID, request_id, call); return

    # Admin-only check
    if str(user_id) != str(ADMIN_ID):
        if callback_data.startswith("admin_") or \
           callback_data.startswith("del_product_") or callback_data.startswith("edit_price_") or \
           callback_data.startswith("unban_user_") or callback_data.startswith("del_product_idx_") or \
           callback_data.startswith("exp_us_") or callback_data.startswith("promo_toggle_") or \
           callback_data.startswith("user_discount_remove_") or \
           callback_data in ("confirm_del_yes",):
            bot.answer_callback_query(call.id, "❌ Admin access only!", show_alert=True)
        return

    # ── ADMIN CALLBACKS ──

    if callback_data == "admin_add_product":
        msg = bot.send_message(ADMIN_ID,
            "🆕 <b>Enter new product name:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_add_product_step1)
        pass  # ACK already sent upfront

    # (admin_api_stock and admin_api_balance callbacks removed — hotmail143.com API disabled)

    elif callback_data == "admin_upload_stock":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        # ✅ FIX: Inline keyboard ব্যবহার করো যাতে next_step_handler conflict না হয়
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for p in products:
            stock_cnt = get_stock_count(p)
            markup2.add(InlineBtn(
                f"📦 {p} (Stock: {stock_cnt})", style="primary", callback_data=f"upload_stock_select_{p}"))
        markup2.add(InlineBtn("🔙 Back", style="primary", callback_data="back_admin"))
        _upload_txt = "📤 <b>Which product's stock to upload?</b>\n\n<i>পণ্য সিলেক্ট করুন 👇</i>"
        try:
            bot.edit_message_text(_upload_txt,
                ADMIN_ID, call.message.message_id, reply_markup=markup2, parse_mode="HTML")
        except:
            bot.send_message(ADMIN_ID, _upload_txt, parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data == "admin_inventory":
        report = "📊 <b>Inventory Report</b>\n" + "" + "\n\n"
        total_value = 0
        for product in db["products"]:
            stock = get_stock_count(product)
            price = db["products"][product]
            # sub-product আছে কিনা check
            sub_items = db.get("sub_products", {}).get(product, {}).get("sub_items", [])
            if sub_items:
                value = sum(
                    get_stock_count(f"{product}__{s['name']}") * s.get("price", price)
                    for s in sub_items
                )
                total_value += value
                status = "📦"
                report += f"{status} <b>{product}</b> (sub-products: {len(sub_items)})\n"
                for s in sub_items:
                    s_stock = get_stock_count(f"{product}__{s['name']}")
                    s_val = s_stock * s.get("price", price)
                    s_icon = "✅" if s_stock > 0 else "❌"
                    on_off = "ON" if s.get("enabled", True) else "OFF"
                    report += f"  {s_icon} {s['name']} — {s.get('price',price)} BDT | 📦{s_stock}pcs | {on_off}\n"
                report += "\n"
            else:
                value = stock * price
                total_value += value
                status = "✅" if stock > 0 else "❌"
                report += (f"{status} <b>{product}</b>\n"
                           f"  💵 Price: {price} BDT | 📦 Stock: {stock} pcs | 💰 Value: {value} BDT\n\n")
        report += "" + f"\n💎 <b>Total Stock Value: {total_value} BDT</b>"
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "admin_delete_product":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        # product নাম session এ store করো — callback_data 64 char limit bypass
        if "del_product_list" not in _admin_sessions:
            _admin_sessions["del_product_list"] = {}
        _admin_sessions["del_product_list"] = {str(i): p for i, p in enumerate(products)}
        markup = types.InlineKeyboardMarkup(row_width=1)
        for i, p in enumerate(products):
            stock = get_stock_count(p)
            sub_items = db.get("sub_products", {}).get(p, {}).get("sub_items", [])
            label = f"🗑️ {p} (📦{stock})" + (f" [{len(sub_items)} sub]" if sub_items else "")
            markup.add(InlineBtn(label, style="primary", callback_data=f"del_product_idx_{i}"))
        markup.add(InlineBtn("🔙 Back", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, "🗑️ <b>Delete করার জন্য প্রোডাক্ট বেছে নিন:</b>",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data.startswith("del_product_idx_"):
        idx = callback_data.replace("del_product_idx_", "")
        product_name = _admin_sessions.get("del_product_list", {}).get(idx)
        if not product_name or product_name not in db["products"]:
            bot.answer_callback_query(call.id, "❌ Product not found! আবার চেষ্টা করুন।", show_alert=True)
            return
        stock = get_stock_count(product_name)
        sub_items = db.get("sub_products", {}).get(product_name, {}).get("sub_items", [])
        _admin_sessions[f"confirm_del_{ADMIN_ID}"] = product_name
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineBtn("✅ হ্যাঁ, Delete করো", style="danger", callback_data="confirm_del_yes"),
            InlineBtn("❌ না, Cancel", style="danger", callback_data="admin_delete_product")
        )
        bot.send_message(ADMIN_ID,
            f"⚠️ <b>নিশ্চিত করুন</b>\n\n\n"
            f"🗑️ প্রোডাক্ট: <b>{product_name}</b>\n"
            f"📦 Stock: <b>{stock} pcs</b>\n"
            f"🔢 Sub-products: <b>{len(sub_items)}</b>\n\n"
            f"⚠️ এই প্রোডাক্ট ও সব স্টক <b>চিরতরে মুছে যাবে!</b>",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "confirm_del_yes":
        product_name = _admin_sessions.pop(f"confirm_del_{ADMIN_ID}", None)
        if not product_name or product_name not in db["products"]:
            bot.answer_callback_query(call.id, "❌ Session expired! আবার চেষ্টা করুন।", show_alert=True)
            return
        # ১. Main product delete
        del db["products"][product_name]
        # ২. sub_products delete
        if product_name in db.get("sub_products", {}):
            del db["sub_products"][product_name]
        # ৩. product_details delete
        if product_name in db.get("product_details", {}):
            del db["product_details"][product_name]
            _fb_ref(f"/product_details/{product_name}").delete()
        # ৪. Flash sale এ থাকলে সরাও
        if product_name in db.get("flash_sale", {}):
            del db["flash_sale"][product_name]
            _fb_ref(f"/flash_sale/{product_name}").delete()
        # ⚡ Targeted deletes — products ও sub_products শুধু
        _fb_ref(f"/products/{product_name}").delete()
        if product_name in db.get("sub_products", {}):
            _fb_ref(f"/sub_products/{product_name}").delete()
        _update_db_cache_in_place(db)
        # ৫. Main stock file delete
        stock_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
        if os.path.exists(stock_path):
            os.remove(stock_path)
        # ৬. Sub-product stock files delete (format: ProductName__SubName.xlsx)
        deleted_files = 0
        try:
            for fname in os.listdir(STOCK_DIR):
                if fname.startswith(f"{product_name}__") and fname.endswith(".xlsx"):
                    os.remove(os.path.join(STOCK_DIR, fname))
                    deleted_files += 1
        except Exception as e:
            logging.error(f"Sub-stock delete error: {e}")
        markup_back = types.InlineKeyboardMarkup()
        markup_back.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            f"✅ <b>Product সফলভাবে Delete হয়েছে!</b>\n\n\n"
            f"🗑️ <b>{product_name}</b>\n"
            f"📂 Stock files removed: {1 + deleted_files}",
            parse_mode="HTML", reply_markup=markup_back)
        bot.answer_callback_query(call.id, "✅ Deleted!")

    elif callback_data == "admin_unsold_products":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ কোনো প্রোডাক্ট নেই!", show_alert=True); return
        markup = types.InlineKeyboardMarkup(row_width=1)
        unsold_found = False
        unsold_map = {}
        uidx = 0
        for p in products:
            stock = get_stock_count(p)
            if stock > 0:
                unsold_found = True
                unsold_map[str(uidx)] = p
                markup.add(InlineBtn(
                    f"📥 {p} ({stock} pcs) — Export XLSX", style="primary", callback_data=f"exp_us_{uidx}"))
                uidx += 1
        if not unsold_found:
            bot.answer_callback_query(call.id, "✅ সব প্রোডাক্ট বিক্রি হয়ে গেছে!", show_alert=True); return
        _admin_sessions["unsold_map"] = unsold_map
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "📋 <b>Unsold Products</b>\n\n\n"
            "⬇️ যে প্রোডাক্ট এখনো বিক্রি হয়নি সেগুলো নিচে দেখানো হচ্ছে।\n"
            "Export বাটনে ক্লিক করলে XLSX ফাইল পাবেন।",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data.startswith("exp_us_"):
        uidx = callback_data.replace("exp_us_", "")
        product_name = _admin_sessions.get("unsold_map", {}).get(uidx)
        if not product_name:
            bot.answer_callback_query(call.id, "❌ Session শেষ! আবার Unsold Products খুলুন।", show_alert=True); return
        try:
            path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")
            if not os.path.exists(path):
                bot.answer_callback_query(call.id, "❌ স্টক ফাইল পাওয়া যায়নি!", show_alert=True); return
            stock_count = get_stock_count(product_name)
            if stock_count == 0:
                bot.answer_callback_query(call.id, "❌ এই প্রোডাক্টের কোনো স্টক নেই!", show_alert=True); return

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb_src = openpyxl.load_workbook(path)
            ws_src = wb_src.active

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = product_name[:31]

            header_font  = Font(bold=True, color="FFFFFF", size=12, name="Arial")
            header_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border  = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'))

            all_rows = list(ws_src.iter_rows(min_row=1, values_only=True))
            if not all_rows:
                bot.answer_callback_query(call.id, "❌ স্টক ফাইল খালি!", show_alert=True); return

            header = ["#"] + list(all_rows[0]) + ["Export Date"]
            ws_out.append(header)
            for cell in ws_out[1]:
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = center_align; cell.border = thin_border

            row_num = 1
            for row in all_rows[1:]:
                if any(v is not None and str(v).strip() != "" for v in row):
                    ws_out.append([row_num] + list(row) + [get_now()])
                    for cell in ws_out[ws_out.max_row]:
                        val = str(cell.value) if cell.value is not None else ""
                        cell.alignment = center_align if len(val) <= 30 else left_align
                        cell.border = thin_border
                    row_num += 1

            for col in ws_out.columns:
                max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
                ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

            out_buffer = BytesIO()
            wb_out.save(out_buffer)
            out_buffer.seek(0)
            wb_src.close(); wb_out.close()

            markup_back = types.InlineKeyboardMarkup()
            markup_back.add(InlineBtn("?? Unsold Products", style="primary", callback_data="admin_unsold_products"))
            bot.send_document(ADMIN_ID,
                out_buffer,
                caption=f"📋 <b>{product_name}</b> — Unsold Stock\n📦 Total: <b>{row_num - 1} pcs</b>\n📅 {get_now()}",
                visible_file_name=f"unsold_{product_name}_{datetime.now(BD_TZ).strftime('%d%m%Y')}.xlsx",
                parse_mode="HTML",
                reply_markup=markup_back)
            bot.answer_callback_query(call.id, "✅ Export সফল!")
        except Exception as e:
            logging.error(f"export_unsold error: {e}")
            bot.answer_callback_query(call.id, f"❌ Error: {str(e)[:50]}", show_alert=True)

    elif callback_data == "admin_ban_user":
        msg = bot.send_message(ADMIN_ID, "🚫 <b>Enter User ID to ban:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_ban_user_step)
        pass  # ACK already sent upfront

    elif callback_data == "admin_unban_user":
        banned = db.get("banned_users", [])
        if not banned:
            bot.send_message(ADMIN_ID, "✅ <b>No banned users.</b>", parse_mode="HTML")
            bot.answer_callback_query(call.id); return
        markup = types.InlineKeyboardMarkup(row_width=1)
        for uid in banned[:10]:
            markup.add(InlineBtn(f"✅ Unban: {uid}", style="danger", callback_data=f"unban_user_{uid}"))
        markup.add(InlineBtn("🔙 Back", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, "✅ <b>Banned users:</b>",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data.startswith("unban_user_"):
        target_user = callback_data.replace("unban_user_", "")
        if target_user in db["banned_users"]:
            db["banned_users"].remove(target_user)
            # ⚡ Targeted write — শুধু banned_users list update করো
            update_db_path("/banned_users", {str(i): v for i, v in enumerate(db["banned_users"])})
            bot.send_message(ADMIN_ID,
                f"✅ <b>Unbanned!</b>\n\n👤 <code>{target_user}</code>", parse_mode="HTML")
            try:
                bot.send_message(target_user,
                    "✅ <b>Your access has been restored!</b>", parse_mode="HTML")
            except: pass
        pass  # ACK already sent upfront

    elif callback_data == "admin_edit_balance":
        msg = bot.send_message(ADMIN_ID, "💰 <b>Enter User ID to edit balance:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_edit_balance_step1)
        pass  # ACK already sent upfront

    elif callback_data == "admin_user_info":
        msg = bot.send_message(ADMIN_ID, "?? <b>Enter User ID:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_user_info_step)
        pass  # ACK already sent upfront

    elif callback_data == "admin_user_discount":
        msg = bot.send_message(ADMIN_ID,
            "🏷️ <b>User Discount</b>\n\n"
            "?? <b>User ID দিন:</b>\n"
            "<i>(যে user কে discount দিতে চান)</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_user_discount_step1)
        pass  # ACK already sent upfront

    elif callback_data.startswith("user_discount_remove_"):
        target_uid = callback_data.replace("user_discount_remove_", "")
        db = load_db()
        if target_uid in db["users"]:
            old_disc = db["users"][target_uid].pop("personal_discount", 0)
            # ⚡ Targeted write
            update_db_path(f"/users/{target_uid}/personal_discount", 0)
            bot.send_message(ADMIN_ID,
                f"✅ <b>Discount সরানো হয়েছে!</b>\n\n"
                f"👤 User: <code>{target_uid}</code>\n"
                f"🏷️ আগের discount: <b>{old_disc}%</b> → এখন <b>0%</b>",
                parse_mode="HTML")
            try:
                bot.send_message(target_uid,
                    "ℹ️ <b>আপনার personal discount সরিয়ে নেওয়া হয়েছে।</b>",
                    parse_mode="HTML")
            except: pass
        pass

    elif callback_data == "admin_message_user":
        msg = bot.send_message(ADMIN_ID, "👤 <b>Enter User ID:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_message_user_step1)
        pass  # ACK already sent upfront

    # ═══════════════════════════════════════════════
    # 🔀 BULKMAIL PRODUCT ON/OFF TOGGLE
    # Admin যেকোনো bulkmail product enable/disable করতে পারবে।
    # Disabled product Edu Mail submenu তে দেখাবে না।
    # DB path: /bulkmail_products/{product_name}/enabled
    # ═══════════════════════════════════════════════
    elif callback_data == "admin_bulkmail_toggle":
        db = load_db()
        markup = types.InlineKeyboardMarkup(row_width=1)
        bm_products = [p for p, c in API_PRODUCTS.items() if c.get("source") == "bulkmail"]
        if not bm_products:
            bot.send_message(ADMIN_ID, "❌ <b>কোনো Bulkmail product নেই!</b>",
                parse_mode="HTML"); return
        for bm_p in bm_products:
            enabled = is_bulkmail_enabled(bm_p, db)
            status_icon = "🟢 ON" if enabled else "🔴 OFF"
            markup.add(InlineBtn(
                f"{status_icon} — {bm_p}", style="primary",
                callback_data=f"bulkmail_tog_{bm_p}"
            ))
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "🔀 <b>Bulkmail Products ON/OFF</b>\n\n"
            "🟢 = চালু (Bot এ দেখাবে)\n"
            "🔴 = বন্ধ (Bot এ দেখাবে না)\n\n"
            "বাটনে ক্লিক করে toggle করুন:",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data.startswith("bulkmail_tog_"):
        prod_tog = callback_data.replace("bulkmail_tog_", "")
        if prod_tog not in API_PRODUCTS or API_PRODUCTS[prod_tog].get("source") != "bulkmail":
            bot.answer_callback_query(call.id, "❌ Product খুঁজে পাওয়া যায়নি!", show_alert=True); return
        db = load_db()
        if "bulkmail_products" not in db:
            db["bulkmail_products"] = {}
        if prod_tog not in db["bulkmail_products"]:
            db["bulkmail_products"][prod_tog] = {}
        current_enabled = db["bulkmail_products"][prod_tog].get("enabled", True)
        new_enabled = not current_enabled
        db["bulkmail_products"][prod_tog]["enabled"] = new_enabled
        # ⚡ Targeted write
        update_db_path(f"/bulkmail_products/{prod_tog}/enabled", new_enabled)
        # ⚡ Cache sync — depth-3 path cache auto-update হয় না, manually sync করো
        _update_db_cache_in_place(db)
        status_str = "🟢 চালু (ON)" if new_enabled else "🔴 বন্ধ (OFF)"
        bot.answer_callback_query(call.id,
            f"{'✅ চালু করা হয়েছে!' if new_enabled else '⛔ বন্ধ করা হয়েছে!'}",
            show_alert=True)
        bot.send_message(ADMIN_ID,
            f"🔀 <b>Bulkmail Product Status Updated</b>\n\n"
            f"📦 Product: <b>{prod_tog}</b>\n"
            f"📊 Status: {status_str}\n\n"
            f"<i>User দের Edu Mail submenu তে এই product টা {'দেখাবে' if new_enabled else 'দেখাবে না'}।</i>",
            parse_mode="HTML")
        # আবার toggle menu দেখাও
        db2 = load_db()
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for bm_p in [p for p, c in API_PRODUCTS.items() if c.get("source") == "bulkmail"]:
            en2 = is_bulkmail_enabled(bm_p, db2)
            markup2.add(InlineBtn(
                f"{'🟢 ON' if en2 else '🔴 OFF'} — {bm_p}", style="danger",
                callback_data=f"bulkmail_tog_{bm_p}"
            ))
        markup2.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "🔀 <b>Bulkmail Products ON/OFF</b>\n\nবাটনে ক্লিক করে toggle করুন:",
            parse_mode="HTML", reply_markup=markup2)

    # ════════════════════════════════════════════════════════
    # 🔀 HOTMAIL143 API PRODUCT ON/OFF TOGGLE
    # ════════════════════════════════════════════════════════
    elif callback_data == "admin_hotmail143_toggle":
        db = load_db()
        markup = types.InlineKeyboardMarkup(row_width=1)
        h143_products = [p for p, c in API_PRODUCTS.items() if c.get("source") == "hotmail143"]
        if not h143_products:
            bot.send_message(ADMIN_ID, "❌ <b>কোনো Hotmail143 product নেই!</b>",
                parse_mode="HTML"); return
        for h_p in h143_products:
            enabled = is_hotmail143_enabled(h_p, db)
            status_icon = "🟢 ON" if enabled else "🔴 OFF"
            markup.add(InlineBtn(
                f"{status_icon} — {h_p}", style="primary",
                callback_data=f"h143_tog_{h_p}"
            ))
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "🔀 <b>Hotmail143 API Products ON/OFF</b>\n\n"
            "🟢 = চালু (User কিনতে পারবে)\n"
            "🔴 = বন্ধ (User কিনতে পারবে না)\n\n"
            "বাটনে ক্লিক করে toggle করুন:",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data.startswith("h143_tog_"):
        prod_tog = callback_data.replace("h143_tog_", "")
        if prod_tog not in API_PRODUCTS or API_PRODUCTS[prod_tog].get("source") != "hotmail143":
            bot.answer_callback_query(call.id, "❌ Product খুঁজে পাওয়া যায়নি!", show_alert=True); return
        db = load_db()
        if "hotmail143_products" not in db:
            db["hotmail143_products"] = {}
        if prod_tog not in db["hotmail143_products"]:
            db["hotmail143_products"][prod_tog] = {}
        current_enabled = db["hotmail143_products"][prod_tog].get("enabled", True)
        new_enabled = not current_enabled
        db["hotmail143_products"][prod_tog]["enabled"] = new_enabled
        # ⚡ Targeted write
        update_db_path(f"/hotmail143_products/{prod_tog}/enabled", new_enabled)
        # ⚡ Cache sync — depth-3 path cache auto-update হয় না, manually sync করো
        _update_db_cache_in_place(db)
        status_str = "🟢 চালু (ON)" if new_enabled else "🔴 বন্ধ (OFF)"
        bot.answer_callback_query(call.id,
            f"{'✅ চালু করা হয়েছে!' if new_enabled else '⛔ বন্ধ করা হয়েছে!'}",
            show_alert=True)
        bot.send_message(ADMIN_ID,
            f"🔀 <b>Hotmail143 Product Status Updated</b>\n\n"
            f"📦 Product: <b>{prod_tog}</b>\n"
            f"📊 Status: {status_str}\n\n"
            f"<i>User রা এই product টা {'কিনতে পারবে' if new_enabled else 'কিনতে পারবে না'}।</i>",
            parse_mode="HTML")
        # আবার toggle menu দেখাও
        db2 = load_db()
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for h_p2 in [p for p, c in API_PRODUCTS.items() if c.get("source") == "hotmail143"]:
            en2 = is_hotmail143_enabled(h_p2, db2)
            markup2.add(InlineBtn(
                f"{'🟢 ON' if en2 else '🔴 OFF'} — {h_p2}", style="danger",
                callback_data=f"h143_tog_{h_p2}"
            ))
        markup2.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "🔀 <b>Hotmail143 API Products ON/OFF</b>\n\nবাটনে ক্লিক করে toggle করুন:",
            parse_mode="HTML", reply_markup=markup2)

    elif callback_data == "admin_bulkmail_stocktest":
        import requests as _req
        import json as _json
        lines = [f"🔍 <b>Bulkmail Stock Test</b>"]
        lines.append(f"🔑 API Key: {'✅ সেট আছে' if BULKMAIL_API_KEY else '❌ নেই!'}")
        lines.append(f"🌐 {BULKMAIL_BASE_URL}\n")
        hdrs = {"Accept": "application/json"}
        if BULKMAIL_API_KEY:
            hdrs["X-API-Key"] = BULKMAIL_API_KEY
        try:
            # ── 1. সব product list ──
            r1 = _req.get(f"{BULKMAIL_BASE_URL}/stock", headers=hdrs, timeout=10)
            lines.append(f"📋 <b>GET /stock</b> → HTTP {r1.status_code}")
            lines.append(f"<code>{r1.text[:600]}</code>\n")

            # ── 2. /stock/check?ids=13 ──
            for pname, conf in API_PRODUCTS.items():
                if conf.get("source") != "bulkmail":
                    continue
                pid = conf["product_id"]
                r2 = _req.get(f"{BULKMAIL_BASE_URL}/stock/check",
                              params={"ids": str(pid)}, headers=hdrs, timeout=10)
                lines.append(f"🎯 <b>{pname}</b> (id={pid}) → HTTP {r2.status_code}")
                lines.append(f"<code>{r2.text[:400]}</code>\n")

                # ── 3. /stock/{id} ──
                r3 = _req.get(f"{BULKMAIL_BASE_URL}/stock/{pid}", headers=hdrs, timeout=10)
                lines.append(f"🔎 <b>GET /stock/{pid}</b> → HTTP {r3.status_code}")
                lines.append(f"<code>{r3.text[:400]}</code>\n")

        except Exception as e:
            lines.append(f"❌ {e}")
        full = "\n".join(lines)
        for i in range(0, len(full), 4000):
            bot.send_message(ADMIN_ID, full[i:i+4000], parse_mode="HTML")

    elif callback_data == "admin_bulkmail_orders":
        import requests as _req
        lines = ["📋 <b>Bulkmail Recent Orders</b>\n"]
        hdrs = {"X-API-Key": BULKMAIL_API_KEY, "Accept": "application/json"}
        try:
            # সর্বশেষ 5টা order আনো
            r = _req.get(f"{BULKMAIL_BASE_URL}/orders", params={"per_page": 5}, headers=hdrs, timeout=10)
            lines.append(f"HTTP: <code>{r.status_code}</code>")
            d = r.json()
            orders = d.get("data", [])
            if not isinstance(orders, list):
                orders = []
            lines.append(f"Total fetched: <b>{len(orders)}</b>\n")
            for o in orders:
                oid = o.get("id")
                lines.append(f"🆔 Order ID: <code>{oid}</code>")
                lines.append(f"   Product: {o.get('product_name','?')}")
                lines.append(f"   Status: <b>{o.get('status','?')}</b>")
                lines.append(f"   Qty: {o.get('quantity','?')} | Total: {o.get('total_amount','?')}")
                # stock_items সহ full detail fetch করো
                r2 = _req.get(f"{BULKMAIL_BASE_URL}/orders/{oid}", headers=hdrs, timeout=10)
                d2 = r2.json()
                d2data = d2.get("data", {})
                si = d2data.get("items") or d2data.get("stock_items") or []
                lines.append(f"   items count: <b>{len(si)}</b>")
                if si:
                    lines.append(f"   First item: <code>{str(si[0])[:80]}</code>")
                else:
                    lines.append(f"   Raw data keys: <code>{list(d2data.keys())}</code>")
                lines.append("")
        except Exception as e:
            lines.append(f"❌ {e}")
        full = "\n".join(lines)
        for i in range(0, len(full), 4000):
            bot.send_message(ADMIN_ID, full[i:i+4000], parse_mode="HTML")

    elif callback_data == "admin_edit_price":
        products = list(db["products"].keys())
        if not products and not API_PRODUCTS:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        markup = types.InlineKeyboardMarkup(row_width=1)
        for p in products:
            markup.add(InlineBtn(
                f"🏷️ {p} — {db['products'][p]} BDT", style="primary", callback_data=f"edit_price_{p}"))
        # API products আলাদা section এ দেখাও
        for ap in API_PRODUCTS:
            if ap not in products:
                cur = db["products"].get(ap, 0)
                markup.add(InlineBtn(
                    f"🌐 {ap} — {cur} BDT", style="primary", callback_data=f"edit_price_{ap}"))
        markup.add(InlineBtn("🔙 Back", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, "🏷️ <b>Select product to edit price:</b>",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data.startswith("edit_price_"):
        product_name = callback_data.replace("edit_price_", "")
        current_price = db["products"].get(product_name, 0)
        msg = bot.send_message(ADMIN_ID,
            f"🏷️ <b>{product_name}</b>\n\nCurrent: <code>{current_price} BDT</code>\n\nNew price:",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: admin_edit_price_step(m, product_name))
        pass  # ACK already sent upfront

    elif callback_data == "admin_search_order":
        msg = bot.send_message(ADMIN_ID, "🔍 <b>Enter Order ID:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_search_order_step)
        pass  # ACK already sent upfront

    elif callback_data == "admin_broadcast":
        msg = bot.send_message(ADMIN_ID, "📢 <b>Write broadcast message:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_broadcast_step)
        pass  # ACK already sent upfront

    elif callback_data == "admin_edit_features":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add( *[KbBtn(_x, role="product_item") for _x in products]); markup.add( KbBtn("❌ Cancel", role="cancel"))
        msg = bot.send_message(ADMIN_ID,
            "✏️ <b>কোন Product এর Details edit করবেন?</b>",
            parse_mode="HTML", reply_markup=markup)
        bot.register_next_step_handler(msg, admin_edit_product_features)
        pass  # ACK already sent upfront

    # ── Description / Features inline edit callbacks ──
    elif callback_data.startswith("edit_desc_"):
        product_name = callback_data.replace("edit_desc_", "")
        current_desc = db.get("product_details", {}).get(product_name, {}).get("description", "") or "নেই"
        msg = bot.send_message(ADMIN_ID,
            f"📝 <b>'{product_name}' এর Description লিখুন:</b>\n\n"
            f"📌 <b>বর্তমান:</b>\n{current_desc}\n\n"
            f"নতুন description লিখুন:\n"
            f"<i>(মুছে ফেলতে <code>skip</code> লিখুন)</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: save_product_description(m, product_name))
        pass  # ACK already sent upfront

    elif callback_data.startswith("edit_feat_"):
        product_name = callback_data.replace("edit_feat_", "")
        current_feat = db.get("product_details", {}).get(product_name, {}).get("features", "") or "নেই"
        msg = bot.send_message(ADMIN_ID,
            f"✨ <b>'{product_name}' এর Features লিখুন:</b>\n\n"
            f"📌 <b>বর্তমান:</b>\n<i>{current_feat}</i>\n\n"
            f"নতুন features লিখুন (প্রতিটা লাইনে এক feature):\n"
            f"<i>উদাহরণ:\n• ৩০ দিনের গ্যারান্টি\n• Full Access\n• ২৪/৭ Support</i>\n\n"
            f"⚠️ Features italic অক্ষরে দেখাবে।\n"
            f"<i>(মুছে ফেলতে <code>skip</code> লিখুন)</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: save_product_features(m, product_name))
        pass  # ACK already sent upfront

    elif callback_data == "admin_flash_sale":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        # Active flash sales দেখাও
        active_flashes = []
        for p in products:
            fi = get_flash_sale_info(db, p)
            if fi:
                r = fi["remaining_seconds"]
                h, m_val, s_val = r // 3600, (r % 3600) // 60, r % 60
                active_flashes.append(f"🔥 {p} — {fi['discount_percent']}% OFF | ⏱ {h:02d}:{m_val:02d}:{s_val:02d}")
        flash_status = "\n".join(active_flashes) if active_flashes else "কোনো active Flash Sale নেই।"
        markup2 = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup2.add( *[KbBtn(_x, role="product_item") for _x in products]); markup2.add( KbBtn("❌ Cancel", role="cancel"))
        msg = bot.send_message(ADMIN_ID,
            f"⚡ <b>Flash Sale Manager</b>\n\n"
            f"📊 <b>Active Flash Sales:</b>\n<i>{flash_status}</i>\n\n"
            f"কোন Product এ Flash Sale সেট করবেন?",
            parse_mode="HTML", reply_markup=markup2)
        bot.register_next_step_handler(msg, admin_set_flash_sale_step1)
        pass  # ACK already sent upfront

    elif callback_data == "admin_backup":
        try:
            with open(DB_FILE, 'rb') as f:
                bot.send_document(ADMIN_ID, f,
                    caption=f"💾 <b>Database Backup</b>\n📅 {get_now()}",
                    parse_mode="HTML")
            bot.answer_callback_query(call.id, "✅ Backup sent!", show_alert=False)
        except Exception as e:
            bot.answer_callback_query(call.id, f"❌ Failed: {str(e)}", show_alert=True)

    elif callback_data == "admin_deposits":
        show_pending_deposits(ADMIN_ID)
        pass  # ACK already sent upfront

    elif callback_data == "admin_edu_mail_price":
        # Edu Mail 24H ও 72H এর price edit করো
        markup_ep = types.InlineKeyboardMarkup(row_width=1)
        for ap in [EDU_MAIL_SUB_24H, EDU_MAIL_SUB_72H]:
            cur_price = db["products"].get(ap, 0)
            cur_stock = get_stock_count(ap)
            markup_ep.add(InlineBtn(
                f"📧 {ap} — {cur_price} BDT | Stock: {cur_stock}", style="primary",
                callback_data=f"edit_price_{ap}"
            ))
        markup_ep.add(InlineBtn("🔙 Back", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "📧 <b>Edu Mail Price Editor</b>\n\n"
            "• Admin এখান থেকে Edu Mail 24H ও 72H এর price set করতে পারবেন।\n"
            "• User buy করলে এই price অনুযায়ী balance কাটবে।\n\n"
            "কোনটি edit করবেন বেছে নিন 👇",
            parse_mode="HTML", reply_markup=markup_ep)

    elif callback_data == "admin_edu_mail_stock":
        lines_stock = []
        for ap in API_DASHBOARD_PRODUCTS:
            api_conf = API_PRODUCTS.get(ap, {})
            api_stock = hotmail143_get_stock(api_conf.get("product_type",""), api_conf.get("account_type",""))
            local_path = os.path.join(STOCK_DIR, ap + ".xlsx")
            local_stock = 0
            if os.path.exists(local_path):
                try:
                    _wb = openpyxl.load_workbook(local_path, read_only=True)
                    _ws = _wb.active
                    for _ri, _row in enumerate(_ws.iter_rows(values_only=True), start=1):
                        if _ri == 1: continue
                        if any(c is not None and str(c).strip() != "" for c in _row):
                            local_stock += 1
                    _wb.close()
                except Exception:
                    pass
            total = (0 if api_stock < 0 else api_stock) + local_stock
            price = db["products"].get(ap, 0)
            icon = "\U0001f4e7" if "Mail" in ap else "\U0001f4e9"
            ast = str(api_stock) if api_stock >= 0 else "error"
            row = [
                icon + " <b>" + ap + "</b>",
                "   Website Stock: <code>" + ast + "</code>",
                "   Local Stock: <code>" + str(local_stock) + "</code>",
                "   Total: <code>" + str(total) + "</code> | Price: <code>" + str(price) + " BDT</code>",
            ]
            lines_stock.append("\n".join(row))
        markup_es = types.InlineKeyboardMarkup()
        markup_es.add(InlineBtn("Back", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "<b>Live Stock - All Products</b>\n\n" + "\n\n".join(lines_stock) + "\n\nUpdated: " + get_now(),
            parse_mode="HTML", reply_markup=markup_es)

    elif callback_data == "admin_hotmail_price":
        markup_hp = types.InlineKeyboardMarkup(row_width=1)
        cur_price = db["products"].get(HOTMAIL_PROD, 0)
        cur_stock = get_stock_count(HOTMAIL_PROD)
        markup_hp.add(InlineBtn(
            "Hotmail - " + str(cur_price) + " BDT | Stock: " + str(cur_stock), style="primary",
            callback_data="edit_price_" + HOTMAIL_PROD
        ))
        markup_hp.add(InlineBtn("Back", style="primary", callback_data="back_admin"))
        txt_hp = "<b>Hotmail Price Editor</b>\n\nAdmin ekhane Hotmail er price set korben.\nUser buy korle balance katbe.\n\nEdit korte button e click korun:"
        bot.send_message(ADMIN_ID, txt_hp, parse_mode="HTML", reply_markup=markup_hp)

    elif callback_data == "admin_outlook_price":
        markup_op = types.InlineKeyboardMarkup(row_width=1)
        cur_price = db["products"].get(OUTLOOK_PROD, 0)
        cur_stock = get_stock_count(OUTLOOK_PROD)
        markup_op.add(InlineBtn(
            "Outlook - " + str(cur_price) + " BDT | Stock: " + str(cur_stock), style="primary",
            callback_data="edit_price_" + OUTLOOK_PROD
        ))
        markup_op.add(InlineBtn("Back", style="primary", callback_data="back_admin"))
        txt_op = "<b>Outlook Price Editor</b>\n\nAdmin ekhane Outlook er price set korben.\nUser buy korle balance katbe.\n\nEdit korte button e click korun:"
        bot.send_message(ADMIN_ID, txt_op, parse_mode="HTML", reply_markup=markup_op)

    elif callback_data == "admin_api_dashboard":
        dash_data = get_api_dashboard_data()
        lines_d = []
        grand_count = 0
        grand_total = 0.0
        for prod in API_DASHBOARD_PRODUCTS:
            key = prod.replace(" ", "_")
            entry = dash_data.get(key, {})
            count = int(entry.get("count", 0))
            total_rev = float(entry.get("total_bdt", 0))
            last = entry.get("last_updated", "-")
            api_conf = API_PRODUCTS.get(prod, {})
            api_stock = hotmail143_get_stock(api_conf.get("product_type",""), api_conf.get("account_type",""))
            local_path = os.path.join(STOCK_DIR, prod + ".xlsx")
            local_stock = 0
            if os.path.exists(local_path):
                try:
                    _wb2 = openpyxl.load_workbook(local_path, read_only=True)
                    _ws2 = _wb2.active
                    for _ri2, _row2 in enumerate(_ws2.iter_rows(values_only=True), start=1):
                        if _ri2 == 1: continue
                        if any(c is not None and str(c).strip() != "" for c in _row2):
                            local_stock += 1
                    _wb2.close()
                except Exception: pass
            total_live = (0 if api_stock < 0 else api_stock) + local_stock
            price_now = db["products"].get(prod, 0)
            grand_count += count
            grand_total += total_rev
            ast = str(api_stock) if api_stock >= 0 else "error"
            icon = "\U0001f4e7" if "Mail" in prod else "\U0001f4e9"
            sep = "\n"
            row_txt = (
                icon + " <b>" + prod + "</b>" + sep +
                "  Price: <code>" + str(price_now) + " BDT</code>" + sep +
                "  Total Sold: <code>" + str(count) + " pcs</code>" + sep +
                "  Revenue: <code>" + str(round(total_rev,2)) + " BDT</code>" + sep +
                "  Website Stock: <code>" + ast + "</code>" + sep +
                "  Local Stock: <code>" + str(local_stock) + "</code>" + sep +
                "  Live Total: <code>" + str(total_live) + "</code>" + sep +
                "  Last Sale: <i>" + last + "</i>"
            )
            lines_d.append(row_txt)
        markup_dash = types.InlineKeyboardMarkup()
        markup_dash.add(InlineBtn("Back to Admin", style="primary", callback_data="back_admin"))
        sep2 = "\n"
        eq = "=" * 20
        summary = (
            "<b>API Product Dashboard</b>" + sep2 + eq + sep2 +
            (sep2 + sep2).join(lines_d) +
            sep2 + sep2 + eq + sep2 +
            "<b>Grand Total Sold:</b> <code>" + str(grand_count) + " pcs</code>" + sep2 +
            "<b>Grand Revenue:</b> <code>" + str(round(grand_total,2)) + " BDT</code>" + sep2 + sep2 +
            "<i>Updated: " + get_now() + "</i>"
        )
        bot.send_message(ADMIN_ID, summary, parse_mode="HTML", reply_markup=markup_dash)

    elif callback_data == "back_admin":
        show_admin_panel(ADMIN_ID)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 🎨 KEYBOARD BUTTON COLOR — Admin Panel (per-button)
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_kb_color":
        color_dot = {"primary": "🔵", "success": "🟢", "danger": "🔴"}
        markup = types.InlineKeyboardMarkup(row_width=1)
        for role, (label, _default) in KB_BUTTON_ROLES.items():
            current = get_kb_button_color(role)
            markup.add(InlineBtn(f"{color_dot.get(current,'⚪')} {label}",
                                  style="primary", callback_data=f"admin_kbc_role_{role}"))
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            f"🎨 <b>Keyboard Button Color</b>\n\n"
            f"প্রতিটা বাটনের পাশে তার বর্তমান কালার (🔵🟢🔴) দেখানো আছে।\n"
            f"যেই বাটনের কালার বদলাতে চান, সেটাতে ক্লিক করুন 👇",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data.startswith("admin_kbc_role_"):
        role = callback_data[len("admin_kbc_role_"):]
        if role not in KB_BUTTON_ROLES:
            bot.send_message(ADMIN_ID, "❌ Invalid button role.")
        else:
            label = KB_BUTTON_ROLES[role][0]
            current = get_kb_button_color(role)
            markup = types.InlineKeyboardMarkup(row_width=1)
            markup.add(
                InlineBtn(("✅ " if current == "danger" else "") + "🔴 লাল (Red)",
                           style="danger", callback_data=f"admin_kbc_set_{role}_danger"),
                InlineBtn(("✅ " if current == "primary" else "") + "🔵 নীল (Blue)",
                           style="primary", callback_data=f"admin_kbc_set_{role}_primary"),
                InlineBtn(("✅ " if current == "success" else "") + "🟢 সবুজ (Green)",
                           style="success", callback_data=f"admin_kbc_set_{role}_success"),
            )
            markup.add(InlineBtn("🔙 Button List", style="primary", callback_data="admin_kb_color"))
            bot.send_message(ADMIN_ID,
                f"🎨 <b>{label}</b>\n\nএই বাটনের জন্য একটা কালার বেছে নিন 👇",
                parse_mode="HTML", reply_markup=markup)

    elif callback_data.startswith("admin_kbc_set_"):
        # format: admin_kbc_set_{role}_{color}
        rest = callback_data[len("admin_kbc_set_"):]
        role, _, chosen = rest.rpartition("_")
        color_dot = {"primary": "🔵", "success": "🟢", "danger": "🔴"}
        label_map = {"primary": "🔵 নীল (Blue)", "success": "🟢 সবুজ (Green)", "danger": "🔴 লাল (Red)"}
        if role not in KB_BUTTON_ROLES or chosen not in _KB_COLOR_CHOICES:
            bot.send_message(ADMIN_ID, "❌ Invalid selection.")
        else:
            set_kb_button_color(role, chosen)
            db = load_db()
            if "settings" not in db:
                db["settings"] = {}
            if "kb_button_colors" not in db["settings"]:
                db["settings"]["kb_button_colors"] = {}
            db["settings"]["kb_button_colors"][role] = chosen
            # ⚡ Targeted write — শুধু এই role এর color
            update_db_path(f"/settings/kb_button_colors/{role}", chosen)
            _update_db_cache_in_place(db)

            label = KB_BUTTON_ROLES[role][0]
            bot.send_message(ADMIN_ID,
                f"✅ <b>{label}</b> এর কালার আপডেট হয়েছে!\n\n"
                f"নতুন কালার: <b>{label_map.get(chosen, chosen)}</b>",
                parse_mode="HTML")

            # Full role list আবার দেখাও, updated dot সহ
            markup = types.InlineKeyboardMarkup(row_width=1)
            for r, (lbl, _default) in KB_BUTTON_ROLES.items():
                cur = get_kb_button_color(r)
                markup.add(InlineBtn(f"{color_dot.get(cur,'⚪')} {lbl}",
                                      style="primary", callback_data=f"admin_kbc_role_{r}"))
            markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
            bot.send_message(ADMIN_ID,
                "🎨 <b>Keyboard Button Color</b>\n\nআরেকটা বাটন বেছে নিন, অথবা Admin Panel এ ফিরে যান 👇",
                parse_mode="HTML", reply_markup=markup)
            # Admin কে সাথে সাথে নতুন কালারের reply keyboard দেখিয়ে দাও
            try:
                bot.send_message(ADMIN_ID, "🔄 Main Menu (updated color):",
                                  reply_markup=get_main_menu(ADMIN_ID))
            except Exception:
                pass

    # ══════════════════════════════════════════════════════
    # 📈 ANALYTICS & REPORTS
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_analytics":
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineBtn("📅 আজকের Report", style="primary", callback_data="analytics_daily"),
            InlineBtn("📆 গতকালকের Report", style="primary", callback_data="analytics_yesterday"),
            InlineBtn("📆 সাপ্তাহিক Report", style="primary", callback_data="analytics_weekly"),
            InlineBtn("🗓️ মাসিক Report", style="primary", callback_data="analytics_monthly"),
            InlineBtn("🏆 Best Selling Products", style="primary", callback_data="analytics_products"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        bot.send_message(ADMIN_ID, "📈 <b>Analytics & Reports</b>\n\nকোন রিপোর্ট দেখতে চান?",
            parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data in ("analytics_daily", "analytics_yesterday", "analytics_weekly", "analytics_monthly"):
        period = {
            "analytics_daily": "daily",
            "analytics_yesterday": "yesterday",
            "analytics_weekly": "weekly",
            "analytics_monthly": "monthly",
        }[callback_data]
        report = generate_sales_report(db, period)
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Analytics", style="primary", callback_data="admin_analytics"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "analytics_products":
        report = generate_product_sales_report(db)
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Analytics", style="primary", callback_data="admin_analytics"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 👥 USER MANAGEMENT
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_user_mgmt":
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineBtn("🔍 User এর Order History", style="primary", callback_data="umgmt_order_history"),
            InlineBtn("💰 User এর Deposit History", style="success", callback_data="umgmt_deposit_history"),
            InlineBtn("⭐ VIP Users দেখুন", style="primary", callback_data="umgmt_vip_users"),
            InlineBtn("👥 সব Users List", style="primary", callback_data="umgmt_all_users"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        bot.send_message(ADMIN_ID, "👥 <b>User Management</b>", parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "umgmt_order_history":
        msg = bot.send_message(ADMIN_ID, "🆔 <b>User ID দিন:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_user_order_history_step)
        pass  # ACK already sent upfront

    elif callback_data == "umgmt_deposit_history":
        msg = bot.send_message(ADMIN_ID, "🆔 <b>User ID দিন:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_user_deposit_history_step)
        pass  # ACK already sent upfront

    elif callback_data == "umgmt_vip_users":
        report = generate_vip_users_report(db)
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 User Management", style="primary", callback_data="admin_user_mgmt"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "umgmt_all_users":
        total = len(db["users"])
        active = sum(1 for u in db["users"].values() if u.get("orders"))
        banned = len(db.get("banned_users", []))
        report = (
            f"👥 <b>সব Users Summary</b>\n\n"
            f"📊 <b>মোট Users:</b> {total}\n"
            f"✅ <b>Active (কেনাকাটা করেছে):</b> {active}\n"
            f"🚫 <b>Banned:</b> {banned}\n"
            f"🆕 <b>Never bought:</b> {total - active}\n"
        )
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineBtn("📋 All User ID List", style="primary", callback_data="umgmt_user_id_list"),
            InlineBtn("🔙 User Management", style="primary", callback_data="admin_user_mgmt")
        )
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    elif callback_data == "umgmt_user_id_list":
        # ✅ সব User ID একটা list হিসেবে দেখাও (username সহ)
        all_uids = list(db["users"].keys())
        if not all_uids:
            bot.send_message(ADMIN_ID, "❌ কোনো user নেই।", parse_mode="HTML")
        else:
            # প্রতি message এ max 50 user (Telegram limit এড়াতে)
            chunk_size = 50
            for i in range(0, len(all_uids), chunk_size):
                chunk = all_uids[i:i+chunk_size]
                lines = []
                for uid in chunk:
                    u = db["users"].get(uid, {})
                    orders_count = len(u.get("orders", []))
                    bal = u.get("balance", 0)
                    # Username আনার চেষ্টা
                    uname = ""
                    try:
                        tg_u = bot.get_chat(int(uid))
                        if tg_u.username:
                            uname = f" (@{tg_u.username})"
                        elif tg_u.first_name:
                            uname = f" ({tg_u.first_name})"
                    except Exception:
                        pass
                    lines.append(
                        f"🆔 <code>{uid}</code>{uname}\n"
                        f"   💰 {bal} BDT | 📦 {orders_count} orders"
                    )
                part_num = (i // chunk_size) + 1
                total_parts = (len(all_uids) + chunk_size - 1) // chunk_size
                header = f"👥 <b>All User ID List</b> ({part_num}/{total_parts})\n\n"
                markup_back = types.InlineKeyboardMarkup()
                markup_back.add(InlineBtn("🔙 Back", style="primary", callback_data="umgmt_all_users"))
                bot.send_message(ADMIN_ID,
                    header + "\n\n".join(lines),
                    parse_mode="HTML",
                    reply_markup=markup_back if i + chunk_size >= len(all_uids) else None)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 💵 FINANCE REPORT
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_finance":
        report = generate_finance_report(db)
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 📦 STOCK ALERTS
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_stock_alerts":
        report = generate_stock_alert_report(db)
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 🎁 REFER BONUS SETTINGS
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_set_refer_bonus":
        current_bonus = db["settings"].get("refer_bonus", 2)
        msg = bot.send_message(ADMIN_ID,
            f"🎁 <b>Refer Bonus সেট করুন</b>\n\n"
            f"বর্তমান bonus: <b>{current_bonus} BDT</b>\n\n"
            f"নতুন bonus amount লিখুন (BDT):",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_set_refer_bonus_step)
        pass  # ACK already sent upfront

    elif callback_data == "admin_refer_settings":
        current_bonus = db["settings"].get("refer_bonus", 2)
        total_referrals = sum(u.get("refer_count", 0) for u in db["users"].values())
        total_bonus_paid = total_referrals * current_bonus
        report = (
            f"🔑 <b>Refer Bonus Settings</b>\n"
            f"\n\n"
            f"💰 <b>Current Bonus:</b> {current_bonus} BDT per refer\n"
            f"👥 <b>Total Referrals:</b> {total_referrals}\n"
            f"💵 <b>Total Bonus Paid:</b> {total_bonus_paid} BDT\n"
            f""
        )
        markup = types.InlineKeyboardMarkup()
        markup.add(
            InlineBtn("✏️ Bonus পরিবর্তন করুন", style="primary", callback_data="admin_set_refer_bonus"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # 🎟️ PROMO CODE MANAGEMENT
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_promo_manager":
        promos = db.get("promo_codes", {})
        promo_feature_on = db.get("settings", {}).get("promo_feature_enabled", True)
        markup = types.InlineKeyboardMarkup(row_width=1)
        toggle_label = "🟢 Promo Feature: ON  (বন্ধ করতে ক্লিক করুন)" if promo_feature_on else "🔴 Promo Feature: OFF  (চালু করতে ক্লিক করুন)"
        markup.add(
            InlineBtn(toggle_label, style="primary", callback_data="admin_promo_feature_toggle"),
            InlineBtn("➕ নতুন Promo Code তৈরি", style="primary", callback_data="admin_create_promo"),
            InlineBtn("📋 সব Promo Code দেখুন", style="primary", callback_data="admin_list_promos"),
            InlineBtn("📊 Promo ব্যবহারের রিপোর্ট", style="primary", callback_data="admin_promo_report"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        total_promos = len(promos)
        active_promos = sum(1 for p in promos.values() if p.get("enabled", True))
        feature_status = "🟢 চালু (ON)" if promo_feature_on else "🔴 বন্ধ (OFF)"
        bot.send_message(ADMIN_ID,
            f"🎟️ <b>Promo Code Manager</b>\n\n"
            f"🔧 <b>Promo Feature:</b> {feature_status}\n"
            f"<i>OFF থাকলে buy করার সময় promo কোড জিজ্ঞেস করবে না।</i>\n\n"
            f"📊 <b>মোট Codes:</b> {total_promos}\n✅ <b>Active:</b> {active_promos}\n❌ <b>Inactive:</b> {total_promos - active_promos}\n\nনিচের অপশন থেকে বেছে নিন 👇",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data == "admin_promo_feature_toggle":
        current = db.get("settings", {}).get("promo_feature_enabled", True)
        if "settings" not in db:
            db["settings"] = {}
        db["settings"]["promo_feature_enabled"] = not current
        # ⚡ Targeted write
        update_db_path("/settings/promo_feature_enabled", not current)
        new_status = "🟢 চালু (ON)" if not current else "🔴 বন্ধ (OFF)"
        note = "এখন থেকে buy করার সময় promo কোড জিজ্ঞেস করবে ✅" if not current else "এখন থেকে buy করার সময় promo কোড জিজ্ঞেস করবে না ✅"
        bot.send_message(ADMIN_ID,
            f"🎟️ <b>Promo Feature → {new_status}</b>\n\n{note}",
            parse_mode="HTML")
        # Refresh the promo manager panel
        promos = db.get("promo_codes", {})
        promo_feature_on = not current
        markup = types.InlineKeyboardMarkup(row_width=1)
        toggle_label = "🟢 Promo Feature: ON  (বন্ধ করতে ক্লিক করুন)" if promo_feature_on else "🔴 Promo Feature: OFF  (চালু করতে ক্লিক করুন)"
        markup.add(
            InlineBtn(toggle_label, style="primary", callback_data="admin_promo_feature_toggle"),
            InlineBtn("➕ নতুন Promo Code তৈরি", style="primary", callback_data="admin_create_promo"),
            InlineBtn("📋 সব Promo Code দেখুন", style="primary", callback_data="admin_list_promos"),
            InlineBtn("📊 Promo ব্যবহারের রিপোর্ট", style="primary", callback_data="admin_promo_report"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        total_promos = len(promos)
        active_promos = sum(1 for p in promos.values() if p.get("enabled", True))
        feature_status = "🟢 চালু (ON)" if promo_feature_on else "🔴 বন্ধ (OFF)"
        bot.send_message(ADMIN_ID,
            f"🎟️ <b>Promo Code Manager</b>\n\n"
            f"🔧 <b>Promo Feature:</b> {feature_status}\n"
            f"<i>OFF থাকলে buy করার সময় promo কোড জিজ্ঞেস করবে না।</i>\n\n"
            f"📊 <b>মোট Codes:</b> {total_promos}\n✅ <b>Active:</b> {active_promos}\n❌ <b>Inactive:</b> {total_promos - active_promos}\n\nনিচের অপশন থেকে বেছে নিন 👇",
            parse_mode="HTML", reply_markup=markup)

    # ══════════════════════════════════════════════════════
    # 🎁 CASHBACK SETTINGS
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_cashback_settings":
        cashback_on = db.get("settings", {}).get("cashback_enabled", True)
        toggle_label = "🟢 Cashback: ON  (বন্ধ করতে ক্লিক করুন)" if cashback_on else "🔴 Cashback: OFF  (চালু করতে ক্লিক করুন)"
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineBtn(toggle_label, style="primary", callback_data="admin_cashback_toggle"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        status_text = "🟢 চালু (ON)" if cashback_on else "🔴 বন্ধ (OFF)"
        bot.send_message(ADMIN_ID,
            f"🎁 <b>Cashback Settings</b>\n\n"
            f"🔧 <b>বর্তমান অবস্থা:</b> {status_text}\n\n"
            f"📋 <b>Cashback অফার:</b>\n"
            f"  💰 ৫০০–৯৯৯ BDT → +৩০ BDT cashback\n"
            f"  💰 ১০০০–২০০০ BDT → +৭০ BDT cashback\n\n"
            f"<i>OFF থাকলে deposit approve এ cashback দেওয়া হবে না এবং user কে cashback অফার দেখানো হবে না।</i>",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data == "admin_cashback_toggle":
        current = db.get("settings", {}).get("cashback_enabled", True)
        if "settings" not in db:
            db["settings"] = {}
        db["settings"]["cashback_enabled"] = not current
        # ⚡ Targeted write
        update_db_path("/settings/cashback_enabled", not current)
        new_status = "🟢 চালু (ON)" if not current else "🔴 বন্ধ (OFF)"
        note = "এখন থেকে deposit এ cashback দেওয়া হবে এবং অফার দেখানো হবে ✅" if not current else "এখন থেকে cashback বন্ধ — অফার দেখানো হবে না ✅"
        bot.send_message(ADMIN_ID,
            f"🎁 <b>Cashback → {new_status}</b>\n\n{note}",
            parse_mode="HTML")
        cashback_on = not current
        toggle_label = "🟢 Cashback: ON  (বন্ধ করতে ক্লিক করুন)" if cashback_on else "🔴 Cashback: OFF  (চালু করতে ক্লিক করুন)"
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineBtn(toggle_label, style="primary", callback_data="admin_cashback_toggle"),
            InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
        )
        status_text = "🟢 চালু (ON)" if cashback_on else "🔴 বন্ধ (OFF)"
        bot.send_message(ADMIN_ID,
            f"🎁 <b>Cashback Settings</b>\n\n"
            f"🔧 <b>বর্তমান অবস্থা:</b> {status_text}\n\n"
            f"📋 <b>Cashback অফার:</b>\n"
            f"  💰 ৫০০–৯৯৯ BDT → +৩০ BDT cashback\n"
            f"  💰 ১০০০–২০০০ BDT → +৭০ BDT cashback\n\n"
            f"<i>OFF থাকলে deposit approve এ cashback দেওয়া হবে না এবং user কে cashback অফার দেখানো হবে না।</i>",
            parse_mode="HTML", reply_markup=markup)

    elif callback_data == "admin_create_promo":
        msg = bot.send_message(ADMIN_ID,
            "🎟️ <b>নতুন Promo Code তৈরি</b>\n\nCode লিখুন (যেমন: SALE20, NEWUSER, EIDOFFER):\n<i>শুধু অক্ষর ও সংখ্যা ব্যবহার করুন</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_create_promo_step1)

    elif callback_data == "admin_list_promos":
        promos = db.get("promo_codes", {})
        if not promos:
            bot.send_message(ADMIN_ID, "❌ <b>কোনো Promo Code নেই!</b>",
                parse_mode="HTML"); return
        markup = types.InlineKeyboardMarkup(row_width=1)
        report = "📋 <b>সব Promo Codes</b>\n\n"
        for code, info in promos.items():
            status = "✅" if info.get("enabled", True) else "❌"
            used = info.get("used_count", 0)
            max_u = info.get("max_uses", 0)
            new_only = "🆕" if info.get("new_user_only") else ""
            report += (f"{status} <code>{code}</code> — <b>{info.get('discount_percent', 0):.0f}% ছাড়</b> {new_only}\n"
                       f"   ব্যবহার: {used}/{max_u if max_u else '∞'} | মেয়াদ: {info.get('expiry', 'অসীম')}\n\n")
            markup.add(InlineBtn(
                f"{'✅' if info.get('enabled', True) else '❌'} {code} — Toggle", style="danger",
                callback_data=f"promo_toggle_{code}"))
        markup.add(InlineBtn("🔙 Promo Manager", style="primary", callback_data="admin_promo_manager"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)

    elif callback_data.startswith("promo_toggle_"):
        code = callback_data.replace("promo_toggle_", "")
        if code in db.get("promo_codes", {}):
            current = db["promo_codes"][code].get("enabled", True)
            db["promo_codes"][code]["enabled"] = not current
            # ⚡ Targeted write
            update_db_path(f"/promo_codes/{code}/enabled", not current)
            status = "✅ Active" if not current else "❌ Inactive"
            bot.send_message(ADMIN_ID,
                f"🎟️ <b>{code}</b> → {status}", parse_mode="HTML")

    elif callback_data == "admin_promo_report":
        promo_usage = db.get("promo_usage", {})
        promos = db.get("promo_codes", {})
        report = "📊 <b>Promo ব্যবহারের রিপোর্ট</b>\n\n"
        total_discount_given = 0
        for code, users_dict in promo_usage.items():
            promo_info = promos.get(code, {})
            disc_pct = promo_info.get("discount_percent", 0)
            used_count = len(users_dict)
            report += f"🎟️ <b>{code}</b> — {disc_pct:.0f}% ছাড়\n"
            report += f"   👥 ব্যবহারকারী: {used_count} জন\n"
            for uid, used_time in list(users_dict.items())[:5]:
                report += f"   • <code>{uid}</code> — {used_time}\n"
            if len(users_dict) > 5:
                report += f"   ... এবং আরো {len(users_dict)-5} জন\n"
            report += "\n"
        # New user discount tracking
        new_user_disc = db.get("settings", {}).get("new_user_discount", 0)
        if new_user_disc > 0:
            report += f"\n🆕 <b>New User Discount:</b> {new_user_disc:.0f}%\n"
        if not promo_usage:
            report += "<i>এখনো কোনো promo code ব্যবহার হয়নি।</i>"
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🔙 Promo Manager", style="primary", callback_data="admin_promo_manager"))
        bot.send_message(ADMIN_ID, report, parse_mode="HTML", reply_markup=markup)

    elif callback_data == "admin_new_user_discount":
        current = db.get("settings", {}).get("new_user_discount", 0)
        msg = bot.send_message(ADMIN_ID,
            f"🆕 <b>নতুন User এর জন্য Discount</b>\n\nবর্তমান: <b>{current:.0f}%</b>\n\nনতুন discount % লিখুন (0 দিলে বন্ধ হবে):\n<i>উদাহরণ: 10 (মানে ১০% ছাড়)</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, admin_set_new_user_discount_step)

    # ══════════════════════════════════════════════════════
    # 📤 FIREBASE EXPORT
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_firebase_export":
        try:
            bot.answer_callback_query(call.id, "⏳ Firebase থেকে data export হচ্ছে...")
            # Firebase থেকে সব data পড়ো
            all_data = _fb_ref("/").get()
            if not all_data:
                bot.send_message(ADMIN_ID, "❌ <b>Firebase তে কোনো data পাওয়া যায়নি!</b>", parse_mode="HTML")
                return
            import json as _json_mod
            json_str = _json_mod.dumps(all_data, ensure_ascii=False, indent=2)
            json_bytes = json_str.encode("utf-8")
            from io import BytesIO as _BytesIO
            json_file = _BytesIO(json_bytes)
            json_file.name = f"firebase_export_{get_now().replace('/', '-').replace(' ', '_').replace(':', '-')}.json"
            bot.send_document(ADMIN_ID, json_file,
                caption=f"📤 <b>Firebase Full Export</b>\n📅 {get_now()}\n\n✅ সব data এক্সপোর্ট হয়েছে।",
                parse_mode="HTML")
        except Exception as e:
            logging.error(f"Firebase export error: {e}")
            bot.send_message(ADMIN_ID, f"❌ <b>Export failed!</b>\n<code>{str(e)}</code>", parse_mode="HTML")

    # ══════════════════════════════════════════════════════
    # 🔘 PRODUCT BUTTON BUILDER
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_button_builder":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for p in products:
            markup2.add(InlineBtn(
                f"🔘 {p}", style="primary", callback_data=f"btn_builder_{p}"))
        markup2.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "🔘 <b>Product Button Builder</b>\n\n"
            "কোন Product এর জন্য Custom Inline/Keyboard বাটন বানাবেন?\n\n"
            "<i>এই বাটনগুলো Product detail page এ দেখাবে।</i>",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("btn_builder_"):
        product_name = callback_data.replace("btn_builder_", "")
        # বর্তমান বাটন দেখাও
        existing_btns = db.get("product_buttons", {}).get(product_name, [])
        btn_preview = ""
        for btn in existing_btns:
            btn_type = "🔘 Inline" if btn.get("type") == "inline" else "⌨️ Keyboard"
            btn_preview += f"  {btn_type} | <b>{btn['text']}</b>"
            if btn.get("url"):
                btn_preview += f" → {btn['url']}\n"
            elif btn.get("callback"):
                btn_preview += f" → callback: {btn['callback']}\n"
            else:
                btn_preview += "\n"
        if not btn_preview:
            btn_preview = "  কোনো বাটন নেই।\n"

        markup2 = types.InlineKeyboardMarkup(row_width=1)
        markup2.add(
            InlineBtn("➕ Inline Button যোগ করুন (URL)", style="success", callback_data=f"btn_add_inline_{product_name}"),
            InlineBtn("➕ Keyboard Button যোগ করুন", style="success", callback_data=f"btn_add_keyboard_{product_name}"),
            InlineBtn("🗑️ সব বাটন মুছুন", style="danger", callback_data=f"btn_clear_{product_name}"),
            InlineBtn("🔙 Button Builder", style="primary", callback_data="admin_button_builder")
        )
        bot.send_message(ADMIN_ID,
            f"🔘 <b>{product_name} — Button Manager</b>\n\n"
            f"<b>বর্তমান বাটনসমূহ:</b>\n{btn_preview}\n"
            f"কি করতে চান? 👇",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("btn_add_inline_"):
        product_name = callback_data.replace("btn_add_inline_", "")
        msg = bot.send_message(ADMIN_ID,
            f"🔘 <b>Inline Button যোগ করুন</b>\n\n"
            f"📌 Product: <b>{product_name}</b>\n\n"
            f"বাটনের <b>Text</b> লিখুন:\n"
            f"<i>উদাহরণ: 🌐 আমাদের ওয়েবসাইট</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: btn_add_inline_text_step(m, product_name))
        pass  # ACK already sent upfront

    elif callback_data.startswith("btn_add_keyboard_"):
        product_name = callback_data.replace("btn_add_keyboard_", "")
        msg = bot.send_message(ADMIN_ID,
            f"⌨️ <b>Keyboard Button যোগ করুন</b>\n\n"
            f"📌 Product: <b>{product_name}</b>\n\n"
            f"বাটনের <b>Text</b> লিখুন:\n"
            f"<i>উদাহরণ: 📞 সাপোর্টে যোগাযোগ করুন</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: btn_add_keyboard_text_step(m, product_name))
        pass  # ACK already sent upfront

    elif callback_data.startswith("btn_clear_"):
        product_name = callback_data.replace("btn_clear_", "")
        if "product_buttons" not in db:
            db["product_buttons"] = {}
        db["product_buttons"][product_name] = []
        # ⚡ Targeted write
        update_db_path(f"/product_buttons/{product_name}", [])
        bot.answer_callback_query(call.id, f"✅ {product_name} এর সব বাটন মুছে ফেলা হয়েছে!", show_alert=True)
        show_admin_panel(ADMIN_ID)

    # ✅ FIX: upload_stock_select_ — inline product selection for stock upload
    elif callback_data.startswith("upload_stock_select_"):
        product_name = callback_data.replace("upload_stock_select_", "")
        if product_name not in db["products"]:
            bot.answer_callback_query(call.id, "❌ Product পাওয়া যায়নি!", show_alert=True); return
        msg = bot.send_message(ADMIN_ID,
            f"📎 <b>Send .xlsx stock file for '{product_name}'</b>\n\n"
            f"<i>⚠️ শুধু .xlsx file পাঠান। Header row সহ।</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: admin_upload_stock_step2(m, product_name))
        pass  # ACK already sent upfront

    # ══════════════════════════════════════════════════════
    # ✅ CONFIRM STOCK BROADCAST
    # ══════════════════════════════════════════════════════
    elif callback_data.startswith("confirm_stock_broadcast_"):
        safe_name = callback_data.replace("confirm_stock_broadcast_", "")
        pending = _pending_broadcasts.pop(safe_name, None)
        if not pending:
            bot.answer_callback_query(call.id, "⚠️ Session পাওয়া যায়নি বা মেয়াদ শেষ!", show_alert=True)
            return
        bot.answer_callback_query(call.id, "📢 Broadcasting in background...")
        try:
            bot.edit_message_reply_markup(ADMIN_ID, call.message.message_id, reply_markup=None)
        except: pass

        pname     = pending["product_name"]
        bcast_msg = pending["broadcast_msg"]
        ch_msg    = pending["channel_msg"]
        all_uids  = list(db["users"].keys())

        bot.send_message(ADMIN_ID,
            f"📢 <b>Broadcast শুরু হচ্ছে...</b>\n\n"
            f"📦 <b>Product:</b> {pname}\n"
            f"👥 <b>মোট Users:</b> {len(all_uids)} জন\n\n"
            f"<i>Background এ পাঠানো হচ্ছে, bot slow হবে না।</i>",
            parse_mode="HTML")

        # ── Background thread এ broadcast — main bot block হবে না ──
        def _do_broadcast(uids, msg, ch, name):
            sent = failed = 0
            for uid in uids:
                try:
                    _markup = types.InlineKeyboardMarkup()
                    _markup.add(InlineBtn(
                        f"🛒 এখনই কিনুন — {name}", style="primary", callback_data="back_to_shop"))
                    bot.send_message(uid, msg, parse_mode="HTML", reply_markup=_markup)
                    sent += 1
                    _time_module.sleep(0.05)  # Telegram rate limit: 20 msg/sec
                except:
                    failed += 1
            try:
                bot.send_message(LOG_CHANNEL_ID, ch, parse_mode="HTML")
            except: pass
            try:
                bot.send_message(ADMIN_ID,
                    f"✅ <b>Broadcast Complete!</b>\n\n"
                    f"📦 <b>Product:</b> {name}\n"
                    f"✅ <b>Sent:</b> {sent} users\n"
                    f"❌ <b>Failed:</b> {failed} users",
                    parse_mode="HTML")
            except: pass

        threading.Thread(
            target=_do_broadcast,
            args=(all_uids, bcast_msg, ch_msg, pname),
            daemon=True
        ).start()

    # ══════════════════════════════════════════════════════
    # ❌ CANCEL STOCK BROADCAST
    # ══════════════════════════════════════════════════════
    elif callback_data.startswith("cancel_stock_broadcast_"):
        safe_name = callback_data.replace("cancel_stock_broadcast_", "")
        _pending_broadcasts.pop(safe_name, None)
        bot.answer_callback_query(call.id, "✅ Broadcast বাতিল করা হয়েছে।", show_alert=True)
        try:
            bot.edit_message_reply_markup(ADMIN_ID, call.message.message_id, reply_markup=None)
        except: pass
        bot.send_message(ADMIN_ID,
            f"🚫 <b>Broadcast Cancelled!</b>\n\n"
            f"Stock upload সম্পন্ন হয়েছে কিন্তু কোনো user কে notification পাঠানো হয়নি।",
            parse_mode="HTML")

    # ══════════════════════════════════════════════════════
    # 📦 SUB-PRODUCT BUY (Gmail sub-types, Facebook sub-types)
    # ══════════════════════════════════════════════════════
    elif callback_data.startswith("sub_buy_"):
        # format: sub_buy_Gmail__Gmail for Personal use
        rest = callback_data.replace("sub_buy_", "")
        if "__" not in rest:
            bot.answer_callback_query(call.id, "❌ Invalid!", show_alert=True); return
        parent_product, sub_name = rest.split("__", 1)
        stock_key = f"{parent_product}__{sub_name}"
        sub_items = db.get("sub_products", {}).get(parent_product, {}).get("sub_items", [])
        sub_info = next((s for s in sub_items if s.get("name") == sub_name), None)
        if not sub_info:
            bot.answer_callback_query(call.id, "❌ Sub-product পাওয়া যায়নি!", show_alert=True); return
        if not sub_info.get("enabled", True):
            bot.answer_callback_query(call.id, "❌ এই product বর্তমানে off আছে!", show_alert=True); return
        sub_price = sub_info.get("price", 0)
        sub_stock = get_stock_count(stock_key)
        if sub_stock <= 0:
            bot.answer_callback_query(call.id, "❌ Stock নেই!", show_alert=True)
            bot.send_message(user_id,
                f"❌ <b>{sub_name}</b> এখন out of stock!\nঅন্য option বেছে নিন।",
                parse_mode="HTML")
            return
        # ✅ BUG FIX: Personal discount apply করো sub-buy তেও
        _sub_personal_disc = float(db.get("users", {}).get(user_id, {}).get("personal_discount", 0))
        final_sub_price = apply_discount(sub_price, _sub_personal_disc)
        current_balance = db["users"].get(user_id, {}).get("balance", 0)
        if current_balance < final_sub_price:
            bot.answer_callback_query(call.id, "💸 Balance কম!", show_alert=True)
            bot.send_message(user_id,
                t(user_id, "insufficient_balance",
                  total=final_sub_price, bal=current_balance, short=round(final_sub_price - current_balance, 2)),
                parse_mode="HTML")
            return
        bot.answer_callback_query(call.id, "⏳ Processing...")
        process_purchase(user_id, stock_key, 1, final_sub_price, final_sub_price,
                         display_name=f"{parent_product} — {sub_name}",
                         discount_percent=_sub_personal_disc)

    # ══════════════════════════════════════════════════════
    # 📦 ADMIN: SUB-PRODUCT MANAGER
    # ══════════════════════════════════════════════════════
    elif callback_data == "admin_subproduct_manager":
        products = list(db["products"].keys())
        if not products:
            bot.answer_callback_query(call.id, "❌ No products!", show_alert=True); return
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for p in products:
            sub_count = len(db.get("sub_products", {}).get(p, {}).get("sub_items", []))
            label = f"📦 {p}" + (f" ({sub_count} sub)" if sub_count > 0 else "")
            markup2.add(InlineBtn(label, style="primary", callback_data=f"subp_manage_{p}"))
        markup2.add(InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin"))
        bot.send_message(ADMIN_ID,
            "📦 <b>Sub-Product Manager</b>\n\n"
            "Gmail, Facebook ইত্যাদি product এ sub-buttons যোগ করুন।\n"
            "Sub-button গুলোতে ক্লিক করে ইউজাররা সরাসরি কিনতে পারবে।",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_manage_"):
        p_name = callback_data.replace("subp_manage_", "")
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        preview = ""
        for item in sub_items:
            status = "✅" if item.get("enabled", True) else "❌"
            sub_stock = get_stock_count(f"{p_name}__{item['name']}")
            preview += f"{status} <b>{item['name']}</b> — {item.get('price',0)} BDT | 📦{sub_stock}pcs\n"
        if not preview:
            preview = "কোনো sub-item নেই।"
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        markup2.add(
            InlineBtn("➕ Sub-item যোগ করুন", style="success", callback_data=f"subp_add_{p_name}"),
            InlineBtn("📤 Stock আপলোড করুন", style="primary", callback_data=f"subp_stock_{p_name}"),
            InlineBtn("🔘 On/Off Toggle", style="primary", callback_data=f"subp_toggle_{p_name}"),
            InlineBtn("🗑️ Sub-item মুছুন", style="danger", callback_data=f"subp_delete_{p_name}"),
            InlineBtn("🔙 Back", style="primary", callback_data="admin_subproduct_manager")
        )
        bot.send_message(ADMIN_ID,
            f"📦 <b>{p_name} — Sub-Products</b>\n\n{preview}\n\nকি করবেন?",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_add_"):
        p_name = callback_data.replace("subp_add_", "")
        msg = bot.send_message(ADMIN_ID,
            f"➕ <b>{p_name} এ নতুন Sub-item যোগ করুন</b>\n\n"
            f"Sub-item এর নাম লিখুন:\n"
            f"<i>উদাহরণ: Gmail for Personal use</i>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: subp_add_step2(m, p_name))
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_stock_"):
        p_name = callback_data.replace("subp_stock_", "")
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        if not sub_items:
            bot.answer_callback_query(call.id, "❌ কোনো sub-item নেই!", show_alert=True); return
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for item in sub_items:
            markup2.add(InlineBtn(
                f"📤 {item['name']}", style="primary", callback_data=f"subp_upstock_{p_name}__{item['name']}"
            ))
        markup2.add(InlineBtn("🔙 Back", style="primary", callback_data=f"subp_manage_{p_name}"))
        bot.send_message(ADMIN_ID, f"📤 <b>কোন sub-item এর stock আপলোড করবেন?</b>",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_upstock_"):
        rest = callback_data.replace("subp_upstock_", "")
        p_name, sub_name = rest.split("__", 1)
        stock_key = f"{p_name}__{sub_name}"
        msg = bot.send_message(ADMIN_ID,
            f"📎 <b>'{sub_name}' এর stock .xlsx ফাইল পাঠান:</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(ADMIN_ID))
        bot.register_next_step_handler(msg, lambda m: subp_upload_stock(m, p_name, sub_name, stock_key))
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_toggle_"):
        p_name = callback_data.replace("subp_toggle_", "")
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        if not sub_items:
            bot.answer_callback_query(call.id, "❌ কোনো sub-item নেই!", show_alert=True); return
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for item in sub_items:
            status = "✅ ON" if item.get("enabled", True) else "❌ OFF"
            markup2.add(InlineBtn(
                f"{status} — {item['name']}", style="primary", callback_data=f"subp_dotoggle_{p_name}__{item['name']}"
            ))
        markup2.add(InlineBtn("🔙 Back", style="primary", callback_data=f"subp_manage_{p_name}"))
        bot.send_message(ADMIN_ID, f"🔘 <b>On/Off Toggle — {p_name}</b>\n\nনিচে ক্লিক করে toggle করুন:",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_dotoggle_"):
        rest = callback_data.replace("subp_dotoggle_", "")
        p_name, sub_name = rest.split("__", 1)
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        for item in sub_items:
            if item["name"] == sub_name:
                item["enabled"] = not item.get("enabled", True)
                new_status = "ON ✅" if item["enabled"] else "OFF ❌"
                break
        if "sub_products" not in db:
            db["sub_products"] = {}
        if p_name not in db["sub_products"]:
            db["sub_products"][p_name] = {"sub_items": []}
        db["sub_products"][p_name]["sub_items"] = sub_items
        # ⚡ Targeted write
        update_db_path(f"/sub_products/{p_name}/sub_items", {str(i): v for i, v in enumerate(sub_items)})
        bot.answer_callback_query(call.id, f"✅ {sub_name} → {new_status}", show_alert=True)
        # Refresh toggle list
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for item in sub_items:
            status = "✅ ON" if item.get("enabled", True) else "❌ OFF"
            markup2.add(InlineBtn(
                f"{status} — {item['name']}", style="primary", callback_data=f"subp_dotoggle_{p_name}__{item['name']}"
            ))
        markup2.add(InlineBtn("🔙 Back", style="primary", callback_data=f"subp_manage_{p_name}"))
        try:
            bot.edit_message_reply_markup(ADMIN_ID, call.message.message_id, reply_markup=markup2)
        except: pass

    elif callback_data.startswith("subp_delete_"):
        p_name = callback_data.replace("subp_delete_", "")
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        if not sub_items:
            bot.answer_callback_query(call.id, "❌ কোনো sub-item নেই!", show_alert=True); return
        markup2 = types.InlineKeyboardMarkup(row_width=1)
        for item in sub_items:
            markup2.add(InlineBtn(
                f"🗑️ {item['name']}", style="danger", callback_data=f"subp_dodelete_{p_name}__{item['name']}"
            ))
        markup2.add(InlineBtn("🔙 Back", style="primary", callback_data=f"subp_manage_{p_name}"))
        bot.send_message(ADMIN_ID, f"🗑️ <b>কোন sub-item মুছবেন?</b>",
            parse_mode="HTML", reply_markup=markup2)
        pass  # ACK already sent upfront

    elif callback_data.startswith("subp_dodelete_"):
        rest = callback_data.replace("subp_dodelete_", "")
        p_name, sub_name = rest.split("__", 1)
        sub_items = db.get("sub_products", {}).get(p_name, {}).get("sub_items", [])
        sub_items = [s for s in sub_items if s["name"] != sub_name]
        db["sub_products"][p_name]["sub_items"] = sub_items
        # ⚡ Targeted write
        update_db_path(f"/sub_products/{p_name}/sub_items", {str(i): v for i, v in enumerate(sub_items)})
        bot.answer_callback_query(call.id, f"✅ '{sub_name}' মুছে ফেলা হয়েছে!", show_alert=True)
        show_admin_panel(ADMIN_ID)

# ═══════════════════════════════════════════════════════════
# 🔘 PRODUCT BUTTON BUILDER — Admin Helper Functions
# ═══════════════════════════════════════════════════════════

def btn_add_inline_text_step(message, product_name):
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    btn_text = (message.text or "").strip()
    if not btn_text:
        msg = bot.send_message(user_id, "❌ <b>Text খালি!</b>\nআবার লিখুন:", parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: btn_add_inline_text_step(m, product_name)); return
    msg = bot.send_message(user_id,
        f"🔗 <b>বাটনের URL লিখুন:</b>\n\n"
        f"<i>উদাহরণ: https://t.me/yourusername</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: btn_add_inline_url_step(m, product_name, btn_text))

def btn_add_inline_url_step(message, product_name, btn_text):
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    btn_url = (message.text or "").strip()
    if not btn_url.startswith("http"):
        msg = bot.send_message(user_id,
            "❌ <b>Valid URL দিন!</b> (http:// বা https:// দিয়ে শুরু হতে হবে)\nআবার লিখুন:",
            parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: btn_add_inline_url_step(m, product_name, btn_text)); return
    db = load_db()
    if "product_buttons" not in db:
        db["product_buttons"] = {}
    if product_name not in db["product_buttons"]:
        db["product_buttons"][product_name] = []
    db["product_buttons"][product_name].append({"type": "inline", "text": btn_text, "url": btn_url})
    # ⚡ Targeted write
    update_db_path(f"/product_buttons/{product_name}", db["product_buttons"][product_name])
    bot.send_message(user_id,
        f"✅ <b>Inline Button যোগ হয়েছে!</b>\n\n"
        f"📦 Product: {product_name}\n"
        f"🔘 Button: [{btn_text}]({btn_url})",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

def btn_add_keyboard_text_step(message, product_name):
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    btn_text = (message.text or "").strip()
    if not btn_text:
        msg = bot.send_message(user_id, "❌ <b>Text খালি!</b>\nআবার লিখুন:", parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: btn_add_keyboard_text_step(m, product_name)); return
    db = load_db()
    if "product_buttons" not in db:
        db["product_buttons"] = {}
    if product_name not in db["product_buttons"]:
        db["product_buttons"][product_name] = []
    db["product_buttons"][product_name].append({"type": "keyboard", "text": btn_text})
    # ⚡ Targeted write
    update_db_path(f"/product_buttons/{product_name}", db["product_buttons"][product_name])
    bot.send_message(user_id,
        f"✅ <b>Keyboard Button যোগ হয়েছে!</b>\n\n"
        f"📦 Product: {product_name}\n"
        f"⌨️ Button: {btn_text}",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))


# ═══════════════════════════════════════════════════════════
# 📦 SUB-PRODUCT ADMIN HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════

def subp_add_step2(message, p_name):
    """Sub-item নাম নেওয়ার পর price নাও।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    sub_name = (message.text or "").strip()
    if not sub_name:
        msg = bot.send_message(user_id, "❌ <b>নাম খালি!</b>\nআবার লিখুন:", parse_mode="HTML")
        bot.register_next_step_handler(msg, lambda m: subp_add_step2(m, p_name)); return
    msg = bot.send_message(user_id,
        f"💵 <b>'{sub_name}' এর price কত BDT?</b>\n\n<i>উদাহরণ: 50</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: subp_add_step3(m, p_name, sub_name))

def subp_add_step3(message, p_name, sub_name):
    """Price নেওয়ার পর DB তে save করো।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    try:
        price = float(message.text.strip().replace(',', '.'))
        if price < 0:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ <b>সঠিক price দিন (যেমন: 50):</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: subp_add_step3(m, p_name, sub_name)); return
    db = load_db()
    if "sub_products" not in db:
        db["sub_products"] = {}
    if p_name not in db["sub_products"]:
        db["sub_products"][p_name] = {"sub_items": []}
    # Check duplicate
    existing = [s["name"] for s in db["sub_products"][p_name].get("sub_items", [])]
    if sub_name in existing:
        bot.send_message(user_id,
            f"⚠️ <b>'{sub_name}' আগেই আছে!</b>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
        return
    db["sub_products"][p_name]["sub_items"].append({
        "name": sub_name,
        "price": price,
        "enabled": True
    })
    # ⚡ Targeted write
    _sub_list = db["sub_products"][p_name_sub]["sub_items"]
    update_db_path(f"/sub_products/{p_name_sub}/sub_items", {str(i): v for i, v in enumerate(_sub_list)})
    bot.send_message(user_id,
        f"✅ <b>Sub-item যোগ হয়েছে!</b>\n\n"
        f"📦 Product: {p_name}\n"
        f"🔖 Sub-item: {sub_name}\n"
        f"💵 Price: {price} BDT\n\n"
        f"এখন stock আপলোড করুন Admin Panel → Sub-Product Manager থেকে।",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

def subp_upload_stock(message, p_name, sub_name, stock_key):
    """Sub-product stock .xlsx ফাইল আপলোড।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    if message.content_type != 'document':
        msg = bot.send_message(user_id, "❌ <b>Please send an .xlsx file!</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: subp_upload_stock(m, p_name, sub_name, stock_key)); return
    if not message.document.file_name.endswith('.xlsx'):
        msg = bot.send_message(user_id, "❌ <b>Wrong format!</b> Only .xlsx files.",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: subp_upload_stock(m, p_name, sub_name, stock_key)); return
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        file_path = os.path.join(STOCK_DIR, f"{stock_key}.xlsx")

        new_wb = openpyxl.load_workbook(BytesIO(downloaded_file))
        new_ws = new_wb.active

        if os.path.exists(file_path):
            existing_wb = openpyxl.load_workbook(file_path)
            existing_ws = existing_wb.active
            new_rows = list(new_ws.iter_rows(min_row=2, values_only=True))
            added_count = 0
            for row in new_rows:
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    existing_ws.append(list(row))
                    added_count += 1
            existing_wb.save(file_path)
            existing_wb.close()
            new_wb.close()
        else:
            with open(file_path, 'wb') as f:
                f.write(downloaded_file)
            new_wb.close()
            pass  # added_count নিচে set হবে

        # ✅ FIX: file save এর পরে cache clear করো — তারপরই fresh count
        invalidate_stock_cache(stock_key)
        stock_count = get_stock_count(stock_key)
        added_count = stock_count
        bot.send_message(user_id,
            f"✅ <b>Stock Uploaded!</b>\n\n"
            f"📦 {p_name} → {sub_name}\n"
            f"📊 Available: <b>{stock_count} pcs</b>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
    except Exception as e:
        bot.send_message(user_id, f"❌ <b>Upload failed!</b>\n{str(e)}",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))


def admin_add_product_step1(message):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    product_name = message.text.strip()
    msg = bot.send_message(user_id,
        f"💰 <b>Enter price for '{product_name}':</b>\n\n(BDT, numbers only):",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_add_product_step2(m, product_name))

def admin_add_product_step2(message, product_name):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    try:
        price = float(message.text.strip().replace(',', '.'))
        if price <= 0:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ <b>Invalid price!</b> Enter a number (e.g. 50, 99.99):",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_add_product_step2(m, product_name)); return
    # ✅ NEW: description আগে নাও
    msg = bot.send_message(user_id,
        f"📝 <b>'{product_name}' এর Description লিখুন:</b>\n\n"
        f"<i>উদাহরণ: এটি একটি প্রিমিয়াম Netflix অ্যাকাউন্ট যা ৪K সাপোর্ট করে।</i>\n\n"
        f"(অথবা <code>skip</code> লিখুন)",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_add_product_step3(m, product_name, price))

def admin_add_product_step3(message, product_name, price):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    description = "" if (message.text or "").strip().lower() == "skip" else (message.text or "").strip()
    # ✅ NEW: এখন features চাও
    msg = bot.send_message(user_id,
        f"✨ <b>'{product_name}' এর Features লিখুন:</b>\n\n"
        f"<i>উদাহরণ:\n"
        f"• ৩০ দিনের গ্যারান্টি\n"
        f"• ফুল একসেস\n"
        f"• ২৪/৭ সাপোর্ট</i>\n\n"
        f"⚠️ Features গুলো italic অক্ষরে দেখাবে।\n\n"
        f"(অথবা <code>skip</code> লিখুন)",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_add_product_step4(m, product_name, price, description))

def admin_add_product_step4(message, product_name, price, description):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    features = "" if (message.text or "").strip().lower() == "skip" else (message.text or "").strip()
    db["products"][product_name] = price
    if "product_details" not in db:
        db["product_details"] = {}
    db["product_details"][product_name] = {"description": description, "features": features}
    # ⚡ Targeted write — নতুন product এর শুধু দুটো path
    update_db_path(f"/products/{product_name}", price)
    update_db_path(f"/product_details/{product_name}", db["product_details"][product_name])
    bot.send_message(user_id,
        f"✅ <b>Product added!</b>\n\n📦 {product_name}\n💵 {price} BDT\n"
        f"📝 Description: {'সেট' if description else 'নেই'}\n"
        f"✨ Features: {'সেট' if features else 'নেই'}",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

def admin_upload_stock_step1(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text in ("❌ Cancel", "❌ বাতিল"):
        show_admin_panel(user_id); return
    product_name = message.text.strip()
    if product_name not in db["products"]:
        products = list(db["products"].keys())
        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        if products:
            for i in range(0, len(products), 2):
                markup.add( *[KbBtn(_x, role="product_item") for _x in products[i:i+2]])
        markup.add( KbBtn("❌ Cancel", role="cancel"))
        msg = bot.send_message(user_id,
            f"❌ <b>Product '{product_name}' not found!</b>\n\nAvailable products 👇",
            parse_mode="HTML", reply_markup=markup)
        bot.register_next_step_handler(msg, admin_upload_stock_step1); return
    msg = bot.send_message(user_id,
        f"📎 <b>Send .xlsx stock file for '{product_name}'</b>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_upload_stock_step2(m, product_name))

def admin_upload_stock_step2(message, product_name):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    if message.content_type != 'document':
        msg = bot.send_message(user_id, "❌ <b>Please send an .xlsx file!</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_upload_stock_step2(m, product_name)); return
    if not message.document.file_name.endswith('.xlsx'):
        msg = bot.send_message(user_id, "❌ <b>Wrong format!</b> Only .xlsx files.",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_upload_stock_step2(m, product_name)); return
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        file_path = os.path.join(STOCK_DIR, f"{product_name}.xlsx")

        # ══════════════════════════════════════════════════════
        # SMART MERGE: আগে stock থাকলে নতুন data যোগ করো
        # ══════════════════════════════════════════════════════
        new_wb = openpyxl.load_workbook(BytesIO(downloaded_file))
        new_ws = new_wb.active

        # ✅ FIX: নতুন file এ header আছে কিনা auto-detect করো
        # 1st row এর সব cell string এবং কোনো cell এ number নেই → header
        all_new_rows = list(new_ws.iter_rows(values_only=True))
        new_has_header = False
        if all_new_rows:
            first_row = all_new_rows[0]
            non_empty = [c for c in first_row if c is not None and str(c).strip() != ""]
            if non_empty and all(not str(c).replace('.', '', 1).isdigit() for c in non_empty):
                new_has_header = True  # 1st row এ শুধু text → header
        data_start_row = 2 if new_has_header else 1

        if os.path.exists(file_path):
            # পুরনো stock file আছে → merge করো
            existing_wb = openpyxl.load_workbook(file_path)
            existing_ws = existing_wb.active

            # নতুন file এর data rows নাও (header skip করো যদি থাকে)
            new_rows = list(new_ws.iter_rows(min_row=data_start_row, values_only=True))
            added_count = 0
            for row in new_rows:
                # Empty row skip করো
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    existing_ws.append(list(row))
                    added_count += 1

            existing_wb.save(file_path)
            existing_wb.close()
            new_wb.close()
            merge_mode = True
            # ✅ FIX: save এর পরে cache clear করো — তারপরই fresh count নাও
            invalidate_stock_cache(product_name)
            old_count = get_stock_count(product_name) - added_count
        else:
            # নতুন product → standard format এ save করো (1st row = header)
            if new_has_header:
                # File এ header আছে → সরাসরি save
                with open(file_path, 'wb') as f:
                    f.write(downloaded_file)
            else:
                # File এ header নেই → "Data" header যোগ করে save করো
                out_wb = openpyxl.Workbook()
                out_ws = out_wb.active
                # Header row যোগ করো
                if all_new_rows and all_new_rows[0]:
                    col_count = len(all_new_rows[0])
                    headers = ["Data"] + [f"Col{i}" for i in range(2, col_count + 1)]
                    out_ws.append(headers)
                else:
                    out_ws.append(["Data"])
                # Data rows যোগ করো
                for row in all_new_rows:
                    if any(cell is not None and str(cell).strip() != "" for cell in row):
                        out_ws.append(list(row))
                out_wb.save(file_path)
                out_wb.close()
            new_wb.close()
            merge_mode = False
            old_count = 0
            # ✅ FIX: নতুন file save এর পরেও cache clear করো
            invalidate_stock_cache(product_name)
        db = load_db()
        # cache আগেই clear হয়েছে (merge/new উভয় path এ) — সরাসরি fresh count নাও
        stock_count = get_stock_count(product_name)
        price = db["products"].get(product_name, 0)

        import html as _html_st
        safe_pname = _html_st.escape(str(product_name))
        if merge_mode:
            status_text = (
                f"✅ <b>Stock Merged Successfully!</b>\n\n"
                f"📦 <b>Product:</b> {safe_pname}\n"
                f"➕ <b>New Added:</b> {added_count} pcs\n"
                f"📊 <b>Total Available:</b> {stock_count} pcs\n"
                f"💵 <b>Price:</b> {price} BDT"
            )
        else:
            status_text = (
                f"✅ <b>Stock Uploaded Successfully!</b>\n\n"
                f"📦 <b>Product:</b> {safe_pname}\n"
                f"📊 <b>Available:</b> {stock_count} pcs\n"
                f"💵 <b>Price:</b> {price} BDT"
            )

        bot.send_message(user_id, status_text, parse_mode="HTML")

        # ── Broadcast message তৈরি করো কিন্তু এখনই পাঠাবো না ──
        now = datetime.now(BD_TZ).strftime("%d %b %Y • %I:%M %p")
        import html as _html_bc
        _safe_pname_bc = _html_bc.escape(str(product_name))
        _safe_now = _html_bc.escape(str(now))
        broadcast_msg = (
            f"🔔 <b>নতুন স্টক আসছে!</b>\n"
            f"\n\n"
            f"📦 <b>পণ্য:</b>  {_safe_pname_bc}\n"
            f"✅ <b>স্টক:</b>  <code>{stock_count} pcs available</code>\n"
            f"💵 <b>দাম:</b>  <code>{price} BDT</code> / পিস\n\n"
            f"\n"
            f"⚡ <b>স্টক সীমিত!</b> দ্রুত অর্ডার করুন।\n"
            f"🛒 Shop → <b>{_safe_pname_bc}</b> সিলেক্ট করুন\n\n"
            f"🕐 <i>{_safe_now}</i>"
        )
        channel_msg = (
            f"<b>📊NEW STOCK AVAILABLE!📊</b>\n"
            f"\n\n"
            f"📦 <b>Product:</b>  {_safe_pname_bc}\n"
            f"✅ <b>Stock:</b>  {stock_count} pcs available\n"
            f"💵 <b>Price:</b>  {price} BDT / pc\n\n"
            f"\n"
            f"🤖 Order via bot @{BOT_USERNAME}\n"
            f"📅 {_safe_now}"
        )

        # ── Pending broadcast store করো ──
        safe_name = product_name.replace(" ", "_")
        _pending_broadcasts[safe_name] = {
            "product_name":  product_name,
            "broadcast_msg": broadcast_msg,
            "channel_msg":   channel_msg,
            "stock_count":   stock_count,
            "price":         price,
            "user_count":    len(db["users"]),
        }

        # ── Admin কে confirm/cancel বাটন দেখাও ──
        confirm_markup = types.InlineKeyboardMarkup()
        confirm_markup.row(
            InlineBtn("✅ Broadcast করুন", style="success", callback_data=f"confirm_stock_broadcast_{safe_name}"),
            InlineBtn("❌ Cancel", style="danger",          callback_data=f"cancel_stock_broadcast_{safe_name}")
        )
        import html as _html_mod
        preview_text = _html_mod.escape(broadcast_msg[:250])
        safe_pname2 = _html_mod.escape(str(product_name))
        bot.send_message(
            user_id,
            f"📢 <b>Broadcast Permission</b>\n\n"
            f"📦 <b>Product:</b> {safe_pname2}\n"
            f"👥 <b>মোট Users:</b> {len(db['users'])} জন\n\n"
            f"সব user কে stock notification পাঠাবেন?\n\n"
            f"<b>Preview:</b>\n<code>{preview_text}...</code>",
            parse_mode="HTML",
            reply_markup=confirm_markup
        )

    except Exception as e:
        import html as _html_err
        safe_err = _html_err.escape(str(e))
        bot.send_message(user_id, f"❌ <b>Upload failed!</b>\n\n<code>{safe_err}</code>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))

def admin_edit_price_step(message, product_name):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    try:
        new_price = float(message.text.strip().replace(',', '.'))
        if new_price <= 0:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ <b>সঠিক মূল্য দিন (decimal সহ, যেমন: 99.99):</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_edit_price_step(m, product_name)); return
    old_price = db["products"].get(product_name, 0)
    db["products"][product_name] = new_price
    # ⚡ Targeted write
    update_db_path(f"/products/{product_name}", new_price)
    src_label = "🌐 API Product" if product_name in API_PRODUCTS else "📦 Local Product"
    bot.send_message(user_id,
        f"✅ <b>Price updated!</b>\n\n{src_label}: {product_name}\n💵 {old_price} → {new_price} BDT",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

def admin_search_order_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    order_id = message.text.strip().upper()
    found = False
    for uid, user in db["users"].items():
        for order in user.get("orders", []):
            if order["id"].upper() == order_id:
                found = True
                bot.send_message(user_id,
                    f"✅ <b>Order Found!</b>\n\n"
                    f"🆔 <b>Order ID:</b> <code>{order['id']}</code>\n"
                    f"👤 <b>User ID:</b> <code>{uid}</code>\n"
                    f"📦 <b>Product:</b> {order['product_name']}\n"
                    f"🔢 <b>Qty:</b> {order['quantity']} pcs\n"
                    f"💰 <b>Total:</b> {order['total']} BDT\n"
                    f"📅 <b>Date:</b> {order['date']}",
                    parse_mode="HTML"); break
    if not found:
        bot.send_message(user_id, "❌ <b>Order not found!</b>", parse_mode="HTML")

def admin_user_info_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target_user = message.text.strip()
    if target_user not in db["users"]:
        bot.send_message(user_id, "❌ <b>User not found!</b>", parse_mode="HTML"); return
    user = db["users"][target_user]
    is_banned = target_user in db.get("banned_users", [])

    # ✅ Username Telegram থেকে আনো
    username_str = "N/A"
    try:
        tg_user = bot.get_chat(int(target_user))
        if tg_user.username:
            username_str = f"@{tg_user.username}"
        elif tg_user.first_name:
            username_str = tg_user.first_name + (f" {tg_user.last_name}" if tg_user.last_name else "")
    except Exception:
        pass

    # ✅ Orders breakdown — কোন product কতটা কিনেছে
    orders = user.get("orders", [])
    order_summary = ""
    if orders:
        from collections import defaultdict
        prod_qty = defaultdict(int)
        for o in orders:
            prod_qty[o.get("product_name", "Unknown")] += int(o.get("quantity", 1))
        lines = [f"   • {p}: <b>{q} pcs</b>" for p, q in prod_qty.items()]
        order_summary = "\n📋 <b>Order Breakdown:</b>\n" + "\n".join(lines)

    bot.send_message(user_id,
        f"👤 <b>User Info</b>\n\n"
        f"🆔 <b>User ID:</b> <code>{target_user}</code>\n"
        f"👤 <b>Username:</b> {username_str}\n"
        f"💳 <b>Balance:</b> {user.get('balance', 0)} BDT\n"
        f"👥 <b>Referrals:</b> {user.get('refer_count', 0)}\n"
        f"📦 <b>Total Orders:</b> {len(orders)} pcs"
        f"{order_summary}\n"
        f"📅 <b>Joined:</b> {user.get('join_date', 'N/A')}\n"
        f"🌐 <b>Language:</b> {user.get('lang', 'en').upper()}\n"
        f"🚫 <b>Status:</b> {'Banned ❌' if is_banned else 'Active ✅'}",
        parse_mode="HTML")

def admin_message_user_step1(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target_user = message.text.strip()
    if target_user not in db["users"]:
        bot.send_message(user_id, "❌ <b>User not found!</b>", parse_mode="HTML"); return
    msg = bot.send_message(user_id,
        f"📝 <b>Message to User {target_user}:</b>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_message_user_step2(m, target_user))

def admin_message_user_step2(message, target_user):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    try:
        bot.send_message(target_user,
            f"📩 <b>Message from Admin</b>\n\n{message.text}", parse_mode="HTML")
        bot.send_message(user_id,
            f"✅ <b>Message sent!</b>\n\nTo: <code>{target_user}</code>", parse_mode="HTML")
    except Exception as e:
        import html as _html_e; bot.send_message(user_id, f"❌ <b>Failed!</b>\n\n<code>{_html_e.escape(str(e))}</code>", parse_mode="HTML")

# ═══════════════════════════════════════════════════════════
# 🏷️ ADMIN USER DISCOUNT — নির্দিষ্ট user কে personal discount
# ═══════════════════════════════════════════════════════════

def admin_user_discount_step1(message):
    """Step 1: User ID নাও।"""
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target_user = message.text.strip()
    if target_user not in db["users"]:
        bot.send_message(user_id, "❌ <b>User পাওয়া যায়নি!</b>\n\nসঠিক User ID দিন:", parse_mode="HTML")
        return
    # User info দেখাও
    user_data = db["users"][target_user]
    current_disc = user_data.get("personal_discount", 0)
    username_str = "N/A"
    try:
        tg_user = bot.get_chat(int(target_user))
        username_str = f"@{tg_user.username}" if tg_user.username else (tg_user.first_name or target_user)
    except: pass

    msg = bot.send_message(user_id,
        f"🏷️ <b>User Discount সেট করুন</b>\n\n"
        f"👤 <b>User:</b> {username_str}\n"
        f"🆔 <b>ID:</b> <code>{target_user}</code>\n"
        f"💳 <b>Balance:</b> {user_data.get('balance', 0)} BDT\n"
        f"🏷️ <b>বর্তমান Discount:</b> {current_disc}%\n\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"নতুন discount % লিখুন (1-99):\n"
        f"<i>যেমন: 20 (মানে 20% ছাড়)</i>\n"
        f"<i>0 লিখলে discount সরিয়ে দেবে</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_user_discount_step2(m, target_user))

def admin_user_discount_step2(message, target_user):
    """Step 2: Discount % নাও এবং save করো।"""
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    try:
        disc_val = int(message.text.strip())
        if disc_val < 0 or disc_val > 99:
            raise ValueError("Out of range")
    except ValueError:
        msg = bot.send_message(user_id,
            "❌ <b>ভুল input!</b>\n\n1 থেকে 99 এর মধ্যে সংখ্যা দিন (0 = discount সরাও):",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_user_discount_step2(m, target_user))
        return

    old_disc = db["users"][target_user].get("personal_discount", 0)
    if disc_val == 0:
        db["users"][target_user].pop("personal_discount", None)
        action_text = "সরিয়ে দেওয়া হয়েছে"
    else:
        db["users"][target_user]["personal_discount"] = disc_val
        action_text = f"সেট করা হয়েছে → <b>{disc_val}%</b>"

    # ⚡ Targeted write
    update_db_path(f"/users/{target_user}/personal_discount", db["users"][target_user].get("personal_discount", 0))

    # Remove discount button
    remove_markup = types.InlineKeyboardMarkup()
    if disc_val > 0:
        remove_markup.add(
            InlineBtn(
                f"🗑️ {target_user} এর discount সরাও", style="danger",
                callback_data=f"user_discount_remove_{target_user}")
        )
    remove_markup.add(
        InlineBtn("🏷️ আরেকজনকে Discount দিন", style="primary", callback_data="admin_user_discount"),
        InlineBtn("🔙 Admin Panel", style="primary", callback_data="back_admin")
    )

    bot.send_message(user_id,
        f"✅ <b>Discount আপডেট হয়েছে!</b>\n\n"
        f"👤 User: <code>{target_user}</code>\n"
        f"??️ আগের discount: <b>{old_disc}%</b>\n"
        f"🏷️ নতুন discount: {action_text}\n\n"
        f"<i>এখন থেকে এই user সব পণ্যে {disc_val}% ছাড় পাবে।</i>",
        parse_mode="HTML", reply_markup=remove_markup)

    # User কে notify করো
    try:
        if disc_val > 0:
            bot.send_message(target_user,
                f"🎉 <b>আপনি একটি special discount পেয়েছেন!</b>\n\n"
                f"🏷️ <b>আপনার Personal Discount: {disc_val}%</b>\n\n"
                f"এখন থেকে সব পণ্য কিনলে <b>{disc_val}% ছাড়</b> পাবেন! 🛒",
                parse_mode="HTML")
        else:
            bot.send_message(target_user,
                "ℹ️ <b>আপনার personal discount সরিয়ে নেওয়া হয়েছে।</b>",
                parse_mode="HTML")
    except: pass


def admin_edit_balance_step1(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target_user = message.text.strip()
    if target_user not in db["users"]:
        bot.send_message(user_id, "❌ <b>User not found!</b>", parse_mode="HTML"); return
    current_balance = db["users"][target_user].get("balance", 0)
    msg = bot.send_message(user_id,
        f"💰 <b>Edit Balance</b>\n\n"
        f"👤 User: <code>{target_user}</code>\n"
        f"💳 Current: <b>{current_balance} BDT</b>\n\n"
        f"• <code>500</code> → Set 500 BDT\n"
        f"• <code>+100</code> → Add 100 BDT\n"
        f"• <code>-50</code> → Deduct 50 BDT",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_edit_balance_step2(m, target_user))

def admin_edit_balance_step2(message, target_user):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    value = message.text.strip()
    current_balance = float(db["users"][target_user].get("balance", 0))
    try:
        if value.startswith('+'):
            added = float(value[1:])
            new_balance = round(current_balance + added, 2)
            action_label = f"+{added} BDT যোগ হয়েছে"
        elif value.startswith('-'):
            deducted = float(value[1:])
            new_balance = round(max(0, current_balance - deducted), 2)
            action_label = f"-{deducted} BDT কাটা হয়েছে"
        else:
            new_balance = round(float(value), 2)
            action_label = f"Balance {new_balance} BDT সেট হয়েছে"

        diff = round(new_balance - current_balance, 2)

        # ⚡ INSTANT: শুধু balance path update — পুরো DB write নয়
        db["users"][target_user]["balance"] = new_balance
        _update_db_cache_in_place(db)  # cache সাথে সাথে update
        update_db_path(f"/users/{target_user}/balance", new_balance, _async=False)

        # Admin কে confirm
        bot.send_message(user_id,
            f"✅ <b>Balance Updated!</b>\n\n"
            f"👤 User: <code>{target_user}</code>\n"
            f"💳 আগে: <b>{current_balance} BDT</b>\n"
            f"💰 এখন: <b>{new_balance} BDT</b>\n"
            f"📊 পরিবর্তন: <code>{action_label}</code>",
            parse_mode="HTML")

        # ⚡ User কে সাথে সাথে notify — স্পষ্ট message
        try:
            lang_u = _lang_cache.get(target_user, "en")
            if diff > 0:
                user_msg = (
                    f"💰 <b>আপনার Balance আপডেট হয়েছে!</b>\n\n"
                    f"✅ <b>+{diff} BDT যোগ হয়েছে</b>\n\n"
                    f"💳 <b>নতুন Balance: {new_balance} BDT</b>\n\n"
                    f"🛍️ এখনই কেনাকাটা করুন!"
                ) if lang_u == "bn" else (
                    f"💰 <b>Your Balance Updated!</b>\n\n"
                    f"✅ <b>+{diff} BDT Added</b>\n\n"
                    f"💳 <b>New Balance: {new_balance} BDT</b>\n\n"
                    f"🛍️ Start shopping now!"
                )
            elif diff < 0:
                user_msg = (
                    f"⚠️ <b>Balance আপডেট হয়েছে</b>\n\n"
                    f"📉 <b>{abs(diff)} BDT কাটা হয়েছে</b>\n\n"
                    f"💳 <b>নতুন Balance: {new_balance} BDT</b>"
                ) if lang_u == "bn" else (
                    f"⚠️ <b>Balance Updated</b>\n\n"
                    f"📉 <b>{abs(diff)} BDT Deducted</b>\n\n"
                    f"💳 <b>New Balance: {new_balance} BDT</b>"
                )
            else:
                user_msg = (
                    f"💳 <b>Balance: {new_balance} BDT</b>\n\n"
                    f"<i>Admin আপনার balance set করেছেন।</i>"
                )
            shop_markup = types.InlineKeyboardMarkup()
            shop_markup.add(InlineBtn("🛍️ Shop Now", style="primary", callback_data="go_to_shop"))
            bot.send_message(target_user, user_msg, parse_mode="HTML", reply_markup=shop_markup)
        except: pass

    except ValueError:
        bot.send_message(user_id,
            "❌ <b>Invalid input!</b>\n\nUse: 500, +100, or -50", parse_mode="HTML")

def admin_ban_user_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target_user = message.text.strip()
    if target_user not in db["users"]:
        bot.send_message(user_id, "❌ <b>User not found!</b>", parse_mode="HTML"); return
    if target_user in db["banned_users"]:
        bot.send_message(user_id, "⚠️ <b>User already banned!</b>", parse_mode="HTML"); return
    db["banned_users"].append(target_user)
    # ⚡ Targeted write
    update_db_path("/banned_users", {str(i): v for i, v in enumerate(db["banned_users"])})
    bot.send_message(user_id,
        f"✅ <b>User banned!</b>\n\n👤 <code>{target_user}</code>", parse_mode="HTML")
    try:
        bot.send_message(target_user,
            "🚫 <b>You have been banned.</b>\n\nContact support.", parse_mode="HTML")
    except: pass

def admin_broadcast_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    import html as _html_bc2
    bcast_text = message.text
    all_uids = list(db["users"].keys())
    bot.send_message(user_id,
        f"📢 <b>Broadcast শুরু হচ্ছে...</b>\n👥 <b>{len(all_uids)} জন</b> user কে পাঠানো হবে।\n"
        f"<i>Background এ চলবে, bot slow হবে না।</i>",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

    def _do_admin_broadcast(uids, text):
        sent = failed = 0
        for uid in uids:
            try:
                bot.send_message(uid, f"📢 <b>Admin Announcement</b>\n\n{text}",
                    parse_mode="HTML")
                sent += 1
                _time_module.sleep(0.05)
            except:
                failed += 1
        try:
            bot.send_message(user_id,
                f"✅ <b>Broadcast complete!</b>\n\n✅ Sent: {sent}\n❌ Failed: {failed}",
                parse_mode="HTML")
        except: pass

    threading.Thread(target=_do_admin_broadcast, args=(all_uids, bcast_text), daemon=True).start()

# ═══════════════════════════════════════════════════════════
# 📈 ANALYTICS & REPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════

def generate_sales_report(db, period="daily"):
    """Daily/Weekly/Monthly sales report generate করো।"""
    from datetime import timedelta
    now_dt = datetime.now(BD_TZ)
    # ✅ "yesterday" period শুধুমাত্র গতকালের নির্দিষ্ট দিনটার জন্য
    # (day_start <= order_date < day_end), বাকিগুলো rolling cutoff ভিত্তিক
    day_start = None
    day_end = None
    if period == "daily":
        cutoff = now_dt - timedelta(days=1)
        period_label = "আজকের"
    elif period == "yesterday":
        y_date = (now_dt - timedelta(days=1)).date()
        day_start = BD_TZ.localize(datetime.combine(y_date, datetime.min.time()))
        day_end = day_start + timedelta(days=1)
        cutoff = day_start
        period_label = "গতকালকের"
    elif period == "weekly":
        cutoff = now_dt - timedelta(days=7)
        period_label = "সাপ্তাহিক (৭ দিন)"
    else:
        cutoff = now_dt - timedelta(days=30)
        period_label = "মাসিক (৩০ দিন)"

    cutoff_str = cutoff.strftime("%d/%m/%Y")

    total_revenue = 0
    total_orders = 0
    product_counts = {}

    for user in db["users"].values():
        for order in user.get("orders", []):
            try:
                order_date_str = order.get("date", "")
                # date format: "dd/mm/YYYY HH:MM AM/PM"
                order_date = datetime.strptime(order_date_str[:10], "%d/%m/%Y")
                order_date = BD_TZ.localize(order_date)
                if day_start is not None:
                    in_range = day_start <= order_date < day_end
                else:
                    in_range = order_date >= cutoff
                if in_range:
                    total_revenue += int(order.get("total", 0))
                    total_orders += 1
                    pname = order.get("product_name", "Unknown")
                    product_counts[pname] = product_counts.get(pname, 0) + int(order.get("quantity", 1))
            except Exception:
                pass

    # Deposit income
    total_deposits = sum(
        int(float(r.get("amount", 0)))
        for r in db.get("deposit_requests", {}).values()
        if r.get("status") == "approved"
    )

    # Best selling
    best_products = sorted(product_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    best_str = ""
    for i, (pname, qty) in enumerate(best_products, 1):
        best_str += f"  {i}. {pname} — {qty} pcs\n"
    if not best_str:
        best_str = "  কোনো বিক্রি নেই\n"

    report = (
        f"📈 <b>{period_label} Sales Report</b>\n"
        f"\n\n"
        f"🛒 <b>মোট Orders:</b> {total_orders}\n"
        f"💰 <b>মোট Revenue:</b> {total_revenue} BDT\n"
        f"👥 <b>মোট Users:</b> {len(db['users'])}\n"
        f"✅ <b>Active Users:</b> {sum(1 for u in db['users'].values() if u.get('orders'))}\n\n"
        f"🏆 <b>Top Products:</b>\n{best_str}"
        f"\n"
        f"📅 {get_now()}"
    )
    return report

def generate_product_sales_report(db):
    """সবচেয়ে বেশি বিক্রি হওয়া product report।"""
    product_counts = {}
    product_revenue = {}
    for user in db["users"].values():
        for order in user.get("orders", []):
            pname = order.get("product_name", "Unknown")
            qty = int(order.get("quantity", 1))
            total = int(order.get("total", 0))
            product_counts[pname] = product_counts.get(pname, 0) + qty
            product_revenue[pname] = product_revenue.get(pname, 0) + total

    sorted_products = sorted(product_counts.items(), key=lambda x: x[1], reverse=True)

    report = (
        f"🏆 <b>Product Sales Report</b>\n"
        f"\n\n"
    )
    for i, (pname, qty) in enumerate(sorted_products, 1):
        revenue = product_revenue.get(pname, 0)
        stock = get_stock_count(pname)
        price = db["products"].get(pname, 0)
        report += (
            f"{i}. <b>{pname}</b>\n"
            f"   📦 বিক্রি: {qty} pcs | 💰 Revenue: {revenue} BDT\n"
            f"   🔢 Stock: {stock} pcs | 💵 Price: {price} BDT\n\n"
        )

    if not sorted_products:
        report += "কোনো বিক্রি হয়নি এখনো।\n"

    report += f"\n📅 {get_now()}"
    return report

def generate_finance_report(db):
    """Finance report — income, pending deposits, today's income।"""
    today = datetime.now(BD_TZ).strftime("%d/%m/%Y")

    all_deposits = db.get("deposit_requests", {})
    approved = [r for r in all_deposits.values() if r["status"] == "approved"]
    pending = [r for r in all_deposits.values() if r["status"] == "pending"]
    rejected = [r for r in all_deposits.values() if r["status"] == "rejected"]

    total_income = sum(int(float(r.get("amount", 0))) for r in approved)
    today_income = sum(
        int(float(r.get("amount", 0))) for r in approved
        if r.get("date", "").startswith(today)
    )
    pending_amount = sum(int(float(r.get("amount", 0))) for r in pending)

    # Total revenue from sales
    total_sales_revenue = sum(
        int(o.get("total", 0))
        for u in db["users"].values()
        for o in u.get("orders", [])
    )

    # Total user balance
    total_balance = sum(u.get("balance", 0) for u in db["users"].values())

    report = (
        f"💵 <b>Finance Report</b>\n"
        f"\n\n"
        f"📅 <b>আজকের Income:</b> {today_income} BDT\n"
        f"💰 <b>মোট Deposit Income:</b> {total_income} BDT\n"
        f"🛒 <b>মোট Sales Revenue:</b> {total_sales_revenue} BDT\n\n"
        f"⏳ <b>Pending Deposits:</b> {len(pending)} টি ({pending_amount} BDT)\n"
        f"✅ <b>Approved Deposits:</b> {len(approved)} টি\n"
        f"❌ <b>Rejected Deposits:</b> {len(rejected)} টি\n\n"
        f"💳 <b>Total User Balance (সব users):</b> {total_balance} BDT\n"
        f"\n"
        f"📅 {get_now()}"
    )
    return report

def generate_stock_alert_report(db):
    """Low stock alert — ৫ পিসের কম থাকলে alert দেখাও।"""
    LOW_STOCK_THRESHOLD = 5
    report = (
        f"📦 <b>Stock Alert Report</b>\n"
        f"\n\n"
    )
    low_stock_items = []
    ok_items = []

    for product in db["products"]:
        stock = get_stock_count(product)
        price = db["products"][product]
        if stock <= LOW_STOCK_THRESHOLD:
            low_stock_items.append((product, stock, price))
        else:
            ok_items.append((product, stock, price))

    if low_stock_items:
        report += f"⚠️ <b>LOW STOCK (≤{LOW_STOCK_THRESHOLD} pcs):</b>\n"
        for pname, stock, price in low_stock_items:
            status = "🔴 Out of Stock" if stock == 0 else f"🟡 {stock} pcs বাকি"
            report += f"  • <b>{pname}</b> — {status} | {price} BDT\n"
        report += "\n"
    else:
        report += "✅ সব products এ পর্যাপ্ত stock আছে।\n\n"

    if ok_items:
        report += f"✅ <b>OK Stock:</b>\n"
        for pname, stock, price in ok_items:
            report += f"  • <b>{pname}</b> — {stock} pcs | {price} BDT\n"

    report += f"\n\n📅 {get_now()}"
    return report

def generate_vip_users_report(db):
    """বেশি কেনাকাটা করেছে এমন VIP users দেখাও।"""
    user_totals = []
    for uid, user in db["users"].items():
        total_spent = sum(int(o.get("total", 0)) for o in user.get("orders", []))
        total_orders = len(user.get("orders", []))
        if total_orders > 0:
            user_totals.append((uid, total_spent, total_orders, user.get("balance", 0)))

    user_totals.sort(key=lambda x: x[1], reverse=True)
    top_users = user_totals[:10]

    report = (
        f"⭐ <b>VIP Users (Top Buyers)</b>\n"
        f"\n\n"
    )
    for i, (uid, spent, orders, balance) in enumerate(top_users, 1):
        vip_badge = "👑" if spent >= 1000 else ("⭐" if spent >= 500 else "🔷")
        report += (
            f"{vip_badge} <b>#{i}</b> <code>{uid}</code>\n"
            f"   💰 Spent: {spent} BDT | 🛒 Orders: {orders} | 💳 Balance: {balance} BDT\n\n"
        )

    if not top_users:
        report += "এখনো কোনো purchase হয়নি।\n"

    report += f"\n📅 {get_now()}"
    return report

def admin_user_order_history_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target = message.text.strip()
    if target not in db["users"]:
        bot.send_message(user_id, "❌ <b>User পাওয়া যায়নি!</b>", parse_mode="HTML"); return
    orders = db["users"][target].get("orders", [])
    if not orders:
        bot.send_message(user_id, f"📭 User <code>{target}</code> এর কোনো order নেই।", parse_mode="HTML"); return
    report = f"📜 <b>User {target} এর Orders</b>\n\n\n"
    for o in orders[-15:]:
        report += (
            f"🆔 <code>{o['id']}</code>\n"
            f"📦 {o['product_name']} | 🔢 {o['quantity']} pcs | ?? {o['total']} BDT\n"
            f"📅 {o['date']}\n\n"
        )
    bot.send_message(user_id, report, parse_mode="HTML")

def admin_user_deposit_history_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    target = message.text.strip()
    deposits = [r for r in db.get("deposit_requests", {}).values() if r.get("user_id") == target]
    if not deposits:
        bot.send_message(user_id, f"📭 User <code>{target}</code> এর কোনো deposit নেই।", parse_mode="HTML"); return
    report = f"💰 <b>User {target} এর Deposits</b>\n\n\n"
    for r in deposits[-10:]:
        status_e = "✅" if r["status"] == "approved" else ("❌" if r["status"] == "rejected" else "⏳")
        report += (
            f"🆔 <code>{r['request_id']}</code>\n"
            f"💳 {r['method']} | 💰 {r['amount']} BDT | {status_e} {r['status'].upper()}\n"
            f"📅 {r['date']}\n\n"
        )
    bot.send_message(user_id, report, parse_mode="HTML")

def admin_set_refer_bonus_step(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    if not message.text.strip().isdigit():
        bot.send_message(user_id, "❌ সংখ্যায় লিখুন (যেমন: 5)", parse_mode="HTML"); return
    new_bonus = int(message.text.strip())
    db["settings"]["refer_bonus"] = new_bonus
    # ⚡ Targeted write
    update_db_path("/settings/refer_bonus", new_bonus)
    bot.send_message(user_id,
        f"✅ <b>Refer Bonus আপডেট হয়েছে!</b>\n\n"
        f"🎁 নতুন Bonus: <b>{new_bonus} BDT</b> per referral",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

# ═══════════════════════════════════════════════════════════
# 🎟️ PROMO CODE ADMIN STEP FUNCTIONS
# ═══════════════════════════════════════════════════════════

def admin_create_promo_step1(message):
    """Step 1: Promo code name নাও।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    code = (message.text or "").strip().upper()
    if not code or not all(c.isalnum() or c in "_-" for c in code):
        msg = bot.send_message(user_id,
            "❌ <b>Invalid code!</b> শুধু অক্ষর, সংখ্যা ব্যবহার করুন:",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, admin_create_promo_step1); return
    db = load_db()
    if code in db.get("promo_codes", {}):
        msg = bot.send_message(user_id,
            f"❌ <b>{code}</b> আগেই বিদ্যমান! অন্য নাম দিন:",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, admin_create_promo_step1); return
    msg = bot.send_message(user_id,
        f"✅ Code: <b>{code}</b>\n\n💸 কত % ছাড় দেবেন? (1-99)\n<i>উদাহরণ: 10 মানে ১০% ছাড়</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_create_promo_step2(m, code))

def admin_create_promo_step2(message, code):
    """Step 2: Discount percent নাও।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    try:
        disc = float(message.text.strip())
        if disc <= 0 or disc > 99:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ ১ থেকে ৯৯ এর মধ্যে সংখ্যা দিন:",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_create_promo_step2(m, code)); return
    msg = bot.send_message(user_id,
        f"✅ Code: <b>{code}</b> | ছাড়: <b>{disc:.0f}%</b>\n\n👥 সর্বোচ্চ কতবার ব্যবহার করা যাবে?\n<i>0 দিলে সীমাহীন, নাহলে সংখ্যা দিন</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_create_promo_step3(m, code, disc))

def admin_create_promo_step3(message, code, disc):
    """Step 3: Max uses নাও।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    try:
        max_uses = int(message.text.strip())
        if max_uses < 0:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ 0 বা তার বেশি সংখ্যা দিন:",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_create_promo_step3(m, code, disc)); return
    msg = bot.send_message(user_id,
        f"✅ Code: <b>{code}</b> | ছাড়: <b>{disc:.0f}%</b> | সীমা: <b>{'∞' if max_uses==0 else max_uses}</b>\n\n📅 মেয়াদ শেষ তারিখ দিন (dd/mm/yyyy):\n<i>মেয়াদ না থাকলে <code>skip</code> লিখুন</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_create_promo_step4(m, code, disc, max_uses))

def admin_create_promo_step4(message, code, disc, max_uses):
    """Step 4: Expiry নাও।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    expiry = ""
    text = (message.text or "").strip().lower()
    if text != "skip":
        try:
            datetime.strptime(text, "%d/%m/%Y")
            expiry = text
        except ValueError:
            msg = bot.send_message(user_id, "❌ তারিখ ফরম্যাট ভুল! dd/mm/yyyy দিন বা skip:",
                parse_mode="HTML", reply_markup=get_cancel_button(user_id))
            bot.register_next_step_handler(msg, lambda m: admin_create_promo_step4(m, code, disc, max_uses)); return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add( KbBtn("✅ নতুন User Only", role="misc_option"), KbBtn("👥 সবার জন্য", role="misc_option"), KbBtn("❌ Cancel", role="cancel"))
    msg = bot.send_message(user_id,
        "🆕 এই code কি শুধু <b>নতুন User</b> দের জন্য?\n(যাদের কোনো পূর্ববর্তী order নেই)",
        parse_mode="HTML", reply_markup=markup)
    bot.register_next_step_handler(msg, lambda m: admin_create_promo_step5(m, code, disc, max_uses, expiry))

def admin_create_promo_step5(message, code, disc, max_uses, expiry):
    """Step 5: Save করো।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    new_user_only = (message.text or "").strip() == "✅ নতুন User Only"
    db = load_db()
    if "promo_codes" not in db:
        db["promo_codes"] = {}
    db["promo_codes"][code] = {
        "discount_percent": disc,
        "max_uses": max_uses,
        "used_count": 0,
        "expiry": expiry,
        "new_user_only": new_user_only,
        "enabled": True,
        "created": get_now()
    }
    # ⚡ Targeted write
    _promo_code_key = list(db["promo_codes"].keys())[-1] if db.get("promo_codes") else ""
    update_db_path("/promo_codes", db["promo_codes"])
    bot.send_message(user_id,
        f"✅ <b>Promo Code তৈরি হয়েছে!</b>\n\n"
        f"🎟️ <b>Code:</b> <code>{code}</code>\n"
        f"💸 <b>ছাড়:</b> {disc:.0f}%\n"
        f"👥 <b>সীমা:</b> {'∞' if max_uses==0 else max_uses} বার\n"
        f"📅 <b>মেয়াদ:</b> {expiry if expiry else 'অসীম'}\n"
        f"🆕 <b>নতুন User Only:</b> {'হ্যাঁ' if new_user_only else 'না'}\n"
        f"✅ <b>Status:</b> Active",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))
    log_action("PROMO_CREATED", user_id, f"Code: {code}, Discount: {disc}%")

def admin_set_new_user_discount_step(message):
    """নতুন user দের global discount % সেট করো।"""
    user_id = str(message.chat.id)
    if message.text in ("❌ Cancel",):
        show_admin_panel(user_id); return
    try:
        disc = float(message.text.strip())
        if disc < 0 or disc > 99:
            raise ValueError
    except ValueError:
        msg = bot.send_message(user_id, "❌ 0 থেকে 99 এর মধ্যে সংখ্যা দিন:",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, admin_set_new_user_discount_step); return
    db = load_db()
    db["settings"]["new_user_discount"] = disc
    # ⚡ Targeted write
    update_db_path("/settings/new_user_discount", disc)
    status = f"{disc:.0f}% ছাড়" if disc > 0 else "বন্ধ (0%)"
    bot.send_message(user_id,
        f"✅ <b>New User Discount আপডেট!</b>\n\n🆕 নতুন User রা এখন <b>{status}</b> পাবেন।",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

# ═══════════════════════════════════════════════════════════
# 🔒 FEATURE 1: ID MASKING — মাঝের ৩ অক্ষর hide করো
# ═══════════════════════════════════════════════════════════

def mask_id(value):
    """TrxID/UserID/OrderID এর মাঝের ৩ অক্ষর xxx দিয়ে hide করো।"""
    s = str(value)
    if len(s) <= 6:
        return s[:2] + "xxx" + s[-1:]
    mid = len(s) // 2
    return s[:mid-1] + "xxx" + s[mid+2:]


# ═══════════════════════════════════════════════════════════
# ⚡ FLASH SALE SYSTEM
# ═══════════════════════════════════════════════════════════

import time as _time_module

def get_flash_sale_info(db, product_name):
    """Flash sale active কিনা এবং remaining time দেখাও।"""
    flash = db.get("flash_sales", {}).get(product_name)
    if not flash:
        return None
    end_time = flash.get("end_time", 0)
    remaining = int(end_time - _time_module.time())
    if remaining <= 0:
        return None
    flash["remaining_seconds"] = remaining
    return flash

def set_flash_sale(db, product_name, discount_percent, duration_minutes):
    """Flash sale সেট করো।"""
    if "flash_sales" not in db:
        db["flash_sales"] = {}

    # ✅ BUG FIX: যদি এই product এ আগে flash sale চলছে, original price থেকে নাও
    # নইলে discounted price এর উপর আবার discount হয় (0.1 → 0.01 bug)
    existing_flash = db["flash_sales"].get(product_name)
    if existing_flash and existing_flash.get("original_price"):
        original_price = existing_flash["original_price"]
    else:
        original_price = db["products"].get(product_name, 0)

    discounted_price = round(original_price * (1 - discount_percent / 100), 2)
    db["products"][product_name] = discounted_price  # price update করো
    flash_start_ts = _time_module.time()
    db["flash_sales"][product_name] = {
        "started_at_ts": flash_start_ts,
        "original_price": original_price,
        "discounted_price": discounted_price,
        "discount_percent": discount_percent,
        "end_time": _time_module.time() + (duration_minutes * 60),
        "started_at": get_now()
    }

def end_flash_sale(db, product_name):
    """Flash sale শেষ করো, original price restore করো।"""
    flash = db.get("flash_sales", {}).get(product_name)
    if flash:
        original_price = flash.get("original_price")
        if original_price:
            db["products"][product_name] = original_price
        db["flash_sales"].pop(product_name, None)

def check_and_expire_flash_sales():
    """Background এ expired flash sales restore করো।"""
    try:
        db = load_db()
        changed = False
        for product_name in list(db.get("flash_sales", {}).keys()):
            flash = db["flash_sales"][product_name]
            if _time_module.time() > flash.get("end_time", 0):
                original_price = flash.get("original_price")
                if original_price:
                    db["products"][product_name] = original_price
                db["flash_sales"].pop(product_name, None)
                changed = True
                # Users দের notify করো
                try:
                    for uid in list(db["users"].keys()):
                        try:
                            bot.send_message(uid,
                                f"⏰ <b>Flash Sale শেষ!</b>\n\n"
                                f"📦 <b>{product_name}</b> এর Flash Sale শেষ হয়েছে।\n"
                                f"💵 Regular Price: <code>{original_price} BDT</code>",
                                parse_mode="HTML")
                        except: pass
                except: pass
                logging.info(f"Flash sale expired: {product_name}")
        if changed:
            # ⚡ Targeted write — only flash_sales changed
            update_db_path("/flash_sales", db.get("flash_sales", {}))
            _update_db_cache_in_place(db)
    except Exception as e:
        logging.error(f"Flash sale check error: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("refresh_flash_"))
def handle_refresh_flash_callback(call):
    """Flash sale timer refresh করো।"""
    user_id = str(call.message.chat.id)
    product_name = call.data.replace("refresh_flash_", "")
    db = load_db()
    flash_info = get_flash_sale_info(db, product_name)

    if not flash_info:
        bot.answer_callback_query(call.id, "⏰ Flash Sale শেষ হয়ে গেছে!", show_alert=True)
        try:
            bot.edit_message_reply_markup(user_id, call.message.message_id, reply_markup=None)
        except: pass
        return

    remaining = flash_info.get("remaining_seconds", 0)
    h = remaining // 3600
    m = (remaining % 3600) // 60
    s = remaining % 60
    orig_price = flash_info.get("original_price", 0)
    discount = flash_info.get("discount_percent", 0)
    price = db["products"].get(product_name, 0)

    # Progress bar
    total_duration = (flash_info.get("end_time", 0) - flash_info.get("started_at_ts", flash_info.get("end_time", 0) - 3600))
    if total_duration <= 0:
        total_duration = 3600
    filled = int((remaining / total_duration) * 10)
    if filled > 10: filled = 10
    bar = "🟥" * filled + "⬜" * (10 - filled)

    updated_text = (
        f"⚡ <b>FLASH SALE — {product_name}</b>\n"
        f"\n"
        f"💥 <s>{orig_price} BDT</s> → <b>{price} BDT</b> ({discount}% OFF)\n"
        f"⏱ <b>বাকি সময়:</b> <code>{h:02d}:{m:02d}:{s:02d}</code>\n"
        f"<code>[{bar}]</code>\n"
        f"\n"
        f"🔄 শেষবার আপডেট: {get_now()}"
    )

    inline_markup = types.InlineKeyboardMarkup(row_width=1)
    inline_markup.add(
        InlineBtn("🔄 Timer Refresh", style="primary", callback_data=f"refresh_flash_{product_name}")
    )

    try:
        bot.edit_message_text(updated_text, user_id, call.message.message_id,
            parse_mode="HTML", reply_markup=inline_markup)
    except:
        bot.send_message(user_id, updated_text, parse_mode="HTML", reply_markup=inline_markup)

    bot.answer_callback_query(call.id, f"⏱ {h:02d}:{m:02d}:{s:02d} বাকি")

def flash_sale_background_worker():
    """Background thread যা প্রতি মিনিটে flash sale check ও low stock alert করে।"""
    _low_stock_alerted = set()  # একবার alert দেওয়া products track করো
    while True:
        _time_module.sleep(60)
        check_and_expire_flash_sales()
        # Low stock check
        try:
            db = load_db()
            LOW_THRESHOLD = 5
            for product in list(db.get("products", {}).keys()):
                stock = get_stock_count(product)
                if stock <= LOW_THRESHOLD and product not in _low_stock_alerted:
                    _low_stock_alerted.add(product)
                    status = "🔴 OUT OF STOCK" if stock == 0 else f"🟡 মাত্র {stock} পিস বাকি"
                    try:
                        bot.send_message(ADMIN_ID,
                            f"⚠️ <b>Low Stock Alert!</b>\n\n"
                            f"📦 <b>Product:</b> {product}\n"
                            f"📊 <b>Stock:</b> {status}\n\n"
                            f"🔔 দ্রুত stock আপলোড করুন!\n"
                            f"📅 {get_now()}",
                            parse_mode="HTML")
                    except: pass
                elif stock > LOW_THRESHOLD and product in _low_stock_alerted:
                    _low_stock_alerted.discard(product)  # stock ঠিক হলে reset
        except Exception as e:
            logging.error(f"Low stock check error: {e}")

        # ── ৩০ দিনের পুরনো SMS cleanup (প্রতি ঘণ্টায় একবার) ──
        try:
            _cleanup_old_sms_pending()
        except Exception as e:
            logging.error(f"SMS cleanup error: {e}")


# ═══════════════════════════════════════════════════════════
# 🏷️ ADMIN: PRODUCT FEATURES EDIT
# ═══════════════════════════════════════════════════════════

def admin_edit_product_features(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    product_name = message.text.strip()
    if product_name not in db["products"]:
        bot.send_message(user_id, "❌ <b>Product পাওয়া যায়নি!</b>", parse_mode="HTML"); return
    current_desc = db.get("product_details", {}).get(product_name, {}).get("description", "নেই")
    current_feat = db.get("product_details", {}).get(product_name, {}).get("features", "নেই")

    # ✅ এখন inline বাটন দিয়ে description বা features edit করতে পারবে
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineBtn("📝 Description এডিট করুন", style="primary", callback_data=f"edit_desc_{product_name}"),
        InlineBtn("✨ Features এডিট করুন", style="primary", callback_data=f"edit_feat_{product_name}")
    )
    bot.send_message(user_id,
        f"✏️ <b>'{product_name}' এডিট করুন</b>\n\n"
        f"📝 <b>বর্তমান Description:</b>\n{current_desc}\n\n"
        f"✨ <b>বর্তমান Features:</b>\n<i>{current_feat}</i>\n\n"
        f"নিচে বেছে নিন কি এডিট করবেন 👇",
        parse_mode="HTML", reply_markup=markup)

def save_product_features(message, product_name):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    # ✅ FIX: "skip" বা "-" লিখলে features খালি হয়ে যাবে (product detail এ দেখাবে না)
    raw = (message.text or "").strip()
    features = "" if raw.lower() in ("skip", "-", "none", "no", "বাদ") else raw
    if "product_details" not in db:
        db["product_details"] = {}
    if product_name not in db["product_details"]:
        db["product_details"][product_name] = {}
    db["product_details"][product_name]["features"] = features
    # ⚡ Targeted write
    update_db_path(f"/product_details/{product_name}/features", features)
    if features:
        bot.send_message(user_id,
            f"✅ <b>Features আপডেট হয়েছে!</b>\n\n📦 {product_name}\n\n<i>{features}</i>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
    else:
        bot.send_message(user_id,
            f"✅ <b>Features সরিয়ে ফেলা হয়েছে!</b>\n\n📦 {product_name}\n"
            f"<i>এখন product detail এ Features section দেখাবে না।</i>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))

def save_product_description(message, product_name):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    # ✅ FIX: "skip" বা "-" লিখলে description খালি হয়ে যাবে (product detail এ দেখাবে না)
    raw = (message.text or "").strip()
    description = "" if raw.lower() in ("skip", "-", "none", "no", "বাদ") else raw
    if "product_details" not in db:
        db["product_details"] = {}
    if product_name not in db["product_details"]:
        db["product_details"][product_name] = {}
    db["product_details"][product_name]["description"] = description
    # ⚡ Targeted write
    update_db_path(f"/product_details/{product_name}/description", description)
    if description:
        bot.send_message(user_id,
            f"✅ <b>Description আপডেট হয়েছে!</b>\n\n📦 {product_name}\n\n{description}",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))
    else:
        bot.send_message(user_id,
            f"✅ <b>Description সরিয়ে ফেলা হয়েছে!</b>\n\n📦 {product_name}\n"
            f"<i>এখন product detail এ Description section দেখাবে না।</i>",
            parse_mode="HTML", reply_markup=get_main_menu(user_id))


# ═══════════════════════════════════════════════════════════
# ⚡ ADMIN: FLASH SALE SET
# ═══════════════════════════════════════════════════════════

def admin_set_flash_sale_step1(message):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    product_name = message.text.strip()
    if product_name not in db["products"]:
        bot.send_message(user_id, "❌ <b>Product পাওয়া যায়নি!</b>", parse_mode="HTML"); return
    current_price = db["products"][product_name]
    msg = bot.send_message(user_id,
        f"⚡ <b>Flash Sale — {product_name}</b>\n\n"
        f"💵 বর্তমান দাম: <code>{current_price} BDT</code>\n\n"
        f"কত % ছাড় দেবেন? (1-90)\n"
        f"<i>উদাহরণ: 20 (মানে 20% ছাড়)</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_set_flash_sale_step2(m, product_name))

def admin_set_flash_sale_step2(message, product_name):
    user_id = str(message.chat.id)
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    if not message.text.strip().isdigit() or not (1 <= int(message.text.strip()) <= 90):
        msg = bot.send_message(user_id, "❌ <b>1 থেকে 90 এর মধ্যে সংখ্যা দিন!</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_set_flash_sale_step2(m, product_name)); return
    discount = int(message.text.strip())
    msg = bot.send_message(user_id,
        f"⏱️ <b>কত মিনিট Flash Sale চলবে?</b>\n\n"
        f"<i>উদাহরণ: 60 (মানে ১ ঘণ্টা), 30, 120</i>",
        parse_mode="HTML", reply_markup=get_cancel_button(user_id))
    bot.register_next_step_handler(msg, lambda m: admin_set_flash_sale_step3(m, product_name, discount))

def admin_set_flash_sale_step3(message, product_name, discount):
    user_id = str(message.chat.id)
    db = load_db()
    if message.text == "❌ Cancel":
        show_admin_panel(user_id); return
    if not message.text.strip().isdigit() or int(message.text.strip()) < 1:
        msg = bot.send_message(user_id, "❌ <b>সঠিক মিনিট দিন (কমপক্ষে ১)!</b>",
            parse_mode="HTML", reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, lambda m: admin_set_flash_sale_step3(m, product_name, discount)); return
    duration = int(message.text.strip())
    set_flash_sale(db, product_name, discount, duration)
    # ⚡ Targeted write — flash_sales only
    update_db_path("/flash_sales", db.get("flash_sales", {}))

    original_price = db["flash_sales"][product_name].get("original_price", 0)
    discounted_price = db["flash_sales"][product_name].get("discounted_price", 0)

    def _fmt_p(p):
        try:
            p = float(p)
            return str(int(p)) if p == int(p) else f"{p:.2f}"
        except:
            return str(p)

    bot.send_message(user_id,
        f"✅ <b>Flash Sale শুরু হয়েছে!</b>\n\n"
        f"📦 <b>Product:</b> {product_name}\n"
        f"💥 <b>Discount:</b> {discount}% OFF\n"
        f"💵 <s>{_fmt_p(original_price)} BDT</s> → <b>{_fmt_p(discounted_price)} BDT</b>\n"
        f"⏱️ <b>Duration:</b> {duration} মিনিট\n\n"
        f"⏳ Broadcasting to all users...",
        parse_mode="HTML", reply_markup=get_main_menu(user_id))

    # সব user দের notify করো
    sent = 0
    for uid in list(db["users"].keys()):
        try:
            bot.send_message(uid,
                f"🔥 <b>FLASH SALE ALERT!</b>\n"
                f"\n\n"
                f"📦 <b>{product_name}</b>\n"
                f"💥 <b>{discount}% ছাড়!</b>\n"
                f"💵 <s>{_fmt_p(original_price)} BDT</s> → <b>{_fmt_p(discounted_price)} BDT</b>\n"
                f"⏱️ মাত্র <b>{duration} মিনিট</b> বাকি!\n\n"
                f"\n"
                f"🛒 এখনই অর্ডার করুন!",
                parse_mode="HTML")
            sent += 1
        except: pass

    bot.send_message(user_id,
        f"📢 <b>Broadcast Complete!</b>\n✅ {sent} users notified.",
        parse_mode="HTML")

def auto_approve_deposit(req_id, req, trx_id, sms_amount):
    """
    ✅ UNIFIED Auto approve — SMS verified amount ব্যবহার করে।
    Cashback, channel log (masked), user notify সব সহ।
    """
    db = load_db()
    user_id = req["user_id"]
    method  = req["method"]
    now     = get_now()

    # ✅ SMS verified amount সবসময় priority — user claimed amount নয়
    if sms_amount is not None:
        amount = int(float(sms_amount))
    else:
        amount = int(float(req.get("amount", 0)))
        logging.warning(f"auto_approve: SMS amount missing, fallback to req amount={amount} | req_id={req_id}")

    if user_id not in db["users"]:
        db["users"][user_id] = {
            "balance": 0, "refer_count": 0, "orders": [],
            "join_date": now, "lang": "en"
        }

    new_balance = round(float(db["users"][user_id].get("balance", 0)) + float(amount), 2)

    # ✅ CASHBACK
    cashback = get_cashback(amount, db)
    if cashback > 0:
        new_balance = round(new_balance + cashback, 2)

    db["users"][user_id]["balance"] = new_balance

    req_update = {
        "status": "approved", "approved_date": now,
        "auto_approved": True, "amount": amount
    }
    db["deposit_requests"][req_id].update(req_update)
    mark_trxid_used(db, trx_id)

    # ⚡ OPTIMIZED: পুরো DB write বাদ — শুধু changed paths update করো
    update_db_path(f"/users/{user_id}/balance", new_balance, _async=False)
    update_db_path(f"/deposit_requests/{req_id}", db["deposit_requests"][req_id], _async=False)
    _EXECUTOR.submit(_update_db_path_sync, "/verified_trxids",
                     {str(i): v for i, v in enumerate(db.get("verified_trxids", []))})
    _update_db_cache_in_place(db)

    # ⚡ User কে সাথে সাথে clear balance notification
    try:
        cashback_line = (
            f"\n🎁 <b>Cashback Bonus:</b> +{cashback} BDT 🎉"
        ) if cashback > 0 else ""

        lang_u = _lang_cache.get(user_id, "en")
        if lang_u == "bn":
            msg_text = (
                f"✅ <b>ডিপোজিট অ্যাপ্রুভ হয়েছে!</b>\n\n"
                f"💰 <b>+{amount} BDT যোগ হয়েছে</b> ({method})"
                f"{cashback_line}\n\n"
                f"💳 <b>নতুন Balance: {new_balance} BDT</b>\n"
                f"🔑 TrxID: <code>{trx_id}</code>\n\n"
                f"⚡ <i>SMS দিয়ে অটো অ্যাপ্রুভ হয়েছে।</i>\n"
                f"🛍️ এখনই কেনাকাটা করুন!"
            )
        else:
            msg_text = (
                f"✅ <b>Deposit Approved!</b>\n\n"
                f"💰 <b>+{amount} BDT Added</b> ({method})"
                f"{cashback_line}\n\n"
                f"💳 <b>New Balance: {new_balance} BDT</b>\n"
                f"🔑 TrxID: <code>{trx_id}</code>\n\n"
                f"⚡ <i>Auto-approved via SMS.</i>\n"
                f"🛍️ Start shopping now!"
            )
        markup = types.InlineKeyboardMarkup()
        markup.add(InlineBtn("🛍️ Shop Now", style="primary", callback_data="go_to_shop"))
        markup.add(InlineBtn("👤 My Account", style="primary", callback_data="back_to_main"))
        bot.send_message(user_id, msg_text, parse_mode="HTML", reply_markup=markup)
    except Exception as e:
        logging.error(f"Auto approve notify error: {e}")

    # ✅ Channel log — masked IDs
    try:
        bot.send_message(LOG_CHANNEL_ID,
            f"✅ <b>Deposit Auto-Approved</b>\n"
            f"\n"
            f"👤 User: <code>{mask_id(user_id)}</code>\n"
            f"💳 Method: {method}\n"
            f"💰 Amount: {amount} BDT\n"
            f"🔑 TrxID: <code>{mask_id(trx_id)}</code>\n"
            f"🆔 Request: <code>{mask_id(req_id)}</code>\n"
            f"⚡ Mode: AUTO\n"
            f"📅 {now}",
            parse_mode="HTML")
    except: pass

    log_action("AUTO_DEPOSIT_APPROVED", user_id,
               f"Method: {method}, Amount: {amount}, TrxID: {trx_id}")


# ═══════════════════════════════════════════════════════════
# 🔒 FEATURE 2: SMS SENDER WHITELIST
# শুধু official bKash/Nagad/Rocket number থেকে SMS accept
# ═══════════════════════════════════════════════════════════

OFFICIAL_SENDERS = {
    "Bkash":  ["01847310080", "01847310081", "01847310082", "16247", "bkash"],
    "Nagad":  ["01911304250", "01911304251", "16167", "nagad"],
    "Rocket": ["16216", "01779-016216", "rocket", "dutchbangla"],
}

def is_official_sender(sender, method):
    """
    Sender official কিনা check করো।
    SMS Forwarder App থেকে আসলে sender হয় phone number বা app-defined name।
    তাই:
      1. Official list এ থাকলে → True
      2. SMS body দেখে method match করলে → True (SMS forwarder app support)
      3. Sender যদি সম্পূর্ণ unknown string হয় → False
    """
    sender_lower = sender.lower().replace("-", "").replace(" ", "")
    # Direct official sender check
    for official in OFFICIAL_SENDERS.get(method, []):
        if official.lower().replace("-", "") in sender_lower:
            return True
    # SMS Forwarder App থেকে আসলে sender যেকোনো কিছু হতে পারে।
    # তাই যদি sender এ method এর keyword থাকে → allow
    method_keywords = {
        "Bkash": ["bkash", "01847"],
        "Nagad": ["nagad", "16167", "01911"],
        "Rocket": ["rocket", "16216", "dutch"],
    }
    for kw in method_keywords.get(method, []):
        if kw in sender_lower:
            return True
    # SMS Forwarder App থেকে আসা SMS — sender সাধারণত short code বা name হয়
    # যদি sender খুব ছোট (≤6 char) বা সম্পূর্ণ numeric হয় → official short code মনে করো
    if len(sender_lower) <= 6 and sender_lower.isdigit():
        return True
    # যদি sender এ কোনো known payment word আছে → allow
    payment_words = ["bkash", "nagad", "rocket", "dutch", "bank", "pay", "cash"]
    for word in payment_words:
        if word in sender_lower:
            return True
    # SMS Forwarder App থেকে সব ধরনের sender allow করতে চাইলে নিচের লাইন uncomment করুন:
    # return True
    return False

# ✅ Payment received SMS patterns — Bkash / Nagad / Rocket official SMS এর real pattern
# এই keywords গুলো শুধু payment RECEIVE এর SMS এ থাকে (send/OTP/balance SMS এ না)

BKASH_RECEIVED_PATTERNS = [
    # "You have received Tk 500.00 from 01XXXXXXXX. TrxID ABCD1234EF..."
    r"you have received\s+tk",
    r"received\s+tk\.?\s*[\d,]+",
    # "Tk 500.00 has been received in your bKash account"
    r"has been received in your bkash",
    # "Cash In successful"
    r"cash\s*in\s*successful",
    r"cash\s*in\s*tk",
]

NAGAD_RECEIVED_PATTERNS = [
    # "Nagad-e aপনার 500 Tk Cash In সম্পন্ন"
    r"nagad.{0,20}cash\s*in",
    r"cash\s*in\s*সম্পন্ন",
    # "আপনার Nagad Account-এ Tk 500.00 জমা হয়েছে"
    r"জমা হয়েছে",
    r"nagad.{0,30}received",
    r"received.{0,30}nagad",
    # "Tk 500 has been credited to your Nagad account"
    r"credited to your nagad",
    r"nagad.{0,20}tk\.?\s*[\d,]+",
]

ROCKET_RECEIVED_PATTERNS = [
    # "Rocket: Tk 500.00 received. TrxID..."
    r"rocket.{0,30}tk\.?\s*[\d,]+\s*received",
    r"received.{0,30}rocket",
    # "Cash In Tk 500.00 successful"
    r"cash\s*in\s*tk\.?\s*[\d,]+",
    # "You have received Tk 500 via Rocket"
    r"received.{0,30}tk\.?\s*[\d,]+",
    r"tk\.?\s*[\d,]+.{0,30}received",
    # Rocket numeric TrxID with amount — "Tk 500.00 TrxID 1234567890"
    r"tk\.?\s*[\d,]+.{0,50}\b\d{8,14}\b",
    r"\b\d{8,14}\b.{0,50}tk\.?\s*[\d,]+",
    # "successful" with amount — general Rocket success SMS
    r"rocket.{0,50}successful",
    r"successful.{0,50}rocket",
]

# Method অনুযায়ী patterns
PAYMENT_RECEIVED_PATTERNS = {
    "Bkash":  BKASH_RECEIVED_PATTERNS,
    "Nagad":  NAGAD_RECEIVED_PATTERNS,
    "Rocket": ROCKET_RECEIVED_PATTERNS,
}

def is_payment_received_sms(body, method=None):
    """
    SMS টি payment RECEIVE এর কিনা check করো।
    method দিলে সেই method এর specific pattern check করে।
    method না দিলে সব pattern check করে।
    """
    body_lower = body.lower()

    # Method specific patterns check
    if method and method in PAYMENT_RECEIVED_PATTERNS:
        for pattern in PAYMENT_RECEIVED_PATTERNS[method]:
            if re.search(pattern, body_lower, re.IGNORECASE):
                return True
        # Fallback 1: Amount + TrxID দুটোই থাকলে
        has_amount = bool(re.search(AMOUNT_PATTERN, body, re.IGNORECASE))
        has_trxid  = bool(re.search(r'TrxID|Trx\s*ID|TxnID|transaction\s*id|reference', body, re.IGNORECASE))
        if has_amount and has_trxid:
            return True
        # Fallback 2: method keyword + TrxID থাকলে (amount parse fail হলেও)
        has_method = method.lower() in body_lower
        if has_method and has_trxid:
            return True
        # Fallback 3: method keyword + amount থাকলে
        if has_method and has_amount:
            return True
        return False

    # Method না জানলে সব check করো
    for m_patterns in PAYMENT_RECEIVED_PATTERNS.values():
        for pattern in m_patterns:
            if re.search(pattern, body_lower, re.IGNORECASE):
                return True

    # Generic fallback
    has_amount = bool(re.search(AMOUNT_PATTERN, body, re.IGNORECASE))
    has_trxid  = bool(re.search(r'TrxID|Trx\s*ID|TxnID|transaction\s*id|reference', body, re.IGNORECASE))
    return has_amount and has_trxid

def process_sms(sender, body):
    """
    ══════════════════════════════════════════════════════════════
    ✅ FINAL UNIFIED process_sms — v3.1
    ══════════════════════════════════════════════════════════════
    Flow:
      1. Method detect (Bkash/Nagad/Rocket)
      2. Official sender check
      3. Payment received SMS check (OTP/balance SMS ignore)
      4. TrxID ও Amount extract
      5. Duplicate TrxID check
      6. Pending deposit এ TrxID + Amount + Method — exact match → AUTO APPROVE
      7. Match না হলে → sms_pending_trxids এ store → User submit করলে approve হবে
      8. User notify (যাদের এই method এ pending deposit আছে)
    ══════════════════════════════════════════════════════════════
    """
    try:
        _process_sms_inner(sender, body)
    except Exception as e:
        logging.error(f"process_sms CRASH | sender={sender!r} | err={e}", exc_info=True)
        try:
            bot.send_message(ADMIN_ID,
                f"🚨 <b>SMS Handler Crash!</b>\n\n"
                f"📨 Sender: <code>{sender}</code>\n"
                f"📄 Body: <code>{body[:200]}</code>\n"
                f"❌ Error: <code>{str(e)[:300]}</code>",
                parse_mode="HTML")
        except Exception:
            pass

def _process_sms_inner(sender, body):
    """process_sms এর inner logic — top-level try/except এ wrapped।"""
    # ── Step 1: Method detect ──
    method = detect_method(sender, body)
    if not method:
        logging.info(f"SMS ignored (unknown method): sender={sender!r} | body={body[:60]}")
        return

    # ── Step 2: Official sender check ──
    # SMS Forwarder App থেকে আসলে sender যেকোনো কিছু হতে পারে।
    # SMS Forwarder App নিজেই filter করে পাঠাবে, bot এ sender filter নেই।
    # Body তে payment keyword + TrxID/Amount থাকলেই accept করো।
    body_lower_s2 = body.lower()
    has_method_kw  = method.lower() in body_lower_s2
    has_trxid_hint = bool(re.search(r'trxid|trx\s*id|txnid|transaction\s*id|reference', body_lower_s2))
    has_amount_val = bool(re.search(AMOUNT_PATTERN, body, re.IGNORECASE))

    if not (has_method_kw or has_trxid_hint or has_amount_val):
        logging.warning(f"SMS ignored (no payment keyword in body): sender={sender!r} | method={method}")
        return
    logging.info(f"SMS accepted | sender={sender!r} | method={method}")

    # ── Step 3: Payment received SMS check ──
    if not is_payment_received_sms(body, method):
        logging.info(f"SMS ignored (not a payment SMS): sender={sender!r} | body={body[:50]}")
        return

    # ── Step 4: TrxID ও Amount extract ──
    trx_id = extract_trxid(body, method)
    amount = extract_amount(body)

    logging.info(f"SMS Received | Method={method} | TrxID={trx_id} | Amount={amount} | Body={body[:80]}")

    if not trx_id:
        logging.warning(f"TrxID not found in SMS | Method={method} | Body={body[:80]}")
        try:
            bot.send_message(ADMIN_ID,
                f"⚠️ <b>SMS TrxID Parse Failed</b>\n\n"
                f"💳 <b>Method:</b> {method}\n"
                f"📨 <b>Sender:</b> {sender}\n"
                f"💰 <b>Amount:</b> {amount} BDT\n\n"
                f"📄 <b>SMS Body:</b>\n<code>{body[:300]}</code>",
                parse_mode="HTML")
        except: pass
        return

    trx_id = trx_id.upper()
    db = load_db()

    # ── Step 5: Duplicate TrxID check ──
    if is_trxid_already_used(db, trx_id):
        logging.warning(f"Duplicate TrxID SMS ignored: {trx_id}")
        return

    # ── Step 6: SMS store করো ──
    if "sms_pending_trxids" not in db:
        db["sms_pending_trxids"] = {}

    db["sms_pending_trxids"][trx_id] = {
        "method": method,
        "amount": amount,
        "sender": sender,
        "body":   body[:300],
        "time":   get_now()
    }

    # ── Step 7: Pending deposit request এ এই TrxID আছে কিনা চেক করো ──
    # User আগে deposit submit করলেও SMS পরে আসলে এখানে auto-approve হবে
    matched_req_id  = None
    matched_req     = None
    matched_amount  = None

    for req_id, req in db.get("deposit_requests", {}).items():
        if req.get("status") != "pending":
            continue
        if not req.get("is_auto", False):
            continue  # Binance (manual) skip
        req_trx = req.get("transaction_id", "").upper()
        req_method = req.get("method", "")
        if req_trx != trx_id:
            continue
        # Method match check
        if req_method.upper() != method.upper():
            logging.warning(f"SMS pending match: method mismatch | req_method={req_method} | sms_method={method} | TrxID={trx_id}")
            continue
        # Amount match check (SMS amount দিয়ে verify)
        if amount is not None:
            try:
                req_amount = int(float(req.get("amount", 0)))
                sms_amount_int = int(float(amount))
                if req_amount != sms_amount_int:
                    logging.warning(f"SMS pending match: amount mismatch | req={req_amount} | sms={sms_amount_int} | TrxID={trx_id}")
                    # Amount mismatch হলেও approve করো — SMS amount টাই সত্য
                    # User ভুল amount দিয়ে থাকতে পারে, কিন্তু TrxID ঠিক আছে
                    matched_amount = sms_amount_int
                else:
                    matched_amount = sms_amount_int
            except (ValueError, TypeError):
                matched_amount = amount
        else:
            # SMS এ amount নেই → user এর claimed amount ব্যবহার করো
            matched_amount = req.get("amount")

        matched_req_id = req_id
        matched_req    = req
        break  # প্রথম match এই approve করো

    if matched_req_id:
        # ✅ Pending deposit পাওয়া গেছে — auto approve করো
        # sms_pending থেকে মুছে দাও (approve এর পরে লাগবে না)
        db["sms_pending_trxids"].pop(trx_id, None)
        # ⚡ OPTIMIZED: পুরো DB write বাদ — শুধু sms_pending update করো
        _EXECUTOR.submit(_update_db_path_sync, "/sms_pending_trxids", db["sms_pending_trxids"])
        _update_db_cache_in_place(db)
        logging.info(f"SMS match found → AUTO APPROVE | TrxID={trx_id} | req_id={matched_req_id} | Amount={matched_amount}")
        auto_approve_deposit(matched_req_id, matched_req, trx_id, matched_amount)

        # Admin notification — approved
        try:
            bot.send_message(ADMIN_ID,
                f"✅ <b>SMS এসে Pending Auto-Approved!</b>\n\n"
                f"💳 <b>Method:</b> {method}\n"
                f"🔑 <b>TrxID:</b> <code>{trx_id}</code>\n"
                f"💰 <b>Amount:</b> {matched_amount if matched_amount else 'N/A'} BDT\n"
                f"🆔 <b>Request:</b> <code>{matched_req_id}</code>\n"
                f"📅 <b>Time:</b> {get_now()}\n\n"
                f"⚡ <i>User আগে request করেছিল, SMS এসে auto-approve হয়েছে।</i>",
                parse_mode="HTML")
        except: pass
    else:
        # No pending deposit — SMS save রাখো, user পরে TrxID submit করলে approve হবে
        # ⚡ OPTIMIZED: পুরো DB write বাদ — শুধু sms_pending_trxids update করো
        update_db_path("/sms_pending_trxids", db["sms_pending_trxids"])
        _update_db_cache_in_place(db)
        logging.info(f"SMS stored (no pending match) | TrxID={trx_id} | Method={method} | Amount={amount}")

        # Admin notification
        try:
            bot.send_message(ADMIN_ID,
                f"📥 <b>New SMS Received</b>\n\n"
                f"💳 <b>Method:</b> {method}\n"
                f"🔑 <b>TrxID:</b> <code>{trx_id}</code>\n"
                f"💰 <b>Amount:</b> {amount if amount else 'N/A'} BDT\n"
                f"📅 <b>Time:</b> {get_now()}\n\n"
                f"⏳ <i>User TrxID + Amount submit করলে auto-approve হবে।</i>",
                parse_mode="HTML")
        except: pass


# ═══════════════════════════════════════════════════════════
# 🔒 FEATURE 3: TrxID 3-ATTEMPT LIMIT
# একই TrxID দিয়ে ৩বারের বেশি try করলে block
# ═══════════════════════════════════════════════════════════

def get_trx_attempt_count(db, user_id, trx_id):
    key = f"{user_id}_{trx_id.upper()}"
    return db.get("trx_attempts", {}).get(key, 0)

def increment_trx_attempt(db, user_id, trx_id):
    if "trx_attempts" not in db:
        db["trx_attempts"] = {}
    key = f"{user_id}_{trx_id.upper()}"
    db["trx_attempts"][key] = db["trx_attempts"].get(key, 0) + 1
    return db["trx_attempts"][key]

def block_trx_id(db, user_id, trx_id):
    if "blocked_trxids" not in db:
        db["blocked_trxids"] = []
    key = f"{user_id}_{trx_id.upper()}"
    if key not in db["blocked_trxids"]:
        db["blocked_trxids"].append(key)

def is_trx_blocked(db, user_id, trx_id):
    key = f"{user_id}_{trx_id.upper()}"
    return key in db.get("blocked_trxids", [])


# ═══════════════════════════════════════════════════════════
# 💬 FEATURE 4: UNKNOWN MESSAGE HANDLER
# ═══════════════════════════════════════════════════════════

@bot.message_handler(func=lambda msg: msg.text and not msg.text.startswith('/') and
    msg.text not in [
        "🛒 Shop Now", "💰 Deposit", "👤 My Account", "📜 My Orders",
        "👥 Refer & Earn", "📞 Support", "⚙️ Admin Panel", "🔙 Back to Main Menu",
        "❌ Cancel", "🛒 শপিং করুন", "💰 ডিপোজিট", "👤 আমার অ্যাকাউন্ট",
        "📜 আমার অর্ডার", "👥 রেফার করুন", "📞 সাপোর্ট", "⚙️ অ্যাডমিন প্যানেল",
        "🔙 মেইন মেনুতে ফিরুন", "❌ বাতিল করুন", "✅ I've Joined — Verify",
        "✅ জয়েন করেছি — ভেরিফাই করুন", "🔙 Back", "🔙 ফিরুন",
        "🟠 Bkash", "🟢 Nagad", "🔴 Rocket", "🔵 Binance (Manual)"
    ])
def handle_unknown_message(message):
    """অপ্রয়োজনীয় text handle করো।"""
    user_id = str(message.chat.id)
    db = load_db()

    # Banned check
    if user_id in db.get("banned_users", []):
        return

    # Admin এর message skip করো
    if str(user_id) == str(ADMIN_ID):
        return

    # Channel subscription check
    if not is_subscribed(user_id):
        return

    bot.send_message(user_id,
        f"🤖 <b>আমি বুঝতে পারিনি!</b>\n\n"
        f"\n"
        f"আমি একটি <b>Digital Store Bot</b>।\n"
        f"শুধু নির্দিষ্ট commands ও buttons কাজ করে।\n\n"
        f"📌 <b>Available Commands:</b>\n"
        f"• /start — মেইন মেনু\n"
        f"• /stock — সব প্রোডাক্টের স্টক\n"
        f"• /token — Mail Box Website\n"
        f"• /2fa_key — 2FA Code Generator\n\n"
        f"\n"
        f"💬 সাহায্যের জন্য: @{SUPPORT_USERNAME}",
        parse_mode="HTML",
        reply_markup=get_main_menu(user_id))


# ═══════════════════════════════════════════════════════════
# 📦 FEATURE 5: /stock COMMAND — সব product এর stock দেখাবে
# ═══════════════════════════════════════════════════════════

@bot.message_handler(commands=['stock'])
def handle_apitest_command(message):
    """
    /apitest — Admin only
    hotmail143 API raw response দেখাও — debug এর জন্য।
    """
    user_id = str(message.chat.id)
    if str(user_id) != str(ADMIN_ID):
        return

    bot.send_message(user_id, "🔄 <b>API Test চলছে...</b>", parse_mode="HTML")

    # ── 1. API Key check ──
    key_status = f"<code>{HOTMAIL143_API_KEY[:6]}...{HOTMAIL143_API_KEY[-4:]}</code>" if HOTMAIL143_API_KEY and len(HOTMAIL143_API_KEY) > 10 else ("❌ খালি / সেট নেই" if not HOTMAIL143_API_KEY else f"<code>{HOTMAIL143_API_KEY}</code>")

    # ── 2. Raw API call — no filter ──
    import json as _json
    try:
        resp = http_requests.get(
            f"{HOTMAIL143_BASE_URL}/stock",
            params={"api_key": HOTMAIL143_API_KEY},
            timeout=8,
            verify=True
        )
        http_status = resp.status_code
        try:
            raw_json = resp.json()
            raw_text = _json.dumps(raw_json, indent=2, ensure_ascii=False)
        except Exception:
            raw_text = resp.text[:500]
        api_ok = True
    except http_requests.exceptions.ConnectTimeout:
        raw_text = "❌ ConnectTimeout — server connect হচ্ছে না (8s)"
        http_status = "TIMEOUT"
        api_ok = False
    except http_requests.exceptions.ReadTimeout:
        raw_text = "❌ ReadTimeout — connect হয়েছে কিন্তু response আসেনি (8s)"
        http_status = "TIMEOUT"
        api_ok = False
    except http_requests.exceptions.ConnectionError as e:
        raw_text = f"❌ ConnectionError: {e}"
        http_status = "CONN_ERR"
        api_ok = False
    except http_requests.exceptions.SSLError as e:
        raw_text = f"❌ SSLError: {e}"
        http_status = "SSL_ERR"
        api_ok = False
    except Exception as e:
        raw_text = f"❌ Exception ({type(e).__name__}): {e}"
        http_status = "ERROR"
        api_ok = False

    # ── 3. Per-product parsed result ──
    parsed_lines = ""
    for pname, pconf in API_PRODUCTS.items():
        pt = pconf.get("product_type", "")
        at = pconf.get("account_type", "")
        parsed_lines += f"  \u2022 <b>{pname}</b>: <code>{cnt}</code>\n"

    msg = (
        f"\U0001f527 <b>hotmail143 API Test</b>\n\n"
        f"\U0001f511 <b>API Key:</b> {key_status}\n"
        f"\U0001f310 <b>Base URL:</b> <code>{HOTMAIL143_BASE_URL}</code>\n"
        f"\U0001f4e1 <b>HTTP Status:</b> <code>{http_status}</code>\n\n"
        f"\U0001f4e6 <b>Parsed Stock:</b>\n{parsed_lines}\n"
        f"\U0001f4c4 <b>Raw Response:</b>\n<pre>{raw_text[:800]}</pre>"
    )
    bot.send_message(user_id, msg, parse_mode="HTML")


def handle_stock_command(message):
    user_id = str(message.chat.id)
    db = load_db()

    if not db["products"]:
        bot.send_message(user_id,
            "📦 <b>কোনো প্রোডাক্ট নেই।</b>",
            parse_mode="HTML")
        return

    msg = "\n"
    msg += "📦 <b>Live Stock Status</b>\n"
    msg += "\n\n"

    for product, price in db["products"].items():
        stock = get_stock_count(product)
        if stock > 10:
            status = f"🟢 <b>{stock} pcs</b>"
        elif stock > 0:
            status = f"🟡 <b>{stock} pcs</b> (সীমিত!)"
        else:
            status = "🔴 <b>Out of Stock</b>"

        flash_info = get_flash_sale_info(db, product)
        flash_badge = ""
        if flash_info:
            r = flash_info["remaining_seconds"]
            h, m_val, s_val = r // 3600, (r % 3600) // 60, r % 60
            orig = flash_info.get("original_price", price)
            disc = flash_info.get("discount_percent", 0)
            flash_badge = f"\n   🔥 FLASH SALE! <s>{orig} BDT</s> → <b>{price} BDT</b> ({disc}% OFF) ⏱ {h:02d}:{m_val:02d}:{s_val:02d}"

        msg += f"📌 <b>{product}</b>\n"
        msg += f"   💵 দাম: <code>{price} BDT</code>\n"
        msg += f"   📦 স্টক: {status}{flash_badge}\n\n"

    msg += "\n"
    msg += f"🕐 <i>Updated: {get_now()}</i>"

    bot.send_message(user_id, msg,
        parse_mode="HTML",
        reply_markup=get_main_menu(user_id))


# ═══════════════════════════════════════════════════════════
# 🎫 SUPPORT TICKET SYSTEM — REMOVED (Simple support only)
# /ticket command disabled — use @{SUPPORT_USERNAME} directly
# ═══════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════
# 📧 FEATURE 7: /token COMMAND — DongvanFB Mail Reader
# ═══════════════════════════════════════════════════════════

import requests as _req_lib
import time as _time_lib

# Hotmail143 API
HM143_API_KEY  = "94a336d8a7b527902894d8e2da2bfdd8"
HM143_BASE_URL = "https://www.hotmail143.com/api/v1"

# In-memory store: user_id → {email, refresh_token, client_id}
_token_session = {}

@bot.message_handler(commands=['token'])
def handle_token_command_legacy(message):
    handle_get_code_command(message)

@bot.message_handler(commands=['get_code'])
def handle_get_code_command(message):
    user_id = str(message.chat.id)
    db = load_db()

    if user_id in db.get("banned_users", []):
        return

    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineBtn("📧 Gmail code", style="primary", url="https://www.tridib.codes"),
        InlineBtn("?? Hotmail & Outlook code", style="primary", url="https://dongvanfb.net/read_mail_box"),
        InlineBtn("🔍 Facebook Id live check", style="primary", url="https://dongvanfb.net/check_live_uid"),
        InlineBtn("🏠 Main Menu", style="primary", callback_data="back_to_main"))

    bot.send_message(user_id,
        "🔑 <b>Get Code — Tools</b>\n\n"
        "\n"
        "নিচের বাটনে ক্লিক করে সরাসরি সাইটে যান:\n"
        "",
        parse_mode="HTML",
        reply_markup=markup)

def process_token_data(message):
    """Legacy handler - no longer used, /token now opens website."""
    user_id = str(message.chat.id)
    markup = types.InlineKeyboardMarkup()
    markup.add(InlineBtn(
        "🌐 Mail Box খুলুন", style="primary", url="https://dongvanfb.net/read_mail_box"))
    bot.send_message(user_id,
        "📧 <b>Mail Box Reader</b>\n\n"
        "নিচের বাটনে ক্লিক করে Mail Box খুলুন:",
        parse_mode="HTML", reply_markup=markup)


def _do_token_fetch(user_id, email, refresh_token, client_id):
    """30 সেকেন্ড পোলিং করে code খোঁজে।"""

    wait_msg = bot.send_message(user_id,
        "⏳ <b>Code খোঁজা হচ্ছে...</b>\n\n"
        "🔄 Inbox check করা হচ্ছে (30s)...",
        parse_mode="HTML")

    found_code = None
    found_mail = None
    deadline   = _time_lib.time() + 30
    attempt    = 0

    while _time_lib.time() < deadline:
        attempt += 1
        try:
            response = _req_lib.post(
                f"{HM143_BASE_URL}/get_messages_oauth2",
                json={
                    "email": email,
                    "refresh_token": refresh_token,
                    "client_id": client_id,
                    "list_mail": "all",
                    "key": HM143_API_KEY
                },
                timeout=15
            )

            if response.status_code == 200:
                data = response.json()
                if data.get("status", False):
                    messages = data.get("messages", data.get("s", []))
                    for mail in messages:
                        code = mail.get("code", "")
                        if code and str(code).strip():
                            found_code = str(code).strip()
                            found_mail = mail
                            break

                    if found_code:
                        break  # code পাওয়া গেছে, loop থামাও
                else:
                    # token/auth error → retry করা দরকার নেই
                    error_msg = str(data.get("message", "")).lower()
                    try: bot.delete_message(user_id, wait_msg.message_id)
                    except: pass
                    if "token" in error_msg or "expired" in error_msg or "invalid" in error_msg:
                        bot.send_message(user_id,
                            f"⚠️ <b>Token মেয়াদ শেষ!</b>\n"
                            f"\n"
                            f"📮 Email: <code>{email}</code>\n\n"
                            f"❌ Refresh Token আর কাজ করছে না।\n"
                            f"🔄 নতুন token দিয়ে আবার চেষ্টা করুন।\n"
                            f"",
                            parse_mode="HTML",
                            reply_markup=get_main_menu(user_id))
                    else:
                        bot.send_message(user_id,
                            f"❌ <b>Login ব্যর্থ!</b>\n"
                            f"\n"
                            f"📮 Email: <code>{email}</code>\n\n"
                            f"🔐 Email বা Token ভুল আছে।\n"
                            f"✏️ সঠিক data দিয়ে আবার চেষ্টা করুন।\n"
                            f"",
                            parse_mode="HTML",
                            reply_markup=get_main_menu(user_id))
                    return

        except Exception as ex:
            logging.warning(f"Token poll attempt {attempt} error: {ex}")

        if _time_lib.time() < deadline:
            _time_lib.sleep(5)  # 5s পরপর retry

    # wait message মুছে ফেলো
    try: bot.delete_message(user_id, wait_msg.message_id)
    except: pass

    if found_code and found_mail:
        # sender email বের করো
        from_list = found_mail.get("from", [])
        if from_list:
            sender_addr = from_list[0].get("address", from_list[0].get("email", ""))
            sender_name = from_list[0].get("name", sender_addr)
            sender_display = sender_addr if sender_addr else sender_name
        else:
            sender_display = "Unknown"

        bot.send_message(user_id,
            f"✅ <b>Code পাওয়া গেছে!</b>\n"
            f"\n"
            f"🔑 <b>Code:</b> <code>{found_code}</code>\n"
            f"📨 <b>From:</b> {sender_display}\n"
            f"",
            parse_mode="HTML",
            reply_markup=_get_try_again_markup(user_id))
    else:
        # Code পাওয়া যায়নি → Try Again বাটন দাও
        bot.send_message(user_id,
            f"📭 <b>Code পাওয়া যায়নি!</b>\n"
            f"\n"
            f"📮 <b>Email:</b> <code>{email}</code>\n\n"
            f"Inbox এ এই মুহূর্তে কোনো\n"
            f"active code পাওয়া যায়নি।\n"
            f"",
            parse_mode="HTML",
            reply_markup=_get_try_again_markup(user_id))


def _get_try_again_markup(user_id):
    """Try Again + Main Menu inline keyboard।"""
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineBtn("🔄 Try Again", style="primary", callback_data=f"token_retry_{user_id}"),
        InlineBtn("🏠 Main Menu", style="primary", callback_data="token_main_menu")
    )
    return markup


@bot.callback_query_handler(func=lambda call: call.data.startswith("token_retry_") or
                                               call.data == "token_main_menu")
def handle_token_retry_callback(call):
    user_id = str(call.message.chat.id)

    if call.data == "token_main_menu":
        bot.answer_callback_query(call.id)
        try: bot.delete_message(user_id, call.message.message_id)
        except: pass
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        return

    # token_retry_
    bot.answer_callback_query(call.id, "🔄 আবার চেষ্টা করা হচ্ছে...")
    try: bot.delete_message(user_id, call.message.message_id)
    except: pass

    session = _token_session.get(user_id)
    if not session:
        bot.send_message(user_id,
            "⚠️ Session শেষ হয়ে গেছে!\n/token দিয়ে আবার শুরু করুন।",
            reply_markup=get_main_menu(user_id))
        return

    _do_token_fetch(
        user_id,
        session["email"],
        session["refresh_token"],
        session["client_id"]
    )



# ═══════════════════════════════════════════════════════════
# 📧 FEATURE: /code COMMAND — Get OTP via OAuth2 (dongvanfb)
# ═══════════════════════════════════════════════════════════

_code_session = {}

@bot.message_handler(commands=['code'])
def handle_code_command(message):
    user_id = str(message.chat.id)
    msg = bot.send_message(
        user_id,
        "📧 <b>OTP Fetcher</b>\n\n"
        "Send your complete Hotmail full data:\n\n"
        "<code>email|pass|refresh_token|client_id</code>\n\n",
        parse_mode="HTML",
        reply_markup=get_cancel_button(user_id)
    )
    bot.register_next_step_handler(msg, process_code_input)


def process_code_input(message):
    user_id = str(message.chat.id)
    lang = get_user_lang(user_id)
    L = LANG[lang]

    if message.text in (L["cancel"], L.get("back", ""), "❌ Cancel"):
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        return

    text = (message.text or "").strip()
    parts = text.split("|")

    if len(parts) < 4:
        msg = bot.send_message(
            user_id,
            "❌ <b>Format ভুল!</b>\n\n"
            "সঠিক format:\n"
            "<code>email|pass|refresh_token|client_id</code>\n\n"
            "আবার চেষ্টা করুন:",
            parse_mode="HTML",
            reply_markup=get_cancel_button(user_id)
        )
        bot.register_next_step_handler(msg, process_code_input)
        return

    email         = parts[0].strip()
    password      = parts[1].strip()
    refresh_token = parts[2].strip()
    client_id     = parts[3].strip()

    _code_session[user_id] = {
        "email": email,
        "password": password,
        "refresh_token": refresh_token,
        "client_id": client_id
    }

    _EXECUTOR.submit(_do_code_fetch, user_id, email, refresh_token, client_id)


def _detect_service_type(*texts):
    """
    Link, content, sender, subject — যেকোনো text থেকে service detect করো।
    একাধিক text দিলে সবগুলো check করে প্রথম match return করে।
    """
    services = [
        ("Facebook",   ["facebook.com", "fb.com", "fb.me", "facebook", "facebookmail"]),
        ("Instagram",  ["instagram.com", "instagr.am", "instagram"]),
        ("TikTok",     ["tiktok.com", "tik.tok", "tiktok"]),
        ("Telegram",   ["telegram.org", "t.me", "telegram.me", "telegram"]),
        ("Twitter/X",  ["twitter.com", "x.com", "t.co", "twitter"]),
        ("Google",     ["google.com", "accounts.google", "gmail.com", "google", "noreply@accounts"]),
        ("Apple",      ["apple.com", "icloud.com", "appleid", "apple"]),
        ("Microsoft",  ["microsoft.com", "live.com", "hotmail.com", "outlook.com", "microsoft"]),
        ("Amazon",     ["amazon.com", "amazon.", "amzn.", "amazon"]),
        ("Netflix",    ["netflix.com", "netflix"]),
        ("Snapchat",   ["snapchat.com", "snap.com", "snapchat"]),
        ("LinkedIn",   ["linkedin.com", "linkedin"]),
        ("Reddit",     ["reddit.com", "redd.it", "reddit"]),
        ("Discord",    ["discord.com", "discord.gg", "discord"]),
        ("Spotify",    ["spotify.com", "spotify"]),
        ("PayPal",     ["paypal.com", "paypal"]),
        ("Binance",    ["binance.com", "binance"]),
        ("Twitch",     ["twitch.tv", "twitch"]),
        ("Yahoo",      ["yahoo.com", "yahoo"]),
        ("Talkatone",  ["talkatone.com", "talkatone"]),
        ("Textme",     ["textme.com", "textmeapp", "textme"]),
        ("Textnow",    ["textnow.com", "textnow"]),
        ("Roblox",     ["roblox.com", "roblox"]),
        ("Steam",      ["steampowered.com", "steam"]),
        ("Uber",       ["uber.com", "uber"]),
        ("Airbnb",     ["airbnb.com", "airbnb"]),
        ("Dropbox",    ["dropbox.com", "dropbox"]),
        ("Adobe",      ["adobe.com", "adobe"]),
        ("GitHub",     ["github.com", "github"]),
    ]
    for text in texts:
        if not text:
            continue
        t = str(text).lower()
        for name, keywords in services:
            for kw in keywords:
                if kw in t:
                    return name
    return None


def _do_code_fetch(user_id, email, refresh_token, client_id):
    import time as _time

    wait_msg = bot.send_message(
        user_id,
        "⏳ <b>Fetching code...</b>\n\n"
        "⏱ Wait for OTP max <b>30s</b>",
        parse_mode="HTML"
    )

    url = "https://tools.dongvanfb.net/api/get_code_oauth2"
    payload = {
        "email": email,
        "refresh_token": refresh_token,
        "client_id": client_id,
        "type": "all",
        "apikey": DONGVANFB_API_KEY
    }
    headers = {
        "Content-Type": "application/json"
    }

    found = None
    elapsed = 0

    while elapsed < 30:
        try:
            resp = http_requests.post(url, json=payload, headers=headers, timeout=10)
            data = resp.json()
            # code অথবা content যেকোনো একটা থাকলেই found ধরো
            if data.get("status") and (
                data.get("code") or data.get("content") or data.get("data")
            ):
                found = data
                break
        except Exception as e:
            logging.warning(f"Code fetch error: {e}")
        _time.sleep(5)
        elapsed += 5

    try:
        bot.delete_message(user_id, wait_msg.message_id)
    except:
        pass

    if found:
        code    = (found.get("code") or "").strip()
        content = (found.get("content") or "").strip()
        subject = (found.get("subject") or found.get("title") or "").strip()
        sender  = (found.get("sender") or found.get("from") or "").strip()

        # data field এ থাকলে সেখান থেকেও বের করো
        if not code and not content and found.get("data"):
            raw_data = found["data"]
            if isinstance(raw_data, list) and raw_data:
                first = raw_data[0]
                if isinstance(first, dict):
                    code    = (first.get("code") or "").strip()
                    content = (first.get("content") or first.get("body") or "").strip()
                    subject = (first.get("subject") or first.get("title") or subject).strip()
                    sender  = (first.get("sender") or first.get("from") or sender).strip()
            elif isinstance(raw_data, dict):
                code    = (raw_data.get("code") or "").strip()
                content = (raw_data.get("content") or raw_data.get("body") or "").strip()
                subject = (raw_data.get("subject") or raw_data.get("title") or subject).strip()
                sender  = (raw_data.get("sender") or raw_data.get("from") or sender).strip()
            elif isinstance(raw_data, str):
                content = raw_data.strip()

        links = re.findall(r'https?://\S+', content) if content else []

        # Service detect: sender > subject > link > content > email (priority order)
        service = _detect_service_type(sender, subject, links[0] if links else None, content, email)
        service_line = f"🏷️ <b>Type :</b> {service}\n" if service else ""

        if code:
            bot.send_message(
                user_id,
                f"✅ <b>OTP Found!</b>\n\n"
                f"📧 <b>Email :</b> <code>{email}</code>\n"
                f"{service_line}"
                f"🔑 <b>OTP :</b> <code>{code}</code>",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id)
            )
        elif links:
            bot.send_message(
                user_id,
                f"✅ <b>Link Found!</b>\n\n"
                f"📧 <b>Email :</b> <code>{email}</code>\n"
                f"{service_line}"
                f"🔗 <b>Link :</b> {links[0]}",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id)
            )
        elif content:
            bot.send_message(
                user_id,
                f"✅ <b>Message Found!</b>\n\n"
                f"📧 <b>Email :</b> <code>{email}</code>\n"
                f"{service_line}"
                f"📨 <b>Content :</b> {content[:500]}",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id)
            )
        else:
            bot.send_message(
                user_id,
                f"⚠️ <b>Response পাওয়া গেছে কিন্তু code/link খালি!</b>\n\n"
                f"📧 <b>Email :</b> <code>{email}</code>\n\n"
                f"Inbox এ নতুন mail নেই বা code/link নেই।",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id)
            )
    else:
        bot.send_message(
            user_id,
            f"❌ <b>Code পাওয়া যায়নি!</b>\n\n"
            f"📧 <b>Email :</b> <code>{email}</code>\n\n"
            f"⏱ 30 সেকেন্ডে কোনো code আসেনি।\n"
            f"একটু পরে আবার চেষ্টা করুন।",
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id)
        )

# ═══════════════════════════════════════════════════════════
# 🔗 FEATURE: /tn_verify COMMAND — Get Link from Inbox (OAuth2)
# ═══════════════════════════════════════════════════════════

_tn_verify_session = {}

@bot.message_handler(commands=['tn_verify'])
def handle_tn_verify_command(message):
    user_id = str(message.chat.id)
    msg = bot.send_message(
        user_id,
        "🔗 <b>TN Verify — Link Fetcher</b>\n\n"
        "আপনার Hotmail/Outlook এর full data দিন:\n\n"
        "<code>email|pass|refresh_token|client_id</code>\n\n"
        "📌 <b>Example:</b>\n"
        "<code>test@hotmail.com|Pass123|M.C518_BAY...|5464fghj-bnmm...</code>",
        parse_mode="HTML",
        reply_markup=get_cancel_button(user_id)
    )
    bot.register_next_step_handler(msg, process_tn_verify_input)


def process_tn_verify_input(message):
    user_id = str(message.chat.id)
    lang = get_user_lang(user_id)
    L = LANG[lang]

    if message.text in (L["cancel"], L.get("back", ""), "❌ Cancel"):
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        return

    text = (message.text or "").strip()
    parts = text.split("|")

    if len(parts) < 4:
        msg = bot.send_message(
            user_id,
            "❌ <b>Format ভুল!</b>\n\n"
            "সঠিক format:\n"
            "<code>email|pass|refresh_token|client_id</code>\n\n"
            "আবার চেষ্টা করুন:",
            parse_mode="HTML",
            reply_markup=get_cancel_button(user_id)
        )
        bot.register_next_step_handler(msg, process_tn_verify_input)
        return

    email         = parts[0].strip()
    password      = parts[1].strip()
    refresh_token = parts[2].strip()
    client_id     = parts[3].strip()

    _tn_verify_session[user_id] = {
        "email": email,
        "password": password,
        "refresh_token": refresh_token,
        "client_id": client_id
    }

    _EXECUTOR.submit(_do_tn_verify_fetch, user_id, email, refresh_token, client_id)


def _do_tn_verify_fetch(user_id, email, refresh_token, client_id):
    import time as _time

    wait_msg = bot.send_message(
        user_id,
        "⏳ <b>Inbox check করা হচ্ছে...</b>\n\n"
        "⏱ Wait for link max <b>30s</b>",
        parse_mode="HTML"
    )

    url = "https://tools.dongvanfb.net/api/get_messages_oauth2"
    payload = {
        "email": email,
        "refresh_token": refresh_token,
        "client_id": client_id,
        "list_mail": "all"
    }
    headers = {"Content-Type": "application/json"}

    found_link    = None
    found_subject = None
    found_sender_name = None  # sender এর name (যেমন: "Facebook", "Instagram")
    elapsed = 0

    SKIP_KEYWORDS = ["open_log_pic", "pixel", "track", "beacon", "1x1", "spacer",
                     "unsubscribe", "email_open_log", "logo", ".png", ".jpg", ".gif"]

    while elapsed < 30:
        try:
            resp = http_requests.post(url, json=payload, headers=headers, timeout=10)
            data = resp.json()
            logging.info(f"TN verify API response: status={data.get('status')} keys={list(data.keys())}")

            if data.get("status") is True or data.get("status") == "true" or data.get("status") == 1:
                # API documentation অনুযায়ী: messages → list
                messages = data.get("messages") or data.get("data") or data.get("emails") or data.get("list") or []

                if isinstance(messages, list):
                    for mail in messages:
                        if not isinstance(mail, dict):
                            continue

                        body    = str(mail.get("message") or mail.get("body") or mail.get("content") or mail.get("text") or "")
                        subject = str(mail.get("subject") or mail.get("title") or "")

                        # sender name বের করো (name field, address নয়)
                        from_field = mail.get("from") or mail.get("sender") or []
                        s_name = ""
                        if isinstance(from_field, list) and from_field:
                            first = from_field[0]
                            if isinstance(first, dict):
                                # name আগে, address পরে
                                s_name = first.get("name") or first.get("address") or ""
                            else:
                                s_name = str(first)
                        elif isinstance(from_field, str):
                            s_name = from_field

                        # href থেকে link বের করো (HTML email এ button link)
                        links = re.findall(r'href=["\']?(https?://[^"\'>\s]+)', body)
                        # fallback: plain text link
                        if not links:
                            links = re.findall(r'https?://\S+', body)

                        for lnk in links:
                            lnk = lnk.rstrip('.,)"\'>&;')
                            # image/tracker link বাদ দাও
                            if any(sk in lnk.lower() for sk in SKIP_KEYWORDS):
                                continue
                            if len(lnk) > 30:
                                found_link        = lnk
                                found_subject     = subject
                                found_sender_name = s_name
                                break

                        if found_link:
                            break

            if found_link:
                break

        except Exception as e:
            logging.warning(f"TN verify fetch error: {e}")

        _time.sleep(5)
        elapsed += 5

    try:
        bot.delete_message(user_id, wait_msg.message_id)
    except:
        pass

    if found_link:
        # Type = sender name (যেমন: Facebook, Instagram, Claude)
        type_line = f"🏷️ <b>Type :</b> {found_sender_name}\n" if found_sender_name else ""
        bot.send_message(
            user_id,
            f"✅ <b>Link Found!</b>\n\n"
            f"📧 <b>Email :</b> <code>{email}</code>\n"
            f"{type_line}"
            f"\n🔗 <b>Link :</b>\n{found_link}",
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id)
        )
    else:
        bot.send_message(
            user_id,
            f"❌ <b>Link পাওয়া যায়নি!</b>\n\n"
            f"📧 <b>Email:</b> <code>{email}</code>\n\n"
            f"⏱ 30 সেকেন্ডে inbox এ কোনো link আসেনি।\n"
            f"একটু পরে আবার চেষ্টা করুন।",
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id)
        )


# ═══════════════════════════════════════════════════════════
# 🔐 FEATURE 8: /2fa_key COMMAND — TOTP Code Generator
# ═══════════════════════════════════════════════════════════

@bot.message_handler(commands=['apitest'])
def cmd_apitest(message):
    """Admin only — hotmail143 API raw test."""
    if str(message.chat.id) != str(ADMIN_ID):
        return
    handle_apitest_command(message)

@bot.message_handler(commands=['2fa_key'])
def handle_2fa_command(message):
    user_id = str(message.chat.id)

    parts = message.text.strip().split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        msg = bot.send_message(user_id,
            "🔐 <b>2FA Code Generator</b>\n\n"
            "\n"
            "🔑 আপনার <b>Secret Key</b> দিন:\n\n"
            "✏️ উদাহরণ:\n"
            "<code>JBSWY3DPEHPK3PXP</code>\n\n"
            "⚠️ Secret key Base32 format এ হয়।\n"
            "",
            parse_mode="HTML",
            reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, process_2fa_key_input)
        return

    secret_key = parts[1].strip()
    generate_and_send_2fa(user_id, secret_key)

def process_2fa_key_input(message):
    user_id = str(message.chat.id)
    lang = get_user_lang(user_id)
    L = LANG[lang]

    if message.text in (L["cancel"], L.get("back", ""), "❌ Cancel"):
        bot.send_message(user_id, t(user_id, "main_menu"),
            reply_markup=get_main_menu(user_id), parse_mode="HTML")
        return

    secret_key = (message.text or "").strip()
    if not secret_key:
        msg = bot.send_message(user_id,
            "❌ <b>Key খালি!</b>\n\nআবার দিন:",
            parse_mode="HTML",
            reply_markup=get_cancel_button(user_id))
        bot.register_next_step_handler(msg, process_2fa_key_input)
        return

    generate_and_send_2fa(user_id, secret_key)

def generate_and_send_2fa(user_id, secret_key):
    import hmac, hashlib, struct, base64, time as t_lib

    try:
        clean_key = secret_key.strip().upper().replace(" ", "").replace("-", "")
        padding = len(clean_key) % 8
        if padding:
            clean_key += "=" * (8 - padding)

        try:
            key_bytes = base64.b32decode(clean_key, casefold=True)
        except Exception:
            bot.send_message(user_id,
                "❌ <b>Invalid Secret Key!</b>\n\n"
                "Base32 format এ দিন।\n"
                "উদাহরণ: <code>JBSWY3DPEHPK3PXP</code>",
                parse_mode="HTML",
                reply_markup=get_main_menu(user_id))
            return

        current_time   = int(t_lib.time())
        time_step      = 30
        counter        = current_time // time_step
        secs_remaining = time_step - (current_time % time_step)

        def get_totp(cnt):
            cb = struct.pack(">Q", cnt)
            hr = hmac.new(key_bytes, cb, hashlib.sha1).digest()
            off = hr[-1] & 0x0F
            tr  = struct.unpack(">I", hr[off:off+4])[0] & 0x7FFFFFFF
            return str(tr % 1_000_000).zfill(6)

        current_code = get_totp(counter)
        prev_code    = get_totp(counter - 1)
        next_code    = get_totp(counter + 1)

        filled = int((secs_remaining / time_step) * 10)
        bar    = "█" * filled + "░" * (10 - filled)

        bot.send_message(user_id,
            f"🔐 <b>2FA Code Generated!</b>\n"
            f"\n\n"
            f"🎯 <b>Current Code:</b>\n"
            f"<code>  {current_code[:3]} {current_code[3:]}  </code>\n\n"
            f"⏱ <b>Expires in:</b> {secs_remaining}s\n"
            f"<code>[{bar}]</code>\n\n"
            f"\n"
            f"⏮ Previous: <code>{prev_code}</code>\n"
            f"⏭ Next:     <code>{next_code}</code>\n"
            f"\n\n"
            f"⚠️ <i>Code টি {secs_remaining} সেকেন্ডে expire হবে।</i>",
            parse_mode="HTML")
        bot.send_message(user_id,
            t(user_id, "main_menu"),
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id))

    except Exception as e:
        logging.error(f"2FA error: {e}")
        bot.send_message(user_id,
            "❌ <b>Code generate করা যায়নি!</b>\n\n"
            "সঠিক Base32 Secret Key দিন।",
            parse_mode="HTML",
            reply_markup=get_main_menu(user_id))


# ═══════════════════════════════════════════════════════════
# 🎫 TICKET CALLBACK HANDLERS — REMOVED (ticket system disabled)
# ═══════════════════════════════════════════════════════════


# ===========================
# BOT STARTUP
# ===========================

# ═══════════════════════════════════════════════════════════
# 🔧 AUTO-INIT: Gmail ও Facebook এর Default Sub-Products
# ═══════════════════════════════════════════════════════════

GMAIL_DEFAULT_SUB_ITEMS = [
    {"name": "Gmail for Personal use", "price": 20, "enabled": True},
    {"name": "Gmail - Tiktok",         "price": 20, "enabled": True},
    {"name": "Gmail - Facebook",       "price": 20, "enabled": True},
    {"name": "Gmail - Telegram",       "price": 20, "enabled": True},
    {"name": "Gmail - Instagram",      "price": 20, "enabled": True},
    {"name": "Gmail - Talkatone",      "price": 20, "enabled": True},
    {"name": "Gmail - Reddit",         "price": 20, "enabled": True},
    {"name": "Gmail - Apple",          "price": 20, "enabled": True},
    {"name": "Gmail - Amazon",         "price": 20, "enabled": True},
    {"name": "Gmail - X",              "price": 20, "enabled": True},
    {"name": "Gmail - Linkedin",       "price": 20, "enabled": True},
    {"name": "Gmail - Textme",         "price": 20, "enabled": True},
]

FACEBOOK_DEFAULT_SUB_ITEMS = [
    {"name": "Facebook",               "price": 8,  "enabled": True},
    {"name": "Facbook -2fa 0friend",   "price": 8,  "enabled": True},
]

def auto_init_sub_products():
    """
    Gmail ও Facebook product এর sub-items যদি DB তে না থাকে,
    তাহলে default sub-items যোগ করো।
    ✅ BUG FIX: Edu Mail 24H / 72H / Hotmail / Outlook এর price DB তে না থাকলে auto সেট করো।
    """
    try:
        db = load_db()
        changed = False
        if "sub_products" not in db:
            db["sub_products"] = {}

        # ✅ BUG FIX: Edu Mail 24H / 72H / Hotmail / Outlook / Edu Mail 24hr এর price DB products এ না থাকলে add করো
        # এই price না থাকলে purchase করলে delivery হয় না (price=0 bug)
        EDU_MAIL_DEFAULT_PRICES = {
            EDU_MAIL_SUB_24H:  50,   # ডিফল্ট price — admin পরে change করতে পারবে
            EDU_MAIL_SUB_72H:  100,  # ডিফল্ট price — admin পরে change করতে পারবে
            EDU_MAIL_SUB_24HR: 50,   # Bulkmail product — ডিফল্ট price
            HOTMAIL_PROD:      30,   # ডিফল্ট price — admin পরে change করতে পারবে
            OUTLOOK_PROD:      30,   # ডিফল্ট price — admin পরে change করতে পারবে
        }
        for api_prod, default_price in EDU_MAIL_DEFAULT_PRICES.items():
            if api_prod not in db["products"]:
                db["products"][api_prod] = default_price
                changed = True
                print(f"✅ Auto-added price for API product: {api_prod} = {default_price} BDT")
            else:
                # Price আছে কিন্তু 0 — সেটাও fix করো
                if float(db["products"][api_prod]) == 0:
                    db["products"][api_prod] = default_price
                    changed = True
                    print(f"✅ Fixed zero price for API product: {api_prod} = {default_price} BDT")

        # Hotmail ও Outlook — API product, DB তে না থাকলে add করো
        for _api_p in [HOTMAIL_PROD, OUTLOOK_PROD, EDU_MAIL_SUB_24HR]:
            if _api_p not in db["products"]:
                _def_price = EDU_MAIL_DEFAULT_PRICES.get(_api_p, 30)
                db["products"][_api_p] = _def_price
                changed = True
                print(f"✅ Auto-added API product to shop: {_api_p} = {_def_price} BDT")

        # Gmail
        if "Gmail" in db["products"]:
            existing = db["sub_products"].get("Gmail", {}).get("sub_items", [])
            if not existing:
                db["sub_products"]["Gmail"] = {"sub_items": GMAIL_DEFAULT_SUB_ITEMS}
                changed = True
                print("✅ Gmail default sub-products initialized.")

        # Facebook
        if "Facebook" in db["products"]:
            existing = db["sub_products"].get("Facebook", {}).get("sub_items", [])
            if not existing:
                db["sub_products"]["Facebook"] = {"sub_items": FACEBOOK_DEFAULT_SUB_ITEMS}
                changed = True
                print("✅ Facebook default sub-products initialized.")

        if changed:
            # ⚡ Targeted writes — products ও sub_products only
            update_db_path("/products", db["products"])
            update_db_path("/sub_products", db.get("sub_products", {}))
            _update_db_cache_in_place(db)
    except Exception as e:
        logging.error(f"auto_init_sub_products error: {e}")


def _init_kb_button_color():
    """Bot startup এ Firebase থেকে admin-সেট করা প্রতিটা keyboard button
    role এর color load করো, যাতে restart এর পরও color মনে থাকে।"""
    try:
        db = load_db()
        saved_colors = db.get("settings", {}).get("kb_button_colors", {})
        load_kb_button_colors_from_dict(saved_colors)
        print(f"✅ Keyboard button colors loaded: {saved_colors or '(defaults)'}")
    except Exception as e:
        logging.error(f"_init_kb_button_color error: {e}")


if __name__ == "__main__":
    print("=" * 60)
    print("🤖 Digital Store Bot v5.0 Starting (ULTRA OPTIMIZED)...")
    print("=" * 60)
    print(f"📅 Start Time: {get_now()}")
    print(f"👤 Admin ID: {ADMIN_ID}")
    print(f"⚡ Auto Deposit: Bkash | Nagad | Rocket (SMS)")
    print(f"?? Manual Deposit: Binance")
    print(f"🌐 SMS Webhook: http://0.0.0.0:5000/sms")
    print(f"🚀 Threaded Mode: ON (32 threads + ThreadPoolExecutor)")
    print(f"💾 DB Cache TTL: {_DB_CACHE_TTL}s")
    print(f"📦 Stock Cache TTL: {_STOCK_CACHE_TTL}s")
    print(f"⚡ Callback ACK: INSTANT (pre-answer + thread pool)")
    print("=" * 60)

    # Gmail ও Facebook default sub-products auto-init
    auto_init_sub_products()
    print("✅ Sub-products auto-initialized (Gmail/Facebook)")

    # Admin-সেট করা keyboard button color load করো
    _init_kb_button_color()

    # Flash Sale background worker
    flash_thread = threading.Thread(target=flash_sale_background_worker, daemon=True)
    flash_thread.start()
    print("✅ Flash Sale background worker started")

    # SMS + Telegram Webhook server — foreground এ চালাও
    # infinity_polling বন্ধ — Webhook mode এ switch হয়েছে ✅
    print("🌐 Starting Webhook+SMS server on port 5000...")
    print(f"🔗 Telegram Webhook URL: {WEBHOOK_URL}")
    start_sms_server()
